import sys

with open('app.py', 'r', encoding='utf-8') as f:
    text = f.read()

# Fix Detalle ARBA/AGIP Formulas (Inconsistent Formula Warnings)
text = text.replace("""                                    if is_sub:
                                        ws2.cell(row=row_i, column=idx_match_arba).value = (
                                            f'=IFERROR(VLOOKUP(A{row_i}, \\'Cruce x CUIT\\'!$A$6:$F${len(cuits_sorted)+5}, 6, 0), "")'
                                        )
                                        ws2.cell(row=row_i, column=idx_cant_arba).value = (
                                            f'=COUNTIFS($A$6:$A${len(df_arba_det)+5}, A{row_i}, $B$6:$B${len(df_arba_det)+5}, "<>*(SUBTOTAL)*")'
                                        )
                                    else:
                                        ws2.cell(row=row_i, column=idx_match_arba).value = (
                                            f'=IF(COUNTIFS(\\'Detalle Mendez\\'!$A:$A, A{row_i}, '
                                            f'\\'Detalle Mendez\\'!$G:$G, G{row_i})>0, "✓ Ok", "⚠ Falta en Mendez")'
                                        )""", """                                    ws2.cell(row=row_i, column=idx_match_arba).value = (
                                        f'=IF(COUNTIF(B{row_i}, "*(SUBTOTAL)*")>0, '
                                        f'IFERROR(VLOOKUP(A{row_i}, \\'Cruce x CUIT\\'!$A$6:$F${len(cuits_sorted)+5}, 6, 0), ""), '
                                        f'IF(COUNTIFS(\\'Detalle Mendez\\'!$A:$A, A{row_i}, \\'Detalle Mendez\\'!$G:$G, G{row_i})>0, "✓ Ok", "⚠ Falta en Mendez"))'
                                    )
                                    ws2.cell(row=row_i, column=idx_cant_arba).value = (
                                        f'=IF(COUNTIF(B{row_i}, "*(SUBTOTAL)*")>0, '
                                        f'COUNTIFS($A$6:$A${len(df_arba_det)+5}, A{row_i}, $B$6:$B${len(df_arba_det)+5}, "<>*(SUBTOTAL)*"), '
                                        f'"")'
                                    )""")

text = text.replace("""                                    if is_sub:
                                        ws3.cell(row=row_i, column=idx_match_men).value = (
                                            f'=IFERROR(VLOOKUP(A{row_i}, \\'Cruce x CUIT\\'!$A$6:$F${len(cuits_sorted)+5}, 6, 0), "")'
                                        )
                                        ws3.cell(row=row_i, column=idx_cant_men).value = (
                                            f'=COUNTIFS($A$6:$A${len(df_mendez_det)+5}, A{row_i}, $B$6:$B${len(df_mendez_det)+5}, "<>*(SUBTOTAL)*")'
                                        )
                                    else:
                                        ws3.cell(row=row_i, column=idx_match_men).value = (
                                            f'=IF(COUNTIFS(\\'Detalle {organismo}\\'!$A:$A, A{row_i}, '
                                            f'\\'Detalle {organismo}\\'!$G:$G, G{row_i})>0, "✓ Ok", "⚠ Falta en {organismo}")'
                                        )""", """                                    ws3.cell(row=row_i, column=idx_match_men).value = (
                                        f'=IF(COUNTIF(B{row_i}, "*(SUBTOTAL)*")>0, '
                                        f'IFERROR(VLOOKUP(A{row_i}, \\'Cruce x CUIT\\'!$A$6:$F${len(cuits_sorted)+5}, 6, 0), ""), '
                                        f'IF(COUNTIFS(\\'Detalle {organismo}\\'!$A:$A, A{row_i}, \\'Detalle {organismo}\\'!$G:$G, G{row_i})>0, "✓ Ok", "⚠ Falta en {organismo}"))'
                                    )
                                    ws3.cell(row=row_i, column=idx_cant_men).value = (
                                        f'=IF(COUNTIF(B{row_i}, "*(SUBTOTAL)*")>0, '
                                        f'COUNTIFS($A$6:$A${len(df_mendez_det)+5}, A{row_i}, $B$6:$B${len(df_mendez_det)+5}, "<>*(SUBTOTAL)*"), '
                                        f'"")'
                                    )""")

# Add conditional formatting imports 
if "from openpyxl.formatting.rule import CellIsRule" not in text:
    text = text.replace("from openpyxl.styles import Font, PatternFill, Alignment, Border, Side", 
                        "from openpyxl.styles import Font, PatternFill, Alignment, Border, Side\n                            from openpyxl.formatting.rule import CellIsRule, FormulaRule")

# Insert CF logic
cf_inject = """                                _autofit_ws(ws1, n1)
                                
                                # Añadir formato condicional amarillo a diferencias
                                FILL_YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                                rule_diff_cruce = FormulaRule(formula=['$F6<>"✓ OK"'], stopIfTrue=False, fill=FILL_YELLOW)
                                ws1.conditional_formatting.add(f'A6:F{len(cuits_sorted)+5}', rule_diff_cruce)
"""
text = text.replace("                                _autofit_ws(ws1, n1)", cf_inject)

# Also apply YELLOW to matrix diferencias
if "FILL_YELLOW = PatternFill(" not in text:
    pass

cf_inject_ws2 = """                                _autofit_ws(ws2, n2)
                                
                                rule_diff_arr = FormulaRule(formula=[f'LEFT(${openpyxl.utils.get_column_letter(idx_match_arba)}6, 1)="⚠"'], stopIfTrue=False, fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'))
                                ws2.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n2)}{len(df_arba_det)+5}', rule_diff_arr)
"""
text = text.replace("                                _autofit_ws(ws2, n2)", cf_inject_ws2)

cf_inject_ws3 = """                                _autofit_ws(ws3, n3)
                                
                                rule_diff_men = FormulaRule(formula=[f'LEFT(${openpyxl.utils.get_column_letter(idx_match_men)}6, 1)="⚠"'], stopIfTrue=False, fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'))
                                ws3.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n3)}{len(df_mendez_det)+5}', rule_diff_men)
"""
text = text.replace("                                _autofit_ws(ws3, n3)", cf_inject_ws3)

with open('app.py', 'w', encoding='utf-8') as f:
    f.write(text)
print('Done!')
