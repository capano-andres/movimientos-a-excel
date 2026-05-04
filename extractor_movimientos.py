import re
import sys
import io
import pandas as pd

# Eliminar el wrapping global de sys.stdout que causa error en Streamlit
# (Se movió al bloque main())
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONCEPTOS_MAP = {
    "1": "Mercaderia c/iva", "2": "mercaderia s/iva", "3": "perecederos", "4": "carnes",
    "5": "verduras", "6": "huevos", "7": "pollos", "8": "no perecederos",
    "9": "materia prima c/iva", "10": "materia prima s/iva", "11": "materiales c/iva",
    "12": "materiales s/iva", "13": "productos varios", "14": "alimentos balanceados",
    "15": "bienes de cambio", "16": "Combustible para la venta", "18": "gs de impo/expo",
    "19": "gastos de prestamo", "20": "gastos generales c/iva", "21": "gastos generales s/iva",
    "22": "gastos bancarios c/iva", "23": "gastos bancarios s/iva", "24": "gastos adm. C/iva",
    "25": "gastos adm. S/iva", "26": "gs.comercializacion c/iva", "27": "gs.comercializacion S/iva",
    "28": "servicios varios", "29": "imp. Tasas y contribuciones", "30": "serv x cta de 3° c/iva",
    "31": "serv x cta de 3° s/iva", "32": "gastos despachantes", "33": "honorarios c/ivA",
    "34": "honorarios s/iva", "35": "derechos de importacion", "36": "prestamos",
    "37": "leasing", "38": "intereses", "39": "gastos de tarjeta", "40": "insumos",
    "41": "material de embalaje", "42": "seguros comerciales", "43": "seguro de vida",
    "44": "seguro de vehiculo", "45": "Gastos de vehiculo c/iva", "46": "Gastos de vehiculo S/iva",
    "47": "combustible", "48": "fletes c/iva", "49": "fletes s/iva", "50": "alquiler con iva",
    "51": "alquiler sin iva", "52": "gsts ch/rechazados", "53": "comisiones pagadas",
    "54": "mant y rep bs. De uso", "55": "mant y rep edificio", "56": "alquiler maquinarias",
    "57": "descuentos otorgados", "58": "indumentaria", "59": "anticipo de materiales",
    "60": "hipoteca", "61": "comitentes", "62": "inmobiliario", "63": "descuentos obtenidos",
    "64": "rodados", "65": "instalaciones", "66": "maquinarias", "67": "sistemas informaticos",
    "68": "compra de mue y utiles c/iva", "69": "compra de mue y utiles s/iva",
    "70": "compra de bs uso c/iva", "71": "compra de bs de uso s/iva", "72": "mejoras",
    "73": "vacunos", "74": "equinos", "75": "conejos", "76": "gallinas ponedoras",
    "77": "mejoras inmuebles ajenos", "78": "moldes y matrices", "79": "restitucion de gastos",
    "80": "venta de mercaderia c/iva", "81": "venta de mercaderia s/iva", "82": "venta cons final",
    "83": "venta resumen del dia", "84": "venta bs de uso c/iva", "85": "venta bs de uso s/iva",
    "86": "venta de combustible", "87": "prestacion de servicios", "88": "honorarios c/iva",
    "89": "honorarios s/iva", "90": "alquiler inmuebles", "91": "alquiler de vehiculos",
    "92": "comisiones cobradas", "93": "liquidacion verduleria", "94": "liquidacion carniceria",
    "95": "liquidacion panaderia", "96": "montajes", "97": "venta mayorista", "98": "fabricacion",
    "99": "ch/rechazados", "100": "intereses por prestamos", "101": "recibo anulado",
    "102": "venta de maquinarias", "103": "liquidacion perfumeria", "104": "venta ganado x cta de 3°",
    "105": "pastaje", "106": "venta de exportacion", "107": "liquidacion agropecuaria",
    "109": "toros", "110": "licencia", "111": "donaciones", "112": "diferencia de cambio",
    "113": "gastos financieros", "114": "arrendamientos", "115": "dto. De valores",
    "116": "mano de obra de 3°", "117": "negociacion de valores", "118": "liquidacion dto de cheques",
    "119": "patentes de vehiculos", "120": "transporte", "121": "alquiler particular c/iva",
    "122": "alquiler particular s/iva", "123": "alquiler comercial c/iva",
    "124": "alquiler comercial s/iva", "125": "utiles y herramientas", "126": "premios",
    "127": "reconocimientos", "128": "publicidad y propaganda", "129": "gastos de seguridad",
    "130": "servivio de transporte", "131": "comprobante anulado", "132": "anticipos",
    "133": "gastos carrera", "134": "Mejora inmueble Propio", "135": "Gastos de medicina",
    "136": "Tasa de Fondeadero", "137": "Seguros Leasing", "138": "Repuestos e Insumos",
    "139": "Ofrendas y Limosnas", "140": "Gastos de Comedor", "141": "Ganado propio c/iva",
    "142": "Compra de Ganado", "143": "Vta. Carne Vacuna", "144": "Envases y Accesorios",
    "145": "Venta de Ganado", "146": "Ajuste Contable", "147": "Fondo de Comercio",
    "148": "Servicios Personales", "149": "COMPRA DE CARNE", "150": "Insumos Papas",
    "151": "Gastos de Arrendamiento", "152": "Venta de Vehiculo", "153": "Compra de Vehiculo",
    "154": "Obras en Curso", "155": "Boletos y Pasajes", "156": "Alquiler Barco",
    "157": "Materiales de Decoracion", "158": "Alquiler y Expensas", "159": "Alquiler de Herramientas",
    "160": "Viandas", "161": "Intereses", "162": "Seguros de Caucion", "163": "Impresiones",
    "164": "Gastos de Producción", "165": "Prestadores", "166": "Devolucion de Mercaderias",
    "167": "Alquiler de Maquinarias", "168": "Certificados Revisión Técnica",
    "169": "Registro Control Modelo", "170": "Camara Arg. De Talleres", "171": "Honorarios Directores",
    "172": "Fondo de Reparo", "173": "Gastos de Sanidad", "174": "Plan de ahorro",
    "175": "Alquiler Temporario", "176": "Alquiler y Logistica", "177": "Alquiler Bs. Muebles",
    "178": "Gastos de Capacitación", "179": "Maq. Y equipos medicos", "180": "gastos de organización",
    "181": "equipos de comunicación", "182": "Gas para la venta", "183": "Venta Flete Internacional",
    "184": "Flete Internacional", "185": "Gastos de Obra", "186": "Gastos de Desarrollo",
    "187": "Embarcaciones", "188": "Gastos de embarcacion", "189": "Venta de Papa",
    "190": "Utiles y elementos de cocina", "191": "cubiertos y vajillas", "192": "elementos ortopedicos",
    "193": "pines", "194": "golosinas", "195": "rotary internacional", "196": "distrito rotario 4825",
    "197": "ret- seguridad e higiene a", "198": "gastos de representacion",
    "199": "ativo de caja  (compra + v", "201": "C.M.", "202": "ativo de caja  (compra + v",
    "203": "Alimentos", "204": "Enfriado", "205": "REGALOS EMPRESARIALES"
}


# ──────────────────────────────────────────────────────────────
# Regex
# ──────────────────────────────────────────────────────────────

# Regex para la línea principal de una transacción
# Ejemplo: " 1 FC 05009-07466844A AUTOPISTAS URBANAS S A Ins. 30-57487647-4  45 B Exento           743,65          0,00          0,00        743,65"
# El día puede ser 1 o 2 dígitos, el tipo 2-3 chars, el número de comprobante variable
RE_MAIN = re.compile(
    r'^\s*(\d{1,2})\s+'                            # Dia
    r'(FC|NC|ND|TF|TK|Li)\s+'                       # Tipo comprobante
    r'(\d{5}-\d{1,12}[A-Z ]?)\s*'                   # Numero (más flexible para exportación)
    r'(.+?)\s+'                                     # Proveedor (Flexible hasta Cond IVA o CUIT)
    r'(?:(Ins\.|Mono|Monot|Exe |Exe\.|C\.F\.|Exp\.|Resp\.|SNC)\s+)?'  # Cond IVA OPCIONAL (a veces no viene)
    r'([\d\- ]{1,13})?\s+'                          # CUIT/DNI (Opcional. Permitimos espacios internos por DNIs mal tipeados)
    r'(\d{1,3})\s+'                                 # Concepto
    r'([A-Z0-9])\s+'                                # Jurisdicción (Letra A-Z o 0 para exportación)
    r'(.+)$'                                        # Resto (tasa + montos)
)

# Regex para líneas de continuación (sub-conceptos)
# Ejemplo: "                                                                       Imp.Inter        385,94          0,00          0,00       5802,89"
RE_CONT = re.compile(
    r'^\s{50,}'                                # Gran cantidad de espacios
    r'(\S.+)$'                                 # contenido
)

# Regex para extraer montos (formato argentino: 1.234,56 o -1.234,56)
RE_MONTO = re.compile(r'-?[\d]+(?:\.[\d]{3})*,\d{2}')

# Líneas a ignorar
RE_IGNORE = re.compile(
    r'^\s*$|'
    r'^\s*Pag\.:|'
    r'^\s*CLASIFICADORURAL|'
    r'^\s*ESTADOS UNIDOS|'
    r'^\s+Numero de CUIT|'
    r'^\s*[A-Z ]?\s*IVA COMPRAS|'
    r'^\s*[A-Z ]?\s*Desde el|'
    r'^\s*Dia\s+Numero|'
    r'^\s*TC\s+|'
    r'^-- --|'
    r'^-{10,}|'
    r'==>|'                                    # Cualquier línea con flecha (subtotales)
    r'TOTALES\s+POR|'                          # Encabezados de tablas de resumen
    r'^\s*TOTAL\s+GENERAL|'
    r'^Cod\s+Concepto|'
    r'^Cod\s+Detalle|'
    r'^\s*\d+\s+Factura|'
    r'^\s*\d+\s+Nota de|'
    r'^\s*\d+\s+Tiquet|'
    r'^\s*[A-Z]\s+(Exento|Resp\.|Resp\.)|'
    r'^I: Valor neto|'
    r'^\x0c|'          # Form feed
    r'^\x0f|'          # Control chars
    r'^\x1b',          # ESC sequences
    re.IGNORECASE
)


def limpiar_control(texto: str) -> str:
    """Elimina caracteres de control y escape del texto."""
    texto = re.sub(r'\x1b[A-Za-z@]', '', texto)   # ESC + letra
    texto = re.sub(r'\x1b[A-Z]', '', texto)         
    texto = re.sub(r'[\x00-\x09\x0b-\x0c\x0e-\x1f]', '', texto)
    return texto.rstrip('\r\n')


def parse_monto(s: str) -> float:
    """Convierte un string '1.234,56' o '-1.234,56' al float correspondiente."""
    s = s.strip()
    s = s.replace('.', '').replace(',', '.')
    return float(s)


def extraer_montos_resto(resto: str):
    """
    Del 'resto' de la línea principal, extrae la Tasa y los 4 montos.
    Retorna: (tasa_str, neto, iva, percepcion, total)
    """
    montos = RE_MONTO.findall(resto)
    
    # Determinar la tasa
    tasa_str = resto.split(montos[0])[0].strip() if montos else resto.strip()
    
    neto = parse_monto(montos[0]) if len(montos) >= 1 else 0.0
    iva = parse_monto(montos[1]) if len(montos) >= 2 else 0.0
    percepcion = parse_monto(montos[2]) if len(montos) >= 3 else 0.0
    total = parse_monto(montos[3]) if len(montos) >= 4 else 0.0
    
    return tasa_str, neto, iva, percepcion, total


def extraer_montos_continuacion(contenido: str):
    """
    De una línea de continuación extrae concepto y montos.
    Retorna: (concepto, neto, iva, percepcion, total_parcial)
    """
    montos = RE_MONTO.findall(contenido)
    concepto = contenido.split(montos[0])[0].strip() if montos else contenido.strip()
    
    neto = parse_monto(montos[0]) if len(montos) >= 1 else 0.0
    iva = parse_monto(montos[1]) if len(montos) >= 2 else 0.0
    percepcion = parse_monto(montos[2]) if len(montos) >= 3 else 0.0
    total = parse_monto(montos[3]) if len(montos) >= 4 else 0.0
    
    return concepto, neto, iva, percepcion, total


def limpiar_para_excel(texto: str) -> str:
    """Elimina caracteres de control no permitidos en Excel/XML."""
    if not texto: return ""
    # Quitar caracteres de control ASCII 0-31 (excepto newline si fuera necesario, pero aqui no)
    # y otros caracteres no imprimibles detectados
    return re.sub(r'[\x00-\x1f\x7f-\x9f]', '', texto).strip()


def parsear_archivo(path: Path = None, content: str = None) -> tuple[list[dict], dict]:
    """Lee el archivo .txt (desde path o contenido directo) y extrae transacciones y metadata."""
    if content is None and path is not None:
        with open(path, 'r', encoding='ansi') as f:
            content = f.read()
    
    if not content:
        return [], {}

    lines = content.splitlines()
    transacciones = []
    current = None
    
    # Metadata del contribuyente
    meta = {
        'razon_social': '',
        'cuit_empresa': '',
        'periodo': ''
    }
    
    # Extraer metadata de las primeras líneas con limpieza
    if len(lines) > 5:
        meta['razon_social'] = limpiar_para_excel(lines[1])
        cuit_match = re.search(r'CUIT:([\d-]+)', lines[3])
        if cuit_match:
            meta['cuit_empresa'] = cuit_match.group(1)
        
        # El tipo de reporte (IVA COMPRAS / IVA VENTAS) suele estar en la linea 5
        reporte_raw = limpiar_para_excel(lines[4])
        # Limpiar prefijos como "E " o "F " que a veces aparecen en el TXT
        meta['tipo_reporte'] = re.sub(r'^[A-Z]\s+', '', reporte_raw).strip()

        # El periodo suele estar en la linea 6. Intentamos captar solo el texto "Desde... hasta..."
        periodo_raw = lines[5]
        p_match = re.search(r'(Desde .* hasta .*)', periodo_raw)
        if p_match:
            meta['periodo'] = limpiar_para_excel(p_match.group(1))
        else:
            meta['periodo'] = limpiar_para_excel(periodo_raw)

    for line in lines:
        linea = limpiar_control(line)
        
        # Ignorar líneas de encabezado, separadores, subtotales, etc.
        if RE_IGNORE.search(linea):
            # NO cerramos la transacción en separadores/subtotales
            # porque puede continuar en la página siguiente
            continue
        
        # Intentar match de línea principal
        m = RE_MAIN.match(linea)
        if m:
            dia = int(m.group(1))
            tipo = m.group(2).strip()
            numero = m.group(3).strip()
            proveedor = m.group(4).strip()
            cond_iva = m.group(5).strip() if m.group(5) else ""
            cuit = m.group(6).strip() if m.group(6) else ""
            concepto = int(m.group(7))
            letra = m.group(8).strip()
            resto = m.group(9)
            
            tasa_str, neto, iva, percepcion, total = extraer_montos_resto(resto)
            
            # Si es el MISMO comprobante (continuación tras salto de página),
            # tratar como sub-concepto en vez de nueva transacción.
            # Agregamos CUIT y Proveedor para evitar agrupar movimientos distintos con mismo número (ej: SIRCREB)
            if (current and
                current['Fecha'] == dia and
                current['Tipo'] == tipo and
                current['Numero'] == numero and
                current['CUIT'] == cuit and
                current['Proveedor'] == proveedor):
                # Es el mismo comprobante (salto de página):
                # Agregamos los montos como sub-conceptos para que se distribuyan 
                # correctamente según su propia 'tasa_str'.
                # NO sumamos a current['Neto']/IVA/Percepcion directamente 
                # porque eso forzaría los valores al bucket de la primera página.
                if total != 0.0:
                    current['Total'] = total
                current['SubConceptos'].append({
                    'Concepto': tasa_str,
                    'Neto': neto,
                    'IVA': iva,
                    'Percepcion': percepcion,
                    'Total': total
                })
                continue
            
            # Es una transacción nueva → guardar la previa
            if current:
                transacciones.append(current)
            
            current = {
                'Fecha': dia,
                'Tipo': tipo,
                'Numero': numero,
                'Proveedor': proveedor,
                'Cond_IVA': cond_iva,
                'CUIT': cuit,
                'Concepto': concepto,
                'Letra': letra,
                'Tasa': tasa_str,
                'Neto': neto,
                'IVA': iva,
                'Percepcion': percepcion,
                'Total': total,
                'SubConceptos': []
            }
            continue
        
        # Intentar match de línea de continuación
        mc = RE_CONT.match(linea)
        if mc and current:
            contenido = mc.group(1)
            concepto_sub, neto_s, iva_s, perc_s, total_s = extraer_montos_continuacion(contenido)
            
            # El total de la última línea de continuación tiene el total correcto
            if total_s != 0.0:
                current['Total'] = total_s
            
            current['SubConceptos'].append({
                'Concepto': concepto_sub,
                'Neto': neto_s,
                'IVA': iva_s,
                'Percepcion': perc_s,
                'Total': total_s
            })
            continue
    
    # Guardar última transacción si existe
    if current:
        transacciones.append(current)
    
    return transacciones, meta


def construir_sistema_aux_set(transacciones: list[dict]) -> set:
    """Construye el set de claves Auxiliar del SISTEMA para cruce con ARCA.

    Misma lógica que la que se usa internamente en crear_excel() al armar
    `sistema_aux_values` (Tipo + ' ' + Letra + PV + Nro + CUIT), replicada acá
    para que se pueda calcular el cruce sin re-ejecutar la generación de Excel.
    """
    aux_set = set()
    for t in transacciones:
        numero_raw = t['Numero']
        pv = numero_raw.split('-')[0] if '-' in numero_raw else numero_raw[:5]
        resto = numero_raw.split('-')[1] if '-' in numero_raw else numero_raw[5:]
        letra = resto[-1] if resto and resto[-1].isalpha() else ''
        nro = resto[:-1] if letra else resto
        pv_int = int(pv) if pv.isdigit() else pv
        nro_val = int(nro) if nro and nro.isdigit() else nro
        cuit_raw = t['CUIT'].replace('-', '') if t['CUIT'] else ''
        cuit_val = int(cuit_raw) if cuit_raw and cuit_raw.isdigit() else cuit_raw
        aux = f"{t['Tipo']} {letra}{pv_int}{nro_val}{cuit_val}"
        aux_set.add(aux)
    return aux_set


def _autofit(ws, n_cols, start_row=6):
    """Ajusta el ancho de todas las columnas de una hoja al contenido.
    Empieza desde start_row para ignorar filas de titulo mergeadas."""
    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                min_row=start_row, max_row=ws.max_row):
            for cell in row:
                val = cell.value
                if val is not None:
                    if isinstance(val, str) and val.startswith('='):
                        text = '($999,999.99)'  # ancho estimado para formulas
                    elif isinstance(val, (int, float)):
                        text = f'${val:,.2f}'
                    else:
                        text = str(val)
                    if len(text) > max_len:
                        max_len = len(text)
        ws.column_dimensions[letter].width = max(max_len + 3, 8)


# ── Constantes y helpers compartidos para la construccion de filas ──
_DESIRED_IVA_ORDER = [
    'Neto IVA 21', 'IVA 21',
    'Neto C.F. 21', 'IVA C.F. 21',
    'Neto IVA 27', 'IVA 27',
    'Neto IVA 10.5', 'IVA 10.5',
    'Neto C.F. 10.5', 'IVA C.F. 10.5',
    'Neto IVA 5', 'IVA 5',
    'Neto IVA 2.5', 'IVA 2.5',
    'Neto Monot. 21', 'IVA Monot. 21',
    'Neto Monot. 10.5', 'IVA Monot. 10.5',
    'Neto Imp. 21', 'IVA Imp. 21',
    'Neto Imp. 10.5', 'IVA Imp. 10.5',
    'Exento', 'Neto C.F.', 'Monotributo',
]

_DEDUCCION_KW = ("PERC", "PER.", "PER ", "RET", "SIRCREB", "SIRTAC")


def _es_deduccion(nombre: str) -> bool:
    nu = nombre.upper()
    return any(kw in nu for kw in _DEDUCCION_KW)


def _iva_rates_map(meta: dict) -> dict:
    """Mapa de tasa-string -> (col Neto, col IVA o None). Expande monotributo en ventas."""
    iva_rates = {
        'Tasa 21%':  ('Neto IVA 21',   'IVA 21'),
        'T.21%':     ('Neto IVA 21',   'IVA 21'),
        'C.F.21%':   ('Neto C.F. 21',  'IVA C.F. 21'),
        'Tasa 27%':  ('Neto IVA 27',   'IVA 27'),
        'T.27%':     ('Neto IVA 27',   'IVA 27'),
        'Tasa 10.5%': ('Neto IVA 10.5', 'IVA 10.5'),
        'Tasa 10,5%': ('Neto IVA 10.5', 'IVA 10.5'),
        'T.10.5%':   ('Neto IVA 10.5', 'IVA 10.5'),
        'T.10,5%':   ('Neto IVA 10.5', 'IVA 10.5'),
        'C.F.10.5%': ('Neto C.F. 10.5', 'IVA C.F. 10.5'),
        'C.F.10,5%': ('Neto C.F. 10.5', 'IVA C.F. 10.5'),
        'Tasa 5%':   ('Neto IVA 5',    'IVA 5'),
        'TASA 5%':   ('Neto IVA 5',    'IVA 5'),
        'T.5%':      ('Neto IVA 5',    'IVA 5'),
        'Tasa 2.5%': ('Neto IVA 2.5',  'IVA 2.5'),
        'Tasa 2,5%': ('Neto IVA 2.5',  'IVA 2.5'),
        'T.2.5%':    ('Neto IVA 2.5',  'IVA 2.5'),
        'T.2,5%':    ('Neto IVA 2.5',  'IVA 2.5'),
        'T.IMP 21%': ('Neto Imp. 21',  'IVA Imp. 21'),
        'T.IMP 10%': ('Neto Imp. 10.5','IVA Imp. 10.5'),
        'Exento':    ('Exento',    None),
        'Cons.Fin.': ('Neto C.F.', None),
    }
    if 'VENTA' in meta.get('tipo_reporte', '').upper():
        iva_rates['R.Monot21'] = ('Neto Monot. 21', 'IVA Monot. 21')
        iva_rates['R.Mont.10'] = ('Neto Monot. 10.5', 'IVA Monot. 10.5')
    return iva_rates


def _construir_filas_consolidado(transacciones: list[dict], meta: dict,
                                  con_auxiliar: bool = False, cruce_arca: bool = False):
    """Descubre columnas dinamicas y arma las filas consolidadas (una por comprobante).

    Devuelve (rows, IVA_COL_ORDER, other_cols, IVA_RATES) o (None, None, None, None) si no hay nada.
    """
    if not transacciones:
        return None, None, None, None

    IVA_RATES = _iva_rates_map(meta)

    present_iva_cols = set()
    found_others = []  # Preserva orden de aparicion en TXT
    for t in transacciones:
        tasa = t['Tasa']
        if tasa in IVA_RATES:
            neto_col, iva_col = IVA_RATES[tasa]
            present_iva_cols.add(neto_col)
            if iva_col: present_iva_cols.add(iva_col)
        elif tasa and tasa.strip():
            t_clean = tasa.strip()
            if t_clean not in found_others:
                found_others.append(t_clean)
        for s in t['SubConceptos']:
            conc = s['Concepto']
            if conc in IVA_RATES:
                neto_col, iva_col = IVA_RATES[conc]
                present_iva_cols.add(neto_col)
                if iva_col: present_iva_cols.add(iva_col)
            elif conc and conc.strip():
                c_clean = conc.strip()
                if c_clean not in found_others:
                    found_others.append(c_clean)

    IVA_COL_ORDER = [c for c in _DESIRED_IVA_ORDER if c in present_iva_cols]
    if not IVA_COL_ORDER:
        if not found_others and not present_iva_cols:
            return None, None, None, None
        IVA_COL_ORDER = sorted(list(present_iva_cols))

    # Ordenar: primero no-deducciones (amarillo), luego deducciones (verde)
    other_cols = [c for c in found_others if not _es_deduccion(c)] + \
                 [c for c in found_others if _es_deduccion(c)]

    rows = []
    for t in transacciones:
        numero_raw = t['Numero']
        pv = numero_raw.split('-')[0] if '-' in numero_raw else numero_raw[:5]
        resto_num = numero_raw.split('-')[1] if '-' in numero_raw else numero_raw[5:]
        letra = resto_num[-1] if resto_num and resto_num[-1].isalpha() else ''
        nro = resto_num[:-1] if letra else resto_num

        cuit_raw = t['CUIT'].replace('-', '') if t['CUIT'] else ''
        cuit_val = int(cuit_raw) if cuit_raw and cuit_raw.isdigit() else cuit_raw

        row = {
            'Fecha': t['Fecha'],
            'Tipo': t['Tipo'],
            'PV': int(pv),
            'Nro.': int(nro) if nro.isdigit() else nro,
            'Letra': letra,
            'Proveedor': t['Proveedor'],
            'Cond. IVA': t['Cond_IVA'],
            'CUIT': cuit_val,
            'Concepto': t['Concepto'],
            'Jur.': t['Letra'],
        }
        for col in IVA_COL_ORDER:
            row[col] = 0.0
        for col in other_cols:
            row[col] = 0.0

        tasa = t['Tasa']
        if tasa in IVA_RATES:
            neto_col, iva_col = IVA_RATES[tasa]
            row[neto_col] += t['Neto']
            if iva_col:
                row[iva_col] += t['IVA']
        elif tasa:
            row[tasa] += t['Neto']

        for s in t['SubConceptos']:
            nombre = s['Concepto']
            if not nombre:
                continue
            if nombre in IVA_RATES:
                neto_col, iva_col = IVA_RATES[nombre]
                row[neto_col] += s['Neto']
                if iva_col:
                    row[iva_col] += s['IVA']
            else:
                monto = s['Neto'] if s['Neto'] != 0.0 else s['Percepcion']
                row[nombre] += monto

        if con_auxiliar or cruce_arca:
            row['Auxiliar'] = ''

        row['Total'] = t['Total']

        if t['Tipo'] == 'NC':
            for col in IVA_COL_ORDER + other_cols:
                row[col] = -row[col]
            row['Total'] = -row['Total']

        rows.append(row)

    return rows, IVA_COL_ORDER, other_cols, IVA_RATES


def crear_excel(transacciones: list[dict], meta: dict, output_path, con_resumenes=True, con_auxiliar=False, cruce_arca=False, df_arca=None, con_asiento=False):
    """Crea un Excel formateado. Cada tasa de IVA tiene sus propias columnas
    Neto/IVA y cada percepcion/retencion tiene su propia columna.
    output_path puede ser una ruta o un BytesIO buffer."""

    rows, IVA_COL_ORDER, other_cols, IVA_RATES = _construir_filas_consolidado(
        transacciones, meta, con_auxiliar=con_auxiliar, cruce_arca=cruce_arca
    )
    if rows is None:
        return

    df = pd.DataFrame(rows)

    all_dynamic = IVA_COL_ORDER + other_cols

    print(f"  Total de transacciones parseadas: {len(df)}")
    print(f"   - FC (Factura): {len(df[df['Tipo'] == 'FC'])}")
    print(f"   - NC (Nota Credito): {len(df[df['Tipo'] == 'NC'])}")
    print(f"   - ND (Nota Debito): {len(df[df['Tipo'] == 'ND'])}")
    print(f"   - TF (Ticket Factura): {len(df[df['Tipo'] == 'TF'])}")
    print(f"   - Li (Liquidacion): {len(df[df['Tipo'] == 'Li'])}")

    # ── 3. Escribir Excel ─────────────────────────────────────
    total_cols = len(df.columns)
    last_col_letter = get_column_letter(total_cols)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # ── Estilos Reutilizables ─────────────────────────────
        title_font = Font(bold=True, size=14, color='FFFFFF')
        title_fill = PatternFill('solid', fgColor='2F5496')
        
        header_font = Font(bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        iva_header_fill = PatternFill('solid', fgColor='BF8F00')
        perc_header_fill = PatternFill('solid', fgColor='70AD47')
        
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        zebra_fill = PatternFill('solid', fgColor='D6E4F0')
        money_fmt = '$#,##0.00'
        # Formato contabilidad Peso para cruce ARCA
        accounting_fmt = '_-"$"* #,##0.00_-;-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

        # ── Hoja Movimientos / SISTEMA ──────────────────────────────────
        # startrow=5 significa que los encabezados del DataFrame van en la fila 6
        mov_sheet_name = 'SISTEMA' if cruce_arca else 'Movimientos'
        df.to_excel(writer, sheet_name=mov_sheet_name, index=False, startrow=5)
        ws = writer.sheets[mov_sheet_name]

        # Estilo de Reporte (Rojo/Diferente para resaltar)
        report_type_font = Font(bold=True, size=12, color='C00000') # Rojo oscuro

        ws.merge_cells(f'A1:{last_col_letter}1')
        ws['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_align

        ws.merge_cells(f'A2:{last_col_letter}2')
        ws['A2'] = meta['tipo_reporte'].upper() if meta['tipo_reporte'] else 'REPORTE DE MOVIMIENTOS'
        ws['A2'].font = report_type_font
        ws['A2'].alignment = center_align

        ws.merge_cells(f'A3:{last_col_letter}3')
        ws['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
        ws['A3'].font = Font(bold=True, size=11, color='2F5496')
        ws['A3'].alignment = center_align

        ws.merge_cells(f'A4:{last_col_letter}4')
        ws['A4'] = f'Total: {len(df)} transacciones'
        ws['A4'].font = Font(italic=True, size=10, color='4472C4')
        ws['A4'].alignment = center_align

        # Fila 5 queda vacía como separador

        col_list = list(df.columns)
        iva_set = set(IVA_COL_ORDER)
        other_set = set(other_cols)
        # Deducciones (PERC/PER./RET/SIRCREB) → verde; otros impuestos → amarillo
        deduccion_set = {c for c in other_cols if _es_deduccion(c)}

        # Los encabezados ahora van en la fila 6
        header_row = 6
        data_start_row = 7
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            col_name = col_list[col_idx - 1]
            if col_name in iva_set or (col_name in other_set and col_name not in deduccion_set):
                cell.fill = iva_header_fill
            elif col_name in deduccion_set:
                cell.fill = perc_header_fill
            else:
                cell.fill = header_fill

        money_col_names = IVA_COL_ORDER + other_cols + ['Total']
        money_col_indices = [col_list.index(c) + 1 for c in money_col_names if c in col_list]
        active_money_fmt = accounting_fmt if cruce_arca else money_fmt
        cuit_col_idx = (col_list.index('CUIT') + 1) if 'CUIT' in col_list else None

        first_sum_col = get_column_letter(col_list.index(IVA_COL_ORDER[0]) + 1)
        last_sum_col = get_column_letter(col_list.index(other_cols[-1]) + 1) if other_cols else get_column_letter(col_list.index(IVA_COL_ORDER[-1]) + 1)
        total_col_idx = col_list.index('Total') + 1

        # Aplicamos alignment y money format en una sola pasada por celda; ya NO
        # hacemos la pasada extra de zebra (~325k assignments adicionales) que
        # era la mas costosa proporcionalmente.
        for row in range(data_start_row, len(df) + data_start_row):
            for col_idx in range(1, total_cols + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.alignment = center_align
                if col_idx in money_col_indices:
                    cell.number_format = active_money_fmt

            ws.cell(row=row, column=total_col_idx).value = f'=SUM({first_sum_col}{row}:{last_sum_col}{row})'
            ws.cell(row=row, column=total_col_idx).number_format = active_money_fmt

        # Fila TOTAL GENERAL en Movimientos
        total_row_mov = len(df) + data_start_row
        first_money_idx = col_list.index(IVA_COL_ORDER[0]) + 1
        
        ws.merge_cells(f'A{total_row_mov}:{get_column_letter(first_money_idx-1)}{total_row_mov}')
        ws[f'A{total_row_mov}'] = "TOTAL GENERAL"
        ws[f'A{total_row_mov}'].font = Font(bold=True)
        ws[f'A{total_row_mov}'].alignment = Alignment(horizontal='right')
        
        for col_idx in range(first_money_idx, total_cols + 1):
            col_l = get_column_letter(col_idx)
            cell = ws.cell(row=total_row_mov, column=col_idx)
            cell.value = f'=SUM({col_l}{data_start_row}:{col_l}{total_row_mov-1})'
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style='double'))
            cell.number_format = active_money_fmt
            cell.alignment = center_align

        # ── Formulas Auxiliar (interactivas) ───────────────────
        if (con_auxiliar or cruce_arca) and 'Auxiliar' in col_list:
            aux_col_idx = col_list.index('Auxiliar') + 1
            aux_col_letter = get_column_letter(aux_col_idx)
            tipo_letter = get_column_letter(col_list.index('Tipo') + 1)
            letra_letter = get_column_letter(col_list.index('Letra') + 1)
            pv_letter = get_column_letter(col_list.index('PV') + 1)
            nro_letter = get_column_letter(col_list.index('Nro.') + 1)
            cuit_letter = get_column_letter(col_list.index('CUIT') + 1)
            total_col_letter = get_column_letter(col_list.index('Total') + 1)
            for row in range(data_start_row, len(df) + data_start_row):
                ws.cell(row=row, column=aux_col_idx).value = (
                    f'={tipo_letter}{row}&" "&{letra_letter}{row}&{pv_letter}{row}&{nro_letter}{row}&{cuit_letter}{row}'
                )

        # ── CRUCE + DIFF formulas en SISTEMA ──────────────────────
        if cruce_arca and 'Auxiliar' in col_list and df_arca is not None and not df_arca.empty:
            # Agregar columnas CRUCE y DIFF al final
            cruce_col_idx = total_cols + 1
            diff_col_idx = total_cols + 2
            cruce_col_letter = get_column_letter(cruce_col_idx)
            diff_col_letter = get_column_letter(diff_col_idx)

            # Headers
            ws.cell(row=6, column=cruce_col_idx).value = 'CRUCE'
            ws.cell(row=6, column=cruce_col_idx).font = header_font
            ws.cell(row=6, column=cruce_col_idx).fill = PatternFill('solid', fgColor='7030A0')
            ws.cell(row=6, column=cruce_col_idx).alignment = header_align
            ws.cell(row=6, column=cruce_col_idx).border = thin_border

            ws.cell(row=6, column=diff_col_idx).value = 'DIFF'
            ws.cell(row=6, column=diff_col_idx).font = header_font
            ws.cell(row=6, column=diff_col_idx).fill = PatternFill('solid', fgColor='7030A0')
            ws.cell(row=6, column=diff_col_idx).alignment = header_align
            ws.cell(row=6, column=diff_col_idx).border = thin_border

            # Calcular rango de lookup en ARCA (Auxiliar:Total son las 2 ultimas cols)
            arca_col_list = list(df_arca.columns)
            arca_aux_col_letter = get_column_letter(arca_col_list.index('Auxiliar') + 1) if 'Auxiliar' in arca_col_list else 'A'
            arca_total_col_letter = get_column_letter(arca_col_list.index('Total') + 1) if 'Total' in arca_col_list else 'B'
            arca_total_col_offset = arca_col_list.index('Total') - arca_col_list.index('Auxiliar') + 1 if 'Auxiliar' in arca_col_list and 'Total' in arca_col_list else 2
            arca_last_data_row = len(df_arca) + 6  # data starts at row 7

            for row in range(data_start_row, len(df) + data_start_row):
                # CRUCE: VLOOKUP en ARCA buscando Auxiliar, trayendo Total
                ws.cell(row=row, column=cruce_col_idx).value = (
                    f'=IFERROR(VLOOKUP({aux_col_letter}{row},'
                    f"ARCA!${arca_aux_col_letter}$7:${arca_total_col_letter}${arca_last_data_row},"
                    f'{arca_total_col_offset},FALSE),"NO ENCONTRADO")'
                )
                ws.cell(row=row, column=cruce_col_idx).number_format = money_fmt
                ws.cell(row=row, column=cruce_col_idx).alignment = center_align

                # DIFF: Total - CRUCE (solo si CRUCE es numerico)
                ws.cell(row=row, column=diff_col_idx).value = (
                    f'=IF({cruce_col_letter}{row}="NO ENCONTRADO","",'
                    f'{total_col_letter}{row}-{cruce_col_letter}{row})'
                )
                ws.cell(row=row, column=diff_col_idx).number_format = money_fmt
                ws.cell(row=row, column=diff_col_idx).alignment = center_align

        _autofit(ws, total_cols)
        ws.column_dimensions['A'].width = 8 # Ancho fijo para columna Fecha

        # ── Hojas de Resumen (Solo si se solicita) ────────────
        if con_resumenes:

            # Resto del código de resúmenes...
            resumen = df.copy()
        
            # Separar conceptos en Deducciones (PERC/RET) y Otros (IMP.CIG, etc.)
            deduccion_cols = [c for c in other_cols if _es_deduccion(c)]
            individual_other_cols = [c for c in other_cols if c not in deduccion_cols]
        
            res_header_row = 6
            res_data_start = 7
            n_mov = len(df)
            mov_cuit_col = get_column_letter(col_list.index('CUIT') + 1)
            mov_tipo_col = get_column_letter(col_list.index('Tipo') + 1)
            mov_conc_col = get_column_letter(col_list.index('Concepto') + 1)
            mov_jur_col = get_column_letter(col_list.index('Jur.') + 1)

            # ── Hoja Resumen por Impuesto (INTERACTIVA) ──────────
            res_imp_data = []
            seen_cols = set()
        
            # Tasas estándar de IVA
            for tasa_label, (neto_col, iva_col) in IVA_RATES.items():
                if neto_col in df.columns and (neto_col, iva_col) not in seen_cols:
                    n_idx = col_list.index(neto_col) + 1
                    i_idx = (col_list.index(iva_col) + 1) if (iva_col and iva_col in df.columns) else None
                    res_imp_data.append({
                        'Tasa': tasa_label,
                        'Neto_Col_M': get_column_letter(n_idx),
                        'IVA_Col_M': get_column_letter(i_idx) if i_idx else None,
                        'Ded_Col_M': None
                    })
                    seen_cols.add((neto_col, iva_col))
        
            # Deducciones y otros
            for col in other_cols:
                c_idx = col_list.index(col) + 1
                col_upper = col.upper()
                if "PERC" in col_upper or "RET" in col_upper or "SIRCREB" in col_upper:
                    res_imp_data.append({
                        'Tasa': col, 'Neto_Col_M': None, 'IVA_Col_M': None, 'Ded_Col_M': get_column_letter(c_idx)
                    })
                else:
                    # Todo lo demás que no es IVA (como IMP.CIG, impuestos internos, ajustes, etc.) va a Neto
                    res_imp_data.append({
                        'Tasa': col, 'Neto_Col_M': get_column_letter(c_idx), 'IVA_Col_M': None, 'Ded_Col_M': None
                    })

            # ── Orden de conceptos: impuestos primero (por código), luego deducciones ──
            TASA_ORDER_MAP = {
                'Exento': 1,
                'Tasa 21%': 2, 'T.21%': 2,
                'Tasa 27%': 3, 'T.27%': 3,
                'T.10.5%': 4, 'T.10,5%': 4, 'Tasa 10.5%': 4, 'Tasa 10,5%': 4,
                'Tasa 21+5': 5,
                'Tasa 27+5': 6,
                'Imp.Inter': 7, 'Imp.Inter.': 7,
                'Cons.Fin.': 8,
                'R.Monot21': 9,
                'R.Mont.10': 10,
                'C.F.21%': 11,
                'C.F.10.5%': 12, 'C.F.10,5%': 12,
                'CPTE.ANUL': 13,
                'IMP.COMB.': 14,
                'IMP.CIG.': 15,
                'ABASTO.': 16,
                'L.25413': 17,
                'IMP.SELLO': 18,
                'D.976/01': 19,
                'BONIFIC.': 20,
                'itida en': 21,
                'TJT Prep.': 22,
                'L25413(2)': 23,
                'AJUST RED': 24,
                'IVA PEAJE': 25,
                'TRANPORT': 26,
                'AJUST IVA': 27,
                'DESC OTOR': 28,
                'DESC.S/IV': 29,
                'DESC.10.5': 30,
                'Valor Cri': 31,
                'Incre Iva': 32,
                'S41': 33,
                'DESC CF': 34,
                'DESC.MONO': 35,
                'Tasa 16,5': 36,
                'Tasa 22%': 37,
                'T.IMP 21%': 38,
                'T.IMP 10%': 39,
                'TASA 16,6': 40,
                'Tasa 2.5%': 41, 'Tasa 2,5%': 41, 'T.2.5%': 41, 'T.2,5%': 41,
                'Tasa 0%': 42,
                'TASA 9%': 43,
                'TASA 5%': 44, 'Tasa 5%': 44, 'T.5%': 44,
                'L.27264': 45,
                'IMP.FONDE': 46,
                'R.Mont.27': 47,
                'TurIVA': 48,
                'REC. GAS.': 49,
                'CRI 10,5': 50,
                'CCF 10,5': 51,
                'CRM 10,50': 52,
                'AJ.IMPORT': 53,
                'CM 21%': 54,
                'CM CF 21%': 55,
                'IMP. PAIS': 56,
                'GST IIBB': 57,
                'T21% 4240': 58,
            }

            # ── Orden de deducciones (percepciones/retenciones) ──
            DEDUCCION_ORDER_MAP = {
                'PERC.I.V.A.': 1, 'PERC.IVA': 1, 'PERC IVA': 1,
                'PERC.GCIAS.': 2, 'PERC.GCIAS': 2, 'PERC GCIAS': 2,
                'PERC.IB.CAP.FED.': 3, 'PERC.IB.CAP.FED': 3,
                'PERC.IB.BS.AS.': 4, 'PERC.IB.BS.AS': 4,
                'PERC.IB.CORDOBA': 5, 'PERC.IB.CÓRDOBA': 5,
                'PERC.IB.MENDOZA': 6,
                'PERC.IB.MISIONES': 7,
                'RET.GCIAS': 8, 'RET GCIAS': 8,
                'RET.IB.BS.AS.': 9, 'RET.IB.BS.AS': 9,
                'RET.IB. CAP.FED': 10, 'RET.IB.CAP.FED': 10, 'RET.IB.CAP.FED.': 10,
                'RET.IB.CORDOBA': 11, 'RET.IB.CÓRDOBA': 11,
                'RET.IB.MENDOZA': 12,
                'RET.IB.MISIONES': 13,
                'RET.SIRCREB CORDOBA': 14, 'RET.SIRCREB CÓRDOBA': 14,
                'RET.SIRCREB MENDOZA': 15,
                'RET.SIRCREB JUJUY': 16,
                'RET. SIRCREB C.A.B.A': 17, 'RET.SIRCREB C.A.B.A': 17, 'RET.SIRCREB CABA': 17,
                'RET.SIRCREB R.NEGRO': 18, 'RET.SIRCREB RIO NEGRO': 18,
                'PERC.ADUANERA C.FED.': 19, 'PERC.ADUANERA C.FED': 19,
                'PERC.ADUANERA BSAS': 20, 'PERC.ADUANERA BS.AS.': 20,
                'PERCEP.ADUAN.CORDOBA': 21, 'PERCEP.ADUAN.CÓRDOBA': 21,
                'PERCEP.ADUAN.MENDOZA': 22,
                'SIRCREB CORRIENTES': 23,
                'PERC.ADUANA CORRIENT': 24, 'PERC.ADUANA CORRIENTES': 24,
                'PERC.ADUAN. RIO NEG.': 25, 'PERC.ADUAN. RIO NEG': 25, 'PERC.ADUAN.RIO NEG': 25,
                'PERC ADUANA JUJUY': 26, 'PERC.ADUANA JUJUY': 26,
            }

            def _get_deduccion_code(nombre):
                """Busca el código de deducción, primero exacto, luego por prefijo."""
                if nombre in DEDUCCION_ORDER_MAP:
                    return DEDUCCION_ORDER_MAP[nombre]
                # Buscar por prefijo (para variantes con números entre paréntesis, etc.)
                nombre_limpio = nombre.split('(')[0].strip()
                if nombre_limpio in DEDUCCION_ORDER_MAP:
                    return DEDUCCION_ORDER_MAP[nombre_limpio]
                return 999

            def _tasa_sort_key(item):
                es_deduccion = 1 if item.get('Ded_Col_M') else 0
                if es_deduccion:
                    codigo = _get_deduccion_code(item['Tasa'])
                else:
                    codigo = TASA_ORDER_MAP.get(item['Tasa'], 999)
                return (es_deduccion, codigo, item['Tasa'])

            res_imp_data.sort(key=_tasa_sort_key)

            n_ri_cols = 5
            ws_ri_name = 'Resumen x Impuesto'
            pd.DataFrame([{'Tasa': r['Tasa']} for r in res_imp_data]).to_excel(writer, sheet_name=ws_ri_name, index=False, startrow=5)
            ws_ri = writer.sheets[ws_ri_name]
        
            ws_ri.merge_cells(f'A1:{get_column_letter(n_ri_cols)}1')
            ws_ri['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws_ri['A1'].font = title_font; ws_ri['A1'].fill = title_fill; ws_ri['A1'].alignment = center_align

            ws_ri.merge_cells(f'A2:{get_column_letter(n_ri_cols)}2')
            ws_ri['A2'] = f"{meta['tipo_reporte'].upper()} - RESUMEN POR IMPUESTO"
            ws_ri['A2'].font = report_type_font; ws_ri['A2'].alignment = center_align

            ws_ri.merge_cells(f'A3:{get_column_letter(n_ri_cols)}3')
            ws_ri['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws_ri['A3'].font = Font(bold=True, size=11, color='2F5496'); ws_ri['A3'].alignment = center_align
        
            ri_headers = ['Tasa', 'Neto', 'IVA', 'Deducciones', 'Total']
            for col_idx, h in enumerate(ri_headers):
                cell = ws_ri.cell(row=res_header_row, column=col_idx+1) # Reusing res_header_row
                cell.value = h
                cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border

            for idx, r_data in enumerate(res_imp_data):
                curr_row = res_data_start + idx # Reusing res_data_start
                ws_ri.cell(row=curr_row, column=1).value = r_data['Tasa']
                ws_ri.cell(row=curr_row, column=1).alignment = center_align
            
                # Neto
                if r_data['Neto_Col_M']: ws_ri.cell(row=curr_row, column=2).value = f"={mov_sheet_name}!{r_data['Neto_Col_M']}{total_row_mov}"
                else: ws_ri.cell(row=curr_row, column=2).value = 0.0
            
                # IVA
                if r_data['IVA_Col_M']: ws_ri.cell(row=curr_row, column=3).value = f"={mov_sheet_name}!{r_data['IVA_Col_M']}{total_row_mov}"
                else: ws_ri.cell(row=curr_row, column=3).value = 0.0

                # Deducciones
                if r_data['Ded_Col_M']: ws_ri.cell(row=curr_row, column=4).value = f"={mov_sheet_name}!{r_data['Ded_Col_M']}{total_row_mov}"
                else: ws_ri.cell(row=curr_row, column=4).value = 0.0
            
                # Total
                ws_ri.cell(row=curr_row, column=5).value = f"=B{curr_row}+C{curr_row}+D{curr_row}"
            
                for c in range(2, 6):
                    ws_ri.cell(row=curr_row, column=c).number_format = money_fmt
                    ws_ri.cell(row=curr_row, column=c).alignment = center_align

            total_row_ri = res_data_start + len(res_imp_data) # Reusing res_data_start
            ws_ri[f'A{total_row_ri}'] = "TOTAL GENERAL"
            ws_ri[f'A{total_row_ri}'].font = Font(bold=True); ws_ri[f'A{total_row_ri}'].alignment = Alignment(horizontal='right')
            for col_idx in range(2, 6):
                 col_l = get_column_letter(col_idx)
                 cell = ws_ri.cell(row=total_row_ri, column=col_idx)
                 cell.value = f'=SUM({col_l}{res_data_start}:{col_l}{total_row_ri-1})' # Reusing res_data_start
                 cell.font = Font(bold=True); cell.border = Border(top=Side(style='double'))
                 cell.number_format = money_fmt; cell.alignment = center_align

            _autofit(ws_ri, n_ri_cols)


            # ── Hoja Resumen por Tipo ─────────────────────────────
            res_tipo = resumen.groupby('Tipo').agg(
                **{c: (c, 'sum') for c in IVA_COL_ORDER},
                **{c: (c, 'sum') for c in individual_other_cols},
                Deducciones=('Total', 'count'), # placeholder
                Cantidad=('Total', 'count'),
            ).reset_index()
            res_tipo['Total'] = 0.0
            cols_order_rt = ['Tipo'] + IVA_COL_ORDER + individual_other_cols + ['Deducciones', 'Total', 'Cantidad']
            cols_order_rt = [c for c in cols_order_rt if c in res_tipo.columns]
            res_tipo = res_tipo[cols_order_rt]
            # Sort logic
            sum_cols = IVA_COL_ORDER + individual_other_cols + deduccion_cols
            res_tipo['_sort'] = resumen.groupby('Tipo')[sum_cols].sum().sum(axis=1).values
            res_tipo = res_tipo.sort_values('_sort', ascending=False).drop(columns='_sort')

            n_rt_cols = len(res_tipo.columns)
            # startrow=5 -> fila 6
            res_tipo.to_excel(writer, sheet_name='Resumen x Comprobante', index=False, startrow=5)
            ws3 = writer.sheets['Resumen x Comprobante']
        
            ws3.merge_cells(f'A1:{get_column_letter(n_rt_cols)}1')
            ws3['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws3['A1'].font = title_font
            ws3['A1'].fill = title_fill
            ws3['A1'].alignment = center_align

            ws3.merge_cells(f'A2:{get_column_letter(n_rt_cols)}2')
            ws3['A2'] = meta['tipo_reporte'].upper() if meta['tipo_reporte'] else 'RESUMEN POR TIPO'
            ws3['A2'].font = report_type_font
            ws3['A2'].alignment = center_align

            ws3.merge_cells(f'A3:{get_column_letter(n_rt_cols)}3')
            ws3['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws3['A3'].font = Font(bold=True, size=11, color='2F5496')
            ws3['A3'].alignment = center_align

            ws3.merge_cells(f'A4:{get_column_letter(n_rt_cols)}4')
            ws3['A4'] = f'Total: {len(res_tipo)} tipos'
            ws3['A4'].font = Font(italic=True, size=10, color='4472C4')
            ws3['A4'].alignment = center_align

            for col_idx in range(1, n_rt_cols + 1):
                cell = ws3.cell(row=res_header_row, column=col_idx) # reusando res_header_row=6
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            rt_col_list = list(res_tipo.columns)
            first_iva_idx_rt = rt_col_list.index(IVA_COL_ORDER[0]) + 1
            ded_idx_rt = rt_col_list.index('Deducciones') + 1
            total_idx_rt = rt_col_list.index('Total') + 1
            first_iva_letter_rt = get_column_letter(first_iva_idx_rt)
            ded_letter_rt = get_column_letter(ded_idx_rt)

            for row in range(res_data_start, len(res_tipo) + res_data_start):
                for col_idx in range(1, n_rt_cols + 1):
                    cell = ws3.cell(row=row, column=col_idx)
                    cell.alignment = center_align
                    col_name = rt_col_list[col_idx - 1]
                
                    if first_iva_idx_rt <= col_idx <= total_idx_rt:
                        if col_name in col_list:
                            v_col = get_column_letter(col_list.index(col_name) + 1)
                            # ROUND para evitar drift de centavos al sumar floats no-representables (ej 6629.05).
                            cell.value = f'=ROUND(SUMIFS(Movimientos!${v_col}${7}:${v_col}${n_mov+7-1}, Movimientos!${mov_tipo_col}${7}:${mov_tipo_col}${n_mov+7-1}, $A{row}),2)'
                            cell.number_format = money_fmt
                        elif col_name == 'Deducciones':
                            formula_parts = []
                            for dc in deduccion_cols:
                                v_col = get_column_letter(col_list.index(dc) + 1)
                                formula_parts.append(f'SUMIFS(Movimientos!${v_col}${7}:${v_col}${n_mov+7-1}, Movimientos!${mov_tipo_col}${7}:${mov_tipo_col}${n_mov+7-1}, $A{row})')
                            cell.value = ('=ROUND(' + '+'.join(formula_parts) + ',2)') if formula_parts else 0
                            cell.number_format = money_fmt
                        elif col_name == 'Total':
                            # Sumar desde el primer IVA hasta Deducciones
                            cell.value = f'=SUM({first_iva_letter_rt}{row}:{ded_letter_rt}{row})'
                            cell.number_format = money_fmt
                    elif col_name == 'Cantidad':
                        cell.value = f'=COUNTIFS(Movimientos!${mov_tipo_col}${7}:${mov_tipo_col}${n_mov+7-1}, $A{row})'

            # Fila TOTAL GENERAL
            total_row_rt = len(res_tipo) + res_data_start
            ws3.merge_cells(f'A{total_row_rt}:A{total_row_rt}')
            ws3[f'A{total_row_rt}'] = "TOTAL GENERAL"
            ws3[f'A{total_row_rt}'].font = Font(bold=True)
            ws3[f'A{total_row_rt}'].alignment = Alignment(horizontal='right')
        
            for col_idx in range(first_iva_idx_rt, n_rt_cols + 1):
                col_letter = get_column_letter(col_idx)
                cell = ws3.cell(row=total_row_rt, column=col_idx)
                cell.value = f'=SUM({col_letter}{res_data_start}:{col_letter}{total_row_rt-1})'
                cell.font = Font(bold=True)
                cell.border = Border(top=Side(style='double'))
                if col_idx < n_rt_cols:
                    cell.number_format = money_fmt

            _autofit(ws3, n_rt_cols)

            # ── Hoja Resumen por Concepto ─────────────────────────
            res_conc = resumen.groupby('Concepto').agg(
                **{c: (c, 'sum') for c in IVA_COL_ORDER},
                **{c: (c, 'sum') for c in individual_other_cols},
                Deducciones=('Total', 'count'), # placeholder
                Cantidad=('Total', 'count'),
            ).reset_index()
        
            # Ordenar por Concepto numérico
            res_conc['Concepto_Num'] = pd.to_numeric(res_conc['Concepto'], errors='coerce')
            res_conc = res_conc.sort_values('Concepto_Num').drop(columns='Concepto_Num')
        
            res_conc['Descripcion'] = res_conc['Concepto'].apply(
                lambda x: CONCEPTOS_MAP.get(str(x), "").replace("°", "o.").upper()
            )
        
            res_conc['Total'] = 0.0
            cols_order_rc = ['Concepto', 'Descripcion'] + IVA_COL_ORDER + individual_other_cols + ['Deducciones', 'Total', 'Cantidad']
            cols_order_rc = [c for c in cols_order_rc if c in res_conc.columns]
            res_conc = res_conc[cols_order_rc]

            n_rc_cols = len(res_conc.columns)
            # startrow=5 -> fila 6
            res_conc.to_excel(writer, sheet_name='Resumen x Concepto', index=False, startrow=5)
            ws4 = writer.sheets['Resumen x Concepto']
        
            ws4.merge_cells(f'A1:{get_column_letter(n_rc_cols)}1')
            ws4['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws4['A1'].font = title_font
            ws4['A1'].fill = title_fill
            ws4['A1'].alignment = center_align

            ws4.merge_cells(f'A2:{get_column_letter(n_rc_cols)}2')
            ws4['A2'] = meta['tipo_reporte'].upper() if meta['tipo_reporte'] else 'RESUMEN POR CONCEPTO'
            ws4['A2'].font = report_type_font
            ws4['A2'].alignment = center_align

            ws4.merge_cells(f'A3:{get_column_letter(n_rc_cols)}3')
            ws4['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws4['A3'].font = Font(bold=True, size=11, color='2F5496')
            ws4['A3'].alignment = center_align

            ws4.merge_cells(f'A4:{get_column_letter(n_rc_cols)}4')
            ws4['A4'] = f'Total: {len(res_conc)} conceptos'
            ws4['A4'].font = Font(italic=True, size=10, color='4472C4')
            ws4['A4'].alignment = center_align

            for col_idx in range(1, n_rc_cols + 1):
                cell = ws4.cell(row=res_header_row, column=col_idx) # reusando res_header_row=6
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            rc_col_list = list(res_conc.columns)
            first_iva_idx_rc = rc_col_list.index(IVA_COL_ORDER[0]) + 1
            ded_idx_rc = rc_col_list.index('Deducciones') + 1
            total_idx_rc = rc_col_list.index('Total') + 1
            first_iva_letter_rc = get_column_letter(first_iva_idx_rc)
            ded_letter_rc = get_column_letter(ded_idx_rc)

            for row in range(res_data_start, len(res_conc) + res_data_start):
                for col_idx in range(1, n_rc_cols + 1):
                    cell = ws4.cell(row=row, column=col_idx)
                    cell.alignment = center_align
                    col_name = rc_col_list[col_idx - 1]

                    if first_iva_idx_rc <= col_idx <= total_idx_rc:
                        if col_name in col_list:
                            v_col = get_column_letter(col_list.index(col_name) + 1)
                            # ROUND para evitar drift de centavos al sumar floats no-representables (ej 6629.05).
                            cell.value = f'=ROUND(SUMIFS(Movimientos!${v_col}${7}:${v_col}${n_mov+7-1}, Movimientos!${mov_conc_col}${7}:${mov_conc_col}${n_mov+7-1}, $A{row}),2)'
                            cell.number_format = money_fmt
                        elif col_name == 'Deducciones':
                            formula_parts = []
                            for dc in deduccion_cols:
                                v_col = get_column_letter(col_list.index(dc) + 1)
                                formula_parts.append(f'SUMIFS(Movimientos!${v_col}${7}:${v_col}${n_mov+7-1}, Movimientos!${mov_conc_col}${7}:${mov_conc_col}${n_mov+7-1}, $A{row})')
                            cell.value = ('=ROUND(' + '+'.join(formula_parts) + ',2)') if formula_parts else 0
                            cell.number_format = money_fmt
                        elif col_name == 'Total':
                            cell.value = f'=SUM({first_iva_letter_rc}{row}:{ded_letter_rc}{row})'
                            cell.number_format = money_fmt
                    elif col_name == 'Cantidad':
                        cell.value = f'=COUNTIFS(Movimientos!${mov_conc_col}${7}:${mov_conc_col}${n_mov+7-1}, $A{row})'
        
            # Fila TOTAL GENERAL
            total_row_rc = len(res_conc) + res_data_start
            ws4.merge_cells(f'A{total_row_rc}:B{total_row_rc}')
            ws4[f'A{total_row_rc}'] = "TOTAL GENERAL"
            ws4[f'A{total_row_rc}'].font = Font(bold=True)
            ws4[f'A{total_row_rc}'].alignment = Alignment(horizontal='right')
        
            for col_idx in range(first_iva_idx_rc, n_rc_cols + 1):
                col_letter = get_column_letter(col_idx)
                cell = ws4.cell(row=total_row_rc, column=col_idx)
                cell.value = f'=SUM({col_letter}{res_data_start}:{col_letter}{total_row_rc-1})'
                cell.font = Font(bold=True)
                cell.border = Border(top=Side(style='double'))
                if col_idx < n_rc_cols:
                    cell.number_format = money_fmt

            _autofit(ws4, n_rc_cols)

            # ── Hoja Resumen por Concepto y Jur. (Pivot para CM05) ──
            # 1. Identificar columnas que forman parte del "Neto" (Base Imponible)
            # Incluimos Netos, Exento, Monotributo y cualquier otro que no sea IVA/PERC/RET/DEDUCC
            cm05_neto_cols = [
                c for c in IVA_COL_ORDER 
                if any(x in c for x in ['Neto', 'Exento', 'Monotributo'])
            ]
            otros_adicionales = [
                c for c in other_cols 
                if not _es_deduccion(c)
            ]
            cm05_neto_cols += otros_adicionales
        
            # 2. Obtener Jurisdicciones y Conceptos únicos
            unique_jurs = sorted([str(j) for j in df['Jur.'].unique() if pd.notna(j) and str(j).strip()])
            if not unique_jurs: unique_jurs = ["S/D"]

            # Mapeo de código de jurisdicción (letra) → nombre de provincia (Mendez)
            JUR_NOMBRES = {
                'A': 'Salta',
                'B': 'Buenos Aires',
                'C': 'Capital Federal',
                'D': 'San Luis',
                'E': 'Entre Ríos',
                'F': 'La Rioja',
                'G': 'Santiago del Estero',
                'H': 'Chaco',
                'J': 'San Juan',
                'K': 'Catamarca',
                'L': 'La Pampa',
                'M': 'Mendoza',
                'N': 'Misiones',
                'P': 'Formosa',
                'Q': 'Neuquén',
                'R': 'Río Negro',
                'S': 'Santa Fe',
                'T': 'Tucumán',
                'U': 'Chubut',
                'V': 'Tierra del Fuego',
                'W': 'Corrientes',
                'X': 'Córdoba',
                'Y': 'Jujuy',
                'Z': 'Santa Cruz',
                '0': 'Exterior',
            }

            def _jur_label(jur_code: str) -> str:
                """Devuelve 'X - Nombre Provincia' o solo 'X' si no se reconoce."""
                nombre = JUR_NOMBRES.get(jur_code.upper())
                return f"{jur_code} - {nombre}" if nombre else jur_code

            conceptos_unicos_df = res_conc[['Concepto', 'Descripcion']].copy()
        
            n_rj_cols = 3 + len(unique_jurs) # Concepto, Desc, Jurs..., Total
            res_jur_sheet_name = 'Resumen x Concepto y Jur.'
            if res_jur_sheet_name in writer.book.sheetnames:
                del writer.book[res_jur_sheet_name]
            
            # startrow=5 -> fila 6
            conceptos_unicos_df.to_excel(writer, sheet_name='Resumen x Concepto y Jur.', index=False, startrow=5)
            ws_rj = writer.sheets['Resumen x Concepto y Jur.']
        
            # Titulos y Estilos
            ws_rj.merge_cells(f'A1:{get_column_letter(n_rj_cols)}1')
            ws_rj['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws_rj['A1'].font = title_font
            ws_rj['A1'].fill = title_fill
            ws_rj['A1'].alignment = center_align

            ws_rj.merge_cells(f'A2:{get_column_letter(n_rj_cols)}2')
            ws_rj['A2'] = meta['tipo_reporte'].upper() if meta['tipo_reporte'] else 'RESUMEN POR CONCEPTO Y JUR.'
            ws_rj['A2'].font = report_type_font
            ws_rj['A2'].alignment = center_align

            ws_rj.merge_cells(f'A3:{get_column_letter(n_rj_cols)}3')
            ws_rj['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws_rj['A3'].font = Font(bold=True, size=11, color='2F5496')
            ws_rj['A3'].alignment = center_align

            ws_rj.merge_cells(f'A4:{get_column_letter(n_rj_cols)}4')
            ws_rj['A4'] = f'Total: {len(conceptos_unicos_df)} conceptos x {len(unique_jurs)} jur.'
            ws_rj['A4'].font = Font(italic=True, size=10, color='4472C4')
            ws_rj['A4'].alignment = center_align

            ws_rj.cell(row=res_header_row, column=1).value = 'Concepto'
            ws_rj.cell(row=res_header_row, column=2).value = 'Descripcion'
            for i, jur in enumerate(unique_jurs):
                cell = ws_rj.cell(row=res_header_row, column=3+i)
                cell.value = _jur_label(jur)
            ws_rj.cell(row=res_header_row, column=3+len(unique_jurs)).value = 'TOTAL'
        
            for col_idx in range(1, n_rj_cols + 1):
                cell = ws_rj.cell(row=res_header_row, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            mov_jur_col = get_column_letter(col_list.index('Jur.') + 1)
            for idx_c, (idx_df, row_data) in enumerate(conceptos_unicos_df.iterrows()):
                curr_row = res_data_start + idx_c
                ws_rj.cell(row=curr_row, column=1).value = row_data['Concepto']
                ws_rj.cell(row=curr_row, column=2).value = row_data['Descripcion']
            
                for j_idx, jur in enumerate(unique_jurs):
                    col_target = 3 + j_idx
                    formula_parts = []
                    for n_col in cm05_neto_cols:
                        v_col_l = get_column_letter(col_list.index(n_col) + 1)
                        formula_parts.append(
                            f'SUMIFS(Movimientos!${v_col_l}${7}:${v_col_l}${n_mov+7-1}, '
                            f'Movimientos!${mov_conc_col}${7}:${mov_conc_col}${n_mov+7-1}, $A{curr_row}, '
                            f'Movimientos!${mov_jur_col}${7}:${mov_jur_col}${n_mov+7-1}, "{jur}")'
                        )
                    # ROUND para evitar drift de centavos al sumar floats no-representables (ej 6629.05).
                    ws_rj.cell(row=curr_row, column=col_target).value = ("=ROUND(" + "+".join(formula_parts) + ",2)") if formula_parts else 0
                    ws_rj.cell(row=curr_row, column=col_target).number_format = money_fmt
                    ws_rj.cell(row=curr_row, column=col_target).alignment = center_align
            
                first_jur_letter = get_column_letter(3)
                last_jur_letter = get_column_letter(3 + len(unique_jurs) - 1)
                total_cell = ws_rj.cell(row=curr_row, column=3 + len(unique_jurs))
                total_cell.value = f'=SUM({first_jur_letter}{curr_row}:{last_jur_letter}{curr_row})'
                total_cell.number_format = money_fmt
                total_cell.alignment = center_align
                total_cell.font = Font(bold=True)

            total_row_rj = res_data_start + len(conceptos_unicos_df)
            ws_rj.merge_cells(f'A{total_row_rj}:B{total_row_rj}')
            ws_rj[f'A{total_row_rj}'] = "TOTAL GENERAL"
            ws_rj[f'A{total_row_rj}'].font = Font(bold=True)
            ws_rj[f'A{total_row_rj}'].alignment = Alignment(horizontal='right')
        
            for col_idx in range(3, n_rj_cols + 1):
                col_l = get_column_letter(col_idx)
                cell = ws_rj.cell(row=total_row_rj, column=col_idx)
                cell.value = f'=SUM({col_l}{res_data_start}:{col_l}{total_row_rj-1})'
                cell.font = Font(bold=True)
                cell.border = Border(top=Side(style='double'))
                cell.number_format = money_fmt
                cell.alignment = center_align

            _autofit(ws_rj, n_rj_cols)

            # ── Hoja Resumen por Proveedor (agrupado por CUIT) ────
            res = resumen.groupby('CUIT').agg(
                Proveedor=('Proveedor', 'first'),
                **{c: (c, 'sum') for c in IVA_COL_ORDER},
                **{c: (c, 'sum') for c in individual_other_cols},
                Deducciones=('Total', 'count'), # placeholder
                Cantidad=('Total', 'count'),
            ).reset_index()

            res['Total'] = 0.0
            cols_order = ['CUIT', 'Proveedor'] + IVA_COL_ORDER + individual_other_cols + ['Deducciones', 'Total', 'Cantidad']
            cols_order = [c for c in cols_order if c in res.columns]
            res = res[cols_order]
            # Sort logic
            res['_sort'] = resumen.groupby('CUIT')[sum_cols].sum().sum(axis=1).values
            res = res.sort_values('_sort', ascending=False).drop(columns='_sort')

            res.to_excel(writer, sheet_name='Resumen x Proveedor', index=False, startrow=5)
            ws2 = writer.sheets['Resumen x Proveedor']
            n_res_cols = len(res.columns)

            ws2.merge_cells(f'A1:{get_column_letter(n_res_cols)}1')
            ws2['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws2['A1'].font = title_font; ws2['A1'].fill = title_fill; ws2['A1'].alignment = center_align

            ws2.merge_cells(f'A2:{get_column_letter(n_res_cols)}2')
            ws2['A2'] = meta['tipo_reporte'].upper() if meta['tipo_reporte'] else 'RESUMEN POR PROVEEDOR'
            ws2['A2'].font = report_type_font; ws2['A2'].alignment = center_align
        
            ws2.merge_cells(f'A3:{get_column_letter(n_res_cols)}3')
            ws2['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws2['A3'].font = Font(bold=True, size=11, color='2F5496'); ws2['A3'].alignment = center_align

            ws2.merge_cells(f'A4:{get_column_letter(n_res_cols)}4')
            ws2['A4'] = f'Total: {len(res)} proveedores'
            ws2['A4'].font = Font(italic=True, size=10, color='4472C4'); ws2['A4'].alignment = center_align
        
            for col_idx in range(1, n_res_cols + 1):
                cell = ws2.cell(row=res_header_row, column=col_idx)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin_border

            res_col_list = list(res.columns)
            first_iva_idx_res = res_col_list.index(IVA_COL_ORDER[0]) + 1
            ded_idx_res = res_col_list.index('Deducciones') + 1
            total_idx_res = res_col_list.index('Total') + 1
            first_iva_letter_res = get_column_letter(first_iva_idx_res)
            ded_letter_res = get_column_letter(ded_idx_res)
        
            for row in range(res_data_start, len(res) + res_data_start):
                for col_idx in range(1, n_res_cols + 1):
                    cell = ws2.cell(row=row, column=col_idx)
                    cell.alignment = center_align
                    col_name = res_col_list[col_idx - 1]
                
                    if first_iva_idx_res <= col_idx <= total_idx_res:
                        if col_name in col_list:
                            v_col = get_column_letter(col_list.index(col_name) + 1)
                            # ROUND para evitar drift de centavos al sumar floats no-representables (ej 6629.05).
                            cell.value = f'=ROUND(SUMIFS(Movimientos!${v_col}${7}:${v_col}${n_mov+7-1}, Movimientos!${mov_cuit_col}${7}:${mov_cuit_col}${n_mov+7-1}, $A{row}),2)'
                            cell.number_format = money_fmt
                        elif col_name == 'Deducciones':
                            formula_parts = []
                            for dc in deduccion_cols:
                                v_col = get_column_letter(col_list.index(dc) + 1)
                                formula_parts.append(f'SUMIFS(Movimientos!${v_col}${7}:${v_col}${n_mov+7-1}, Movimientos!${mov_cuit_col}${7}:${mov_cuit_col}${n_mov+7-1}, $A{row})')
                            cell.value = ('=ROUND(' + '+'.join(formula_parts) + ',2)') if formula_parts else 0
                            cell.number_format = money_fmt
                        elif col_name == 'Total':
                            cell.value = f'=SUM({first_iva_letter_res}{row}:{ded_letter_res}{row})'
                            cell.number_format = money_fmt
                    elif col_name == 'Cantidad':
                        cell.value = f'=COUNTIFS(Movimientos!${mov_cuit_col}${7}:${mov_cuit_col}${n_mov+7-1}, $A{row})'
                
                    # Formato CUIT como texto

            total_row = len(res) + res_data_start
            ws2.merge_cells(f'A{total_row}:B{total_row}')
            ws2[f'A{total_row}'] = "TOTAL GENERAL"
            ws2[f'A{total_row}'].font = Font(bold=True); ws2[f'A{total_row}'].alignment = Alignment(horizontal='right')
        
            for col_idx in range(first_iva_idx_res, n_res_cols + 1):
                col_letter = get_column_letter(col_idx)
                cell = ws2.cell(row=total_row, column=col_idx)
                cell.value = f'=SUM({col_letter}{res_data_start}:{col_letter}{total_row-1})'
                cell.font = Font(bold=True); cell.border = Border(top=Side(style='double'))
                if col_idx < n_res_cols: cell.number_format = money_fmt
            
            _autofit(ws2, n_res_cols)

            # ── Hoja Mayor x Proveedor ────────────────────────────
            df_with_idx = df.copy()
            # Original data rows in Movimientos start at Row 7
            df_with_idx['_orig_row'] = range(7, len(df) + 7)
        
            mayor = df_with_idx.sort_values(['CUIT', 'Fecha'])
        
            def format_comp(r):
                pv_s = f"{r['PV']:05d}"
                nro_s = f"{r['Nro.']:08d}" if isinstance(r['Nro.'], int) else str(r['Nro.'])
                return f"{pv_s}-{nro_s}{r['Letra']}"
            
            mayor['Comp.'] = mayor.apply(format_comp, axis=1)
            mayor['Saldo Acumulado'] = mayor.groupby('CUIT')['Total'].cumsum()
        
            cols_mayor = ['CUIT', 'Proveedor', 'Fecha', 'Tipo', 'Comp.', 'Concepto', 'Total', 'Saldo Acumulado', '_orig_row']
            mayor = mayor[cols_mayor]
        
            n_mayor_cols = len(mayor.columns) - 1
            # startrow=5 -> fila 6
            mayor.to_excel(writer, sheet_name='Mayor x Proveedor', index=False, startrow=5)
            ws5 = writer.sheets['Mayor x Proveedor']

            ws5.merge_cells(f'A1:{get_column_letter(n_mayor_cols)}1')
            ws5['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws5['A1'].font = title_font
            ws5['A1'].fill = title_fill
            ws5['A1'].alignment = center_align

            ws5.merge_cells(f'A2:{get_column_letter(n_mayor_cols)}2')
            ws5['A2'] = meta['tipo_reporte'].upper() if meta['tipo_reporte'] else 'MAYOR AUXILIAR'
            ws5['A2'].font = report_type_font
            ws5['A2'].alignment = center_align

            ws5.merge_cells(f'A3:{get_column_letter(n_mayor_cols)}3')
            ws5['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws5['A3'].font = Font(bold=True, size=11, color='2F5496')
            ws5['A3'].alignment = center_align

            ws5.merge_cells(f'A4:{get_column_letter(n_mayor_cols)}4')
            ws5['A4'] = f'Total: {len(mayor)} movimientos'
            ws5['A4'].font = Font(italic=True, size=10, color='4472C4')
            ws5['A4'].alignment = center_align
        
            total_mov_col = get_column_letter(col_list.index('Total') + 1)

            for col_idx in range(1, n_mayor_cols + 1):
                cell = ws5.cell(row=res_header_row, column=col_idx) # reusando res_header_row=6
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            for row_idx in range(res_data_start, len(mayor) + res_data_start): # res_data_start=7
                orig_row_idx = row_idx - res_data_start
                orig_row = mayor.iloc[orig_row_idx]['_orig_row']
                for col_idx in range(1, n_mayor_cols + 1):
                    cell = ws5.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_align
                    if col_idx == 7: # Total
                        cell.value = f'={mov_sheet_name}!{total_mov_col}{orig_row}'
                        cell.number_format = money_fmt
                    elif col_idx == 8: # Saldo Acumulado
                        # Formula unificada: IF(mismo CUIT que arriba, saldo_ant + total_actual, total_actual)
                        if row_idx == res_data_start:
                            cell.value = f'=G{row_idx}'
                        else:
                            cell.value = f'=IF(A{row_idx}=A{row_idx-1}, H{row_idx-1}+G{row_idx}, G{row_idx})'
                        cell.number_format = money_fmt
        
            ws5.delete_cols(n_mayor_cols + 1) # Borrar columna auxiliar
            _autofit(ws5, n_mayor_cols)

        # ── Hoja Asiento Contable ────────────────────────────────
        if con_asiento:
            es_ventas = 'VENTA' in (meta.get('tipo_reporte') or '').upper()

            # --- Columnas de neto (base imponible) y de IVA ---
            _as_neto_base_cols = [
                c for c in IVA_COL_ORDER
                if c.startswith('Neto') or c in ('Exento', 'Monotributo')
            ]
            _as_iva_cols = [c for c in IVA_COL_ORDER if c.startswith('IVA')]
            _as_deducc_cols  = [c for c in other_cols if _es_deduccion(c)]
            _as_individ_cols = [c for c in other_cols if not _es_deduccion(c)]
            _as_all_neto_cols = _as_neto_base_cols + _as_individ_cols

            # --- A. Neto por Concepto ---
            if _as_all_neto_cols:
                _conc_grp = df.groupby('Concepto')[_as_all_neto_cols].sum()
                _conc_grp['_neto'] = _conc_grp.sum(axis=1)
                _conc_total = _conc_grp[['_neto']].reset_index()
                _conc_total = _conc_total[_conc_total['_neto'] != 0].copy()
                _conc_total['Desc'] = _conc_total['Concepto'].apply(
                    lambda x: CONCEPTOS_MAP.get(str(x), f'Concepto {x}').replace('°', 'o.').upper()
                )
                _conc_total['_num'] = pd.to_numeric(_conc_total['Concepto'], errors='coerce')
                _conc_total = _conc_total.sort_values('_num').drop(columns='_num')
            else:
                _conc_total = pd.DataFrame(columns=['Concepto', '_neto', 'Desc'])

            # --- B. IVA total ---
            _iva_total = float(df[_as_iva_cols].sum().sum()) if _as_iva_cols else 0.0

            # --- C. Deducciones en orden (solo las que tienen valor != 0) ---
            _deducc_items = []
            for _c in _as_deducc_cols:
                _v = float(df[_c].sum())
                if _v != 0:
                    _deducc_items.append((_c, _v))

            # --- D. IVA de NCs (para asiento de Restitución) ---
            # NCs ya están negadas en df → abs() recupera el valor positivo
            _nc_iva_total = (
                abs(float(df.loc[df['Tipo'] == 'NC', _as_iva_cols].sum().sum()))
                if _as_iva_cols else 0.0
            )

            # --- Crear hoja con encabezado común (compras y ventas) ---
            ws_as = writer.book.create_sheet('Asiento Contable')

            # Filas 1-3: encabezado (razón social / tipo + título / CUIT + periodo)
            ws_as.merge_cells('A1:C1')
            ws_as['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws_as['A1'].font = title_font
            ws_as['A1'].fill = title_fill
            ws_as['A1'].alignment = center_align

            ws_as.merge_cells('A2:C2')
            _tr_label = (
                f"{meta['tipo_reporte'].upper()} - ASIENTO CONTABLE"
                if meta['tipo_reporte'] else 'ASIENTO CONTABLE'
            )
            ws_as['A2'] = _tr_label
            ws_as['A2'].font = report_type_font
            ws_as['A2'].alignment = center_align

            ws_as.merge_cells('A3:C3')
            ws_as['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws_as['A3'].font = Font(bold=True, size=11, color='2F5496')
            ws_as['A3'].alignment = center_align

            # Headers de columna en fila 5 (fila 4 separadora)
            _as_header_row  = 5
            _as_data_start  = 6
            _as_header_labels = ['DESCRIPCIÓN', 'DEBE', 'HABER']
            _as_col_fills = [
                header_fill,                              # Descripción → azul
                PatternFill('solid', fgColor='2F5496'),   # DEBE → azul oscuro
                PatternFill('solid', fgColor='375623'),   # HABER → verde oscuro
            ]
            for _ci, (_h, _hf) in enumerate(zip(_as_header_labels, _as_col_fills), 1):
                _cell = ws_as.cell(row=_as_header_row, column=_ci)
                _cell.value = _h
                _cell.font = header_font; _cell.fill = _hf
                _cell.alignment = header_align; _cell.border = thin_border

            _curr = _as_data_start
            _desc_font = Font(bold=True, size=10)

            if es_ventas:
                # ═══════════════════════════════════════════════════════
                # Asiento de VENTAS (FC − NC neteado)
                # ═══════════════════════════════════════════════════════

                # ── DEUDORES POR VENTAS → DEBE (col B), va primero ──
                _deud_row = _curr
                ws_as.cell(row=_deud_row, column=1).value = 'DEUDORES POR VENTAS'
                ws_as.cell(row=_deud_row, column=1).font = Font(bold=True, size=11, color='1F3864')
                _du = ws_as.cell(row=_deud_row, column=2)
                _du.number_format = money_fmt; _du.alignment = center_align
                _du.font = Font(bold=True, size=11)
                _curr += 1

                _haber_first = _curr

                # ── Filas de Concepto (Neto) → HABER (col C) ──
                for _, _rd in _conc_total.iterrows():
                    _desc_text = _rd['Desc'] or f"CONCEPTO {_rd['Concepto']}"
                    ws_as.cell(row=_curr, column=1).value = f"A {_desc_text}"
                    ws_as.cell(row=_curr, column=1).font = _desc_font
                    _hc = ws_as.cell(row=_curr, column=3)
                    _hc.value = float(_rd['_neto']); _hc.number_format = money_fmt; _hc.alignment = center_align
                    _curr += 1

                # ── Fila IVA DEBITO → HABER ──
                if _iva_total != 0:
                    ws_as.cell(row=_curr, column=1).value = 'A IVA DEBITO'
                    ws_as.cell(row=_curr, column=1).font = _desc_font
                    _ic = ws_as.cell(row=_curr, column=3)
                    _ic.value = _iva_total; _ic.number_format = money_fmt; _ic.alignment = center_align
                    _curr += 1

                # ── Percepciones IIBB efectuadas → HABER (RET ignorado en ventas) ──
                for _dn, _dv in _deducc_items:
                    _nu = _dn.upper()
                    if 'RET' in _nu and 'SIRCREB' not in _nu and 'SIRTAC' not in _nu:
                        continue  # las RET no se emiten en el asiento de ventas
                    ws_as.cell(row=_curr, column=1).value = f"A {_dn}"
                    ws_as.cell(row=_curr, column=1).font = _desc_font
                    _pc = ws_as.cell(row=_curr, column=3)
                    _pc.value = _dv; _pc.number_format = money_fmt; _pc.alignment = center_align
                    _curr += 1

                _haber_last = _curr - 1

                # Setear la fórmula de DEUDORES ahora que conocemos el rango HABER
                if _haber_last >= _haber_first:
                    _du.value = f'=SUM(C{_haber_first}:C{_haber_last})'
                    # Doble subrayado de cierre en la última fila HABER
                    _last_haber_cell = ws_as.cell(row=_haber_last, column=3)
                    _last_haber_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
                else:
                    _du.value = 0

            else:
                # ═══════════════════════════════════════════════════════
                # Asiento de COMPRAS (lógica original)
                # ═══════════════════════════════════════════════════════
                # Identificar cuáles deducciones son RET fiscales
                # SIRCREB, SIRTAC y retenciones bancarias (BCO/BANCO) quedan fuera:
                # no son retenciones que vayan a DEUDORES POR VENTAS
                _ret_fiscal_set = {
                    n for n, _ in _deducc_items
                    if 'RET' in n.upper()
                    and 'SIRCREB' not in n.upper()
                    and 'SIRTAC' not in n.upper()
                    and 'BCO' not in n.upper()
                    and 'BANCO' not in n.upper()
                }

                # ── Filas de Concepto (Neto) ──
                for _, _rd in _conc_total.iterrows():
                    ws_as.cell(row=_curr, column=1).value = _rd['Desc'] or f"CONCEPTO {_rd['Concepto']}"
                    ws_as.cell(row=_curr, column=1).font = _desc_font
                    _dc = ws_as.cell(row=_curr, column=2)
                    _dc.value = float(_rd['_neto']); _dc.number_format = money_fmt; _dc.alignment = center_align
                    _curr += 1

                # ── Fila IVA ──
                if _iva_total != 0:
                    ws_as.cell(row=_curr, column=1).value = 'IVA'
                    ws_as.cell(row=_curr, column=1).font = _desc_font
                    _ic = ws_as.cell(row=_curr, column=2)
                    _ic.value = _iva_total; _ic.number_format = money_fmt; _ic.alignment = center_align
                    _curr += 1

                # ── Filas de Deducciones (rastrear filas de RET fiscales) ──
                _ret_fiscal_rows = []
                for _dn, _dv in _deducc_items:
                    ws_as.cell(row=_curr, column=1).value = _dn
                    ws_as.cell(row=_curr, column=1).font = _desc_font
                    _dedcc = ws_as.cell(row=_curr, column=2)
                    _dedcc.value = _dv; _dedcc.number_format = money_fmt; _dedcc.alignment = center_align
                    if _dn in _ret_fiscal_set:
                        _ret_fiscal_rows.append(_curr)
                    _curr += 1

                _last_debe_row = _curr - 1

                # ── a PROVEEDORES (col C) = SUM(DEBE) - DEUDORES ── va primero
                _prov_row = _curr
                _pc = ws_as.cell(row=_prov_row, column=1)
                _pc.value = 'a PROVEEDORES'
                _pc.font = Font(bold=True, size=11, color='1F3864')
                _deud_row = _prov_row + 1  # DEUDORES siempre está en la fila siguiente
                _ph = ws_as.cell(row=_prov_row, column=3)
                _ph.value = f'=SUM(B{_as_data_start}:B{_last_debe_row})-C{_deud_row}'
                _ph.number_format = money_fmt; _ph.alignment = center_align
                _ph.font = Font(bold=True, size=11)
                _ph.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
                _curr += 1

                # ── a DEUDORES POR VENTAS (col C) — fórmula de filas B específicas ──
                _dd = ws_as.cell(row=_deud_row, column=1)
                _dd.value = 'a DEUDORES POR VENTAS'
                _dd.font = Font(bold=True, size=11, color='1F3864')
                _dh = ws_as.cell(row=_deud_row, column=3)
                if _ret_fiscal_rows:
                    _dh.value = '=' + '+'.join(f'B{r}' for r in _ret_fiscal_rows)
                else:
                    _dh.value = 0
                _dh.number_format = money_fmt; _dh.alignment = center_align
                _dh.font = Font(bold=True, size=11)
                _dh.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
                _curr += 1

            # ═══════════════════════════════════════════════════════
            # Asiento de Restitución (común a compras y ventas)
            # Sólo si hay IVA de NCs en el período
            # ═══════════════════════════════════════════════════════
            if _nc_iva_total != 0:
                # Fila separadora en blanco
                _curr += 1

                if es_ventas:
                    _titulo      = 'RESTITUCION DE DEBITO'
                    _label_debe  = 'CREDITO FISCAL IVA'
                    _label_haber = 'A DEBITO FISCAL IVA'
                else:
                    _titulo      = 'RESTITUCION DE CREDITO'
                    _label_debe  = 'DEBITO FISCAL IVA'
                    _label_haber = 'A CREDITO FISCAL IVA'

                # Título de sección
                ws_as.cell(row=_curr, column=1).value = _titulo
                ws_as.cell(row=_curr, column=1).font = Font(bold=True, size=11, color='C00000')
                _curr += 1

                # Fila DEBE
                ws_as.cell(row=_curr, column=1).value = _label_debe
                ws_as.cell(row=_curr, column=1).font = _desc_font
                _rdc = ws_as.cell(row=_curr, column=2)
                _rdc.value = _nc_iva_total
                _rdc.number_format = money_fmt
                _rdc.alignment = center_align
                _restit_row = _curr
                _curr += 1

                # Fila HABER (= al DEBE para que se mantengan acoplados)
                ws_as.cell(row=_curr, column=1).value = _label_haber
                ws_as.cell(row=_curr, column=1).font = Font(bold=True, size=11, color='1F3864')
                _rhc = ws_as.cell(row=_curr, column=3)
                _rhc.value = f'=B{_restit_row}'
                _rhc.number_format = money_fmt
                _rhc.alignment = center_align
                _rhc.font = Font(bold=True, size=11)
                _rhc.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

            _autofit(ws_as, 3, start_row=_as_header_row)
            ws_as.column_dimensions['A'].width = 38
            ws_as.column_dimensions['B'].width = 20
            ws_as.column_dimensions['C'].width = 20



        # ── Hoja ARCA (datos del CSV de ARCA) ──────────────────
        if cruce_arca and df_arca is not None and not df_arca.empty:
            df_arca.to_excel(writer, sheet_name='ARCA', index=False, startrow=5)
            ws_arca = writer.sheets['ARCA']
            n_arca_cols = len(df_arca.columns)

            ws_arca.merge_cells(f'A1:{get_column_letter(n_arca_cols)}1')
            ws_arca['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
            ws_arca['A1'].font = title_font; ws_arca['A1'].fill = title_fill; ws_arca['A1'].alignment = center_align

            ws_arca.merge_cells(f'A2:{get_column_letter(n_arca_cols)}2')
            ws_arca['A2'] = f"{meta['tipo_reporte'].upper()} - COMPROBANTES ARCA"
            ws_arca['A2'].font = report_type_font; ws_arca['A2'].alignment = center_align

            ws_arca.merge_cells(f'A3:{get_column_letter(n_arca_cols)}3')
            ws_arca['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
            ws_arca['A3'].font = Font(bold=True, size=11, color='2F5496'); ws_arca['A3'].alignment = center_align

            ws_arca.merge_cells(f'A4:{get_column_letter(n_arca_cols)}4')
            ws_arca['A4'] = f'Total: {len(df_arca)} comprobantes'
            ws_arca['A4'].font = Font(italic=True, size=10, color='4472C4'); ws_arca['A4'].alignment = center_align

            for col_idx in range(1, n_arca_cols + 1):
                cell = ws_arca.cell(row=6, column=col_idx)
                cell.font = header_font; cell.fill = header_fill
                cell.alignment = header_align; cell.border = thin_border

            # Identificar columnas monetarias
            arca_col_list_final = list(df_arca.columns)
            non_money = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'Tipo Doc.', 'CUIT', 'Razon Social', 'Auxiliar'}
            arca_money_indices = []
            for ci, cn in enumerate(arca_col_list_final):
                if cn not in non_money and df_arca[cn].dtype in ('float64', 'int64', 'float32', 'int32'):
                    arca_money_indices.append(ci + 1)

            for row_idx in range(7, len(df_arca) + 7):
                for col_idx in range(1, n_arca_cols + 1):
                    cell = ws_arca.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_align
                    if col_idx in arca_money_indices:
                        cell.number_format = accounting_fmt
                if (row_idx - 7) % 2 == 0:
                    for col_idx in range(1, n_arca_cols + 1):
                        ws_arca.cell(row=row_idx, column=col_idx).fill = zebra_fill

            # ── CRUCE + DIFF en ARCA (busca en SISTEMA) ──────────────
            if 'Auxiliar' in arca_col_list_final and 'Total' in arca_col_list_final and 'Auxiliar' in col_list:
                arca_cruce_col_idx = n_arca_cols + 1
                arca_diff_col_idx = n_arca_cols + 2
                arca_cruce_letter = get_column_letter(arca_cruce_col_idx)
                arca_diff_letter = get_column_letter(arca_diff_col_idx)
                arca_aux_col_idx = arca_col_list_final.index('Auxiliar') + 1
                arca_aux_letter = get_column_letter(arca_aux_col_idx)
                arca_total_col_idx = arca_col_list_final.index('Total') + 1
                arca_total_letter = get_column_letter(arca_total_col_idx)

                # Rango de lookup en SISTEMA
                sys_aux_letter = get_column_letter(col_list.index('Auxiliar') + 1)
                sys_total_letter = get_column_letter(col_list.index('Total') + 1)
                sys_total_offset = col_list.index('Total') - col_list.index('Auxiliar') + 1
                sys_last_data_row = len(df) + 6

                # Headers CRUCE
                ws_arca.cell(row=6, column=arca_cruce_col_idx).value = 'CRUCE'
                ws_arca.cell(row=6, column=arca_cruce_col_idx).font = header_font
                ws_arca.cell(row=6, column=arca_cruce_col_idx).fill = PatternFill('solid', fgColor='7030A0')
                ws_arca.cell(row=6, column=arca_cruce_col_idx).alignment = header_align
                ws_arca.cell(row=6, column=arca_cruce_col_idx).border = thin_border

                ws_arca.cell(row=6, column=arca_diff_col_idx).value = 'DIFF'
                ws_arca.cell(row=6, column=arca_diff_col_idx).font = header_font
                ws_arca.cell(row=6, column=arca_diff_col_idx).fill = PatternFill('solid', fgColor='7030A0')
                ws_arca.cell(row=6, column=arca_diff_col_idx).alignment = header_align
                ws_arca.cell(row=6, column=arca_diff_col_idx).border = thin_border

                for row_idx in range(7, len(df_arca) + 7):
                    ws_arca.cell(row=row_idx, column=arca_cruce_col_idx).value = (
                        f'=IFERROR(VLOOKUP({arca_aux_letter}{row_idx},'
                        f"SISTEMA!${sys_aux_letter}$7:${sys_total_letter}${sys_last_data_row},"
                        f'{sys_total_offset},FALSE),"NO ENCONTRADO")'
                    )
                    ws_arca.cell(row=row_idx, column=arca_cruce_col_idx).number_format = accounting_fmt
                    ws_arca.cell(row=row_idx, column=arca_cruce_col_idx).alignment = center_align

                    ws_arca.cell(row=row_idx, column=arca_diff_col_idx).value = (
                        f'=IF({arca_cruce_letter}{row_idx}="NO ENCONTRADO","",'
                        f'{arca_total_letter}{row_idx}-{arca_cruce_letter}{row_idx})'
                    )
                    ws_arca.cell(row=row_idx, column=arca_diff_col_idx).number_format = accounting_fmt
                    ws_arca.cell(row=row_idx, column=arca_diff_col_idx).alignment = center_align

            _autofit(ws_arca, n_arca_cols + 2)

            # ── Hojas de overflow: DE MAS EN SISTEMA / FALTANTES ARCA ─────
            # Construir sets de auxiliares para comparar
            if 'Auxiliar' in arca_col_list_final and 'Auxiliar' in col_list:
                # Auxiliar de ARCA: valores del df
                arca_aux_set = set(df_arca['Auxiliar'].dropna().astype(str).values)
                # Auxiliar de SISTEMA: construir igual que la formula
                sistema_aux_values = (
                    df['Tipo'].astype(str) + ' ' + df['Letra'].astype(str) +
                    df['PV'].astype(str) + df['Nro.'].astype(str) + df['CUIT'].astype(str)
                )
                sistema_aux_set = set(sistema_aux_values.values)

                # DE MAS EN SISTEMA: filas del SISTEMA no encontradas en ARCA
                mask_extra_sistema = ~sistema_aux_values.isin(arca_aux_set)
                df_extra_sistema = df[mask_extra_sistema].copy()
                if 'Auxiliar' in df_extra_sistema.columns:
                    df_extra_sistema = df_extra_sistema.drop(columns=['Auxiliar'])
                if not df_extra_sistema.empty:
                    df_extra_sistema.to_excel(writer, sheet_name='DE MAS EN SISTEMA', index=False, startrow=5)
                    ws_extra = writer.sheets['DE MAS EN SISTEMA']
                    n_extra_cols = len(df_extra_sistema.columns)
                    ws_extra.merge_cells(f'A1:{get_column_letter(n_extra_cols)}1')
                    ws_extra['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
                    ws_extra['A1'].font = title_font; ws_extra['A1'].fill = title_fill; ws_extra['A1'].alignment = center_align
                    ws_extra.merge_cells(f'A2:{get_column_letter(n_extra_cols)}2')
                    ws_extra['A2'] = 'DE MAS EN SISTEMA'
                    ws_extra['A2'].font = Font(bold=True, size=14, color='FFFFFF')
                    ws_extra['A2'].fill = PatternFill('solid', fgColor='C00000')
                    ws_extra['A2'].alignment = center_align
                    ws_extra.merge_cells(f'A3:{get_column_letter(n_extra_cols)}3')
                    ws_extra['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
                    ws_extra['A3'].font = Font(bold=True, size=11, color='2F5496'); ws_extra['A3'].alignment = center_align
                    ws_extra.merge_cells(f'A4:{get_column_letter(n_extra_cols)}4')
                    ws_extra['A4'] = f'{len(df_extra_sistema)} comprobantes en SISTEMA no encontrados en ARCA'
                    ws_extra['A4'].font = Font(italic=True, size=10, color='C00000')
                    ws_extra['A4'].alignment = center_align
                    for ci in range(1, n_extra_cols + 1):
                        c = ws_extra.cell(row=6, column=ci)
                        c.font = header_font; c.fill = header_fill
                        c.alignment = header_align; c.border = thin_border
                    # Aplicar formato contabilidad Peso a columnas numéricas
                    extra_col_list = list(df_extra_sistema.columns)
                    extra_non_money = {'Fecha', 'Tipo', 'PV', 'Nro.', 'Letra', 'Proveedor', 'Cond. IVA', 'CUIT', 'Concepto', 'Jur.'}
                    # Calcular rango de SUM para la columna Total (misma lógica que Movimientos)
                    extra_total_col_idx = extra_col_list.index('Total') + 1 if 'Total' in extra_col_list else None
                    extra_first_sum = None
                    extra_last_sum = None
                    if extra_total_col_idx:
                        # Buscar primera columna IVA presente
                        for iva_c in IVA_COL_ORDER:
                            if iva_c in extra_col_list:
                                extra_first_sum = get_column_letter(extra_col_list.index(iva_c) + 1)
                                break
                        # Buscar última columna antes de Total (other_cols o última IVA)
                        if other_cols:
                            for oc in reversed(other_cols):
                                if oc in extra_col_list:
                                    extra_last_sum = get_column_letter(extra_col_list.index(oc) + 1)
                                    break
                        if not extra_last_sum:
                            for iva_c in reversed(IVA_COL_ORDER):
                                if iva_c in extra_col_list:
                                    extra_last_sum = get_column_letter(extra_col_list.index(iva_c) + 1)
                                    break
                    for row_idx in range(7, len(df_extra_sistema) + 7):
                        for ci, cn in enumerate(extra_col_list):
                            cell = ws_extra.cell(row=row_idx, column=ci + 1)
                            cell.alignment = center_align
                            if cn not in extra_non_money:
                                cell.number_format = accounting_fmt
                        # Formula SUM en columna Total
                        if extra_total_col_idx and extra_first_sum and extra_last_sum:
                            ws_extra.cell(row=row_idx, column=extra_total_col_idx).value = f'=SUM({extra_first_sum}{row_idx}:{extra_last_sum}{row_idx})'
                            ws_extra.cell(row=row_idx, column=extra_total_col_idx).number_format = accounting_fmt

                    _autofit(ws_extra, n_extra_cols)

                # FALTANTES ARCA: filas de ARCA no encontradas en SISTEMA
                mask_falt_arca = ~df_arca['Auxiliar'].astype(str).isin(sistema_aux_set)
                df_falt_arca = df_arca[mask_falt_arca].copy()
                if 'Auxiliar' in df_falt_arca.columns:
                    df_falt_arca = df_falt_arca.drop(columns=['Auxiliar'])
                if not df_falt_arca.empty:
                    df_falt_arca.to_excel(writer, sheet_name='FALTANTES ARCA', index=False, startrow=5)
                    ws_falt = writer.sheets['FALTANTES ARCA']
                    n_falt_cols = len(df_falt_arca.columns)
                    ws_falt.merge_cells(f'A1:{get_column_letter(n_falt_cols)}1')
                    ws_falt['A1'] = meta['razon_social'].upper() if meta['razon_social'] else 'CONTRIBUYENTE'
                    ws_falt['A1'].font = title_font; ws_falt['A1'].fill = title_fill; ws_falt['A1'].alignment = center_align
                    ws_falt.merge_cells(f'A2:{get_column_letter(n_falt_cols)}2')
                    ws_falt['A2'] = f"Compras Faltantes ({meta['periodo']})"
                    ws_falt['A2'].font = Font(bold=True, size=14, color='FFFFFF')
                    ws_falt['A2'].fill = PatternFill('solid', fgColor='C00000')
                    ws_falt['A2'].alignment = center_align
                    ws_falt.merge_cells(f'A3:{get_column_letter(n_falt_cols)}3')
                    ws_falt['A3'] = f"CUIT: {meta['cuit_empresa']} | Periodo: {meta['periodo']}"
                    ws_falt['A3'].font = Font(bold=True, size=11, color='2F5496'); ws_falt['A3'].alignment = center_align
                    ws_falt.merge_cells(f'A4:{get_column_letter(n_falt_cols)}4')
                    ws_falt['A4'] = f'{len(df_falt_arca)} comprobantes en ARCA no encontrados en SISTEMA'
                    ws_falt['A4'].font = Font(italic=True, size=10, color='C00000')
                    ws_falt['A4'].alignment = center_align
                    for ci in range(1, n_falt_cols + 1):
                        c = ws_falt.cell(row=6, column=ci)
                        c.font = header_font; c.fill = header_fill
                        c.alignment = header_align; c.border = thin_border
                    # Aplicar formato contabilidad Peso a columnas numéricas
                    falt_col_list = list(df_falt_arca.columns)
                    falt_non_money = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'Tipo Doc.', 'CUIT', 'Razon Social'}
                    for row_idx in range(7, len(df_falt_arca) + 7):
                        for ci, cn in enumerate(falt_col_list):
                            cell = ws_falt.cell(row=row_idx, column=ci + 1)
                            cell.alignment = center_align
                            if cn not in falt_non_money:
                                cell.number_format = accounting_fmt
                    _autofit(ws_falt, n_falt_cols)

    print(f"\n  Excel guardado en: {output_path}")


def crear_excel_consolidado_simple(transacciones: list[dict], meta: dict, output_path):
    """Version rapida (xlsxwriter) del consolidado: una sola hoja Movimientos,
    sin resumenes, sin Auxiliar, sin cruce ARCA. ~3-5x mas veloz que crear_excel
    para datasets grandes (>10k transacciones).

    output_path puede ser una ruta o un BytesIO.
    """
    import xlsxwriter
    from xlsxwriter.utility import xl_col_to_name

    rows, IVA_COL_ORDER, other_cols, _ = _construir_filas_consolidado(
        transacciones, meta, con_auxiliar=False, cruce_arca=False
    )
    if rows is None:
        return

    # Orden final de columnas
    base_cols = ['Fecha', 'Tipo', 'PV', 'Nro.', 'Letra', 'Proveedor',
                 'Cond. IVA', 'CUIT', 'Concepto', 'Jur.']
    money_cols = IVA_COL_ORDER + other_cols + ['Total']
    cols = base_cols + IVA_COL_ORDER + other_cols + ['Total']
    n_cols = len(cols)

    iva_set = set(IVA_COL_ORDER)
    other_set = set(other_cols)
    deduccion_set = {c for c in other_cols if _es_deduccion(c)}
    money_idx_set = {cols.index(c) for c in money_cols}
    total_col_idx = cols.index('Total')
    first_money_idx = cols.index(IVA_COL_ORDER[0]) if IVA_COL_ORDER else cols.index(other_cols[0])
    last_sum_idx = cols.index(other_cols[-1]) if other_cols else cols.index(IVA_COL_ORDER[-1])
    first_sum_letter = xl_col_to_name(first_money_idx)
    last_sum_letter = xl_col_to_name(last_sum_idx)

    # ── Workbook (in_memory evita archivos temporales en disco) ──
    wb = xlsxwriter.Workbook(output_path, {'in_memory': True, 'strings_to_formulas': False})
    ws = wb.add_worksheet('Movimientos')

    # ── Formatos: una sola instancia compartida por todas las celdas que matchean ──
    title_fmt = wb.add_format({
        'bold': True, 'font_size': 14, 'font_color': '#FFFFFF',
        'bg_color': '#2F5496', 'align': 'center', 'valign': 'vcenter'
    })
    subtitle_red_fmt = wb.add_format({
        'bold': True, 'font_size': 12, 'font_color': '#C00000',
        'align': 'center', 'valign': 'vcenter'
    })
    subtitle_blue_fmt = wb.add_format({
        'bold': True, 'font_size': 11, 'font_color': '#2F5496',
        'align': 'center', 'valign': 'vcenter'
    })
    subtitle_italic_fmt = wb.add_format({
        'italic': True, 'font_size': 10, 'font_color': '#4472C4',
        'align': 'center', 'valign': 'vcenter'
    })

    header_base = {
        'bold': True, 'font_size': 10, 'font_color': '#FFFFFF',
        'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1,
    }
    header_default_fmt = wb.add_format({**header_base, 'bg_color': '#4472C4'})
    header_iva_fmt = wb.add_format({**header_base, 'bg_color': '#BF8F00'})
    header_perc_fmt = wb.add_format({**header_base, 'bg_color': '#70AD47'})

    money_num_fmt = '$#,##0.00'
    plain_fmt = wb.add_format({'align': 'center', 'valign': 'vcenter'})
    plain_zebra_fmt = wb.add_format({'align': 'center', 'valign': 'vcenter', 'bg_color': '#D6E4F0'})
    money_fmt = wb.add_format({'align': 'center', 'valign': 'vcenter', 'num_format': money_num_fmt})
    money_zebra_fmt = wb.add_format({
        'align': 'center', 'valign': 'vcenter',
        'num_format': money_num_fmt, 'bg_color': '#D6E4F0'
    })

    # top=6 -> double border
    total_label_fmt = wb.add_format({
        'bold': True, 'align': 'right', 'valign': 'vcenter', 'top': 6
    })
    total_money_fmt = wb.add_format({
        'bold': True, 'align': 'center', 'valign': 'vcenter',
        'num_format': money_num_fmt, 'top': 6
    })

    def _header_fmt_for(col_name):
        if col_name in iva_set or (col_name in other_set and col_name not in deduccion_set):
            return header_iva_fmt
        if col_name in deduccion_set:
            return header_perc_fmt
        return header_default_fmt

    # ── Bloque de titulo (filas 0-3 xlsxwriter = 1-4 Excel) ──
    last_col_idx = n_cols - 1
    razon_social = (meta.get('razon_social') or 'CONTRIBUYENTE').upper()
    tipo_reporte = (meta.get('tipo_reporte') or 'REPORTE DE MOVIMIENTOS').upper()

    ws.merge_range(0, 0, 0, last_col_idx, razon_social, title_fmt)
    ws.merge_range(1, 0, 1, last_col_idx, tipo_reporte, subtitle_red_fmt)
    ws.merge_range(2, 0, 2, last_col_idx,
                   f"CUIT: {meta.get('cuit_empresa', '')} | Periodo: {meta.get('periodo', '')}",
                   subtitle_blue_fmt)
    ws.merge_range(3, 0, 3, last_col_idx,
                   f"Total: {len(rows)} transacciones",
                   subtitle_italic_fmt)

    # Fila 4 (Excel 5) vacia. Headers en fila 5 (Excel 6)
    HEADER_ROW = 5
    DATA_START_ROW = 6

    ws.set_row(HEADER_ROW, 30)
    for c_idx, col_name in enumerate(cols):
        ws.write_string(HEADER_ROW, c_idx, col_name, _header_fmt_for(col_name))

    # ── Anchos de columna + formato moneda a nivel columna ──
    # Aplicamos formato moneda via set_column (default de la columna) en vez de
    # por celda. xlsxwriter usa el formato de columna cuando la celda no tiene
    # uno propio, asi mantenemos la presentacion sin pagar el costo de asignar
    # un format object en cada una de las ~650k celdas de datos.
    n_rows = len(rows)
    SAMPLE = min(200, n_rows)
    sample_rows = rows[:SAMPLE]
    for c_idx, col_name in enumerate(cols):
        if col_name == 'Fecha':
            width = 8
        elif c_idx in money_idx_set:
            max_abs = 0.0
            for r in sample_rows:
                v = r.get(col_name, 0.0)
                if isinstance(v, (int, float)) and abs(v) > max_abs:
                    max_abs = abs(v)
            width = max(len(f'${max_abs:,.2f}'), len(col_name)) + 2
        else:
            max_len = len(col_name)
            for r in sample_rows:
                v = r.get(col_name, '')
                s = str(v) if v is not None else ''
                if len(s) > max_len:
                    max_len = len(s)
            width = max_len + 3
        # Format por columna: money (alineacion + formato moneda) para money cols,
        # plain (solo alineacion) para el resto. Asi todas las celdas de datos
        # heredan alineacion centrada y formato moneda donde corresponde, sin
        # asignacion por-celda. Lo unico que se pierde respecto del original es la zebra.
        col_fmt = money_fmt if c_idx in money_idx_set else plain_fmt
        ws.set_column(c_idx, c_idx, max(width, 8), col_fmt)

    # ── Datos: escritura batch sin estilos por celda ──
    # Sin alignment, sin zebra, sin format por celda. El formato moneda viene
    # del column-level. Para datasets grandes es ~3-5x mas rapido que aplicar
    # un Format object en cada celda.
    for r_idx, row_dict in enumerate(rows):
        excel_row = DATA_START_ROW + r_idx
        for c_idx, col_name in enumerate(cols):
            val = row_dict.get(col_name, '')
            if c_idx == total_col_idx:
                cached = float(val) if isinstance(val, (int, float)) else 0.0
                ws.write_formula(
                    excel_row, c_idx,
                    f'=SUM({first_sum_letter}{excel_row + 1}:{last_sum_letter}{excel_row + 1})',
                    None, cached
                )
            elif c_idx in money_idx_set:
                num = float(val) if isinstance(val, (int, float)) else 0.0
                ws.write_number(excel_row, c_idx, num)
            else:
                if val is None or val == '':
                    pass  # celda en blanco sin estilo
                elif isinstance(val, (int, float)):
                    ws.write_number(excel_row, c_idx, val)
                else:
                    ws.write_string(excel_row, c_idx, str(val))

    # ── Fila TOTAL GENERAL (mantiene estilo) ──
    total_row_idx = DATA_START_ROW + n_rows
    if first_money_idx > 0:
        ws.merge_range(total_row_idx, 0, total_row_idx, first_money_idx - 1,
                       'TOTAL GENERAL', total_label_fmt)
    for c_idx in range(first_money_idx, n_cols):
        col_letter = xl_col_to_name(c_idx)
        formula = f'=SUM({col_letter}{DATA_START_ROW + 1}:{col_letter}{total_row_idx})'
        ws.write_formula(total_row_idx, c_idx, formula, total_money_fmt, 0)

    wb.close()
    print(f"\n  Excel consolidado (xlsxwriter) generado: {output_path}")


def generar_sifere_txt(transacciones: list[dict], meta: dict) -> str:
    """Genera un archivo TXT con formato SIFERE para percepciones de IIBB.
    Cada línea: CodJurisdiccion(3) + CUIT(11) + Fecha(DD/MM/YYYY) + PV(4) + Nro(8) + TipoComp(2) + Monto(11)
    """
    # ── Mapeo de nombre de percepción → código de jurisdicción SIFERE ──
    CODIGOS_JURISDICCION = {
        "PERC.IB.CAP.FED.": "901",
        "PERC.IB.CABA C.ELECT": "901",
        "PERC.IB.BS.AS.": "902",
        "PER. IIBB CATAMARCA": "903",
        "PERC.IB.CORDOBA": "904",
        "PERC. CORRIENTES": "905",
        "PERC. IIBB CHACO": "906",
        "PERC IIBB CHUBUT": "907",
        "PERCEP IB ENTRE RIOS": "908",
        "PERC. IIBB FORMOSA": "909",
        "PERC.IIBB JUJUY": "910",
        "PERC.LA PAMPA": "911",
        "PERC.IB.LA RIOJA": "912",
        "PERC.IB.MENDOZA": "913",
        "PERC.IB MISIONES": "914",
        "Perc.IIBB Neuquen": "915",
        "PERC. IB RIO NEGRO": "916",
        "PERC.IB.SALTA": "917",
        "PERC.IB SAN JUAN": "918",
        "PERC. SAN LUIS": "919",
        "PERCEP IIBB STA CRUZ": "920",
        "PERC IIBB SANTA FE": "921",
        "PERC IIBB SGO ESTERO": "922",
        "PERC. TIERRA D.FUEGO": "923",
        "PERCEP IIBB TUCUMAN": "924",
    }

    # ── Mapeo de tipo de comprobante para SIFERE ──
    TIPO_COMP_SIFERE = {
        "FC": "FA",
        "ND": "DA",
        "NC": "CA",
        "TF": "FA",
        "TK": "FA",
        "Li": "FA",
    }

    # ── Tasas IVA (para excluirlas de percepciones) ──
    IVA_RATES = {
        'Tasa 21%', 'T.21%', 'C.F.21%', 'Tasa 27%', 'T.27%',
        'Tasa 10.5%', 'Tasa 10,5%', 'T.10.5%', 'T.10,5%',
        'C.F.10.5%', 'C.F.10,5%', 'Tasa 5%', 'T.5%',
        'Tasa 2.5%', 'Tasa 2,5%', 'T.2.5%', 'T.2,5%',
        'T.IMP 21%', 'T.IMP 10%', 'Exento',
        'R.Monot21', 'R.Mont.10',
    }

    # ── Extraer periodo (mes/año) del meta ──
    periodo_str = meta.get('periodo', '')
    # El periodo viene como "Desde el 01/MM/YYYY hasta el DD/MM/YYYY"
    p_match = re.search(r'(\d{2})/(\d{4})', periodo_str)
    if p_match:
        mes_periodo = p_match.group(1)
        anio_periodo = p_match.group(2)
    else:
        # Fallback: intentar extraer de otra forma
        nums = re.findall(r'\d+', periodo_str)
        if len(nums) >= 5:
            # Formato DD/MM/YYYY → posiciones 1=mes, 2=año
            mes_periodo = nums[1]
            anio_periodo = nums[2]
        else:
            mes_periodo = "01"
            anio_periodo = "2025"

    # ── Recopilar percepciones IIBB de cada transacción ──
    lineas_txt = []

    for t in transacciones:
        # Datos base de la transacción
        dia = t['Fecha']
        tipo = t['Tipo']
        numero_raw = t['Numero']
        cuit_raw = t['CUIT'] if t['CUIT'] else ''
        # Formatear CUIT con guiones: XX-XXXXXXXX-X (13 chars)
        if '-' in cuit_raw:
            cuit_formateado = cuit_raw
        else:
            cuit_limpio = cuit_raw.replace('-', '')
            if len(cuit_limpio) == 11:
                cuit_formateado = f"{cuit_limpio[:2]}-{cuit_limpio[2:10]}-{cuit_limpio[10]}"
            else:
                cuit_formateado = cuit_limpio

        # Separar PV y Nro del número de comprobante
        if '-' in numero_raw:
            pv_str = numero_raw.split('-')[0]
            resto_num = numero_raw.split('-')[1]
        else:
            pv_str = numero_raw[:5]
            resto_num = numero_raw[5:]

        # Quitar letra del final si existe
        letra = resto_num[-1] if resto_num and resto_num[-1].isalpha() else ''
        nro_str = resto_num[:-1] if letra else resto_num

        # Formatear fecha completa
        fecha_completa = f"{int(dia):02d}/{mes_periodo}/{anio_periodo}"

        # Formatear PV y Nro
        pv_formateado = pv_str[-4:].zfill(4)
        nro_formateado = nro_str.zfill(8)

        # Tipo de comprobante SIFERE
        tipo_sifere = TIPO_COMP_SIFERE.get(tipo, tipo)

        # Comprobante = PV + Nro + Tipo
        comprobante_sifere = f"{pv_formateado}{nro_formateado}{tipo_sifere}"

        # ── Recopilar percepciones de esta transacción ──
        percepciones = {}  # nombre_percepcion -> monto

        # Desde la tasa principal (si no es IVA)
        tasa = t['Tasa']
        if tasa and tasa not in IVA_RATES:
            nombre_upper = tasa.upper()
            es_perc_iibb = ("PERC" in nombre_upper or "PER." in nombre_upper or "PER " in nombre_upper)
            if es_perc_iibb and "ADUA" not in nombre_upper and \
               "I.V.A" not in nombre_upper and "GCIAS" not in nombre_upper and \
               "IVA" not in nombre_upper:
                percepciones[tasa] = percepciones.get(tasa, 0.0) + t['Neto']

        # Desde sub-conceptos
        for s in t['SubConceptos']:
            nombre = s['Concepto']
            if not nombre or nombre in IVA_RATES:
                continue
            nombre_upper = nombre.upper()
            es_perc_iibb = ("PERC" in nombre_upper or "PER." in nombre_upper or "PER " in nombre_upper)
            if es_perc_iibb and "ADUA" not in nombre_upper and \
               "I.V.A" not in nombre_upper and "GCIAS" not in nombre_upper and \
               "IVA" not in nombre_upper:
                monto = s['Neto'] if s['Neto'] != 0.0 else s['Percepcion']
                percepciones[nombre] = percepciones.get(nombre, 0.0) + monto

        # ── Generar líneas TXT para cada percepción ──
        for nombre_perc, monto in percepciones.items():
            if monto == 0.0:
                continue

            # Buscar código de jurisdicción
            codigo = CODIGOS_JURISDICCION.get(nombre_perc, None)
            if codigo is None:
                # Intento fuzzy: buscar por contenido parcial
                for key, val in CODIGOS_JURISDICCION.items():
                    if key.upper() in nombre_perc.upper() or nombre_perc.upper() in key.upper():
                        codigo = val
                        break
                if codigo is None:
                    continue  # No se encontró jurisdicción, saltar

            # Invertir signo para NC (el extractor ya invierte, pero el formato
            # SIFERE espera el monto con signo negativo explícito para CA)
            # En nuestro extractor, NC ya tienen montos negativos en los SubConceptos NO,
            # la inversión se hace en crear_excel. Aquí trabajamos con datos crudos.
            monto_final = monto
            es_nc = (tipo == 'NC')

            # Formatear monto
            valor_abs = abs(monto_final)
            parte_entera = int(valor_abs)
            parte_decimal = f"{valor_abs:.2f}".split('.')[1]

            if es_nc:
                monto_formateado = f"-{parte_entera:07d},{parte_decimal}"
            else:
                monto_formateado = f"{parte_entera:08d},{parte_decimal}"

            # Construir línea
            linea = (
                f"{codigo}"
                f"{cuit_formateado}"
                f"{fecha_completa}"
                f"{comprobante_sifere}"
                f"{monto_formateado}"
            )
            lineas_txt.append(linea)

    return "\n".join(lineas_txt)


def generar_sifere_retenciones_txt(transacciones: list[dict], meta: dict) -> str:
    """Genera un archivo TXT con formato SIFERE Formato Nº 1 para retenciones de IIBB.
    Cada línea (79 chars): CodJurisdiccion(3) + CUIT(13) + Fecha(10) + Sucursal(4)
    + NroConstancia(16) + TipoComp(1) + LetraComp(1) + NroCompOriginal(20) + Importe(11)
    """
    # ── Mapeo provincia → código de jurisdicción (reutiliza el de percepciones) ──
    # Palabras clave de provincia extraídas de los nombres de retención
    PROVINCIA_A_JURISDICCION = {
        "CAP.FED": "901", "CABA": "901", "C.A.B.A": "901",
        "BS.AS": "902", "BSAS": "902", "BS AS": "902", "BUENOS AIRES": "902",
        "CATAMARCA": "903",
        "CORDOBA": "904", "CÓRDOBA": "904",
        "CORRIENTES": "905",
        "CHACO": "906",
        "CHUBUT": "907",
        "ENTRE RIOS": "908", "ENTRE RÍOS": "908",
        "FORMOSA": "909",
        "JUJUY": "910",
        "LA PAMPA": "911", "PAMPA": "911",
        "LA RIOJA": "912", "RIOJA": "912",
        "MENDOZA": "913",
        "MISIONES": "914",
        "NEUQUEN": "915", "NEUQUÉN": "915",
        "RIO NEGRO": "916", "RÍO NEGRO": "916", "R.NEGRO": "916",
        "SALTA": "917",
        "SAN JUAN": "918",
        "SAN LUIS": "919",
        "STA CRUZ": "920", "SANTA CRUZ": "920",
        "SANTA FE": "921",
        "SGO ESTERO": "922", "SGO.ESTERO": "922", "SANTIAGO": "922",
        "TIERRA D.FUEGO": "923", "TIERRA DEL FUEGO": "923",
        "TUCUMAN": "924", "TUCUMÁN": "924",
    }

    # ── Mapeo tipo comprobante del sistema → tipo SIFERE retenciones (1 char) ──
    TIPO_COMP_RET = {
        "FC": "F", "TF": "F", "TK": "F", "Li": "F",
        "NC": "C",
        "ND": "D",
    }

    # ── Palabras clave a EXCLUIR de retenciones ──
    EXCLUIR = {"SIRCREB", "SIRTAC", "BCO", "GCIAS", "IVA", "I.V.A", "BANCO", "BANCAR"}

    # ── Tasas IVA (para excluirlas) ──
    IVA_RATES = {
        'Tasa 21%', 'T.21%', 'C.F.21%', 'Tasa 27%', 'T.27%',
        'Tasa 10.5%', 'Tasa 10,5%', 'T.10.5%', 'T.10,5%',
        'C.F.10.5%', 'C.F.10,5%', 'Tasa 5%', 'T.5%',
        'Tasa 2.5%', 'Tasa 2,5%', 'T.2.5%', 'T.2,5%',
        'T.IMP 21%', 'T.IMP 10%', 'Exento',
        'R.Monot21', 'R.Mont.10',
    }

    def _buscar_jurisdiccion(nombre_ret: str) -> str | None:
        """Busca el código de jurisdicción extrayendo la provincia del nombre."""
        nombre_upper = nombre_ret.upper()
        for provincia, codigo in PROVINCIA_A_JURISDICCION.items():
            if provincia in nombre_upper:
                return codigo
        return None

    def _es_retencion_iibb(nombre: str) -> bool:
        """Retorna True si el concepto es una retención IIBB (no bancaria/gcias/iva)."""
        nombre_upper = nombre.upper()
        if "RET" not in nombre_upper:
            return False
        for excl in EXCLUIR:
            if excl in nombre_upper:
                return False
        return True

    # ── Extraer periodo (mes/año) del meta ──
    periodo_str = meta.get('periodo', '')
    p_match = re.search(r'(\d{2})/(\d{4})', periodo_str)
    if p_match:
        mes_periodo = p_match.group(1)
        anio_periodo = p_match.group(2)
    else:
        nums = re.findall(r'\d+', periodo_str)
        if len(nums) >= 5:
            mes_periodo = nums[1]
            anio_periodo = nums[2]
        else:
            mes_periodo = "01"
            anio_periodo = "2025"

    # ── Procesar transacciones ──
    lineas_txt = []

    for t in transacciones:
        dia = t['Fecha']
        tipo = t['Tipo']
        numero_raw = t['Numero']
        cuit_raw = t['CUIT'] if t['CUIT'] else ''
        letra = t.get('Letra', '')

        # CUIT del agente (proveedor) con guiones, 13 chars
        # Si ya tiene guiones, usar directo; si no, formatear XX-XXXXXXXX-X
        if '-' in cuit_raw:
            cuit_formateado = cuit_raw
        else:
            cuit_limpio = cuit_raw.replace('-', '')
            if len(cuit_limpio) == 11:
                cuit_formateado = f"{cuit_limpio[:2]}-{cuit_limpio[2:10]}-{cuit_limpio[10]}"
            else:
                cuit_formateado = cuit_limpio
        # Asegurar 13 chars
        cuit_formateado = cuit_formateado[:13].ljust(13)

        # Separar PV y Nro del número de comprobante
        if '-' in numero_raw:
            pv_str = numero_raw.split('-')[0]
            resto_num = numero_raw.split('-')[1]
        else:
            pv_str = numero_raw[:5]
            resto_num = numero_raw[5:]

        # Quitar letra del final si existe en el número
        if resto_num and resto_num[-1].isalpha():
            letra_comp = resto_num[-1]
            nro_str = resto_num[:-1]
        else:
            letra_comp = letra if letra else 'A'
            nro_str = resto_num

        # Fecha dd/mm/yyyy
        fecha_completa = f"{int(dia):02d}/{mes_periodo}/{anio_periodo}"

        # Sucursal (PV, 4 dígitos, ceros a izquierda) — default 1 si no tiene
        sucursal = pv_str.strip().lstrip('0') or "1"
        sucursal = sucursal[-4:].zfill(4)

        # Nro. Constancia (16 dígitos, ceros a izquierda) = Nro comprobante
        nro_constancia = nro_str.zfill(16)

        # Tipo de comprobante SIFERE retención (1 char) — siempre "O" (Otros)
        tipo_sifere = "O"

        # Letra del comprobante (1 char) — espacio en blanco para retenciones
        letra_sifere = " "

        # Nro. Comprobante Original (20 chars, ceros a izquierda) = mismo nro repetido
        nro_comp_original = nro_str.zfill(20)

        # ── Recopilar retenciones IIBB de esta transacción ──
        retenciones = {}  # nombre_retencion -> monto

        # Desde la tasa principal
        tasa = t['Tasa']
        if tasa and tasa not in IVA_RATES and _es_retencion_iibb(tasa):
            retenciones[tasa] = retenciones.get(tasa, 0.0) + t['Neto']

        # Desde sub-conceptos
        for s in t['SubConceptos']:
            nombre = s['Concepto']
            if not nombre or nombre in IVA_RATES:
                continue
            if _es_retencion_iibb(nombre):
                monto = s['Neto'] if s['Neto'] != 0.0 else s['Percepcion']
                retenciones[nombre] = retenciones.get(nombre, 0.0) + monto

        # ── Generar líneas TXT para cada retención ──
        for nombre_ret, monto in retenciones.items():
            if monto == 0.0:
                continue

            codigo = _buscar_jurisdiccion(nombre_ret)
            if codigo is None:
                continue  # No se encontró jurisdicción

            # Montos negativos solo para NC (tipo C o H)
            monto_final = monto
            es_nc = (tipo in ('NC',))

            valor_abs = abs(monto_final)
            parte_entera = int(valor_abs)
            parte_decimal = f"{valor_abs:.2f}".split('.')[1]

            if es_nc:
                monto_formateado = f"-{parte_entera:07d},{parte_decimal}"
            else:
                monto_formateado = f"{parte_entera:08d},{parte_decimal}"

            # Construir línea Formato 1 (79 chars)
            linea = (
                f"{codigo}"                # pos 1-3:   Jurisdicción (3)
                f"{cuit_formateado}"        # pos 4-16:  CUIT agente (13)
                f"{fecha_completa}"         # pos 17-26: Fecha (10)
                f"{sucursal}"              # pos 27-30: Sucursal (4)
                f"{nro_constancia}"        # pos 31-46: Nro Constancia (16)
                f"{tipo_sifere}"           # pos 47:    Tipo Comprobante (1)
                f"{letra_sifere}"          # pos 48:    Letra Comprobante (1)
                f"{nro_comp_original}"     # pos 49-68: Nro Comp Original (20)
                f"{monto_formateado}"      # pos 69-79: Importe (11)
            )
            lineas_txt.append(linea)

    return "\n".join(lineas_txt)


def generar_percepciones_arba(transacciones: list[dict], meta: dict) -> tuple[str, pd.DataFrame]:
    """Genera un archivo TXT con formato ARBA para percepciones IIBB de ventas, y un DataFrame equivalente.
    Cada línea (71 chars): CUIT(13) + Fecha(10) + TipoComp(1) + Letra(1) + PV(5)
    + NroComp(8) + BaseImponible(14) + Alicuota(5) + ImportePerc(13) + LetraFija(1)
    """
    # ── Mapeo tipo comprobante del sistema → código ARBA (1 char) ──
    TIPO_COMP_ARBA = {
        "FC": "F", "TF": "F", "TK": "F", "Li": "F",
        "NC": "C",
        "ND": "D",
        "RC": "R",
    }

    # ── Tasas IVA (para excluirlas al buscar percepciones) ──
    IVA_RATES = {
        'Tasa 21%', 'T.21%', 'C.F.21%', 'Tasa 27%', 'T.27%',
        'Tasa 10.5%', 'Tasa 10,5%', 'T.10.5%', 'T.10,5%',
        'C.F.10.5%', 'C.F.10,5%', 'Tasa 5%', 'T.5%',
        'Tasa 2.5%', 'Tasa 2,5%', 'T.2.5%', 'T.2,5%',
        'T.IMP 21%', 'T.IMP 10%', 'Exento',
        'R.Monot21', 'R.Mont.10',
    }

    # ── Palabras clave para identificar percepción IIBB Buenos Aires ──
    KEYWORDS_BS_AS = ["BS.AS", "BSAS", "BS AS", "BUENOS AIRES"]

    def _es_percepcion_bs_as(nombre: str) -> bool:
        """Retorna True si el concepto es una percepción IIBB Buenos Aires."""
        nombre_upper = nombre.upper()
        if "PERC" not in nombre_upper:
            return False
        # Excluir aduanera, IVA, ganancias
        if any(x in nombre_upper for x in ("ADUA", "I.V.A", "GCIAS", "IVA")):
            return False
        return any(kw in nombre_upper for kw in KEYWORDS_BS_AS)

    # ── Extraer periodo (mes/año) del meta ──
    periodo_str = meta.get('periodo', '')
    p_match = re.search(r'(\d{2})/(\d{4})', periodo_str)
    if p_match:
        mes_periodo = p_match.group(1)
        anio_periodo = p_match.group(2)
    else:
        nums = re.findall(r'\d+', periodo_str)
        if len(nums) >= 5:
            mes_periodo = nums[1]
            anio_periodo = nums[2]
        else:
            mes_periodo = "01"
            anio_periodo = "2025"

    # ── Procesar transacciones ──
    lineas_txt = []
    filas_excel = []

    for t in transacciones:
        dia = t['Fecha']
        tipo = t['Tipo']
        numero_raw = t['Numero']
        cuit_raw = t['CUIT'] if t['CUIT'] else ''

        # ── CUIT con guiones (13 chars: XX-XXXXXXXX-X) ──
        if '-' in cuit_raw:
            cuit_formateado = cuit_raw
        else:
            cuit_limpio = cuit_raw.replace('-', '')
            if len(cuit_limpio) == 11:
                cuit_formateado = f"{cuit_limpio[:2]}-{cuit_limpio[2:10]}-{cuit_limpio[10]}"
            else:
                cuit_formateado = cuit_limpio
        cuit_formateado = cuit_formateado[:13].ljust(13)

        # ── Fecha completa DD/MM/YYYY ──
        fecha_completa = f"{int(dia):02d}/{mes_periodo}/{anio_periodo}"

        # ── Tipo comprobante ARBA (1 char) ──
        tipo_arba = TIPO_COMP_ARBA.get(tipo, tipo[0] if tipo else "F")

        # ── Separar PV y Nro del número de comprobante ──
        if '-' in numero_raw:
            pv_str = numero_raw.split('-')[0]
            resto_num = numero_raw.split('-')[1]
        else:
            pv_str = numero_raw[:5]
            resto_num = numero_raw[5:]

        # Quitar letra del final si existe → esa es la letra del comprobante
        if resto_num and resto_num[-1].isalpha():
            letra_comp = resto_num[-1]
            nro_str = resto_num[:-1]
        else:
            letra_comp = 'A'
            nro_str = resto_num

        pv_formateado = pv_str[-5:].zfill(5)
        nro_formateado = nro_str.zfill(8)

        # ── Buscar percepción IIBB BS.AS. en esta transacción ──
        monto_percepcion = 0.0

        # Desde la tasa principal
        tasa = t['Tasa']
        if tasa and tasa not in IVA_RATES and _es_percepcion_bs_as(tasa):
            monto_percepcion += t['Neto']

        # Desde sub-conceptos
        for s in t['SubConceptos']:
            nombre = s['Concepto']
            if not nombre or nombre in IVA_RATES:
                continue
            if _es_percepcion_bs_as(nombre):
                monto = s['Neto'] if s['Neto'] != 0.0 else s['Percepcion']
                monto_percepcion += monto

        # Si no hay percepción BS.AS., saltar esta transacción
        if monto_percepcion == 0.0:
            continue

        # ── Base imponible = Neto gravado del movimiento ──
        # Recopilar neto de todas las tasas IVA (excluyendo percepciones/retenciones)
        base_imponible = 0.0

        # Neto de la tasa principal (si es IVA)
        if tasa and tasa in IVA_RATES:
            base_imponible += t['Neto']

        # Neto de sub-conceptos que son tasas IVA
        for s in t['SubConceptos']:
            nombre = s['Concepto']
            if nombre and nombre in IVA_RATES:
                base_imponible += s['Neto']

        # Si la base es 0, intentar usar el Neto principal
        if base_imponible == 0.0:
            base_imponible = t['Neto']

        # ── Calcular alícuota = Percepción / Base * 100 ──
        if base_imponible != 0.0:
            alicuota = abs(monto_percepcion) / abs(base_imponible) * 100
        else:
            alicuota = 0.0

        # ── Determinar si es NC (montos negativos) ──
        es_nc = (tipo == 'NC')

        # ── Formatear Base Imponible (14 chars: 11 enteros + . + 2 decimales) ──
        base_abs = abs(base_imponible)
        if es_nc:
            # Signo negativo reemplaza un cero de relleno
            base_str = f"-{int(base_abs):010d}.{base_abs:.2f}".split('.')
            base_formateada = f"-{int(base_abs):010d}.{f'{base_abs:.2f}'.split('.')[1]}"
        else:
            base_formateada = f"{int(base_abs):011d}.{f'{base_abs:.2f}'.split('.')[1]}"

        # ── Formatear Alícuota (5 chars: 2 enteros + . + 2 decimales) ──
        alic_formateada = f"{int(alicuota):02d}.{f'{alicuota:.2f}'.split('.')[1]}"

        # ── Formatear Importe Percepción (13 chars: 10 enteros + . + 2 decimales) ──
        perc_abs = abs(monto_percepcion)
        if es_nc:
            perc_formateada = f"-{int(perc_abs):09d}.{f'{perc_abs:.2f}'.split('.')[1]}"
        else:
            perc_formateada = f"{int(perc_abs):010d}.{f'{perc_abs:.2f}'.split('.')[1]}"

        # ── Construir línea (71 chars) ──
        linea = (
            f"{cuit_formateado}"          # pos 1-13:  CUIT (13)
            f"{fecha_completa}"           # pos 14-23: Fecha (10)
            f"{tipo_arba}"                # pos 24:    Tipo comprobante (1)
            f"{letra_comp}"               # pos 25:    Letra comprobante (1)
            f"{pv_formateado}"            # pos 26-30: Punto de venta / Sucursal (5)
            f"{nro_formateado}"           # pos 31-38: Nro comprobante (8)
            f"{base_formateada}"          # pos 39-52: Base imponible (14)
            f"{alic_formateada}"          # pos 53-57: Alícuota (5)
            f"{perc_formateada}"          # pos 58-70: Importe percepción (13)
            f"A"                          # pos 71:    Tipo Operación (Alta) (1)
        )
        lineas_txt.append(linea)

        # ── Construir fila para Excel ──
        filas_excel.append({
            'CUIT': cuit_formateado.strip(),
            'Fecha': fecha_completa,
            'Tipo Comprobante': tipo_arba,
            'Letra': letra_comp,
            'Sucursal': int(pv_formateado),
            'Nro Comprobante': int(nro_formateado),
            'Base Imponible': -base_abs if es_nc else base_abs,
            'Alicuota': alicuota,
            'Importe Percepción': -perc_abs if es_nc else perc_abs,
            'Tipo Operación': 'A'
        })

    df = pd.DataFrame(filas_excel)
    return "\n".join(lineas_txt), df


def generar_arba_desde_excel(df: pd.DataFrame) -> str:
    """Convierte un DataFrame con formato pre-establecido de nuevo al TXT posicional ARBA de 71 chars."""
    lineas_txt = []
    
    col_map = {str(c).lower().strip(): c for c in df.columns}
    
    def get_val(row, posible_names, default=""):
        for name in posible_names:
            if name in col_map:
                v = row[col_map[name]]
                return "" if pd.isna(v) else v
        return default

    for _, row in df.iterrows():
        cuit = get_val(row, ["cuit"])
        fecha = get_val(row, ["fecha"])
        tipo = get_val(row, ["tipo comprobante", "tipo"])
        letra = get_val(row, ["letra"])
        sucursal = get_val(row, ["sucursal", "suc", "pv"])
        nro = get_val(row, ["nro comprobante", "numero", "nro"])
        base = get_val(row, ["base imponible", "base", "neto"])
        ali = get_val(row, ["alicuota", "tasa"])
        importe = get_val(row, ["importe percepción", "importe", "percepcion"])

        cuit_str = str(cuit).replace('-', '').replace('.', '').strip()
        if cuit_str.endswith('0'):
            # fix for float .0
            pass
        if len(cuit_str) == 13 and cuit_str[-2:] == ".0":
            cuit_str = cuit_str[:-2]

        if len(cuit_str) == 11:
            cuit_formateado = f"{cuit_str[:2]}-{cuit_str[2:10]}-{cuit_str[10]}".ljust(13)
        else:
            cuit_formateado = cuit_str.ljust(13)

        if isinstance(fecha, pd.Timestamp):
            fecha_str = fecha.strftime("%d/%m/%Y")
        else:
            f_str = str(fecha).strip()
            if len(f_str) >= 10 and '-' in f_str[:10]: # yyyy-mm-dd
                parts = f_str[:10].split('-')
                if len(parts) == 3:
                     fecha_str = f"{parts[2]}/{parts[1]}/{parts[0]}"
                else:
                     fecha_str = f_str[:10]
            else:
                fecha_str = f_str[:10]
        fecha_str = fecha_str.ljust(10)

        tipo_str = str(tipo).strip()[0:1].upper() if pd.notna(tipo) and str(tipo).strip() else "F"
        letra_str = str(letra).strip()[0:1].upper() if pd.notna(letra) and str(letra).strip() else "A"
        
        try:
            suc_val = float(str(sucursal).replace(',','.')) if str(sucursal).strip() else 0
            sucursal_str = str(int(suc_val)).zfill(5)
        except Exception:
            sucursal_str = "00000"
            
        try:
            nro_val = float(str(nro).replace(',','.')) if str(nro).strip() else 0
            nro_str = str(int(nro_val)).zfill(8)
        except Exception:
            nro_str = "00000000"

        # --- Detección y corrección de Formato Porcentaje de Excel ---
        # Si Excel envía 1,50% como una celda porcentaje, pandas lo lee como 0.015
        try:
            b_val = float(str(base).replace(',', '.'))
            i_val = float(str(importe).replace(',', '.'))
            calc_ali = (abs(i_val) / abs(b_val) * 100) if b_val != 0 else 0.0
        except Exception:
            calc_ali = 0.0

        try:
            a_str = str(ali).strip().replace(',', '.')
            if '%' in a_str:
                a_val = float(a_str.replace('%', ''))
            else:
                a_val = float(a_str)
                # Si a_val es fracción y multiplicarlo por 100 coincide con la alícuota calculada teórica
                if a_val < 1.0 and calc_ali > 0.0:
                    if abs(a_val * 100 - calc_ali) < 0.1:
                        a_val = a_val * 100
            ali = a_val
        except Exception:
            if calc_ali > 0.0:
                ali = calc_ali
        # -----------------------------------------------------------

        def fmt_float(val, int_len):
            try:
                v = float(str(val).replace(',', '.'))
                abs_v = abs(v)
                s_int = str(int(abs_v)).zfill(int_len)
                s_dec = ("%.2f" % abs_v).split('.')[1]
                if v < 0:
                    return f"-{str(int(abs_v)).zfill(int_len-1)}.{s_dec}"
                return f"{s_int}.{s_dec}"
            except Exception:
                return ("0" * int_len) + ".00"
                
        base_f = fmt_float(base, 11)
        ali_f = fmt_float(ali, 2)
        imp_f = fmt_float(importe, 10)
        
        linea = (
            f"{cuit_formateado}"          
            f"{fecha_str}"                
            f"{tipo_str}"                 
            f"{letra_str}"                
            f"{sucursal_str}"             
            f"{nro_str}"                  
            f"{base_f}"                   
            f"{ali_f}"                    
            f"{imp_f}"                    
            f"A"                          
        )
        lineas_txt.append(linea)
        
    return "\n".join(lineas_txt)


def _es_retencion_bs_as(nombre: str) -> bool:
    """Detecta si un rubro corresponde a una retención de IIBB Buenos Aires 'normal' (sin SIRTAC/SIRCREB/Bancarias)"""
    n_lower = nombre.lower()
    if any(excl in n_lower for excl in ['sirtac', 'sircreb', 'bco', 'bancari', 'banco', 'bc']):
        return False
    if ('ret' in n_lower or 'rt' in n_lower) and ('bs' in n_lower or 'as' in n_lower or 'ba' in n_lower or 'b.a' in n_lower or 'buenos' in n_lower):
        return True
    return False

def generar_retenciones_arba(movimientos: list, metadata: dict) -> tuple[str, pd.DataFrame]:
    """Genera el archivo TXT para ARBA Retenciones (67 caracteres) a partir de compras Mendez."""
    lineas_txt = []
    filas_excel = []

    for mov in movimientos:
        for t in mov['Transacciones']:
            monto_retencion = 0.0
            
            for imp in t.get('OtrosImpuestos', []):
                if imp['Tipo'] == 'RET' and imp['Monto'] != 0.0:
                    if _es_retencion_bs_as(imp['Nombre']):
                        monto_retencion += abs(imp['Monto'])

            if monto_retencion == 0.0:
                continue

            try:
                nro_str = str(int(t['Numero'])).zfill(20)[:20]
            except Exception:
                nro_str = "0".zfill(20)

            cuit_str = mov['CUIT'].replace('-', '').replace('.', '').strip()
            cuit_formateado = cuit_str.ljust(11)[:11]

            try:
                suc_val = int(t['PuntoVenta'])
                sucursal_str = str(suc_val).zfill(5)[:5]
            except Exception:
                sucursal_str = "00000"

            fecha_str = mov['Fecha'].strftime('%d/%m/%Y')

            importe = monto_retencion
            base = t['Neto']
            try:
                if base != 0: alic = abs(importe) / abs(base) * 100
                else: alic = 0.0
            except Exception:
                alic = 0.0

            def fmt_float(val, int_len):
                try:
                    v = float(str(val).replace(',', '.'))
                    abs_v = abs(v)
                    s_int = str(int(abs_v)).zfill(int_len)
                    s_dec = ("%.2f" % abs_v).split('.')[1]
                    if v < 0:
                        return f"-{str(int(abs_v)).zfill(int_len-1)}.{s_dec}"
                    return f"{s_int}.{s_dec}"
                except Exception:
                    return ("0" * int_len) + ".00"

            ali_f = fmt_float(alic, 2)     # format 99.99 (5)
            base_f = fmt_float(base, 13)   # format 9999999999999.99 (16)

            linea = (
                f"{nro_str}"
                f"{cuit_formateado}"
                f"{sucursal_str}"
                f"{fecha_str}"
                f"{ali_f}"
                f"{base_f}"
            )
            lineas_txt.append(linea)

            filas_excel.append({
                'Transaccion Agente': nro_str,
                'CUIT': mov['CUIT'],
                'Sucursal': sucursal_str,
                'Fecha': fecha_str,
                'Base Imponible': base,
                'Alicuota': alic,
                'Importe': importe
            })

    df = pd.DataFrame(filas_excel)
    return "\n".join(lineas_txt), df


def generar_retenciones_arba_desde_excel(df: pd.DataFrame) -> str:
    """Convierte un DataFrame con formato pre-establecido al TXT posicional ARBA Retenciones (67 chars)."""
    lineas_txt = []
    
    col_map = {str(c).lower().strip(): c for c in df.columns}
    
    def get_val(row, posible_names, default=""):
        for name in posible_names:
            if name in col_map:
                v = row[col_map[name]]
                return "" if pd.isna(v) else v
        return default

    for _, row in df.iterrows():
        nro_trans = get_val(row, ["transaccion", "nro transaccion agente", "transaccion agente", "id"])
        cuit = get_val(row, ["cuit"])
        fecha = get_val(row, ["fecha"])
        sucursal = get_val(row, ["sucursal", "suc", "pv"])
        base = get_val(row, ["base imponible", "base", "neto"])
        ali = get_val(row, ["alicuota", "tasa"])
        importe = get_val(row, ["importe retencion", "importe", "retencion", "ret"])

        try:
            nro_str = str(int(float(nro_trans))).zfill(20)[:20] if str(nro_trans).strip() else "0".zfill(20)
        except Exception:
            nro_str = "0".zfill(20)

        cuit_str = str(cuit).replace('-', '').replace('.', '').strip()
        if cuit_str.endswith('.0') and len(cuit_str) == 13: cuit_str = cuit_str[:-2]
        cuit_formateado = cuit_str.ljust(11)[:11]

        if isinstance(fecha, pd.Timestamp):
            fecha_str = fecha.strftime("%d/%m/%Y")
        else:
            f_str = str(fecha).strip()
            if len(f_str) >= 10 and '-' in f_str[:10]:
                parts = f_str[:10].split('-')
                if len(parts) == 3: fecha_str = f"{parts[2]}/{parts[1]}/{parts[0]}"
                else: fecha_str = f_str[:10]
            else:
                fecha_str = f_str[:10]
        fecha_str = fecha_str.ljust(10)[:10]

        try:
            suc_val = float(str(sucursal).replace(',','.')) if str(sucursal).strip() else 0
            sucursal_str = str(int(suc_val)).zfill(5)[:5]
        except Exception:
            sucursal_str = "00000"

        try:
            b_val = float(str(base).replace(',', '.'))
            i_val = float(str(importe).replace(',', '.'))
            calc_ali = (abs(i_val) / abs(b_val) * 100) if b_val != 0 else 0.0
        except Exception:
            calc_ali = 0.0

        try:
            a_str = str(ali).strip().replace(',', '.')
            if '%' in a_str:
                a_val = float(a_str.replace('%', ''))
            else:
                a_val = float(a_str)
                if a_val < 1.0 and calc_ali > 0.0:
                    if abs(a_val * 100 - calc_ali) < 0.1:
                        a_val = a_val * 100
            ali = a_val
        except Exception:
            if calc_ali > 0.0: ali = calc_ali

        def fmt_float(val, int_len):
            try:
                v = float(str(val).replace(',', '.'))
                abs_v = abs(v)
                s_int = str(int(abs_v)).zfill(int_len)
                s_dec = ("%.2f" % abs_v).split('.')[1]
                if v < 0:
                    return f"-{str(int(abs_v)).zfill(int_len-1)}.{s_dec}"
                return f"{s_int}.{s_dec}"
            except Exception:
                return ("0" * int_len) + ".00"
                
        base_f = fmt_float(base, 13)
        ali_f = fmt_float(ali, 2)
        
        linea = (
            f"{nro_str}"          
            f"{cuit_formateado}"                
            f"{sucursal_str}"                 
            f"{fecha_str}"                
            f"{ali_f}"             
            f"{base_f}"                  
        )
        lineas_txt.append(linea)
        
    return "\n".join(lineas_txt)


def seleccionar_archivo() -> Path:
    """Abre un diálogo para que el usuario seleccione un archivo .txt."""
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()  # Ocultar ventana principal
    root.attributes('-topmost', True)  # Traer diálogo al frente

    archivo = filedialog.askopenfilename(
        title='Seleccionar archivo de movimientos',
        filetypes=[('Archivos de texto', '*.txt'), ('Todos los archivos', '*.*')],
        initialdir=Path(__file__).parent
    )

    root.destroy()

    if not archivo:
        print("❌ No se seleccionó ningún archivo. Saliendo...")
        sys.exit(0)

    return Path(archivo)


# ============================================================================
# CITI Ventas (RG 3685 AFIP) - Generadores VENTAS.txt + ALICUOTAS.txt
# ============================================================================

CITI_TIPO_DESC = {
    1: 'FC A', 2: 'ND A', 3: 'NC A', 4: 'Rec A', 5: 'NVC A',
    6: 'FC B', 7: 'ND B', 8: 'NC B', 9: 'Rec B', 10: 'NVC B',
    11: 'FC C', 12: 'ND C', 13: 'NC C', 15: 'Rec C',
    19: 'FC E', 20: 'ND E', 21: 'NC E', 22: 'FC PES',
    27: 'LUCI A', 28: 'LUCI B', 29: 'LUCI C',
    33: 'LPG',
    34: 'CMP A 1415', 35: 'CMP B 1415', 37: 'ND 1415', 38: 'NC 1415',
    39: 'OTROS A 1415', 40: 'OTROS B 1415',
    43: 'NC LUCI B', 44: 'NC LUCI C', 45: 'ND LUCI A', 46: 'ND LUCI B',
    47: 'ND LUCI C', 48: 'NC LUCI A',
    51: 'FC M', 52: 'ND M', 53: 'NC M', 54: 'Rec M', 55: 'NVC M',
    63: 'Liq A', 64: 'Liq B', 68: 'Liq C',
    81: 'TF A', 82: 'TF B', 83: 'Tique Z',
    110: 'T NC',
    111: 'TF C', 113: 'T NV C', 114: 'T NC C', 115: 'T ND C',
    118: 'T NC B', 119: 'T NC A', 120: 'T ND',
    195: 'FC T', 196: 'ND T', 197: 'NC T',
    201: 'FC A MiPyME', 202: 'ND A MiPyME', 203: 'NC A MiPyME',
    206: 'FC B MiPyME', 207: 'ND B MiPyME', 208: 'NC B MiPyME',
    211: 'FC C MiPyME', 212: 'ND C MiPyME', 213: 'NC C MiPyME',
}


def _citi_desc_tipo(cod) -> str:
    try:
        return CITI_TIPO_DESC.get(int(cod), f'Tipo {int(cod)}')
    except (TypeError, ValueError):
        return str(cod)


CITI_ALIC_CODIGOS = {
    'IVA 0%':    '0003',
    'IVA 2.5%':  '0009',
    'IVA 5%':    '0008',
    'IVA 10.5%': '0004',
    'IVA 21%':   '0005',
    'IVA 27%':   '0006',
}

CITI_ALICUOTAS_ORDEN = ['IVA 0%', 'IVA 2.5%', 'IVA 5%', 'IVA 10.5%', 'IVA 21%', 'IVA 27%']

CITI_HEADERS_ESPERADOS = {
    'fecha': 'Fecha de Emisión',
    'tipo': 'Tipo de Comprobante',
    'pv': 'Punto de Venta',
    'numero': 'Número de Comprobante',
    'numero_hasta': 'Número de Comprobante Hasta',
    'doc_cod': 'Tipo Doc. Comprador',
    'doc_nro': 'Nro. Doc. Comprador',
    'denom': 'Denominación Comprador',
    'fecha_vto': 'Fecha de Vencimiento del Pago',
    'importe_total': 'Importe Total',
    'moneda': 'Moneda Original',
    'tc': 'Tipo de Cambio',
    'no_gravado': 'Importe No Gravado',
    'exento': 'Importe Exento',
    'perc_iva': 'Importe de Per. o Pagos a Cta. de Otros Imp. Nac.',
    'perc_iibb': 'Importe de Percepciones de Ingresos Brutos',
    'perc_mun': 'Importe de Impuestos Municipales',
    'perc_no_categ': 'Percepción a No Categorizados',
    'imp_internos': 'Importe de Impuestos Internos',
    'otros_trib': 'Importe Otros Tributos',
    'neto_0': 'Neto Gravado IVA 0%',
    'neto_2_5': 'Neto Gravado IVA 2,5%',
    'iva_2_5': 'Importe IVA 2,5%',
    'neto_5': 'Neto Gravado IVA 5%',
    'iva_5': 'Importe IVA 5%',
    'neto_10_5': 'Neto Gravado IVA 10,5%',
    'iva_10_5': 'Importe IVA 10,5%',
    'neto_21': 'Neto Gravado IVA 21%',
    'iva_21': 'Importe IVA 21%',
    'neto_27': 'Neto Gravado IVA 27%',
    'iva_27': 'Importe IVA 27%',
    'total_neto': 'Total Neto Gravado',
    'total_iva': 'Total IVA',
}


def _citi_parse_money(v) -> float:
    """Convierte un string de monto formato argentino ('1.234,56' o '1234,56') a float."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    neg = s.startswith('-')
    s = s.lstrip('-').replace('.', '').replace(',', '.')
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return 0.0


def _citi_parse_int(v, default: int = 0) -> int:
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    s = re.sub(r'[^0-9-]', '', s)
    try:
        return int(s)
    except ValueError:
        return default


def _citi_parse_fecha(v) -> str:
    """Convierte fechas en formatos ISO/AR a 'YYYYMMDD'. Devuelve '' si no parsea."""
    if v is None:
        return ''
    s = str(v).strip()
    if not s:
        return ''
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:4] + s[5:7] + s[8:10]
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return f"{m.group(3)}{int(m.group(2)):02d}{int(m.group(1)):02d}"
    if re.match(r'^\d{8}$', s):
        return s
    return ''


def _citi_sanitizar_denom(s: str) -> str:
    """ASCII uppercase, sin tildes ni caracteres no permitidos. Trunca a 30 chars."""
    import unicodedata
    if s is None:
        return ''
    txt = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode('ascii')
    txt = txt.upper()
    txt = re.sub(r'[^A-Z0-9 ./,\-]', ' ', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt[:30]


def normalizar_csv_ventas_arca(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza el DataFrame leído del CSV del Portal IVA ARCA (ventas).
    Devuelve un nuevo DataFrame con columnas casteadas y una columna interna por header esperado.
    """
    faltantes = [h for h in CITI_HEADERS_ESPERADOS.values() if h not in df.columns]
    if faltantes:
        raise ValueError(
            "Faltan columnas en el CSV ARCA. Esperadas pero no encontradas:\n"
            + "\n".join(f"  - {c}" for c in faltantes)
        )

    out = pd.DataFrame()
    out['fecha'] = df[CITI_HEADERS_ESPERADOS['fecha']].apply(_citi_parse_fecha)
    out['tipo'] = df[CITI_HEADERS_ESPERADOS['tipo']].apply(_citi_parse_int)
    out['pv'] = df[CITI_HEADERS_ESPERADOS['pv']].apply(_citi_parse_int)
    out['numero'] = df[CITI_HEADERS_ESPERADOS['numero']].apply(_citi_parse_int)
    out['numero_hasta'] = df[CITI_HEADERS_ESPERADOS['numero_hasta']].apply(_citi_parse_int)
    # Si NumeroHasta < Numero o es 0, usar Numero (vectorizado).
    out['numero_hasta'] = out['numero_hasta'].where(
        out['numero_hasta'] >= out['numero'], out['numero']
    )
    out['doc_cod'] = df[CITI_HEADERS_ESPERADOS['doc_cod']].apply(_citi_parse_int)
    out['doc_nro'] = df[CITI_HEADERS_ESPERADOS['doc_nro']].astype(str).apply(
        lambda s: re.sub(r'[^0-9]', '', s) or '0'
    )
    out['denom'] = df[CITI_HEADERS_ESPERADOS['denom']].apply(_citi_sanitizar_denom)
    out['fecha_vto'] = df[CITI_HEADERS_ESPERADOS['fecha_vto']].apply(_citi_parse_fecha)
    out['moneda'] = df[CITI_HEADERS_ESPERADOS['moneda']].astype(str).str.strip().str.upper().replace('', 'PES')
    out['tc'] = df[CITI_HEADERS_ESPERADOS['tc']].apply(_citi_parse_money)

    monto_keys = [
        'importe_total', 'no_gravado', 'exento', 'perc_iva', 'perc_iibb',
        'perc_mun', 'perc_no_categ', 'imp_internos', 'otros_trib',
        'neto_0', 'neto_2_5', 'iva_2_5', 'neto_5', 'iva_5',
        'neto_10_5', 'iva_10_5', 'neto_21', 'iva_21', 'neto_27', 'iva_27',
        'total_neto', 'total_iva',
    ]
    for k in monto_keys:
        out[k] = df[CITI_HEADERS_ESPERADOS[k]].apply(_citi_parse_money)

    return out


def consolidar_ventas_citi(df_norm: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa por (Fecha, PV, Tipo, DocCod, Doc) — consolidación tipo Ticket Z.
    Tickets a CF (DocCod=99, Doc=99999999) se acumulan en una línea por día/PV/tipo.
    Facturas con CUITs distintos quedan separadas.
    """
    if df_norm.empty:
        return df_norm.copy()

    # Cantidad de comprobantes contemplada en cada fila original (rangos pre-consolidados):
    df = df_norm.copy()
    df['_cant_fila'] = (df['numero_hasta'] - df['numero'] + 1).clip(lower=1)

    grupos = df.groupby(['fecha', 'pv', 'tipo', 'doc_cod', 'doc_nro'], dropna=False, sort=True)

    sum_cols = [
        'importe_total', 'no_gravado', 'exento', 'perc_iva', 'perc_iibb',
        'perc_mun', 'perc_no_categ', 'imp_internos', 'otros_trib',
        'neto_0', 'neto_2_5', 'iva_2_5', 'neto_5', 'iva_5',
        'neto_10_5', 'iva_10_5', 'neto_21', 'iva_21', 'neto_27', 'iva_27',
        'total_neto', 'total_iva',
    ]

    rows_out = []
    for (fecha, pv, tipo, doc_cod, doc_nro), g in grupos:
        row = {
            'fecha': fecha,
            'pv': int(pv),
            'tipo': int(tipo),
            'doc_cod': int(doc_cod),
            'doc_nro': doc_nro,
            'desde': int(g['numero'].min()),
            'hasta': int(g['numero_hasta'].max()),
            'cant_cbtes': int(g['_cant_fila'].sum()),
        }
        denoms_unicas = g['denom'].dropna().unique()
        denoms_unicas = [d for d in denoms_unicas if d]
        if len(denoms_unicas) == 1:
            row['denom'] = denoms_unicas[0]
        elif len(denoms_unicas) == 0:
            row['denom'] = ''
        else:
            row['denom'] = 'VARIOS'
        # Fecha vto: la mínima no vacía, fallback a fecha
        vtos = [v for v in g['fecha_vto'] if v]
        row['fecha_vto'] = min(vtos) if vtos else fecha
        # Moneda y TC: tomar la primera (típicamente PES uniforme)
        row['moneda'] = g['moneda'].iloc[0]
        tc = g['tc'].iloc[0]
        row['tc'] = tc if tc and tc > 0 else 1.0
        for c in sum_cols:
            row[c] = float(g[c].sum())
        rows_out.append(row)

    df_out = pd.DataFrame(rows_out)
    if df_out.empty:
        return df_out
    return df_out.sort_values(['fecha', 'pv', 'tipo', 'desde']).reset_index(drop=True)


def _alicuotas_de_consolidado(row) -> list[tuple[str, float, float]]:
    """Devuelve la lista de (codigo_alicuota_4, neto, iva) con neto>0 o iva>0 para la fila consolidada."""
    candidatos = [
        ('IVA 0%',    row.get('neto_0', 0.0),    0.0),
        ('IVA 2.5%',  row.get('neto_2_5', 0.0),  row.get('iva_2_5', 0.0)),
        ('IVA 5%',    row.get('neto_5', 0.0),    row.get('iva_5', 0.0)),
        ('IVA 10.5%', row.get('neto_10_5', 0.0), row.get('iva_10_5', 0.0)),
        ('IVA 21%',   row.get('neto_21', 0.0),   row.get('iva_21', 0.0)),
        ('IVA 27%',   row.get('neto_27', 0.0),   row.get('iva_27', 0.0)),
    ]
    res = []
    for nombre, neto, iva in candidatos:
        if abs(neto) > 0.001 or abs(iva) > 0.001:
            res.append((CITI_ALIC_CODIGOS[nombre], float(neto), float(iva)))

    if not res:
        # Fallback comprobantes 100% exentos: emitir 1 línea con neto=Exento, código 0% (0003)
        exento = float(row.get('exento', 0.0))
        if abs(exento) > 0.001:
            res.append((CITI_ALIC_CODIGOS['IVA 0%'], exento, 0.0))
        else:
            # Sin nada — emitir 1 línea con neto = importe_total, IVA 0
            res.append((CITI_ALIC_CODIGOS['IVA 0%'], float(row.get('importe_total', 0.0)), 0.0))
    return res


def _citi_n(val: float, length: int) -> str:
    """Codifica monto: abs(val)*100, redondeado, zfill al ancho."""
    cents = int(round(abs(float(val)) * 100))
    return str(cents).zfill(length)[-length:]


def _citi_z(val, length: int) -> str:
    """Entero zfill."""
    try:
        n = int(val)
    except (TypeError, ValueError):
        n = 0
    return str(abs(n)).zfill(length)[-length:]


def _citi_a(val, length: int) -> str:
    """Texto ljust con espacios."""
    s = '' if val is None else str(val)
    return s.ljust(length)[:length]


def generar_citi_ventas_lineas(df_consolidado: pd.DataFrame) -> list[str]:
    """Genera las líneas de VENTAS.txt (266 chars/línea) según RG 3685 REGINFO_CV_VENTAS_CBTE."""
    lineas = []
    for _, r in df_consolidado.iterrows():
        cant_alic = len(_alicuotas_de_consolidado(r))
        moneda = (r.get('moneda') or 'PES').strip().upper()[:3].ljust(3)
        if moneda.strip() == 'PES':
            tc_str = '0001000000'
        else:
            tc_val = float(r.get('tc') or 1.0)
            tc_str = str(int(round(tc_val * 1_000_000))).zfill(10)[-10:]

        linea = (
            _citi_a(r['fecha'], 8)
            + _citi_z(r['tipo'], 3)
            + _citi_z(r['pv'], 5)
            + _citi_z(r['desde'], 20)
            + _citi_z(r['hasta'], 20)
            + _citi_z(r['doc_cod'], 2)
            + _citi_a(str(r['doc_nro']).rjust(20, '0')[-20:], 20)
            + _citi_a(r.get('denom', ''), 30)
            + _citi_n(r['importe_total'], 15)
            + _citi_n(r['no_gravado'], 15)
            + _citi_n(r['perc_no_categ'], 15)
            + _citi_n(r['exento'], 15)
            + _citi_n(r['perc_iva'], 15)
            + _citi_n(r['perc_iibb'], 15)
            + _citi_n(r['perc_mun'], 15)
            + _citi_n(r['imp_internos'], 15)
            + moneda
            + tc_str
            + str(min(max(cant_alic, 0), 9))
            + '0'
            + _citi_n(r['otros_trib'], 15)
            + _citi_a(r.get('fecha_vto') or r['fecha'], 8)
        )
        if len(linea) != 266:
            raise ValueError(f"VENTAS.txt: línea con largo {len(linea)} (debe ser 266) para fila {dict(r)}")
        lineas.append(linea)
    return lineas


def generar_citi_alicuotas_lineas(df_consolidado: pd.DataFrame) -> list[str]:
    """Genera las líneas de ALICUOTAS.txt (62 chars/línea) según RG 3685 REGINFO_CV_VENTAS_ALICUOTAS."""
    lineas = []
    for _, r in df_consolidado.iterrows():
        for cod_alic, neto, iva in _alicuotas_de_consolidado(r):
            linea = (
                _citi_z(r['tipo'], 3)
                + _citi_z(r['pv'], 5)
                + _citi_z(r['desde'], 20)
                + _citi_n(neto, 15)
                + cod_alic
                + _citi_n(iva, 15)
            )
            if len(linea) != 62:
                raise ValueError(f"ALICUOTAS.txt: línea con largo {len(linea)} (debe ser 62)")
            lineas.append(linea)
    return lineas


_CITI_COLUMNAS = [
    ('Fecha', 'fecha', 'fija'),
    ('PV', 'pv', 'fija'),
    ('Tipo', 'tipo', 'fija'),
    ('Desde', 'desde', 'fija'),
    ('Hasta', 'hasta', 'fija'),
    ('Cant.', 'cant_cbtes', 'fija'),
    ('DocCod', 'doc_cod', 'fija'),
    ('Doc', 'doc_nro', 'fija'),
    ('Denominación', 'denom', 'fija'),
    ('Importe Total', 'importe_total', 'monto'),
    ('Neto 0%', 'neto_0', 'iva'),
    ('Neto 2,5%', 'neto_2_5', 'iva'),
    ('IVA 2,5%', 'iva_2_5', 'iva'),
    ('Neto 5%', 'neto_5', 'iva'),
    ('IVA 5%', 'iva_5', 'iva'),
    ('Neto 10,5%', 'neto_10_5', 'iva'),
    ('IVA 10,5%', 'iva_10_5', 'iva'),
    ('Neto 21%', 'neto_21', 'iva'),
    ('IVA 21%', 'iva_21', 'iva'),
    ('Neto 27%', 'neto_27', 'iva'),
    ('IVA 27%', 'iva_27', 'iva'),
    ('Perc IIBB', 'perc_iibb', 'deduc'),
    ('Perc Mun', 'perc_mun', 'deduc'),
    ('Perc IVA', 'perc_iva', 'deduc'),
    ('Perc no Categ', 'perc_no_categ', 'deduc'),
    ('Imp Internos', 'imp_internos', 'deduc'),
    ('Otros Trib', 'otros_trib', 'deduc'),
    ('No Gravado', 'no_gravado', 'monto'),
    ('Exento', 'exento', 'monto'),
    ('Total IVA', 'total_iva', 'monto'),
]


_CITI_LITE_THRESHOLD = 500


def _escribir_hoja_citi(ws, df_subset: pd.DataFrame, titulo: str, subtitulo: str) -> None:
    """Escribe una hoja con el formato CITI (header trifilas + datos + TOTAL GENERAL).
    Las columnas numéricas todo-cero en este subset se eliminan.
    Para hojas con > _CITI_LITE_THRESHOLD filas usa modo lite (sin zebra/borders en cuerpo)."""
    from openpyxl.utils import get_column_letter

    azul = PatternFill('solid', fgColor='1F4E78')
    azul_claro = PatternFill('solid', fgColor='F2F7FB')
    amarillo = PatternFill('solid', fgColor='FFF2CC')
    verde = PatternFill('solid', fgColor='E2EFDA')
    total_fill = PatternFill('solid', fgColor='D9E1F2')
    iva_hdr = PatternFill('solid', fgColor='C09700')
    deduc_hdr = PatternFill('solid', fgColor='548235')
    bold_white = Font(bold=True, color='FFFFFF', name='Calibri')
    bold = Font(bold=True, name='Calibri')
    border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )

    columnas = []
    for titulo_col, key, kind in _CITI_COLUMNAS:
        if kind in ('monto', 'iva', 'deduc') and not df_subset.empty:
            col_vals = df_subset.get(key)
            if col_vals is not None and (col_vals.abs().sum() < 0.005):
                continue
        columnas.append((titulo_col, key, kind))

    n_cols = len(columnas)
    n_rows = len(df_subset)
    lite = n_rows > _CITI_LITE_THRESHOLD

    ws.cell(1, 1, titulo).font = Font(bold=True, color='FFFFFF', size=14, name='Calibri')
    ws.cell(1, 1).fill = azul
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 26

    ws.cell(2, 1, subtitulo).font = Font(bold=True, color='1F4E78', name='Calibri')
    ws.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 18

    HDR_ROW = 4
    for i, (titulo_col, _, kind) in enumerate(columnas, start=1):
        c = ws.cell(HDR_ROW, i, titulo_col)
        c.font = bold_white
        if kind == 'iva':
            c.fill = iva_hdr
        elif kind == 'deduc':
            c.fill = deduc_hdr
        else:
            c.fill = azul
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[HDR_ROW].height = 30

    if lite and n_rows > 0:
        # Modo lite: vectorizar transformaciones, ws.append() en bulk, estilos por columna.
        df_render = df_subset.copy()
        if 'fecha' in df_render.columns:
            df_render['fecha'] = df_render['fecha'].apply(
                lambda v: f"{v[:4]}-{v[4:6]}-{v[6:8]}" if isinstance(v, str) and len(v) == 8 else v
            )
        if 'tipo' in df_render.columns:
            df_render['tipo'] = df_render['tipo'].map(
                lambda v: _citi_desc_tipo(v) if v is not None else ''
            )
        # Reemplazar NaN por 0 en columnas numéricas
        num_keys = [k for _, k, kind in columnas if kind in ('monto', 'iva', 'deduc')]
        for k in num_keys:
            if k in df_render.columns:
                df_render[k] = df_render[k].fillna(0)

        keys_orden = [k for _, k, _ in columnas]
        rows_data = df_render[keys_orden].values.tolist()
        for vals in rows_data:
            ws.append(vals)

        # Aplicar fill IVA/deduc + number_format por columna en una pasada.
        for cidx, (_, key, kind) in enumerate(columnas, start=1):
            if kind not in ('monto', 'iva', 'deduc'):
                continue
            fill_for_col = amarillo if kind == 'iva' else (verde if kind == 'deduc' else None)
            for r in range(HDR_ROW + 1, HDR_ROW + 1 + n_rows):
                cell = ws.cell(r, cidx)
                cell.number_format = '#,##0.00'
                if fill_for_col is not None:
                    cell.fill = fill_for_col
    else:
        # Modo full: cell-by-cell con borders + zebra + fills por celda (hojas chicas).
        for ridx, (_, row) in enumerate(df_subset.iterrows()):
            excel_row = HDR_ROW + 1 + ridx
            zebra = (ridx % 2 == 1)
            for cidx, (_, key, kind) in enumerate(columnas, start=1):
                v = row.get(key)
                if key == 'fecha' and isinstance(v, str) and len(v) == 8:
                    v = f"{v[:4]}-{v[4:6]}-{v[6:8]}"
                if key == 'tipo':
                    v = _citi_desc_tipo(v)
                cell = ws.cell(excel_row, cidx, v if not (isinstance(v, float) and pd.isna(v)) else 0)
                if kind in ('monto', 'iva', 'deduc'):
                    cell.number_format = '#,##0.00'
                cell.border = border
                if zebra:
                    cell.fill = azul_claro
                if kind == 'iva':
                    cell.fill = amarillo
                elif kind == 'deduc':
                    cell.fill = verde

    if n_rows > 0:
        total_row = HDR_ROW + 1 + n_rows
        ws.cell(total_row, 1, 'TOTAL').font = bold
        ws.cell(total_row, 1).fill = total_fill
        ws.cell(total_row, 1).border = border
        for cidx, (_, key, kind) in enumerate(columnas, start=1):
            cell = ws.cell(total_row, cidx)
            cell.border = border
            cell.fill = total_fill
            cell.font = bold
            if kind in ('monto', 'iva', 'deduc'):
                col_letter = get_column_letter(cidx)
                cell.value = f"=SUM({col_letter}{HDR_ROW + 1}:{col_letter}{HDR_ROW + n_rows})"
                cell.number_format = '#,##0.00'

    anchos = {
        'fecha': 12, 'pv': 6, 'tipo': 14, 'desde': 11, 'hasta': 11, 'cant_cbtes': 7,
        'doc_cod': 7, 'doc_nro': 14, 'denom': 28,
    }
    for cidx, (_, key, _) in enumerate(columnas, start=1):
        ws.column_dimensions[get_column_letter(cidx)].width = anchos.get(key, 14)

    ws.freeze_panes = ws.cell(HDR_ROW + 1, 1)


def crear_excel_ventas_citi(
    df_consolidado: pd.DataFrame,
    periodo: str,
    output,
    df_original: pd.DataFrame | None = None,
) -> None:
    """Genera Excel CITI Ventas con hojas: TODOS (consolidado), Original (sin agrupar) y una hoja por (PV, Tipo)."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    if df_consolidado.empty:
        ws = wb.create_sheet('Vacío')
        _escribir_hoja_citi(ws, df_consolidado, 'ARMADO CITI VENTAS', f'Periodo: {periodo}  |  Sin datos')
    else:
        # Hoja principal con TODOS los movimientos consolidados.
        n_total = len(df_consolidado)
        cbtes_total = int(df_consolidado['cant_cbtes'].sum())
        ws_todos = wb.create_sheet('TODOS')
        _escribir_hoja_citi(
            ws_todos, df_consolidado.reset_index(drop=True),
            'CITI VENTAS · TODOS LOS MOVIMIENTOS',
            f'Periodo: {periodo}  |  Consolidados: {n_total}  |  Comprobantes originales: {cbtes_total}',
        )

        # Hoja Original: una fila por comprobante del CSV (sin agrupar).
        if df_original is not None and not df_original.empty:
            df_orig_view = df_original.copy()
            df_orig_view['desde'] = df_orig_view['numero']
            df_orig_view['hasta'] = df_orig_view['numero_hasta']
            df_orig_view['cant_cbtes'] = (
                df_orig_view['numero_hasta'] - df_orig_view['numero'] + 1
            ).clip(lower=1)
            df_orig_view = df_orig_view.sort_values(
                ['fecha', 'pv', 'tipo', 'desde']
            ).reset_index(drop=True)
            ws_orig = wb.create_sheet('Original')
            _escribir_hoja_citi(
                ws_orig, df_orig_view,
                'CITI VENTAS · ORIGINAL (SIN AGRUPAR)',
                f'Periodo: {periodo}  |  Filas CSV: {len(df_orig_view)}  |  Comprobantes originales: {int(df_orig_view["cant_cbtes"].sum())}',
            )

        # Una hoja por (PV, Tipo). Orden: PV asc, luego Tipo asc.
        combos = sorted(
            df_consolidado[['pv', 'tipo']].drop_duplicates().itertuples(index=False, name=None)
        )
        used_names = set()
        for pv, tipo in combos:
            df_grp = df_consolidado[
                (df_consolidado['pv'] == pv) & (df_consolidado['tipo'] == tipo)
            ].reset_index(drop=True)
            n_rows = len(df_grp)
            total_cbtes = int(df_grp['cant_cbtes'].sum())
            tipo_desc = _citi_desc_tipo(tipo)
            sheet_name = f'PV {int(pv):05d} - {tipo_desc}'[:31]
            base_name = sheet_name
            i = 2
            while sheet_name in used_names:
                suf = f' ({i})'
                sheet_name = base_name[:31 - len(suf)] + suf
                i += 1
            used_names.add(sheet_name)
            ws = wb.create_sheet(sheet_name)
            subtitulo = (
                f'Periodo: {periodo}  |  PV {int(pv):05d}  |  Tipo: {tipo_desc} ({int(tipo)})  |  '
                f'Consolidados: {n_rows}  |  Comprobantes originales: {total_cbtes}'
            )
            _escribir_hoja_citi(
                ws, df_grp,
                f'CITI VENTAS · PV {int(pv):05d} · {tipo_desc}',
                subtitulo,
            )

    if hasattr(output, 'write'):
        wb.save(output)
    else:
        wb.save(str(output))


# ════════════════════════════════════════════════════════════════════════════════
#  IMPORTACION DE RETENCIONES IVA / GANANCIAS — XLS ARCA → ZIP Portal IVA
# ════════════════════════════════════════════════════════════════════════════════

# Headers del CSV de salida — replican exactamente la hoja "FORMULA PARA IMPORTAR
# RETENCION" del template Excel original (con doble-encoding mojibake en algunos
# nombres). Mantenerlos byte-equivalentes al template asegura que el sistema
# Mendez los acepte igual que cuando el operador exporta a CSV manualmente.
RETENCIONES_OUTPUT_HEADERS = [
    "Fecha de EmisiÃ³n",
    "Tipo de Comprobante",
    "Punto de Venta",
    "NÃºmero de Comprobante",
    "Tipo Doc. Vendedor",
    "Nro. Doc. Vendedor",
    "DenominaciÃ³n Vendedor",
    "Importe Total",
    "Moneda Original",
    "Tipo de Cambio",
    "Importe No Gravado",
    "Importe Exento",
    "CrÃ©dito Fiscal Computable",
    " Importe de Per. o Pagos a Cta. de Otros Imp. Nac. ",
    "Importe de Percepciones de Ingresos Brutos",
    "Importe de Impuestos Municipales",
    "Importe de Percepciones o Pagos a Cuenta de IVA",
    "Importe de Impuestos Internos",
    "Importe Otros Tributos",
    "Neto Gravado IVA 0%",
    "Neto Gravado IVA 2,5%",
    "Importe IVA 2,5%",
    "Neto Gravado IVA 5%",
    "Importe IVA 5%",
    "Neto Gravado IVA 10,5%",
    "Importe IVA 10,5%",
    "Neto Gravado IVA 21%",
    "Importe IVA 21%",
    "Neto Gravado IVA 27%",
    "Importe IVA 27%",
    "Total Neto Gravado",
    "Total IVA",
]

# Headers obligatorios del XLS de Mis Retenciones/Percepciones de ARCA. El parser
# valida que estén presentes antes de transformar.
RETENCIONES_INPUT_HEADERS = [
    'CUIT Agente Ret./Perc.',
    'Denominación o Razón Social',
    'Fecha Ret./Perc.',
    'Número Certificado',
    'Importe Ret./Perc.',
]


def parsear_arca_retenciones_xls(file_bytes: bytes) -> pd.DataFrame:
    """Lee el XLS de Mis Retenciones/Percepciones de ARCA y valida estructura.

    Devuelve el DataFrame crudo. Lanza ValueError si faltan columnas requeridas
    o si el archivo está vacío.
    """
    df = pd.read_excel(io.BytesIO(file_bytes))
    if df.empty:
        raise ValueError("El archivo XLS está vacío.")
    faltantes = [c for c in RETENCIONES_INPUT_HEADERS if c not in df.columns]
    if faltantes:
        raise ValueError(
            "El archivo no parece ser un XLS de Mis Retenciones/Percepciones de ARCA. "
            f"Faltan columnas: {faltantes}"
        )
    return df


def _fmt_importe_arg(val) -> str:
    """Formatea un número en notación argentina (1.234,56) con 2 decimales fijos."""
    if val is None or pd.isna(val):
        return ''
    s = f"{float(val):,.2f}"
    return s.replace(',', '\x00').replace('.', ',').replace('\x00', '.')


def _fmt_fecha_iso(s) -> str:
    """Convierte 'DD/MM/YYYY' → 'YYYY-MM-DD'. Acepta datetime, string ISO o vacío."""
    if s is None:
        return ''
    if hasattr(s, 'strftime'):
        return s.strftime('%Y-%m-%d')
    if isinstance(s, float) and pd.isna(s):
        return ''
    s = str(s).strip()
    if not s:
        return ''
    if len(s) == 10 and s[2] == '/' and s[5] == '/':
        return f"{s[6:10]}-{s[3:5]}-{s[0:2]}"
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        return s[:10]
    return s


def _str_int_safe(val) -> str:
    """Convierte un valor numérico (int/float) a string sin '.0' al final."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def transformar_retenciones_a_csv_arca(df: pd.DataFrame):
    """Transforma el DataFrame del XLS ARCA al CSV formato Portal IVA.

    Aplica el mapeo de la hoja "FORMULA PARA IMPORTAR RETENCION" del template:
      - Fecha de Emisión = Fecha Ret./Perc. reformateada YYYY-MM-DD
      - Tipo Cbte = 99 (constante)
      - PV = primeros 2 chars del Número Certificado
      - Nro = últimos 8 chars del Número Certificado
      - Tipo Doc Vend = 80 (constante, CUIT)
      - Nro Doc = CUIT Agente Ret./Perc.
      - Denominación = Denominación o Razón Social
      - Importe Total = Importe Ret./Perc.
      - Moneda = PES, TC = 1
      - Importe de Per. o Pagos a Cta. de Otros Imp. Nac. = Importe Ret./Perc.

    Devuelve (csv_text, periodo_yyyymm) donde periodo_yyyymm es el mes/año más
    frecuente de Fecha Ret./Perc.
    """
    from collections import Counter
    from datetime import datetime as _dt

    rows = []
    meses = []
    for _, src in df.iterrows():
        fecha_iso = _fmt_fecha_iso(src['Fecha Ret./Perc.'])
        if len(fecha_iso) == 10:
            meses.append(fecha_iso[:4] + fecha_iso[5:7])  # YYYYMM

        cert_s = _str_int_safe(src['Número Certificado'])
        cuit_s = _str_int_safe(src['CUIT Agente Ret./Perc.'])

        denom = src['Denominación o Razón Social']
        denom_s = '' if (denom is None or (isinstance(denom, float) and pd.isna(denom))) else str(denom).strip()

        importe_s = _fmt_importe_arg(src['Importe Ret./Perc.'])

        rows.append([
            fecha_iso,    # 1  Fecha de Emisión
            '99',         # 2  Tipo de Comprobante
            cert_s[:2],   # 3  Punto de Venta
            cert_s[-8:],  # 4  Número de Comprobante
            '80',         # 5  Tipo Doc. Vendedor
            cuit_s,       # 6  Nro. Doc. Vendedor
            denom_s,      # 7  Denominación Vendedor
            importe_s,    # 8  Importe Total
            'PES',        # 9  Moneda Original
            '1',          # 10 Tipo de Cambio
            '',           # 11 Importe No Gravado
            '',           # 12 Importe Exento
            '',           # 13 Crédito Fiscal Computable
            importe_s,    # 14 Importe de Per. o Pagos a Cta. de Otros Imp. Nac.
        ] + [''] * 18)    # 15-32 vacíos

    if meses:
        periodo = Counter(meses).most_common(1)[0][0]
    else:
        periodo = _dt.now().strftime('%Y%m')

    df_out = pd.DataFrame(rows, columns=RETENCIONES_OUTPUT_HEADERS)
    csv_text = df_out.to_csv(sep=';', index=False, lineterminator='\n')
    return csv_text, periodo


def generar_zip_retenciones_arca(csv_text: str, periodo_yyyymm: str, *, now=None):
    """Empaqueta el CSV en un .zip con el patrón de naming del Portal IVA.

    Patrón: comprobantes_periodo_{YYYYMM}_compras_{YYYYMMDD}_{HHMM}.zip
    El CSV interno usa el mismo nombre base con extensión .csv. Encoding latin-1.
    Devuelve (zip_bytes, zip_name).
    """
    import zipfile
    from datetime import datetime as _dt

    if now is None:
        now = _dt.now()
    timestamp = now.strftime('%Y%m%d_%H%M')
    basename = f"comprobantes_periodo_{periodo_yyyymm}_compras_{timestamp}"
    zip_name = f"{basename}.zip"
    csv_name = f"{basename}.csv"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(csv_name, csv_text.encode('latin-1', errors='replace'))
    return buf.getvalue(), zip_name


def main():
    # Forzar UTF-8 en la consola de Windows (solo cuando se corre como script)
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

    if len(sys.argv) < 2:
        input_file = seleccionar_archivo()
    else:
        input_file = Path(sys.argv[1])

    if not input_file.exists():
        print(f"❌ No se encontró el archivo: {input_file}")
        sys.exit(1)

    output_file = input_file.with_suffix('.xlsx')

    print(f"📖 Leyendo: {input_file}")
    transacciones, meta = parsear_archivo(path=input_file)

    if not transacciones:
        print("❌ No se encontraron transacciones en el archivo.")
        sys.exit(1)

    crear_excel(transacciones, meta, output_file)
    print("✅ Proceso completado exitosamente.")


if __name__ == '__main__':
    main()
