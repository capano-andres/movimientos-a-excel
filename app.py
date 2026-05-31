import sys
sys.setrecursionlimit(10_000)  # default 1000 se queda corto para xlrd parseando .xls en Streamlit Cloud

import streamlit as st
import io
import re
import zipfile
import hashlib
import pandas as pd
import openpyxl
import PyPDF2
from pathlib import Path
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import requests

# Los XLS exportados por Mendez/AddiSyc usan BIFF3 ("Markus Wiederstein Excel Format"),
# limitado por xlrd a 16.384 filas. Subimos el tope al maximo de Excel moderno
# para evitar el AssertionError en put_cell_unragged cuando el archivo es grande.
import xlrd.sheet as _xlrd_sheet
_orig_sheet_init = _xlrd_sheet.Sheet.__init__
def _patched_sheet_init(self, *args, **kwargs):
    _orig_sheet_init(self, *args, **kwargs)
    self.utter_max_rows = 1048576
_xlrd_sheet.Sheet.__init__ = _patched_sheet_init

from extractor_movimientos import parsear_archivo, crear_excel, crear_excel_consolidado_simple, generar_sifere_txt, generar_sifere_retenciones_txt, generar_percepciones_arba, generar_arba_desde_excel, generar_retenciones_arba, generar_retenciones_arba_desde_excel, construir_sistema_aux_set, CONCEPTOS_MAP, normalizar_csv_ventas_arca, consolidar_ventas_citi, generar_citi_ventas_lineas, generar_citi_alicuotas_lineas, crear_excel_ventas_citi, parsear_arca_retenciones_xls, transformar_retenciones_a_csv_arca, generar_zip_retenciones_arca, asignar_mes_por_xls_mendez, crear_excel_asiento_anual

@st.cache_data(show_spinner=False)
def obtener_razon_social_cuitonline(cuit):
    clean_cuit = re.sub(r'[^0-9]', '', str(cuit)).strip()
    url = f"https://www.cuitonline.com/detalle/{clean_cuit}/"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "es-AR,es;q=0.9",
    }
    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            html = response.text
            if "detailResults" not in html and "Actividades" not in html:
                return None
            
            # Buscar en H1
            h1_match = re.search(r'<h1[^>]*>([^<]+)</h1>', html, re.IGNORECASE)
            if h1_match:
                return h1_match.group(1).strip()
            
            # Fallback en Title
            title_match = re.search(r'<title>([^(]+)\(', html)
            if title_match:
                return title_match.group(1).strip()
    except Exception:
        pass
    return None


# --- Page Config ---
st.set_page_config(
    page_title="ADDISYC ETL",
    page_icon="📗",
    layout="centered"
)

# --- Styling ---
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;600;800&display=swap');

:root {
    --bg:        #0d0f14;
    --surface:   #141720;
    --border:    #252935;
    --accent:    #e8c84a;
    --accent2:   #4ae8a0;
    --text:      #e4e8f0;
    --muted:     #6b7280;
    --danger:    #f87171;
    --radius:    10px;
}

*, *::before, *::after { box-sizing: border-box; }

.stApp {
    background-color: var(--bg) !important;
    font-family: 'Syne', sans-serif;
    color: var(--text);
}

.block-container {
    padding-top: 2.5rem !important;
    padding-bottom: 3rem !important;
    max-width: 860px !important;
}

h1, h2, h3, h4, p, span, div, label {
    color: var(--text) !important;
}

/* Header */
.etl-logo {
    font-family: 'Space Mono', monospace;
    font-size: 0.7rem;
    letter-spacing: 0.35em;
    color: var(--accent) !important;
    text-transform: uppercase;
    text-align: center;
    margin-bottom: 0.8rem;
}
.etl-title {
    font-family: 'Syne', sans-serif !important;
    font-weight: 800;
    font-size: 3.4rem !important;
    line-height: 1.4;
    color: var(--text) !important;
    text-align: center;
    margin: 0 0 0.5rem !important;
}
.etl-title span { color: var(--accent) !important; }
.etl-subtitle {
    font-size: 0.85rem;
    color: var(--muted) !important;
    font-family: 'Space Mono', monospace;
    letter-spacing: 0.05em;
    text-align: center;
}
.divider {
    border: none;
    border-top: 1px solid var(--border);
    margin: 1.8rem 0;
}

/* Cards */
.card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 1.6rem 1.8rem;
    margin-bottom: 1.2rem;
    position: relative;
    overflow: hidden;
}
.card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, var(--accent), transparent);
}
.card-label {
    font-family: 'Space Mono', monospace;
    font-size: 0.75rem;
    font-weight: 700;
    letter-spacing: 0.2em;
    color: var(--accent) !important;
    text-transform: uppercase;
    margin-bottom: 1rem;
}

/* File uploader */
[data-testid="stFileUploader"] > div,
[data-testid="stFileUploader"] > div > div,
[data-testid="stFileUploader"] section,
[data-testid="stFileUploader"] section > div,
[data-testid="stFileUploadDropzone"],
[data-testid="stFileDropzoneInstructions"],
.stFileUploader > div,
.stFileUploader section {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border: 1.5px dashed var(--border) !important;
    border-radius: var(--radius) !important;
    transition: border-color 0.2s ease;
}
[data-testid="stFileUploader"] > div:hover,
[data-testid="stFileUploadDropzone"]:hover,
.stFileUploader > div:hover {
    border-color: var(--accent) !important;
}
.stFileUploader label, [data-testid="stFileUploader"] label {
    color: var(--muted) !important;
}
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] span,
[data-testid="stFileDropzoneInstructions"] span,
[data-testid="stFileDropzoneInstructions"] small,
[data-testid="stFileDropzoneInstructions"] div {
    color: var(--muted) !important;
}
.stFileUploader button, [data-testid="stFileUploader"] button {
    background: var(--surface) !important;
    color: var(--accent) !important;
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 0.75rem !important;
}
.stFileUploader button:hover, [data-testid="stFileUploader"] button:hover {
    border-color: var(--accent) !important;
}

/* Traducir textos del file uploader al español */
/* Ocultar todo el texto original dentro del dropzone instructions */
[data-testid="stFileUploaderDropzoneInstructions"] span,
[data-testid="stFileUploaderDropzoneInstructions"] small,
[data-testid="stFileUploaderDropzoneInstructions"] div,
[data-testid="stFileDropzoneInstructions"] span,
[data-testid="stFileDropzoneInstructions"] small,
[data-testid="stFileDropzoneInstructions"] div {
    font-size: 0 !important;
    line-height: 0 !important;
    color: transparent !important;
}
/* Restaurar el ícono SVG */
[data-testid="stFileUploaderDropzoneInstructions"] svg,
[data-testid="stFileDropzoneInstructions"] svg {
    font-size: initial !important;
    line-height: initial !important;
    color: var(--muted) !important;
    fill: var(--muted) !important;
}
/* Inyectar texto en español en el contenedor principal */
[data-testid="stFileUploaderDropzoneInstructions"]::after,
[data-testid="stFileDropzoneInstructions"]::after {
    content: "Arrastrá y soltá el archivo acá\A Límite de 200MB por archivo";
    white-space: pre;
    font-family: 'Syne', sans-serif;
    font-size: 0.9rem;
    line-height: 1.4;
    color: var(--muted) !important;
    margin-left: 0.8rem;
    display: inline-block;
    vertical-align: middle;
}
/* Botón "Browse files" → "Buscar archivo" (SOLO el del dropzone, NO el de eliminar archivo) */
[data-testid="stFileUploaderDropzone"] button {
    font-size: 0 !important;
    color: transparent !important;
    position: relative;
    min-width: 140px !important;
    width: auto !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    white-space: nowrap !important;
}
[data-testid="stFileUploaderDropzone"] button * {
    font-size: 0 !important;
    color: transparent !important;
}
[data-testid="stFileUploaderDropzone"] button::after {
    content: "Buscar archivo";
    font-size: 0.75rem !important;
    color: var(--accent) !important;
    font-family: 'Space Mono', monospace !important;
    letter-spacing: 0.04em;
    position: absolute;
    inset: 0;
    display: flex;
    align-items: center;
    justify-content: center;
}

/* Checkbox */
.stCheckbox label span { color: var(--text) !important; }
[data-testid="stCheckbox"] > label > div {
    border-color: var(--border) !important;
}

/* Main action button */
.stButton > button {
    width: 100% !important;
    background: var(--accent) !important;
    color: #0a0c10 !important;
    border: none !important;
    border-radius: var(--radius) !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 800 !important;
    font-size: 1rem !important;
    letter-spacing: 0.08em;
    height: 3.2em !important;
    margin-top: 0.5rem;
    transition: all 0.18s ease !important;
    box-shadow: 0 0 20px rgba(232,200,74,0.15);
    text-shadow: none !important;
    -webkit-text-fill-color: #0a0c10 !important;
}
.stButton > button:hover {
    background: #f5d84e !important;
    box-shadow: 0 0 30px rgba(232,200,74,0.3) !important;
    transform: translateY(-1px);
}
.stButton > button:active { transform: translateY(0); }

/* Download button */
[data-testid="stDownloadButton"] > button {
    background: transparent !important;
    color: var(--accent2) !important;
    border: 1.5px solid var(--accent2) !important;
    border-radius: var(--radius) !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.06em;
    width: 100% !important;
    height: 3em !important;
    margin-top: 0.8rem;
    transition: all 0.18s ease !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: rgba(74,232,160,0.08) !important;
    box-shadow: 0 0 20px rgba(74,232,160,0.2) !important;
}

/* Alerts */
[data-testid="stAlert"] {
    border-radius: var(--radius) !important;
}
.stSuccess > div {
    background: rgba(74,232,160,0.07) !important;
    border: 1px solid rgba(74,232,160,0.25) !important;
}
.stSuccess p, .stSuccess span, .stSuccess strong { color: var(--accent2) !important; }

.stError > div {
    background: rgba(248,113,113,0.07) !important;
    border: 1px solid rgba(248,113,113,0.3) !important;
}
.stError p, .stError span { color: var(--danger) !important; }

.stWarning > div {
    background: rgba(232,200,74,0.07) !important;
    border: 1px solid rgba(232,200,74,0.25) !important;
}
.stWarning p, .stWarning span { color: var(--accent) !important; }

.stInfo > div {
    background: rgba(99,122,255,0.07) !important;
    border: 1px solid rgba(99,122,255,0.3) !important;
}
.stInfo p, .stInfo span, .stInfo strong { color: #a5b4fc !important; }

/* Spinner */
.stSpinner > div { border-top-color: var(--accent) !important; }

/* Stats row */
.stats-row {
    display: flex;
    gap: 0.8rem;
    margin-top: 1rem;
}
.stat-chip {
    flex: 1;
    background: #0a0c10;
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 0.7rem 0.5rem;
    text-align: center;
}
.stat-chip .stat-val {
    font-family: 'Space Mono', monospace;
    font-size: 1.3rem;
    font-weight: 700;
    color: var(--accent) !important;
    display: block;
}
.stat-chip .stat-lbl {
    font-size: 0.65rem;
    letter-spacing: 0.1em;
    color: var(--muted) !important;
    text-transform: uppercase;
    display: block;
    margin-top: 0.2rem;
}

/* Scrollbar */
::-webkit-scrollbar { width: 5px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 99px; }

/* Navbar / Header bar */
header[data-testid="stHeader"],
.stAppHeader,
header.stAppHeader {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border-bottom: 1px solid var(--border) !important;
}
header[data-testid="stHeader"] *,
.stAppHeader * {
    color: var(--accent) !important;
}

/* Selectbox */
[data-testid="stSelectbox"] > div > div,
.stSelectbox > div > div {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border: 1.5px solid var(--border) !important;
    border-radius: var(--radius) !important;
    color: var(--accent) !important;
}
[data-testid="stSelectbox"] > div > div:hover,
.stSelectbox > div > div:hover {
    border-color: var(--accent) !important;
}
[data-testid="stSelectbox"] span,
[data-testid="stSelectbox"] div[data-baseweb="select"] span,
.stSelectbox span {
    color: var(--accent) !important;
}
[data-testid="stSelectbox"] svg {
    fill: var(--accent) !important;
}
[data-testid="stSelectbox"] label {
    color: var(--muted) !important;
}
/* Selectbox dropdown menu */
[data-baseweb="popover"],
[data-baseweb="popover"] > div,
[data-baseweb="menu"],
ul[role="listbox"],
div[data-baseweb="popover"] div,
div[data-baseweb="popover"] ul {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border-color: var(--border) !important;
}
[data-baseweb="popover"] {
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
}
ul[role="listbox"] li,
[data-baseweb="menu"] li,
[data-baseweb="popover"] li,
li[role="option"] {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    color: var(--text) !important;
}
ul[role="listbox"] li:hover,
[data-baseweb="menu"] li:hover,
[data-baseweb="popover"] li:hover,
li[role="option"]:hover,
ul[role="listbox"] li[aria-selected="true"],
[data-baseweb="menu"] li[aria-selected="true"],
li[role="option"][aria-selected="true"] {
    background: var(--surface) !important;
    background-color: var(--surface) !important;
    color: var(--accent) !important;
}

/* Radio buttons */
[data-testid="stRadio"] > div > div > label > div {
    background: #1a1d24 !important;
}

/* Footer */
.etl-footer {
    text-align: center;
    padding-top: 2rem;
    font-family: 'Space Mono', monospace;
    font-size: 0.62rem;
    color: var(--muted) !important;
    letter-spacing: 0.15em;
}
</style>
""", unsafe_allow_html=True)


# ─── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<!-- Dani ♥ -->
<div>
    <h1 class="etl-title">Transformación<span> Datos Mendez</span></h1>
    <p class="etl-subtitle">TXT  →  XLSX</p>
    <span aria-hidden="true" style="position:absolute;left:-9999px;top:auto;width:1px;height:1px;overflow:hidden;">Dani ♥</span>
</div>
<hr class="divider">
""", unsafe_allow_html=True)


# ─── Selector de herramienta ────────────────────────────────────────────────────────────
TOOL_MOVIMIENTOS = "Listado por fecha TXT Mendez a Excel limpio"
TOOL_PORTAL_IVA = "Archivo .zip PORTAL IVA"
TOOL_SIFERE = "Archivos SIFERE (.txt)"
TOOL_LIQUIDACIONES = "Liquidaciones Tarjeta FISERV (.pdf)"
TOOL_DEDUCCIONES = "Limpieza Excel Deducciones IVA/Ganancias"
TOOL_ARBA = "Agentes de Recaudacion ARBA"
TOOL_CRUCE_CONCEPTO = "Excel Mendez + TXT Mendez"
TOOL_CM05 = "Papeles de Trabajo CM05"
TOOL_CRUCE_DEDUCCIONES = "Cruce de Deducciones"
TOOL_IMPORTACION = "Importacion Compras (TXT + ZIP ARCA -> ZIPs por Concepto)"
TOOL_VENTAS_CITI = "Armado .zip Importacion Ventas / CITI (ZIP ARCA -> VENTAS.txt + ALICUOTAS.txt)"
TOOL_RETENCIONES = "Importacion Retenciones IVA / Ganancias (XLS ARCA -> .zip Portal IVA)"

herramienta = st.selectbox(
    "Seleccioná la herramienta:",
    options=[TOOL_MOVIMIENTOS, TOOL_PORTAL_IVA, TOOL_SIFERE, TOOL_ARBA, TOOL_LIQUIDACIONES, TOOL_DEDUCCIONES, TOOL_CRUCE_CONCEPTO, TOOL_CM05, TOOL_CRUCE_DEDUCCIONES, TOOL_IMPORTACION, TOOL_VENTAS_CITI, TOOL_RETENCIONES],
    index=0,
)

if herramienta in (TOOL_MOVIMIENTOS, TOOL_PORTAL_IVA):
    with st.expander("📋 Códigos de comprobantes ARCA"):
        st.markdown("""
| Código | Tipo | Código | Tipo | Código | Tipo |
|--------|------|--------|------|--------|------|
| 1 | FC A | 2 | ND A | 3 | NC A |
| 6 | FC B | 7 | ND B | 8 | NC B |
| 11 | FC C | 12 | ND C | 13 | NC C |
| 51 | FC M | 52 | ND M | 53 | NC M |
| 19 | FC | 20 | ND | 21 | NC |
| 22 | FC | 37 | ND | 38 | NC |
| 195 | FC T | 196 | ND T | 197 | NC T |
| 201 | FC A | 202 | ND A | 203 | NC A |
| 206 | FC B | 207 | ND B | 208 | NC B |
| 211 | FC C | 212 | ND C | 213 | NC C |
| 81 | TF A | 45 | ND A | 48 | NC A |
| 82 | TF B | 46 | ND B | 43 | NC B |
| 111 | TF C | 47 | ND C | 44 | NC C |
| 118 | TF M | 90 | NC | 83 | TK |
| 109 | TK C | 110 | TK | 112 | TK A |
| 113 | TK B | 114 | TK C | 115 | TK A |
| 116 | TK B | 117 | TK C | 119 | TK M |
| 120 | TK M | 4 | RC A → FC A | 9 | RC B → FC B |
| 15 | RC C → FC C | | | | |

**FC** = Factura · **NC** = Nota de Crédito · **ND** = Nota de Débito · **TF** = Tique Factura · **TK** = Tique · **RC** = Recibo (se trata como FC)
        """)

# ─── Parser de Liquidaciones VISA Santander (PDF) ──────────────────────────────────
# Estructura del PDF: bloques "FECHA DE PAGO DD/MM" con uno o más Liq. N°/Lote N°,
# detalle de ventas y descuentos por línea (Arancel, Serv.Costos Financieros, Servicio
# PAYWAY, Deduc.Impositivas) más una línea "Total del día" con monto presentado, total
# descuentos y neto percibido. Al final del PDF: "DESGLOSE DE DESCUENTOS" con montos
# consolidados (arancel por tipo, plan cuotas por modalidad, base imponible IVA,
# deducciones impositivas, percep./retenc. AFIP-DGI).
RE_SANT_BLOQUE_FECHA = re.compile(
    r'FECHA DE PAGO\s+(\d{2}/\d{2})(.*?)(?=FECHA DE PAGO\s+\d{2}/\d{2}|SE ACREDITO EN|DESGLOSE DE DESCUENTOS|Fin de la informaci|\Z)',
    re.DOTALL,
)
RE_SANT_FECHA_PRES = re.compile(r'Fecha de presentaci[oó]n\s+(\d{2}/\d{2})')
RE_SANT_LIQ_LOTE = re.compile(r'Liq\.\s*N[°º]\s*(\d+)\s*-\s*Lote\s*N[°º]\s*(\d+)')
RE_SANT_VENTA = re.compile(r'(\d+\s+Vent[as]+\s+[^\n$]+?)\$\s*([\d.]+,\d{2})')
RE_SANT_TOTAL_DIA = re.compile(
    r'Total del d[ií]a\s*\$?\s*([\d.]+,\d{2})\s*\$?\s*([\d.]+,\d{2})\s*\$?\s*([\d.]+,\d{2})',
    re.DOTALL,
)
RE_SANT_DESC_LINE = re.compile(
    r'(Arancel|Serv\.Costos Financieros|Servicio PAYWAY|Serv\.Cobro Anticipado|Deduc\.Impositivas)\s*\$\s*([\d.]+,\d{2})',
    re.IGNORECASE,
)
RE_SANT_LINEA_MONTO = re.compile(r'^(.*?)\$\s*(-?[\d.]+,\d{2})\s*$')
RE_SANT_TASA_LINE = re.compile(r'Tasa\s+([\d,]+\s*%)')
RE_SANT_FECHA_EMI = re.compile(r'FECHA DE EMISION:\s*\d{2}/\d{2}/(\d{4})')

_SANT_SECTION_HEADERS = {
    "arancel": re.compile(r'^\s*Arancel\s*$', re.IGNORECASE),
    "scf": re.compile(r'^\s*Servicio Costos Financieros\s*$', re.IGNORECASE),
    "payway": re.compile(r'^\s*Servicio\s+PAYWAY\b', re.IGNORECASE),
    "base_iva": re.compile(r'^\s*Base Imponible IVA\b', re.IGNORECASE),
    "deduc": re.compile(r'^\s*Deducciones Impositivas\s*$', re.IGNORECASE),
    "afip": re.compile(r'^\s*Percep\.\/?\s*Retenc\.\s*AFIP', re.IGNORECASE),
    "ecommerce": re.compile(r'^\s*Serv\.\s*Ecommerce\s*$', re.IGNORECASE),
}


def _sant_parse_monto(s: str) -> float:
    return float(s.strip().replace(".", "").replace(",", "."))


def _sant_parsear_desglose(texto: str) -> dict:
    """Parsea la sección DESGLOSE DE DESCUENTOS (o Serv.Ecommerce en PDFs vacíos)
    con un state machine que tolera labels multi-línea y headers mezclados."""
    res = {
        "arancel": [],
        "serv_costos_financieros": [],
        "servicio_payway": [],
        "base_imp_iva": None,
        "deducciones_impositivas": [],
        "afip_dgi": None,
    }
    m = re.search(
        r'DESGLOSE DE DESCUENTOS(.*?)(?:Fin de la informaci|Constancia Mensual|SR\. COMERCIANTE|\Z)',
        texto, re.DOTALL,
    )
    cuerpo = m.group(1) if m else ""

    if not cuerpo:
        m_eco = re.search(
            r'(Serv\.\s*Ecommerce.*?)(?:SR\. COMERCIANTE|Fin de la informaci|\Z)',
            texto, re.DOTALL,
        )
        cuerpo = m_eco.group(1) if m_eco else ""
        if cuerpo:
            etiqueta_buffer = []
            for raw in cuerpo.split("\n"):
                line = raw.strip()
                if not line or _SANT_SECTION_HEADERS["ecommerce"].match(line):
                    continue
                m_l = RE_SANT_LINEA_MONTO.match(line)
                if m_l:
                    label = " ".join([*etiqueta_buffer, m_l.group(1).strip()]).strip() or "Cargo por Servicio"
                    res["servicio_payway"].append((label, _sant_parse_monto(m_l.group(2))))
                    etiqueta_buffer = []
                else:
                    etiqueta_buffer.append(line)
        return res

    state = None
    sub_label = ""
    etiqueta_buffer = []
    base_iva_tasa = None

    for raw in cuerpo.split("\n"):
        line = raw.strip()
        if not line:
            continue

        cambio = None
        for sec, rx in _SANT_SECTION_HEADERS.items():
            if rx.match(line):
                cambio = sec
                break
        if cambio:
            state = cambio
            sub_label = ""
            etiqueta_buffer = []
            if cambio == "afip":
                m_l = RE_SANT_LINEA_MONTO.match(line)
                if m_l:
                    res["afip_dgi"] = (
                        "Percep./Retenc. AFIP-DGI (RG 796)",
                        _sant_parse_monto(m_l.group(2)),
                    )
            continue

        # SCF subcategoría: PyPDF2 a veces colapsa subcategoría + primera entrada
        # ("-Plan Cuotas       10 Ventas en  3 cuotas $ 534.808,00") en una sola línea.
        if state == "scf" and line.startswith("-"):
            m_l = RE_SANT_LINEA_MONTO.match(line)
            if m_l:
                rest = m_l.group(1).lstrip("- ").strip()
                sm = re.match(r'(.+?)\s{2,}(\d+\s+Vent[as]+.*)', rest) \
                    or re.match(r'(.+?)\s+(\d+\s+Vent[as]+.*)', rest)
                if sm:
                    sub_label = sm.group(1).strip()
                    full = f"{sub_label} - {sm.group(2).strip()}"
                else:
                    sub_label = rest
                    full = rest
                res["serv_costos_financieros"].append((full, _sant_parse_monto(m_l.group(2))))
                etiqueta_buffer = []
            else:
                sub_label = line.lstrip("- ").strip()
                etiqueta_buffer = []
            continue

        if state == "base_iva":
            t = RE_SANT_TASA_LINE.search(line)
            if t:
                base_iva_tasa = t.group(1).strip()
                m_l = RE_SANT_LINEA_MONTO.match(line)
                if m_l:
                    res["base_imp_iva"] = (
                        f"Tasa {base_iva_tasa} (Monto Gravado)",
                        _sant_parse_monto(m_l.group(2)),
                    )
                continue
            m_l = RE_SANT_LINEA_MONTO.match(line)
            if m_l and base_iva_tasa:
                res["base_imp_iva"] = (
                    f"Tasa {base_iva_tasa} (Monto Gravado)",
                    _sant_parse_monto(m_l.group(2)),
                )
                continue

        m_l = RE_SANT_LINEA_MONTO.match(line)
        if m_l:
            label = " ".join([*etiqueta_buffer, m_l.group(1).strip()]).strip()
            monto = _sant_parse_monto(m_l.group(2))
            if state == "arancel":
                lbl = re.sub(r'^Arancel\s+', '', label).strip() or label
                res["arancel"].append((lbl, monto))
            elif state == "scf":
                lbl = f"{sub_label} - {label}" if sub_label else label
                res["serv_costos_financieros"].append((lbl, monto))
            elif state == "payway":
                res["servicio_payway"].append((label or "Cargo por Servicio", monto))
            elif state == "deduc":
                res["deducciones_impositivas"].append((label, monto))
            elif state == "afip":
                res["afip_dgi"] = ("Percep./Retenc. AFIP-DGI (RG 796)", monto)
            etiqueta_buffer = []
        else:
            etiqueta_buffer.append(line)

    return res


def _sant_extract_rate(label: str) -> tuple:
    """Extrae la tasa numérica de un label como 'IVA  21,00 %' → (0.21, '21,00')."""
    m = re.search(r'(\d+(?:,\d+)?)\s*%', label)
    if not m:
        return (0.0, "")
    rate_str = m.group(1)
    return (float(rate_str.replace(",", ".")) / 100, rate_str)


def _sant_parsear_constancia(texto: str) -> list:
    """Parsea la Constancia RG 796 - Percepción IVA (Reg.493) al final del PDF.
    Devuelve lista de dicts {fecha_pago, liq, importe}; vacío si no hay constancia."""
    m = re.search(
        r'Constancia Mensual.*?(?=Fin de la informaci|\Z)',
        texto, re.DOTALL,
    )
    if not m:
        return []
    body = m.group(0)
    items = []
    for fecha, liq, imp in re.findall(
        r'(\d{2}/\d{2}/\d{4})\s+(\d+)\s+\$\s*([\d.]+,\d{2})',
        body,
    ):
        items.append({
            "fecha_pago": fecha,
            "liq": liq,
            "importe": _sant_parse_monto(imp),
        })
    return items


def parsear_pdf_santander(texto: str) -> dict:
    """Parsea un PDF de liquidación VISA Santander. Devuelve dict con `liquidaciones`
    (una fila por FECHA DE PAGO con IVA/Perc.IB/SIRTAC calculados con las alícuotas
    del desglose), `desglose` (totales del periodo) y `meta` (razón social, periodo,
    nro de resumen, etiquetas de tasas)."""
    desglose = _sant_parsear_desglose(texto)

    # Alícuotas: leer del desglose; si no aparecen, usar defaults razonables.
    tasa_iva, lbl_iva = 0.21, "21,00"
    tasa_perc_ib, lbl_perc_ib = 0.04, "4,00"
    tasa_sirtac, lbl_sirtac = 0.015, "1,50"
    for label, _ in desglose.get("deducciones_impositivas", []):
        rate, rate_str = _sant_extract_rate(label)
        if rate <= 0:
            continue
        up = label.upper()
        if "SIRTAC" in up:
            tasa_sirtac, lbl_sirtac = rate, rate_str
        elif "BUENOS AIRES" in up or "PERC. IB" in up or "PERC IB" in up:
            tasa_perc_ib, lbl_perc_ib = rate, rate_str
        elif "IVA" in up and "PERC" not in up:
            tasa_iva, lbl_iva = rate, rate_str

    # Año del periodo: priorizar la FECHA DE EMISION (encabezado), no la primera línea
    fecha_emi_match = re.search(
        r'FECHA DE EMISION:\s*\n.*?(\d{2})/(\d{2})/(\d{4})',
        texto, re.DOTALL,
    )
    año = fecha_emi_match.group(3) if fecha_emi_match else "2025"

    # Constancia RG 796 → AFIP-DGI por fecha de pago.  La Constancia lista cada
    # percepción IVA (RG 2408 Reg.493) con su Liq. N° y fecha. Sumamos por fecha
    # para asignar a cada bloque "FECHA DE PAGO" del cuerpo.
    constancia = _sant_parsear_constancia(texto)
    afip_por_fecha = {}
    for it in constancia:
        afip_por_fecha[it["fecha_pago"]] = afip_por_fecha.get(it["fecha_pago"], 0.0) + it["importe"]

    liquidaciones = []
    for m in RE_SANT_BLOQUE_FECHA.finditer(texto):
        fecha_pago_dm = m.group(1)
        cuerpo = m.group(2)

        fecha_pres_match = RE_SANT_FECHA_PRES.search(cuerpo)
        fecha_pres_dm = fecha_pres_match.group(1) if fecha_pres_match else ""

        liqs = RE_SANT_LIQ_LOTE.findall(cuerpo)
        ventas = RE_SANT_VENTA.findall(cuerpo)

        total_match = RE_SANT_TOTAL_DIA.search(cuerpo)
        if total_match:
            monto_pres = _sant_parse_monto(total_match.group(1))
            neto = _sant_parse_monto(total_match.group(3))
        else:
            monto_pres = sum(_sant_parse_monto(v[1]) for v in ventas)
            neto = 0.0

        arancel = scf = payway = deduc_imp = 0.0
        for label, monto_s in RE_SANT_DESC_LINE.findall(cuerpo):
            monto = _sant_parse_monto(monto_s)
            lbl_l = label.lower()
            if "arancel" in lbl_l:
                arancel += monto
            elif "deduc" in lbl_l:
                deduc_imp += monto
            elif "payway" in lbl_l:
                payway += monto
            else:
                scf += monto  # Costos Financieros + Cobro Anticipado

        # Impuestos por liquidación. IVA y SIRTAC se calculan con fórmula (sus
        # bases son inequívocas). Perc.IB se calcula por **residuo** sobre la línea
        # Deduc.Impositivas del PDF para que el resultado refleje exactamente lo
        # que el banco descontó (correctamente 0 en días puro Tj.Débito, etc.). Si
        # no hay Constancia (no podemos extraer AFIP), fallback a fórmula directa.
        fecha_pago_full = f"{fecha_pago_dm}/{año}"
        base_aplicable = arancel + scf + payway
        iva = round(base_aplicable * tasa_iva, 2)
        sirtac = round(monto_pres * tasa_sirtac, 2)
        afip_dgi = afip_por_fecha.get(fecha_pago_full, 0.0)
        if afip_por_fecha:
            perc_ib = round(deduc_imp - iva - sirtac - afip_dgi, 2)
            if perc_ib < 0:
                perc_ib = 0.0  # safety: redondeos / ediciones del banco
        else:
            perc_ib = round(base_aplicable * tasa_perc_ib, 2)

        if liqs:
            por_liq = {}
            for liq, lote in liqs:
                por_liq.setdefault(liq, []).append(lote)
            liq_str = ", ".join(f"{liq}/L{'+'.join(lotes)}" for liq, lotes in por_liq.items())
        else:
            liq_str = ""

        detalle = " + ".join(re.sub(r'\s+', ' ', v[0]).strip() for v in ventas) if ventas else ""

        liquidaciones.append({
            "Fecha Pago": fecha_pago_full,
            "Fecha Pres.": f"{fecha_pres_dm}/{año}" if fecha_pres_dm else "",
            "Liquidaciones": liq_str,
            "Detalle": detalle,
            "Monto Presentado": monto_pres,
            "Arancel": arancel,
            "Serv. Costos Financieros": scf,
            "Servicio PAYWAY": payway,
            "IVA": iva,
            "Perc.IB": perc_ib,
            "SIRTAC": sirtac,
            "AFIP-DGI": afip_dgi,
            "Neto Percibido": neto,
        })

    # ─── Metadata desde el encabezado ───────────────────────────────────────
    fecha_emision = ""
    periodo = ""
    nro_resumen = ""
    if fecha_emi_match:
        dd, mm, yyyy = fecha_emi_match.group(1), fecha_emi_match.group(2), fecha_emi_match.group(3)
        fecha_emision = f"{dd}/{mm}/{yyyy}"
        periodo = f"{mm}/{yyyy}"
        # Nro de resumen: primer número largo (≥8 dígitos) después de la fecha de emisión
        m_nro = re.search(
            r'\n\s*(\d{8,})',
            texto[fecha_emi_match.end():fecha_emi_match.end() + 200],
        )
        if m_nro:
            nro_resumen = m_nro.group(1)

    # Razón Social: sigue al label "Razón Social" en línea separada y precede a "Establecimiento"
    razon_match = re.search(
        r'Raz[óo]n Social\s*\n\s*([^\n]+?)\s*\n\s*Establecimiento',
        texto,
    )
    razon_social = razon_match.group(1).strip() if razon_match else ""

    meta = {
        "nro_resumen": nro_resumen,
        "razon_social": razon_social,
        "periodo": periodo,
        "fecha_emision": fecha_emision,
        "tasa_iva_label": lbl_iva,
        "tasa_perc_ib_label": lbl_perc_ib,
        "tasa_sirtac_label": lbl_sirtac,
    }

    return {"liquidaciones": liquidaciones, "desglose": desglose, "meta": meta}


if herramienta == TOOL_MOVIMIENTOS:
        # ─── Card 01: Archivo ──────────────────────────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">01 · Archivo fuente</div>', unsafe_allow_html=True)
        uploaded_file = st.file_uploader(
            "Arrastrá tu archivo o hacé click para seleccionarlo",
            type=["txt", "prn"],
            label_visibility="visible"
        )
        st.markdown('</div>', unsafe_allow_html=True)


        # ─── Card 02: Opciones ─────────────────────────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">02 · Opciones de exportación</div>', unsafe_allow_html=True)
        OPT_SOLO = "Solo Movimientos"
        OPT_AUXILIAR = "Exportar con columna Auxiliar"
        OPT_RESUMENES = "Incluir hojas de resumen"
        OPT_ARCA = "Cruce de comprobantes con ARCA"
        OPT_ASIENTO = "Asiento Contable"
        OPT_ASIENTO_ANUAL = "Asiento Contable Anualizado (1 hoja x mes)"

        modo_export = st.radio(
            "Seleccioná el modo de exportación:",
            options=[OPT_SOLO, OPT_AUXILIAR, OPT_RESUMENES, OPT_ARCA, OPT_ASIENTO, OPT_ASIENTO_ANUAL],
            index=0,
            help="Solo se puede elegir una opción a la vez."
        )
        con_auxiliar  = modo_export == OPT_AUXILIAR
        con_resumenes = modo_export == OPT_RESUMENES
        cruce_arca    = modo_export == OPT_ARCA
        con_asiento   = modo_export == OPT_ASIENTO
        asiento_anual = modo_export == OPT_ASIENTO_ANUAL
        st.markdown('</div>', unsafe_allow_html=True)

        # ─── Card 02b: Archivo ARCA (condicional) ──────────────────────────────────────
        df_arca = None
        df_arca_raw = None       # CSV crudo (str), para round-trip a .zip de faltantes
        arca_csv_basename = None # Nombre del CSV interno del .zip
        arca_sep = None          # Separator del CSV original
        arca_zip_name = None     # Nombre del .zip subido
        if cruce_arca:
            st.markdown('<div class="card"><div class="card-label">02b · Archivo ARCA (.zip)</div>', unsafe_allow_html=True)
            uploaded_arca = st.file_uploader(
                "Subí el .zip descargado de ARCA con los comprobantes",
                type=["zip"],
                label_visibility="visible",
                key="arca_zip"
            )
            if uploaded_arca:
                try:
                    with zipfile.ZipFile(io.BytesIO(uploaded_arca.getvalue())) as zf:
                        all_files = [f for f in zf.namelist() if not f.endswith('/')]
                        if all_files:
                            target_file = all_files[0]
                            with zf.open(target_file) as data_file:
                                raw = data_file.read()
                            csv_text = raw.decode('latin-1')
                            sep = ';' if csv_text.count(';') > csv_text.count(',') else ','
                            df_arca = pd.read_csv(
                                io.StringIO(csv_text), sep=sep, on_bad_lines='skip'
                            )
                            # Vista cruda para round-trip al .zip de faltantes
                            df_arca_raw = pd.read_csv(
                                io.StringIO(csv_text), sep=sep,
                                dtype=str, keep_default_na=False, on_bad_lines='skip'
                            )
                            arca_csv_basename = target_file
                            arca_sep = sep
                            arca_zip_name = uploaded_arca.name
                            # Mapear códigos de comprobante ARCA a tipos del sistema (con letra)
                            ARCA_TIPO_MAP = {
                                # Facturas
                                1: 'FC A', 6: 'FC B', 11: 'FC C', 51: 'FC M',
                                19: 'FC', 22: 'FC', 195: 'FC T',
                                201: 'FC A', 206: 'FC B', 211: 'FC C',
                                # Recibos (se tratan como FC)
                                4: 'FC A', 9: 'FC B', 15: 'FC C',
                                # Notas de Débito
                                2: 'ND A', 7: 'ND B', 12: 'ND C', 52: 'ND M',
                                20: 'ND', 37: 'ND', 196: 'ND T',
                                45: 'ND A', 46: 'ND B', 47: 'ND C',
                                202: 'ND A', 207: 'ND B', 212: 'ND C',
                                # Notas de Crédito
                                3: 'NC A', 8: 'NC B', 13: 'NC C', 53: 'NC M',
                                21: 'NC', 38: 'NC', 90: 'NC', 197: 'NC T',
                                43: 'NC B', 44: 'NC C', 48: 'NC A',
                                203: 'NC A', 208: 'NC B', 213: 'NC C',
                                # Tique Factura
                                81: 'TF A', 82: 'TF B', 111: 'TF C', 118: 'TF M',
                                # Tique
                                83: 'TK', 109: 'TK C', 110: 'TK',
                                112: 'TK A', 113: 'TK B', 114: 'TK C',
                                115: 'TK A', 116: 'TK B', 117: 'TK C',
                                119: 'TK M', 120: 'TK M',
                            }
                            col_tipo = 'Tipo de Comprobante'
                            if col_tipo in df_arca.columns:
                                df_arca[col_tipo] = pd.to_numeric(df_arca[col_tipo], errors='coerce').astype('Int64')
                                df_arca[col_tipo] = df_arca[col_tipo].map(ARCA_TIPO_MAP).fillna(df_arca[col_tipo].astype(str))

                            # ── Limpieza de columnas ARCA ──────────────────────────────
                            # Renombrar columnas (usa partial match para encodings rotos)
                            RENAME_RULES = [
                                (['fecha', 'emisi'], 'Fecha'),
                                (['tipo', 'comprobante'], 'Comprobante'),
                                (['punto', 'venta'], 'PV'),
                                (['mero', 'comprobante'], 'Nro.'),
                                (['tipo', 'doc', 'vendedor'], 'Tipo Doc.'),
                                (['nro', 'doc', 'vendedor'], 'CUIT'),
                                (['denominaci', 'vendedor'], 'Razon Social'),
                                (['importe', 'total'], 'Total'),
                        (['moneda', 'original'], 'Moneda'),
                        (['tipo', 'cambio'], 'Tipo Cambio'),
                                (['importe', 'no', 'gravado'], 'No Gravado'),
                                (['importe', 'exento'], 'Exento'),
                                (['pagos', 'cta', 'otros'], 'Otras Perc.'),
                                (['percepciones', 'ingresos', 'brutos'], 'Perc IIBB'),
                                (['impuestos', 'municipales'], 'Impuestos Munic.'),
                                (['percepciones', 'pagos', 'cuenta', 'iva'], 'Perc. IVA'),
                                (['impuestos', 'internos'], 'Imp. Int.'),
                                (['importe', 'otros', 'tributos'], 'Otros. Trib.'),
                                (['neto', 'gravado', 'iva', '0'], 'IVA 0%'),
                                (['neto', 'gravado', 'iva', '21'], 'Gravado IVA 21'),
                                (['importe', 'iva', '21'], 'IVA 21'),
                                (['neto', 'gravado', 'iva', '27'], 'Gravado IVA 27'),
                                (['importe', 'iva', '27'], 'IVA 27'),
                                (['neto', 'gravado', 'iva', '10'], 'Gravado IVA 10,5'),
                                (['importe', 'iva', '10'], 'IVA 10,5'),
                                (['neto', 'gravado', 'iva', '2'], 'Gravado IVA 2,5'),
                                (['importe', 'iva', '2'], 'IVA 2,5'),
                                (['neto', 'gravado', 'iva', '5%'], 'Gravado IVA 5'),
                                (['importe', 'iva', '5%'], 'IVA 5'),
                            ]
                            rename_map = {}
                            for keywords, new_name in RENAME_RULES:
                                for c in df_arca.columns:
                                    cl = c.strip().lower()
                                    if all(k in cl for k in keywords) and c not in rename_map:
                                        rename_map[c] = new_name
                                        break
                            df_arca = df_arca.rename(columns=rename_map)

                            # Convertir fecha de aaaa-mm-dd a dd/mm/aaaa
                            if 'Fecha' in df_arca.columns:
                                df_arca['Fecha'] = df_arca['Fecha'].astype(str).apply(
                                    lambda x: '/'.join(x.split('-')[::-1]) if '-' in x else x
                                )

                            # Eliminar columnas no deseadas
                            DROP_KEYWORDS = [
                                ['dito', 'fiscal', 'computable'],
                                ['total', 'neto', 'gravado'],
                                ['total', 'iva'],
                                ['tipo', 'doc'],
                            ]
                            cols_to_drop = []
                            for keywords in DROP_KEYWORDS:
                                for c in df_arca.columns:
                                    cl = c.strip().lower()
                                    if all(k in cl for k in keywords):
                                        cols_to_drop.append(c)
                                        break
                            df_arca = df_arca.drop(columns=[c for c in cols_to_drop if c in df_arca.columns], errors='ignore')

                            # Mover Total al final
                            if 'Total' in df_arca.columns:
                                total_data = df_arca.pop('Total')
                                df_arca['Total'] = total_data

                            # Columna Auxiliar: Tipo + PV + Nro Comprobante + Nro Doc Vendedor
                            def find_col(df, keywords):
                                """Busca columna que contenga todas las keywords (case-insensitive)."""
                                for c in df.columns:
                                    cl = c.strip().lower()
                                    if all(k in cl for k in keywords):
                                        return c
                                return None

                            # Crear columna Auxiliar con nombres ya renombrados
                            aux_cols = ['Comprobante', 'PV', 'Nro.', 'CUIT']
                            if all(c in df_arca.columns for c in aux_cols):
                                df_arca['Auxiliar'] = (
                                    df_arca['Comprobante'].astype(str) +
                                    df_arca['PV'].astype(str) +
                                    df_arca['Nro.'].astype(str) +
                                    df_arca['CUIT'].astype(str)
                                )
                                # Mover Auxiliar justo antes de Total
                                cols = list(df_arca.columns)
                                cols.remove('Auxiliar')
                                total_pos = cols.index('Total') if 'Total' in cols else len(cols)
                                cols.insert(total_pos, 'Auxiliar')
                                df_arca = df_arca[cols]

                            # Columnas monetarias: desde 'No Gravado' en adelante (excluyendo Auxiliar)
                            all_cols = list(df_arca.columns)
                            ng_idx = all_cols.index('No Gravado') if 'No Gravado' in all_cols else None
                            if ng_idx is not None:
                                money_cols = [c for c in all_cols[ng_idx:] if c != 'Auxiliar']
                                for c in money_cols:
                                    # Convertir formato argentino: 1.234,56 -> 1234.56
                                    df_arca[c] = df_arca[c].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                                    df_arca[c] = pd.to_numeric(df_arca[c], errors='coerce').fillna(0)
                                # Eliminar columnas monetarias que son todo cero
                                empty_money = [c for c in money_cols if (df_arca[c] == 0).all()]
                                df_arca = df_arca.drop(columns=empty_money)

                            st.success(f"**{target_file}** · {len(df_arca)} comprobantes leídos de ARCA")
                        else:
                            st.error("El .zip está vacío")
                except Exception as e:
                    st.error(f"Error al leer el .zip: {str(e)}")
            else:
                st.info("Subí el archivo .zip de ARCA para continuar")
            st.markdown('</div>', unsafe_allow_html=True)

        # ─── Card 02c: Excel del Sistema Mendez (condicional al modo Anualizado) ─────
        xls_mendez_bytes = None
        xls_mendez_name = None
        if asiento_anual:
            st.markdown('<div class="card"><div class="card-label">02c · Excel del Sistema Mendez (.XLS)</div>', unsafe_allow_html=True)
            uploaded_mendez = st.file_uploader(
                "Subí el .XLS del sistema Mendez con las fechas completas por comprobante",
                type=["xls", "xlsx"],
                label_visibility="visible",
                key="mendez_xls",
                help="El TXT anual sólo trae el día; el XLS aporta mes/año por cruce de Tipo+PV+Nro+CUIT."
            )
            if uploaded_mendez:
                xls_mendez_bytes = uploaded_mendez.getvalue()
                xls_mendez_name = uploaded_mendez.name
                st.success(f"**{xls_mendez_name}** listo para cruzar")
            else:
                st.info("Subí el .XLS de Mendez para continuar")
            st.markdown('</div>', unsafe_allow_html=True)


        if uploaded_file is not None:
            filename = Path(uploaded_file.name).stem
            st.success(f"**{uploaded_file.name}** listo para procesar")

            st.markdown('<div class="card"><div class="card-label">03 · Procesar</div>', unsafe_allow_html=True)

            if st.button("⬡  Procesar Archivo"):
                try:
                    with st.spinner("Analizando información..."):
                        content = uploaded_file.getvalue().decode("latin-1")
                        transacciones, meta = parsear_archivo(content=content)

                    if not transacciones:
                        st.error("No se encontraron transacciones. Verificá el formato del archivo.")
                    elif asiento_anual and xls_mendez_bytes is None:
                        st.error("Subí el .XLS de Mendez para poder discriminar los meses.")
                    else:
                        with st.spinner("Generando Excel..."):
                            output = io.BytesIO()
                            if asiento_anual:
                                con_mes, sin_mes = asignar_mes_por_xls_mendez(
                                    transacciones, io.BytesIO(xls_mendez_bytes)
                                )
                                stats_anual = crear_excel_asiento_anual(con_mes, sin_mes, meta, output)
                                if stats_anual['sin_asignar'] > 0:
                                    st.warning(
                                        f"⚠ {stats_anual['sin_asignar']} comprobantes no se cruzaron "
                                        f"con el XLS Mendez — se volcaron en la hoja 'Sin Asignar'."
                                    )
                            else:
                                crear_excel(transacciones, meta, output,
                                            con_resumenes=con_resumenes,
                                            con_auxiliar=con_auxiliar,
                                            cruce_arca=cruce_arca,
                                            df_arca=df_arca,
                                            con_asiento=con_asiento)
                            output.seek(0)

                        st.success("✓  Proceso completado con éxito")

                        # Stats chips
                        from collections import Counter
                        tipos = Counter(t['Tipo'] for t in transacciones)
                        st.markdown(f"""
                        <div class="stats-row">
                            <div class="stat-chip">
                                <span class="stat-val">{len(transacciones)}</span>
                                <span class="stat-lbl">Total</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{tipos.get('FC', 0)}</span>
                                <span class="stat-lbl">Facturas</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{tipos.get('NC', 0)}</span>
                                <span class="stat-lbl">Notas Cred.</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{tipos.get('ND', 0) + tipos.get('TF', 0) + tipos.get('TK', 0)}</span>
                                <span class="stat-lbl">Otros</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                        st.info(
                            f"**{meta.get('tipo_reporte', 'N/A')}** · "
                            f"{meta.get('razon_social', 'Contribuyente')} · "
                            f"{meta.get('periodo', '')}"
                        )

                        if cruce_arca:
                            excel_filename = "Cruce Compras.xlsx"
                        elif asiento_anual:
                            excel_filename = f"{filename}_asiento_anual.xlsx"
                        elif con_asiento:
                            excel_filename = f"{filename}_asiento.xlsx"
                        else:
                            excel_filename = f"{filename}_procesado.xlsx"
                        st.download_button(
                            label="↓  Descargar Excel",
                            data=output,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

                        # ── Cruce ARCA: .zip de comprobantes faltantes (mismo formato que ARCA) ──
                        if cruce_arca and df_arca is not None and df_arca_raw is not None and 'Auxiliar' in df_arca.columns:
                            sistema_aux_set = construir_sistema_aux_set(transacciones)
                            mask_falt = ~df_arca['Auxiliar'].astype(str).isin(sistema_aux_set)
                            falt_idx = df_arca.index[mask_falt]
                            df_falt_raw = df_arca_raw.loc[df_arca_raw.index.intersection(falt_idx)]

                            if len(df_falt_raw) > 0:
                                # Serializar a CSV con el mismo separator/encoding del ARCA original
                                csv_io = io.StringIO()
                                df_falt_raw.to_csv(csv_io, sep=arca_sep, index=False, lineterminator='\n')
                                csv_bytes = csv_io.getvalue().encode('latin-1', errors='replace')

                                # Empaquetar en .zip preservando el nombre del CSV interno
                                zip_buf = io.BytesIO()
                                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                                    zout.writestr(arca_csv_basename, csv_bytes)

                                # Nombre del .zip de salida: derivado del original
                                base_zip_name = Path(arca_zip_name).stem if arca_zip_name else 'ARCA'
                                falt_zip_name = f"{base_zip_name}_FALTANTES.zip"

                                st.download_button(
                                    label=f"↓  Descargar .zip de Faltantes ({len(df_falt_raw)} comprobantes)",
                                    data=zip_buf.getvalue(),
                                    file_name=falt_zip_name,
                                    mime="application/zip",
                                    use_container_width=True,
                                )

                except Exception as e:
                    st.error(f"Error al procesar el archivo: {str(e)}")
                    st.exception(e)

            st.markdown('</div>', unsafe_allow_html=True)

        else:
            st.markdown("""
            <div style="
                text-align: center;
                padding: 2rem 1rem;
                font-family: 'Space Mono', monospace;
                font-size: 0.72rem;
                color: #6b7280;
                letter-spacing: 0.12em;
            ">
                ESPERANDO ARCHIVO · PASO 01
            </div>
            """, unsafe_allow_html=True)


elif herramienta == TOOL_PORTAL_IVA:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Archivo .zip PORTAL IVA (modos: Limpiar / Edición .zip)
    # ───────────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Archivo ARCA (.zip)</div>', unsafe_allow_html=True)
    uploaded_zip_iva = st.file_uploader(
        "Subí el .zip descargado del Portal IVA de ARCA",
        type=["zip"],
        label_visibility="visible",
        key="portal_iva_zip"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_zip_iva:
        st.success(f"**{uploaded_zip_iva.name}** listo para procesar")

        st.markdown('<div class="card"><div class="card-label">02 · Modo</div>', unsafe_allow_html=True)
        MODO_PORTAL_LIMPIAR = "Limpiar"
        MODO_PORTAL_EDITAR = "Edición .zip"
        modo_portal = st.radio(
            "Elegí qué hacer con el .zip:",
            options=[MODO_PORTAL_LIMPIAR, MODO_PORTAL_EDITAR],
            horizontal=True,
            key="portal_iva_modo",
            help="Limpiar: genera un Excel formateado con los movimientos. Edición .zip: round-trip a Excel para editar masivamente el CSV de ARCA y devolverlo en su mismo formato .zip."
        )
        st.markdown('</div>', unsafe_allow_html=True)

        if modo_portal == MODO_PORTAL_EDITAR:
            # ── Modo Edición .zip: round-trip a Excel ─────────────────────────
            EDIT_DATOS_SHEET = "Comprobantes"
            EDIT_META_SHEET = "_meta"

            st.markdown('<div class="card"><div class="card-label">03 · Generar Excel para edición</div>', unsafe_allow_html=True)
            st.caption(
                "Genera un .xlsx con todas las filas y columnas del CSV interno del .zip. "
                "Editalo en Excel (sort, filter, autofill, Ctrl+Enter, fórmulas, copy/paste de rangos), "
                "después subilo en el paso 04 y la app rearma el .zip con el formato original."
            )

            source_id_edit = f"{uploaded_zip_iva.name}|{uploaded_zip_iva.size}"
            if st.session_state.get('edit_source_id') != source_id_edit:
                st.session_state['edit_source_id'] = source_id_edit
                st.session_state.pop('edit_xlsx_bytes', None)
                st.session_state.pop('edit_xlsx_name', None)

            if st.button("⬡  Generar Excel editable", use_container_width=True, key='edit_btn_make_xlsx'):
                try:
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter

                    with st.spinner("Leyendo ZIP..."):
                        with zipfile.ZipFile(io.BytesIO(uploaded_zip_iva.getvalue())) as zf_in:
                            archivos_in = [f for f in zf_in.namelist() if not f.endswith('/')]
                            if not archivos_in:
                                st.error("El .zip está vacío.")
                                st.stop()
                            target_in = archivos_in[0]
                            raw_in = zf_in.open(target_in).read()

                    csv_text_in = raw_in.decode('latin-1')
                    sep_in = ';' if csv_text_in.count(';') > csv_text_in.count(',') else ','
                    df_in = pd.read_csv(
                        io.StringIO(csv_text_in),
                        sep=sep_in,
                        dtype=str,
                        keep_default_na=False,
                        na_values=[],
                        on_bad_lines='skip',
                    )

                    with st.spinner("Generando Excel..."):
                        wb_e = openpyxl.Workbook()
                        ws_e = wb_e.active
                        ws_e.title = EDIT_DATOS_SHEET
                        cols_in = list(df_in.columns)
                        ws_e.append(cols_in)

                        header_fill_e = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
                        header_font_e = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                        header_align_e = Alignment(horizontal="center", vertical="center")
                        for col_idx in range(1, len(cols_in) + 1):
                            cell = ws_e.cell(row=1, column=col_idx)
                            cell.fill = header_fill_e
                            cell.font = header_font_e
                            cell.alignment = header_align_e

                        for _, row in df_in.iterrows():
                            ws_e.append([row[c] for c in cols_in])

                        ws_e.freeze_panes = "A2"
                        for i, col_name in enumerate(cols_in, start=1):
                            sample_len = max([len(str(col_name))] + [len(str(v)) for v in df_in[col_name].head(50)])
                            ws_e.column_dimensions[get_column_letter(i)].width = min(max(sample_len + 2, 10), 40)

                        ws_meta_e = wb_e.create_sheet(EDIT_META_SHEET)
                        ws_meta_e.append(["csv_basename", Path(target_in).name])
                        ws_meta_e.append(["sep", sep_in])
                        ws_meta_e.append(["zip_basename", Path(uploaded_zip_iva.name).name])
                        ws_meta_e.sheet_state = "hidden"

                        out_xlsx_e = io.BytesIO()
                        wb_e.save(out_xlsx_e)
                        st.session_state['edit_xlsx_bytes'] = out_xlsx_e.getvalue()
                        st.session_state['edit_xlsx_name'] = f"{Path(uploaded_zip_iva.name).stem}_EDITAR.xlsx"

                    st.success(f"Excel generado · {len(df_in)} filas · {len(cols_in)} columnas")
                except Exception as e:
                    st.error(f"Error al generar el Excel: {e}")

            if 'edit_xlsx_bytes' in st.session_state:
                st.download_button(
                    label=f"↓  Descargar {st.session_state['edit_xlsx_name']}",
                    data=st.session_state['edit_xlsx_bytes'],
                    file_name=st.session_state['edit_xlsx_name'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key='edit_btn_download_xlsx',
                )
            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="card"><div class="card-label">04 · Subir Excel editado</div>', unsafe_allow_html=True)
            uploaded_xlsx_edit = st.file_uploader(
                "Subí el Excel editado para reconstruir el .zip",
                type=["xlsx"],
                key="edit_xlsx_upload"
            )
            st.markdown('</div>', unsafe_allow_html=True)

            if uploaded_xlsx_edit is not None:
                st.markdown('<div class="card"><div class="card-label">05 · Generar .zip modificado</div>', unsafe_allow_html=True)
                if st.button("⬡  Generar .zip modificado", use_container_width=True, key='edit_btn_gen_zip'):
                    try:
                        xlsx_bytes_e = uploaded_xlsx_edit.getvalue()

                        try:
                            df_meta_e = pd.read_excel(
                                io.BytesIO(xlsx_bytes_e),
                                sheet_name=EDIT_META_SHEET,
                                header=None, dtype=str, keep_default_na=False,
                            )
                            meta_dict_e = dict(zip(df_meta_e.iloc[:, 0], df_meta_e.iloc[:, 1]))
                            sep_out = meta_dict_e.get('sep', ';')
                            csv_basename_out = meta_dict_e.get('csv_basename', 'comprobantes.csv')
                            zip_basename_out = meta_dict_e.get('zip_basename', 'modificado.zip')
                        except Exception:
                            st.error(f"El Excel no contiene la hoja `{EDIT_META_SHEET}` con la metadata. Asegurate de subir el archivo generado por esta misma herramienta.")
                            st.stop()

                        df_out = pd.read_excel(
                            io.BytesIO(xlsx_bytes_e),
                            sheet_name=EDIT_DATOS_SHEET,
                            dtype=str, keep_default_na=False,
                        )

                        csv_io_out = io.StringIO()
                        df_out.to_csv(csv_io_out, sep=sep_out, index=False, lineterminator='\n')
                        csv_bytes_out = csv_io_out.getvalue().encode('latin-1', errors='replace')

                        zip_buf_out = io.BytesIO()
                        with zipfile.ZipFile(zip_buf_out, 'w', zipfile.ZIP_DEFLATED) as zout_e:
                            zout_e.writestr(csv_basename_out, csv_bytes_out)

                        nombre_zip_out = f"{Path(zip_basename_out).stem}_EDITADO.zip"
                        st.success(f".zip generado · {len(df_out)} filas · CSV interno: {csv_basename_out}")
                        st.download_button(
                            label=f"↓  Descargar {nombre_zip_out}",
                            data=zip_buf_out.getvalue(),
                            file_name=nombre_zip_out,
                            mime="application/zip",
                            use_container_width=True,
                            key='edit_btn_download_zip',
                        )
                    except Exception as e:
                        st.error(f"Error al reconstruir el .zip: {e}")
                st.markdown('</div>', unsafe_allow_html=True)

            st.stop()

        st.markdown('<div class="card"><div class="card-label">03 · Datos del contribuyente</div>', unsafe_allow_html=True)
        nombre_contribuyente = st.text_input("Nombre / Razón Social del contribuyente", value="", key="nombre_portal_iva")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><div class="card-label">04 · Procesar</div>', unsafe_allow_html=True)

        if st.button("⬡  Procesar ZIP"):
            if not nombre_contribuyente.strip():
                st.error("Ingresá el nombre del contribuyente para continuar.")
            else:
              try:
                with st.spinner("Leyendo archivo ARCA..."):
                    with zipfile.ZipFile(io.BytesIO(uploaded_zip_iva.getvalue())) as zf:
                        all_files = [f for f in zf.namelist() if not f.endswith('/')]
                        if not all_files:
                            st.error("El .zip está vacío")
                            st.stop()
                        target_file = all_files[0]
                        with zf.open(target_file) as data_file:
                            raw = data_file.read()

                    csv_text = raw.decode('latin-1')
                    sep = ';' if csv_text.count(';') > csv_text.count(',') else ','
                    df_iva = pd.read_csv(io.StringIO(csv_text), sep=sep, on_bad_lines='skip')

                    # Detectar tipo (Compras/Ventas), CUIT y periodo del nombre del zip
                    import re as _re
                    zip_name_raw = uploaded_zip_iva.name.upper()
                    es_ventas = 'VENTA' in zip_name_raw
                    es_compras = 'COMPRA' in zip_name_raw
                    tipo_portal = 'VENTAS' if es_ventas else ('COMPRAS' if es_compras else 'PORTAL IVA')

                    # Buscar CUIT (11 dígitos) y periodo (YYYYMM o YYYY-MM)
                    cuit_match = _re.search(r'(\d{11})', zip_name_raw)
                    cuit_portal = cuit_match.group(1) if cuit_match else ''
                    periodo_match = _re.search(r'(\d{4})(\d{2})(?!\d)', zip_name_raw)
                    if periodo_match:
                        meses = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
                                 'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
                        m_num = int(periodo_match.group(2))
                        periodo_portal = f"{meses[m_num]} {periodo_match.group(1)}" if 1 <= m_num <= 12 else ''
                    else:
                        periodo_portal = ''

                    # Mapear códigos de comprobante
                    ARCA_TIPO_MAP = {
                        1: 'FC A', 6: 'FC B', 11: 'FC C', 51: 'FC M',
                        19: 'FC', 22: 'FC', 195: 'FC T',
                        201: 'FC A', 206: 'FC B', 211: 'FC C',
                        # Recibos (se tratan como FC)
                        4: 'FC A', 9: 'FC B', 15: 'FC C',
                        2: 'ND A', 7: 'ND B', 12: 'ND C', 52: 'ND M',
                        20: 'ND', 37: 'ND', 196: 'ND T',
                        45: 'ND A', 46: 'ND B', 47: 'ND C',
                        202: 'ND A', 207: 'ND B', 212: 'ND C',
                        3: 'NC A', 8: 'NC B', 13: 'NC C', 53: 'NC M',
                        21: 'NC', 38: 'NC', 90: 'NC', 197: 'NC T',
                        43: 'NC B', 44: 'NC C', 48: 'NC A',
                        203: 'NC A', 208: 'NC B', 213: 'NC C',
                        81: 'TF A', 82: 'TF B', 111: 'TF C', 118: 'TF M',
                        83: 'TK', 109: 'TK C', 110: 'TK',
                        112: 'TK A', 113: 'TK B', 114: 'TK C',
                        115: 'TK A', 116: 'TK B', 117: 'TK C',
                        119: 'TK M', 120: 'TK M',
                    }

                    def find_col_iva(df, keywords):
                        for c in df.columns:
                            cl = c.strip().lower()
                            if all(k in cl for k in keywords):
                                return c
                        return None

                    col_tipo_iva = find_col_iva(df_iva, ['tipo', 'comprobante'])
                    if col_tipo_iva:
                        df_iva[col_tipo_iva] = pd.to_numeric(df_iva[col_tipo_iva], errors='coerce').astype('Int64')
                        df_iva[col_tipo_iva] = df_iva[col_tipo_iva].map(ARCA_TIPO_MAP).fillna(df_iva[col_tipo_iva].astype(str))

                    # Renombrar columnas (funciona para compras y ventas)
                    RENAME_RULES = [
                        (['fecha', 'emisi'], 'Fecha'),
                        (['tipo', 'comprobante'], 'Comprobante'),
                        (['punto', 'venta'], 'PV'),
                        (['mero', 'comprobante', 'hasta'], 'Nro. Hasta'),
                        (['mero', 'comprobante'], 'Nro.'),
                        (['tipo', 'doc'], 'Tipo Doc.'),
                        (['nro', 'doc', 'vendedor'], 'CUIT'),
                        (['nro', 'doc', 'comprador'], 'CUIT'),
                        (['denominaci', 'vendedor'], 'Razon Social'),
                        (['denominaci', 'comprador'], 'Razon Social'),
                        (['fecha', 'vencimiento'], 'Fecha Vto. Pago'),
                        (['importe', 'total'], 'Total'),
                        (['moneda', 'original'], 'Moneda'),
                        (['tipo', 'cambio'], 'Tipo Cambio'),
                        (['importe', 'no', 'gravado'], 'No Gravado'),
                        (['importe', 'exento'], 'Exento'),
                        (['pagos', 'cta', 'otros'], 'Otras Perc.'),
                        (['percepciones', 'ingresos', 'brutos'], 'Perc IIBB'),
                        (['impuestos', 'municipales'], 'Impuestos Munic.'),
                        (['percepciones', 'pagos', 'cuenta', 'iva'], 'Perc. IVA'),
                        (['percepci', 'no', 'categorizados'], 'Perc. No Cat.'),
                        (['impuestos', 'internos'], 'Imp. Int.'),
                        (['importe', 'otros', 'tributos'], 'Otros. Trib.'),
                        (['neto', 'gravado', 'iva', '0'], 'IVA 0%'),
                        (['neto', 'gravado', 'iva', '21'], 'Gravado IVA 21'),
                        (['importe', 'iva', '21'], 'IVA 21'),
                        (['neto', 'gravado', 'iva', '27'], 'Gravado IVA 27'),
                        (['importe', 'iva', '27'], 'IVA 27'),
                        (['neto', 'gravado', 'iva', '10'], 'Gravado IVA 10,5'),
                        (['importe', 'iva', '10'], 'IVA 10,5'),
                        (['neto', 'gravado', 'iva', '2'], 'Gravado IVA 2,5'),
                        (['importe', 'iva', '2'], 'IVA 2,5'),
                        (['neto', 'gravado', 'iva', '5%'], 'Gravado IVA 5'),
                        (['importe', 'iva', '5%'], 'IVA 5'),
                    ]
                    rename_map = {}
                    for keywords, new_name in RENAME_RULES:
                        for c in df_iva.columns:
                            cl = c.strip().lower()
                            if all(k in cl for k in keywords) and c not in rename_map:
                                rename_map[c] = new_name
                                break
                    df_iva = df_iva.rename(columns=rename_map)

                    # Convertir fecha de aaaa-mm-dd a dd/mm/aaaa
                    if 'Fecha' in df_iva.columns:
                        df_iva['Fecha'] = df_iva['Fecha'].astype(str).apply(
                            lambda x: '/'.join(x.split('-')[::-1]) if '-' in x else x
                        )

                    # Eliminar columnas no deseadas
                    DROP_KW = [
                        ['dito', 'fiscal', 'computable'],
                        ['total', 'neto', 'gravado'],
                        ['total', 'iva'],
                        ['tipo', 'doc'],
                        ['nro.', 'hasta'],
                        ['fecha', 'vto'],
                    ]
                    cols_to_drop = []
                    for kws in DROP_KW:
                        for c in df_iva.columns:
                            cl = c.strip().lower()
                            if all(k in cl for k in kws):
                                cols_to_drop.append(c)
                                break
                    df_iva = df_iva.drop(columns=[c for c in cols_to_drop if c in df_iva.columns], errors='ignore')

                    # Mover Total al final
                    if 'Total' in df_iva.columns:
                        total_data = df_iva.pop('Total')
                        df_iva['Total'] = total_data

                    # Columnas monetarias: convertir y limpiar
                    all_cols_iva = list(df_iva.columns)
                    non_money = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'CUIT', 'Razon Social', 'Moneda', 'Tipo Cambio'}
                    money_cols_iva = [c for c in all_cols_iva if c not in non_money and c in df_iva.select_dtypes(include='object').columns]
                    for c in money_cols_iva:
                        df_iva[c] = df_iva[c].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                        df_iva[c] = pd.to_numeric(df_iva[c], errors='coerce').fillna(0)
                    # Rellenar NaN restantes en columnas numéricas
                    for c in all_cols_iva:
                        if c not in non_money and df_iva[c].dtype in ('float64', 'int64'):
                            df_iva[c] = df_iva[c].fillna(0)
                    # Eliminar columnas monetarias todo cero
                    empty_cols = [c for c in all_cols_iva if c not in non_money and c in df_iva.columns and df_iva[c].dtype in ('float64', 'int64') and (df_iva[c] == 0).all()]
                    df_iva = df_iva.drop(columns=empty_cols)

                with st.spinner("Generando Excel..."):
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_iva.to_excel(writer, sheet_name=tipo_portal, index=False, startrow=5)
                        ws = writer.sheets[tipo_portal]
                        n_cols = len(df_iva.columns)

                        title_font = Font(bold=True, size=14, color='FFFFFF')
                        title_fill = PatternFill('solid', fgColor='2F5496')
                        header_font = Font(bold=True, size=10, color='FFFFFF')
                        header_fill = PatternFill('solid', fgColor='4472C4')
                        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        center_align = Alignment(horizontal='center', vertical='center')
                        thin_border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        zebra_fill = PatternFill('solid', fgColor='D6E4F0')
                        money_fmt = '$#,##0.00'

                        ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
                        ws['A1'] = nombre_contribuyente.strip().upper()
                        ws['A1'].font = title_font; ws['A1'].fill = title_fill
                        ws['A1'].alignment = center_align

                        ws.merge_cells(f'A2:{get_column_letter(n_cols)}2')
                        sub_parts = [p for p in [f'CUIT: {cuit_portal}' if cuit_portal else '', f'{len(df_iva)} comprobantes'] if p]
                        ws['A2'] = ' | '.join(sub_parts)
                        ws['A2'].font = Font(bold=True, size=11, color='2F5496')
                        ws['A2'].alignment = center_align

                        ws.merge_cells(f'A5:{get_column_letter(n_cols)}5')
                        ws['A5'] = f'{tipo_portal} {periodo_portal}'.strip()
                        ws['A5'].font = Font(bold=True, size=12, color='2F5496')
                        ws['A5'].alignment = center_align

                        col_list = list(df_iva.columns)
                        non_money_set = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'CUIT', 'Razon Social', 'Auxiliar'}
                        money_indices = [i + 1 for i, c in enumerate(col_list) if c not in non_money_set and df_iva[c].dtype in ('float64', 'int64')]

                        for col_idx in range(1, n_cols + 1):
                            cell = ws.cell(row=6, column=col_idx)
                            cell.font = header_font; cell.fill = header_fill
                            cell.alignment = header_align; cell.border = thin_border

                        for row_idx in range(7, len(df_iva) + 7):
                            for col_idx in range(1, n_cols + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.alignment = center_align
                                if col_idx in money_indices:
                                    cell.number_format = money_fmt
                            if (row_idx - 7) % 2 == 0:
                                for col_idx in range(1, n_cols + 1):
                                    ws.cell(row=row_idx, column=col_idx).fill = zebra_fill

                        # Fila TOTAL con fórmulas SUM
                        total_row = len(df_iva) + 7
                        col_list = list(df_iva.columns)
                        non_money_set2 = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'CUIT', 'Razon Social', 'Moneda', 'Tipo Cambio'}
                        for col_idx in range(1, n_cols + 1):
                            cell = ws.cell(row=total_row, column=col_idx)
                            col_name = col_list[col_idx - 1] if col_idx - 1 < len(col_list) else ''
                            if col_name not in non_money_set2 and col_idx in money_indices:
                                letter = get_column_letter(col_idx)
                                cell.value = f'=SUM({letter}7:{letter}{total_row - 1})'
                                cell.number_format = money_fmt
                            elif col_idx == 1:
                                cell.value = 'TOTAL'
                            cell.font = Font(bold=True, size=10, color='FFFFFF')
                            cell.fill = PatternFill('solid', fgColor='2F5496')
                            cell.alignment = center_align

                        for col_idx in range(1, n_cols + 1):
                            max_len = max(
                                len(str(ws.cell(row=r, column=col_idx).value or ''))
                                for r in range(6, min(len(df_iva) + 7, 50))
                            )
                            letter = get_column_letter(col_idx)
                            ws.column_dimensions[letter].width = max(max_len + 3, 8)

                    output.seek(0)

                st.success("✓  Proceso completado con éxito")

                from collections import Counter
                tipos_iva = Counter(df_iva['Comprobante']) if 'Comprobante' in df_iva.columns else {}
                fc_count = sum(v for k, v in tipos_iva.items() if str(k).startswith('FC'))
                nc_count = sum(v for k, v in tipos_iva.items() if str(k).startswith('NC'))
                otros_count = len(df_iva) - fc_count - nc_count

                st.markdown(f"""
                <div class="stats-row">
                    <div class="stat-chip">
                        <span class="stat-val">{len(df_iva)}</span>
                        <span class="stat-lbl">Total</span>
                    </div>
                    <div class="stat-chip">
                        <span class="stat-val">{fc_count}</span>
                        <span class="stat-lbl">Facturas</span>
                    </div>
                    <div class="stat-chip">
                        <span class="stat-val">{nc_count}</span>
                        <span class="stat-lbl">Notas Cred.</span>
                    </div>
                    <div class="stat-chip">
                        <span class="stat-val">{otros_count}</span>
                        <span class="stat-lbl">Otros</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                excel_name = 'Compras' if tipo_portal == 'COMPRAS' else ('Ventas' if tipo_portal == 'VENTAS' else Path(uploaded_zip_iva.name).stem)
                st.download_button(
                    label="↓  Descargar Excel",
                    data=output,
                    file_name=f"{excel_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

              except Exception as e:
                st.error(f"Error al procesar: {str(e)}")
                st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            ESPERANDO ARCHIVO .ZIP · PASO 01
        </div>
        """, unsafe_allow_html=True)


elif herramienta == TOOL_SIFERE:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Archivos SIFERE (TXT)
    # ───────────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Tipo de archivo SIFERE</div>', unsafe_allow_html=True)
    tipo_sifere = st.radio(
        "¿Qué tipo de archivo SIFERE querés generar?",
        options=["Percepciones", "Retenciones"],  # "Percepciones AGIP (RentasCiudad.xls)" deshabilitado temporalmente
        horizontal=True,
        key="sifere_tipo"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    # # ── Rama AGIP desde XLS ──────────────────────────────────────────────────────
    # if tipo_sifere == "Percepciones AGIP (RentasCiudad.xls)":
    # st.markdown('<div class="card"><div class="card-label">02 · Archivo RentasCiudad (.xls / .csv)</div>', unsafe_allow_html=True)
    # uploaded_agip_xls = st.file_uploader(
    # "Subí el archivo exportado de Rentas Ciudad AGIP",
    # type=["xls", "csv"],
    # label_visibility="visible",
    # key="sifere_agip_xls"
    # )
    # st.markdown('</div>', unsafe_allow_html=True)
    #
    # st.markdown('<div class="card"><div class="card-label">03 · Datos del período</div>', unsafe_allow_html=True)
    # col_ag1, col_ag2 = st.columns(2)
    # with col_ag1:
    # cuit_agip = st.text_input("CUIT del contribuyente (sin guiones)", value="", placeholder="30631005833", key="sifere_agip_cuit")
    # with col_ag2:
    # periodo_agip = st.text_input("Período (MM/AAAA)", value="", placeholder="03/2026", key="sifere_agip_periodo")
    # st.markdown('</div>', unsafe_allow_html=True)
    #
    # st.markdown('<div class="card"><div class="card-label">03b · Excepciones de Punto de Venta (opcional)</div>', unsafe_allow_html=True)
    # st.caption(
    # "Algunos proveedores tienen PV que termina en ceros (ej: 1000, 200) y el sistema no puede detectarlo automáticamente. "
    # "Ingresá los casos especiales en formato `CUIT:PV` (uno por línea, el CUIT sin guiones)."
    # )
    # pv_overrides_raw = st.text_area(
    # "Excepciones CUIT → PV (sin guiones · sin ceros iniciales)",
    # value="",
    # placeholder="30123456789:1000\n30987654321:200\n20111222333:10",
    # height=100,
    # key="sifere_agip_pv_overrides"
    # )
    # # Parsear las excepciones → dict {cuit_sin_guiones: pv_str_zfill4}
    # pv_overrides = {}
    # for line in pv_overrides_raw.strip().splitlines():
    # line = line.strip()
    # if ':' in line:
    # parts = line.split(':', 1)
    # cuit_ov = re.sub(r'\D', '', parts[0].strip())
    # pv_ov   = parts[1].strip().lstrip('0') or '0'
    # if cuit_ov:
    # pv_overrides[cuit_ov] = pv_ov.zfill(4)
    # if pv_overrides:
    # st.info(f"✓ {len(pv_overrides)} excepción(es) cargada(s): " + ", ".join(f"{k}→{v}" for k, v in pv_overrides.items()))
    # st.markdown('</div>', unsafe_allow_html=True)
    #
    # if uploaded_agip_xls:
    # agip_filename = Path(uploaded_agip_xls.name).stem
    # st.success(f"**{uploaded_agip_xls.name}** listo para procesar")
    # st.markdown('<div class="card"><div class="card-label">04 · Generar TXT SIFERE AGIP</div>', unsafe_allow_html=True)
    #
    # if st.button("⬡  Generar SIFERE desde AGIP"):
    # if not re.match(r'^(0[1-9]|1[0-2])/\d{4}$', periodo_agip.strip()):
    # st.error("El período debe tener formato MM/AAAA (ej: 03/2026).")
    # else:
    # try:
    # with st.spinner("Leyendo archivo AGIP..."):
    # raw_agip = uploaded_agip_xls.getvalue()
    # texto_agip = raw_agip.decode('latin-1', errors='replace')
    # if texto_agip.lstrip().startswith('sep='):
    # df_agip = pd.read_csv(io.StringIO(texto_agip), sep=',', encoding='latin-1', skiprows=1, on_bad_lines='skip')
    # else:
    # df_agip = pd.read_csv(io.StringIO(texto_agip), sep=',', encoding='latin-1', on_bad_lines='skip')
    #
    # # El CSV de RentasCiudad usa el CUIT como índice; resetear
    # df_agip = df_agip.reset_index()
    # df_agip.columns = [str(c).strip() for c in df_agip.columns]
    #
    # # Columnas por posición (por encoding la col 0 es el CUIT real):
    # # 0=CUIT prov, 1=NAdher, 2=RazonSocial, 3=NroCert, 4=Norma,
    # # 5=FechaPerc, 6=TipoComp(F/C/D), 7=NroComp("A0  XXXX"),
    # # 8=FechaComp(DD/MM/YYYY), 9=MontoPercibido, 10=BaseCalculo
    # idx_cuit    = 0
    # idx_tipo    = 6
    # idx_nrocomp = 7
    # idx_fecha   = 8
    # idx_monto   = 9
    #
    # per_parts = periodo_agip.strip().split('/')
    # mes_agip  = per_parts[0].zfill(2)
    # anio_agip = per_parts[1]
    #
    # TIPO_AGIP_SIFERE = {'F': 'FA', 'f': 'FA', 'C': 'CA', 'c': 'CA', 'D': 'DA', 'd': 'DA'}
    # lineas = []
    # errores = []
    #
    # for _, fila in df_agip.iterrows():
    # try:
    # cuit_prov_raw = str(fila.iloc[idx_cuit]).strip().replace('.', '').replace('-', '')
    # nro_comp_raw  = str(fila.iloc[idx_nrocomp]).strip()
    # tipo_raw      = str(fila.iloc[idx_tipo]).strip()
    # fecha_raw     = str(fila.iloc[idx_fecha]).strip()
    # monto_raw     = fila.iloc[idx_monto]
    #
    # # Parsear "A0   711000008388283" → PV(4) + Nro(8)
    # # AGIP concatena el punto de venta (sin ceros iniciales)
    # # seguido de ceros de relleno y luego el Nro (8 dígitos).
    # # Regla: últimos 8 dígitos = Nro, el resto quitando
    # # los ceros de relleno del final = PV.
    # # EXCEPCIÓN: si el PV termina en ceros (ej: 1000) el
    # # rstrip falla → se usa el override manual si está definido.
    # partes_comp = nro_comp_raw.split()
    # if len(partes_comp) >= 2:
    # numero_largo = re.sub(r'\D', '', partes_comp[1])
    # else:
    # numero_largo = re.sub(r'\D', '', nro_comp_raw[2:])
    #
    # cuit_prov_clean_key = re.sub(r'\D', '', str(fila.iloc[idx_cuit]).strip())
    # if cuit_prov_clean_key in pv_overrides:
    # # Override manual: PV conocido, Nro = últimos 8
    # pv_sif  = pv_overrides[cuit_prov_clean_key]
    # nro_sif = numero_largo[-8:].zfill(8)
    # elif len(numero_largo) > 8:
    # nro_sif = numero_largo[-8:].zfill(8)
    # pv_raw  = numero_largo[:-8].rstrip('0') or '0'
    # pv_sif  = pv_raw[-4:].zfill(4)
    # else:
    # nro_sif = numero_largo.zfill(8)
    # pv_sif  = '0000'
    #
    #
    # tipo_sif = TIPO_AGIP_SIFERE.get(tipo_raw, 'FA')
    #
    # cuit_prov_clean = cuit_prov_raw.replace('-', '')
    # if len(cuit_prov_clean) == 11:
    # cuit_prov_fmt = f"{cuit_prov_clean[:2]}-{cuit_prov_clean[2:10]}-{cuit_prov_clean[10]}"
    # else:
    # cuit_prov_fmt = cuit_prov_clean.ljust(13)
    #
    # fecha_sif = fecha_raw if re.match(r'\d{2}/\d{2}/\d{4}', fecha_raw) else f"01/{mes_agip}/{anio_agip}"
    #
    # try:
    # monto_val = float(str(monto_raw).replace(',', '.'))
    # except Exception:
    # monto_val = 0.0
    # if monto_val == 0.0:
    # continue
    #
    # parte_entera  = int(abs(monto_val))
    # parte_decimal = f"{abs(monto_val):.2f}".split('.')[1]
    # monto_fmt = f"-{parte_entera:07d},{parte_decimal}" if tipo_sif == 'CA' else f" {parte_entera:07d},{parte_decimal}"
    #
    # linea = f"901{cuit_prov_fmt}{fecha_sif}{pv_sif}{nro_sif}{tipo_sif}{monto_fmt}"
    # lineas.append(linea)
    #
    # except Exception as ex_fila:
    # errores.append(str(ex_fila))
    # continue
    #
    # txt_agip = "\r\n".join(lineas)
    # st.success(f"✓  SIFERE AGIP generado: **{len(lineas)} líneas**")
    #
    # if errores:
    # st.warning(f"Se omitieron {len(errores)} filas con errores.")
    # with st.expander("Ver errores"):
    # for e in errores[:20]: st.text(e)
    #
    # st.markdown(f"""
    # <div class="stats-row">
    # <div class="stat-chip"><span class="stat-val">{len(df_agip)}</span><span class="stat-lbl">Registros</span></div>
    # <div class="stat-chip"><span class="stat-val">{len(lineas)}</span><span class="stat-lbl">Líneas TXT</span></div>
    # <div class="stat-chip"><span class="stat-val">901</span><span class="stat-lbl">Jurisdicción AGIP</span></div>
    # </div>
    # """, unsafe_allow_html=True)
    #
    # st.download_button(
    # label="↓  Descargar TXT SIFERE AGIP",
    # data=txt_agip.encode("latin-1", errors="replace"),
    # file_name=f"{agip_filename}_sifere_agip.txt",
    # mime="text/plain",
    # use_container_width=True,
    # )
    # except Exception as e:
    # st.error(f"Error al procesar el archivo: {str(e)}")
    # st.exception(e)
    #
    # st.markdown('</div>', unsafe_allow_html=True)
    # else:
    # st.markdown('<div style="text-align:center;padding:2rem 1rem;font-family:\'Space Mono\',monospace;font-size:0.72rem;color:#6b7280;letter-spacing:0.12em;">ESPERANDO ARCHIVO RENTAS CIUDAD · PASO 02</div>', unsafe_allow_html=True)

    # ── Rama TXT Mendez (Percepciones / Retenciones) ──────────────────────────────
    st.markdown('<div class="card"><div class="card-label">02 · Archivo fuente para SIFERE</div>', unsafe_allow_html=True)
    uploaded_sifere = st.file_uploader(
        "Arrastrá tu archivo de movimientos o hacé click para seleccionarlo",
        type=["txt", "prn"],
        label_visibility="visible",
        key="sifere_file"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_sifere:
        sifere_filename = Path(uploaded_sifere.name).stem
        st.success(f"**{uploaded_sifere.name}** listo para procesar")

        tipo_label = tipo_sifere.lower()  # "percepciones" o "retenciones"
        st.markdown(f'<div class="card"><div class="card-label">03 · Generar TXT SIFERE ({tipo_sifere})</div>', unsafe_allow_html=True)

        if st.button(f"⬡  Generar archivo SIFERE ({tipo_sifere})"):
            try:
                with st.spinner("Procesando..."):
                    raw_bytes = uploaded_sifere.getvalue()
                    content_str = raw_bytes.decode('latin-1', errors='replace')
                    movimientos, metadata = parsear_archivo(content=content_str)

                    if tipo_sifere == "Percepciones":
                        txt_sifere = generar_sifere_txt(movimientos, metadata)
                    else:
                        txt_sifere = generar_sifere_retenciones_txt(movimientos, metadata)

                st.success(f"✓  Archivo SIFERE ({tipo_sifere}) generado con éxito")

                # Stats
                st.markdown(f"""
                <div class="stats-row">
                    <div class="stat-chip">
                        <span class="stat-val">{len(movimientos)}</span>
                        <span class="stat-lbl">Movimientos</span>
                    </div>
                    <div class="stat-chip">
                        <span class="stat-val">{len(txt_sifere.splitlines())}</span>
                        <span class="stat-lbl">Líneas TXT</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.download_button(
                    label=f"↓  Descargar TXT SIFERE ({tipo_sifere})",
                    data=txt_sifere.encode("latin-1", errors="replace"),
                    file_name=f"{sifere_filename}_sifere_{tipo_label}.txt",
                    mime="text/plain",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Error al procesar el archivo: {str(e)}")
                st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            ESPERANDO ARCHIVO · PASO 01
        </div>
        """, unsafe_allow_html=True)



elif herramienta == TOOL_LIQUIDACIONES:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Liquidaciones Tarjeta (PDF)
    # ───────────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Archivo PDF de Liquidaciones</div>', unsafe_allow_html=True)
    uploaded_liq = st.file_uploader(
        "Arrastrá tu PDF de liquidaciones o hacé click para seleccionarlo",
        type=["pdf"],
        label_visibility="visible",
        key="liquidaciones_pdf"
    )
    formato_pdf = st.radio(
        "Formato del PDF",
        options=["Fiserv / First Data", "Santander"],
        index=0,
        horizontal=True,
        key="liq_formato_pdf",
        help="Fiserv: tabular con columnas $. Santander: bloques 'FECHA DE PAGO' con desglose final.",
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_liq:
        liq_filename = Path(uploaded_liq.name).stem
        st.success(f"**{uploaded_liq.name}** listo para procesar")

        # ─── Auto-fill desde el encabezado del PDF (sólo Santander) ───────────────
        # Si el archivo cambió, parseamos metadata y pre-cargamos los inputs vía
        # session_state. El usuario puede sobrescribir manualmente — no volvemos a
        # auto-fill en ese caso (porque el file_id ya se registró).
        sant_meta = {}
        if formato_pdf == "Santander":
            file_id = f"{uploaded_liq.name}|{uploaded_liq.size}"
            if st.session_state.get("liq_sant_last_file_id") != file_id:
                try:
                    _rd = PyPDF2.PdfReader(io.BytesIO(uploaded_liq.getvalue()))
                    _txt = "".join(p.extract_text() + "\n" for p in _rd.pages)
                    sant_meta = parsear_pdf_santander(_txt)["meta"]
                    st.session_state["liq_sant_meta"] = sant_meta
                    st.session_state["liq_sant_last_file_id"] = file_id
                    if sant_meta.get("razon_social"):
                        st.session_state["liq_contribuyente"] = sant_meta["razon_social"]
                    if sant_meta.get("periodo"):
                        st.session_state["liq_periodo"] = sant_meta["periodo"]
                except Exception:
                    sant_meta = {}
            else:
                sant_meta = st.session_state.get("liq_sant_meta", {})

        # ─── Card 02: Datos del contribuyente ──────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">02 · Datos del contribuyente</div>', unsafe_allow_html=True)
        col_a, col_b = st.columns(2)
        with col_a:
            nombre_contribuyente = st.text_input(
                "Nombre del contribuyente",
                value="",
                placeholder="Ej: Juan Pérez",
                key="liq_contribuyente"
            )
        with col_b:
            if formato_pdf == "Santander":
                # Santander emite un único resumen consolidado por tarjeta —
                # no se subdivide en Crédito/Débito.
                tipo_tarjeta = st.selectbox(
                    "Tipo de tarjeta / Entidad",
                    options=["Visa", "Mastercard", "American Express", "Maestro",
                             "Cabal", "Naranja", "First Data", "Otra"],
                    index=0,
                    key="liq_tarjeta_santander",
                )
            else:
                tipo_tarjeta = st.selectbox(
                    "Tipo de tarjeta / Entidad",
                    options=["Visa Crédito", "Visa Débito", "Mastercard Crédito", "Mastercard Débito",
                             "American Express Crédito", "American Express Débito",
                             "Maestro Crédito", "Maestro Débito",
                             "Cabal Crédito", "Cabal Débito", "Naranja",
                             "First Data", "Otra"],
                    index=0,
                    key="liq_tarjeta",
                )
        # Si selecciona "Otra", mostrar text input
        if tipo_tarjeta == "Otra":
            tipo_tarjeta_custom = st.text_input(
                "Especificá el tipo de tarjeta / entidad",
                value="",
                placeholder="Ej: Mercado Pago",
                key="liq_tarjeta_custom"
            )
            tipo_tarjeta_final = tipo_tarjeta_custom.strip() if tipo_tarjeta_custom.strip() else "Otra"
        else:
            tipo_tarjeta_final = tipo_tarjeta

        periodo_liq = st.text_input(
            "Periodo (MM/AAAA)",
            value="",
            placeholder="Ej: 04/2025",
            key="liq_periodo"
        )
        if periodo_liq and not re.match(r'^(0[1-9]|1[0-2])/\d{4}$', periodo_liq):
            st.error("El periodo debe tener formato MM/AAAA (ej: 04/2025)")
        st.markdown('</div>', unsafe_allow_html=True)

        # ─── Card 03: Procesar ─────────────────────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">03 · Generar Excel de Liquidaciones</div>', unsafe_allow_html=True)

        btn_procesar = st.button("⬡  Procesar Liquidaciones")

        # ─── Botón Procesar ────────────────────────────────────────────────────────
        if btn_procesar:
            if not nombre_contribuyente.strip():
                st.warning("Ingresá el nombre del contribuyente antes de procesar.")
            elif not periodo_liq.strip():
                st.warning("Ingresá el periodo (MM/AAAA) antes de procesar.")
            elif formato_pdf == "Santander":
                # ─── Rama Santander ──────────────────────────────────────────────
                try:
                    periodo_parts = periodo_liq.strip().split("/")
                    if len(periodo_parts) == 2:
                        mes_liq = periodo_parts[0].zfill(2)
                        anio_liq = periodo_parts[1]
                        periodo_codigo = anio_liq[2:] + mes_liq
                    else:
                        periodo_codigo = "2501"

                    with st.spinner("Leyendo PDF..."):
                        reader = PyPDF2.PdfReader(io.BytesIO(uploaded_liq.getvalue()))
                        texto = "".join(p.extract_text() + "\n" for p in reader.pages)

                    with st.spinner("Extrayendo liquidaciones VISA Santander..."):
                        out = parsear_pdf_santander(texto)

                    if not out["liquidaciones"]:
                        st.error("No se encontraron bloques 'FECHA DE PAGO' en el PDF. Verificá que sea un resumen mensual VISA Santander.")
                    else:
                        meta = out["meta"]
                        # Etiquetas dinámicas con la alícuota leída del desglose
                        col_iva = f"IVA {meta['tasa_iva_label']} %"
                        col_perc_ib = f"Perc. IB {meta['tasa_perc_ib_label']} %"
                        col_sirtac = f"Ret. IB SIRTAC {meta['tasa_sirtac_label']} %"

                        df_liq = pd.DataFrame(out["liquidaciones"]).rename(columns={
                            "IVA": col_iva,
                            "Perc.IB": col_perc_ib,
                            "SIRTAC": col_sirtac,
                        })[[
                            "Fecha Pago", "Fecha Pres.", "Liquidaciones", "Detalle",
                            "Monto Presentado", "Arancel", "Serv. Costos Financieros",
                            "Servicio PAYWAY", col_iva, col_perc_ib, col_sirtac,
                            "AFIP-DGI", "Neto Percibido",
                        ]]

                        # Encabezado: usar Nro de Resumen del PDF (más identificativo que la primera liq)
                        nro_resumen_clean = (meta.get("nro_resumen") or "").lstrip("0") or "0"
                        encabezado_fc = f"FC {periodo_codigo}-{nro_resumen_clean}/A"
                        banco = "BANCO SANTANDER RIO S.A."

                        with st.spinner("Generando Excel..."):
                            output = io.BytesIO()
                            money_fmt = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                            header_fill = PatternFill('solid', fgColor='2E7D32')
                            zebra_fill = PatternFill('solid', fgColor='C8E6C9')
                            base_iva_fill = PatternFill('solid', fgColor='E3F2FD')
                            header_font_white = Font(bold=True, size=11, color='FFFFFF')
                            center_align = Alignment(horizontal='center', vertical='center')
                            thick_side = Side(border_style='thick', color='000000')
                            no_side = Side(border_style=None)

                            def _santander_encabezado(ws):
                                ws.sheet_view.showGridLines = False
                                ws.insert_cols(1)
                                ws.column_dimensions['A'].width = 4
                                ws.insert_rows(1)
                                ws.insert_rows(2, 6)
                                merge_end = "E"
                                cabecera = [
                                    (2, f"LIQUIDACION DE TARJETA: {tipo_tarjeta_final.upper()} SANTANDER", Font(bold=True, size=14, color='FFFFFF'), header_fill),
                                    (3, nombre_contribuyente.upper(), Font(bold=True, size=11, color='2E7D32'), None),
                                    (4, encabezado_fc, Font(bold=True, size=11, color='2E7D32'), None),
                                    (5, banco, Font(italic=True, size=10, color='388E3C'), None),
                                    (6, f"PERIODO: {periodo_liq.strip()}", Font(italic=True, size=10, color='388E3C'), None),
                                ]
                                for r, val, font, fill in cabecera:
                                    ws.merge_cells(f'B{r}:{merge_end}{r}')
                                    cell = ws[f'B{r}']
                                    cell.value = val
                                    cell.font = font
                                    if fill:
                                        cell.fill = fill
                                    cell.alignment = center_align
                                for row_i in range(2, 7):
                                    for col_i in range(2, 6):
                                        c = ws.cell(row=row_i, column=col_i)
                                        t = thick_side if row_i == 2 else no_side
                                        b = thick_side if row_i == 6 else no_side
                                        l = thick_side if col_i == 2 else no_side
                                        r_ = thick_side if col_i == 5 else no_side
                                        c.border = Border(top=t, bottom=b, left=l, right=r_)

                            def formatear_hoja_liq_sant(ws, df):
                                _santander_encabezado(ws)
                                first_data_col = 2
                                last_data_col = len(df.columns) + 1
                                data_header_row = 8
                                data_start_row = 9
                                last_data_row = data_start_row + len(df) - 1
                                total_row = last_data_row + 1
                                for col_idx in range(first_data_col, last_data_col + 1):
                                    cell = ws.cell(row=data_header_row, column=col_idx)
                                    cell.font = header_font_white
                                    cell.fill = header_fill
                                    cell.alignment = center_align
                                text_cols = {2, 3, 4, 5}
                                for row_idx in range(data_start_row, last_data_row + 1):
                                    for col_idx in range(first_data_col, last_data_col + 1):
                                        cell = ws.cell(row=row_idx, column=col_idx)
                                        cell.alignment = center_align
                                        if col_idx not in text_cols and isinstance(cell.value, (int, float)):
                                            cell.number_format = money_fmt
                                    if (row_idx - data_start_row) % 2 == 0:
                                        for col_idx in range(first_data_col, last_data_col + 1):
                                            ws.cell(row=row_idx, column=col_idx).fill = zebra_fill
                                ws.merge_cells(f'B{total_row}:E{total_row}')
                                ws[f'B{total_row}'] = "TOTAL"
                                ws[f'B{total_row}'].font = Font(bold=True, size=11, color='FFFFFF')
                                ws[f'B{total_row}'].fill = header_fill
                                ws[f'B{total_row}'].alignment = center_align
                                for col_idx in range(3, 6):
                                    ws.cell(row=total_row, column=col_idx).fill = header_fill
                                for col_idx in range(6, last_data_col + 1):
                                    cell = ws.cell(row=total_row, column=col_idx)
                                    col_letter = get_column_letter(col_idx)
                                    cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
                                    cell.number_format = money_fmt
                                    cell.font = Font(bold=True, size=10, color='FFFFFF')
                                    cell.fill = header_fill
                                    cell.alignment = center_align
                                for row_i in range(data_header_row, total_row + 1):
                                    is_special = row_i in (data_header_row, total_row)
                                    for col_i in range(first_data_col, last_data_col + 1):
                                        cell = ws.cell(row=row_i, column=col_i)
                                        t = thick_side if row_i == data_header_row else (thick_side if row_i == total_row else no_side)
                                        b = thick_side if row_i == total_row else (thick_side if row_i == data_header_row else no_side)
                                        l = thick_side if (col_i == first_data_col or is_special) else no_side
                                        r_ = thick_side if (col_i == last_data_col or is_special) else no_side
                                        cell.border = Border(top=t, bottom=b, left=l, right=r_)
                                for col_idx in range(first_data_col, last_data_col + 1):
                                    col_letter = get_column_letter(col_idx)
                                    max_len = max(
                                        len(str(ws.cell(row=r, column=col_idx).value or ''))
                                        for r in range(data_header_row, min(total_row + 1, data_header_row + 50))
                                    )
                                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                                ws.column_dimensions['E'].width = max(ws.column_dimensions['E'].width, 38)
                                # ─── Resumen Impositivo (referencia las celdas de la fila TOTAL) ──
                                # NETO 21 = Arancel + SCF + PAYWAY (suma de las 3 columnas, fila TOTAL).
                                # IVA / Perc.IB / SIRTAC / AFIP-DGI = celdas TOTAL de cada columna.
                                df_cols = list(df.columns)
                                idx_arancel = df_cols.index("Arancel") + first_data_col
                                idx_scf = df_cols.index("Serv. Costos Financieros") + first_data_col
                                idx_payway = df_cols.index("Servicio PAYWAY") + first_data_col
                                idx_iva = next((i + first_data_col for i, c in enumerate(df_cols) if c.startswith("IVA ")), None)
                                idx_pib = next((i + first_data_col for i, c in enumerate(df_cols) if c.startswith("Perc. IB ")), None)
                                idx_sir = next((i + first_data_col for i, c in enumerate(df_cols) if "SIRTAC" in c), None)
                                idx_afip = (df_cols.index("AFIP-DGI") + first_data_col) if "AFIP-DGI" in df_cols else None
                                resumen = [(
                                    "NETO 21",
                                    f"={get_column_letter(idx_arancel)}{total_row}+{get_column_letter(idx_scf)}{total_row}+{get_column_letter(idx_payway)}{total_row}",
                                )]
                                if idx_iva:
                                    resumen.append(("IVA 21", f"={get_column_letter(idx_iva)}{total_row}"))
                                if idx_pib:
                                    resumen.append(("PERC. IB BS AS", f"={get_column_letter(idx_pib)}{total_row}"))
                                if idx_sir:
                                    resumen.append(("SIRTAC", f"={get_column_letter(idx_sir)}{total_row}"))
                                if idx_afip:
                                    resumen.append(("PERC. AFIP-DGI", f"={get_column_letter(idx_afip)}{total_row}"))
                                rs = total_row + 2
                                ws.merge_cells(f'B{rs}:D{rs}')
                                ws[f'B{rs}'] = "RESUMEN IMPOSITIVO"
                                ws[f'B{rs}'].font = Font(bold=True, size=11, color='FFFFFF')
                                ws[f'B{rs}'].fill = header_fill
                                ws[f'B{rs}'].alignment = center_align
                                for idx, (concepto, formula) in enumerate(resumen):
                                    r = rs + 1 + idx
                                    ws.merge_cells(f'B{r}:C{r}')
                                    ws[f'B{r}'] = concepto
                                    ws[f'B{r}'].font = Font(bold=True, size=10)
                                    ws[f'B{r}'].alignment = center_align
                                    cv = ws.cell(row=r, column=4)
                                    cv.value = formula
                                    cv.number_format = money_fmt
                                    cv.alignment = center_align
                                    if idx % 2 == 0:
                                        ws[f'B{r}'].fill = zebra_fill
                                        ws.cell(row=r, column=3).fill = zebra_fill
                                        cv.fill = zebra_fill
                                rt = rs + 1 + len(resumen)
                                ws.merge_cells(f'B{rt}:C{rt}')
                                ws[f'B{rt}'] = "TOTAL"
                                ws[f'B{rt}'].font = Font(bold=True, size=11, color='FFFFFF')
                                ws[f'B{rt}'].fill = header_fill
                                ws[f'B{rt}'].alignment = center_align
                                ws.cell(row=rt, column=3).fill = header_fill
                                ct = ws.cell(row=rt, column=4)
                                # TOTAL del resumen = IVA + Perc.IB + SIRTAC + AFIP (suma desde el segundo item, NETO 21 es base imponible).
                                ct.value = f"=SUM(D{rs+2}:D{rt-1})"
                                ct.number_format = money_fmt
                                ct.font = Font(bold=True, size=10, color='FFFFFF')
                                ct.fill = header_fill
                                ct.alignment = center_align
                                for row_i in range(rs, rt + 1):
                                    for col_i in range(2, 5):
                                        c = ws.cell(row=row_i, column=col_i)
                                        t = thick_side if row_i == rs else no_side
                                        b = thick_side if row_i == rt else no_side
                                        l = thick_side if col_i == 2 else no_side
                                        r_ = thick_side if col_i == 4 else no_side
                                        c.border = Border(top=t, bottom=b, left=l, right=r_)

                            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                                df_liq.to_excel(writer, sheet_name="Liquidaciones", index=False)
                                wb = writer.book
                                formatear_hoja_liq_sant(wb["Liquidaciones"], df_liq)

                            output.seek(0)

                        st.success(f"✓  {len(df_liq)} fechas de pago procesadas (formato Santander)")
                        n_constancia = sum(1 for v in (out.get("liquidaciones") or []) if v.get("AFIP-DGI", 0) > 0)
                        st.markdown(f"""
                        <div class="stats-row">
                            <div class="stat-chip">
                                <span class="stat-val">{len(df_liq)}</span>
                                <span class="stat-lbl">Fechas de Pago</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{n_constancia}</span>
                                <span class="stat-lbl">Con Perc. AFIP</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{len(reader.pages)}</span>
                                <span class="stat-lbl">Páginas PDF</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        st.info(
                            f"**{nombre_contribuyente}** · {tipo_tarjeta_final} Santander · "
                            f"**{encabezado_fc}** · {banco}"
                        )
                        st.download_button(
                            label="↓  Descargar Excel de Liquidaciones",
                            data=output,
                            file_name=f"{liq_filename}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                except Exception as e:
                    st.error(f"Error al procesar el archivo: {str(e)}")
                    st.exception(e)
            else:
                try:
                    # Parsear periodo para el encabezado (MM/AAAA -> AAMM)
                    periodo_parts = periodo_liq.strip().split("/")
                    if len(periodo_parts) == 2:
                        mes_liq = periodo_parts[0].zfill(2)
                        anio_liq = periodo_parts[1]
                        periodo_codigo = anio_liq[2:] + mes_liq  # AAMM
                    else:
                        mes_liq = "01"
                        anio_liq = "2025"
                        periodo_codigo = "2501"

                    with st.spinner("Leyendo PDF..."):
                        reader = PyPDF2.PdfReader(io.BytesIO(uploaded_liq.getvalue()))
                        texto = "".join(page.extract_text() + "\n" for page in reader.pages)
                        texto_lines = texto.splitlines()

                    with st.spinner("Extrayendo movimientos..."):
                        capturar = False
                        movimientos = []
                        movimiento = {}
                        # Extraer nombre del banco de la segunda línea del PDF
                        banco = texto_lines[1] if len(texto_lines) > 1 else "Banco desconocido"

                        for linea in texto_lines:
                            if "F.de Pago" in linea:
                                cbu_match = re.search(r"\d{1,3}(\.\d{3})*,\d+\-?\s+(\d+)", linea)
                                if cbu_match:
                                    # Línea de cierre con datos reales de pago → guardar movimiento
                                    capturar = False
                                    todas_fechas = re.findall(r"(\d{2}/\d{2}/\d{4})", linea)
                                    fecha_pres = todas_fechas[0] if len(todas_fechas) >= 1 else "No se encontró fecha"
                                    fecha_pago = todas_fechas[-1] if len(todas_fechas) >= 2 else fecha_pres
                                    movimiento["Liquidacion"] = round(float(cbu_match.group(2)))
                                    movimiento["Fecha Pago"] = fecha_pago
                                    movimiento["Fecha Pres."] = fecha_pres
                                    if movimiento.get("Liquidacion"):
                                        movimientos.append(movimiento.copy())
                                        movimiento = {}
                                    capturar = True  # Activar captura para el bloque siguiente
                                else:
                                    # Línea de encabezado de bloque (sin datos de pago) → activar captura
                                    capturar = True
                                continue  # No procesar líneas F.de Pago como dato

                            if "VENTAS" in linea or "QR" in linea or "AJUSTE" in linea or "ACREDITACIONES PAGO QRD" in linea:
                                capturar = True

                            if capturar:
                                partes = linea.split("$")
                                if len(partes) > 1:
                                    concepto_raw = partes[0].strip()
                                    # Limpiar el +/- del final del nombre del concepto (notación del formato)
                                    concepto = concepto_raw.rstrip("+-").strip()
                                    # Signo: "+" al final del concepto = crédito/reversa = negativo
                                    #        "-" al final del concepto = deducción normal = positivo
                                    #        "-" en el valor (formato original FISERV) = negativo
                                    es_negativo = ("-" in partes[1]) or concepto_raw.endswith("+")
                                    valor = partes[1].strip().replace("Fecha", "").replace("-", "").replace(".", "").replace(",", ".")
                                    if "/" not in valor:
                                        try:
                                            num_val = round(float(valor), 2) * (-1 if es_negativo else 1)
                                            if "ACREDITACIONES PAGO QRD" in concepto:
                                                num_val = -abs(num_val)

                                            # Sumar si el concepto ya existe en el movimiento
                                            if concepto in movimiento and isinstance(movimiento[concepto], (int, float)):
                                                movimiento[concepto] += num_val
                                            else:
                                                movimiento[concepto] = num_val
                                        except ValueError:
                                            continue
                                    else:
                                        movimiento[concepto] = partes[1]

                    if not movimientos:
                        st.error("No se encontraron liquidaciones en el PDF. Verificá el formato del archivo.")
                    else:
                        with st.spinner("Generando Excel..."):
                            df_total = pd.DataFrame(movimientos).fillna(0)
                            
                            # Integrar ACREDITACIONES PAGO QRD a QR RETENCION IIBB
                            col_acred = next((c for c in df_total.columns if "ACREDITACIONES PAGO QRD" in c), None)
                            col_qr_ret = next((c for c in df_total.columns if "QR" in c and "RETENCION" in c and "IIBB" in c), None)
                            
                            if col_acred:
                                if col_qr_ret:
                                    df_total[col_qr_ret] += df_total[col_acred]
                                else:
                                    df_total["QR RETENCION IIBB"] = df_total[col_acred]

                            columnas_qr = [col for col in df_total.columns if col.startswith("QR")]
                            df_qr = df_total[df_total[columnas_qr].sum(axis=1) != 0][["Fecha Pago", "Fecha Pres.", "Liquidacion"] + columnas_qr] if columnas_qr else None
                            columnas_ajuste = [col for col in df_total.columns if "AJUSTE" in col]
                            df_ajuste = df_total[df_total[columnas_ajuste].sum(axis=1) != 0][["Fecha Pago", "Fecha Pres.", "Liquidacion"] + columnas_ajuste] if columnas_ajuste else None

                            df_movimientos = df_total.drop(columns=columnas_qr + columnas_ajuste)
                            columnas_importe_neto = [col for col in df_movimientos.columns if "IMPORTE NETO" in col]
                            columnas_ventas = [col for col in df_movimientos.columns if col.startswith("VENTAS")]
                            columnas_restantes = [col for col in df_movimientos.columns if col not in columnas_importe_neto + columnas_ventas + ["Fecha Pago", "Fecha Pres.", "Liquidacion"]]
                            df_movimientos = df_movimientos[["Fecha Pago", "Fecha Pres.", "Liquidacion"] + columnas_restantes + columnas_importe_neto + columnas_ventas]

                            # Obtener primer numero de liquidacion
                            primer_liq = str(int(df_movimientos["Liquidacion"].iloc[0])) if len(df_movimientos) > 0 else "0"
                            encabezado_fc = f"FC {periodo_codigo}-{primer_liq}/A"

                            # CUIT First Data (hardcoded, no aparece en extractos)
                            CUIT_FIRST_DATA = "30-52221156-3"

                            output = io.BytesIO()
                            border = Border(
                                left=Side(border_style="thin"), right=Side(border_style="thin"),
                                top=Side(border_style="thin"), bottom=Side(border_style="thin")
                            )
                            money_fmt = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                            # Colores VERDES
                            header_fill = PatternFill('solid', fgColor='2E7D32')
                            zebra_fill = PatternFill('solid', fgColor='C8E6C9')
                            header_font_white = Font(bold=True, size=11, color='FFFFFF')
                            center_align = Alignment(horizontal='center', vertical='center')

                            def formatear_hoja_liq(ws, df_hoja, columnas_ignorar, titulo_encabezado=None, nombre_entidad="", mostrar_resumen=True):
                                # Ocultar líneas de cuadrícula
                                ws.sheet_view.showGridLines = False

                                # Insertar columna vacía al principio como separador visual
                                ws.insert_cols(1)
                                ws.column_dimensions['A'].width = 4  # Ancho para espaciado visual

                                # Insertar fila vacía al principio como separador visual
                                ws.insert_rows(1)

                                n_cols = len(df_hoja.columns)
                                # Offset +1 por la columna espaciadora
                                col_offset = 1
                                first_data_col = 1 + col_offset  # Columna B
                                last_data_col = n_cols + col_offset  # Última columna de datos

                                # Encabezado: filas de titulo
                                if titulo_encabezado:
                                    ws.insert_rows(2, 6)  # 5 filas de encabezado + 1 en blanco (después de fila vacía)

                                    # 4 columnas para el encabezado (B:E)
                                    merge_end = get_column_letter(first_data_col + 3)  # 4 columnas desde B

                                    # Fila 2: LIQUIDACION DE TARJETA: (tarjeta)
                                    ws.merge_cells(f'B2:{merge_end}2')
                                    ws['B2'] = f"LIQUIDACION DE TARJETA: {tipo_tarjeta_final.upper()}"
                                    ws['B2'].font = Font(bold=True, size=14, color='FFFFFF')
                                    ws['B2'].fill = header_fill
                                    ws['B2'].alignment = center_align

                                    # Fila 3: Contribuyente
                                    ws.merge_cells(f'B3:{merge_end}3')
                                    ws['B3'] = nombre_contribuyente.upper()
                                    ws['B3'].font = Font(bold=True, size=11, color='2E7D32')
                                    ws['B3'].alignment = center_align

                                    # Fila 4: Comprobante (AAMM-NroLiq/A)
                                    ws.merge_cells(f'B4:{merge_end}4')
                                    ws['B4'] = titulo_encabezado
                                    ws['B4'].font = Font(bold=True, size=11, color='2E7D32')
                                    ws['B4'].alignment = center_align

                                    # Fila 5: Entidad bancaria
                                    ws.merge_cells(f'B5:{merge_end}5')
                                    entidad_display = nombre_entidad if nombre_entidad else banco
                                    ws['B5'] = entidad_display
                                    ws['B5'].font = Font(italic=True, size=10, color='388E3C')
                                    ws['B5'].alignment = center_align

                                    # Fila 6: Periodo
                                    ws.merge_cells(f'B6:{merge_end}6')
                                    ws['B6'] = f"PERIODO: {periodo_liq.strip()}"
                                    ws['B6'].font = Font(italic=True, size=10, color='388E3C')
                                    ws['B6'].alignment = center_align

                                    # ─── Borde negro intenso externo en el encabezado (filas 2-6, cols B:merge_end) ───
                                    thick_side = Side(border_style='thick', color='000000')
                                    no_side = Side(border_style=None)
                                    merge_end_idx = first_data_col + 3
                                    for row_i in range(2, 7):
                                        for col_i in range(first_data_col, merge_end_idx + 1):
                                            cell = ws.cell(row=row_i, column=col_i)
                                            t = thick_side if row_i == 2 else no_side
                                            b = thick_side if row_i == 6 else no_side
                                            l = thick_side if col_i == first_data_col else no_side
                                            r = thick_side if col_i == merge_end_idx else no_side
                                            cell.border = Border(top=t, bottom=b, left=l, right=r)

                                    # Fila 7: en blanco (separador)
                                    data_header_row = 8
                                    data_start_row = 9
                                else:
                                    data_header_row = 2
                                    data_start_row = 3

                                # Ajustar columnas_ignorar con offset (B=col2, C=col3, etc.)
                                columnas_ignorar_offset = [get_column_letter(ord(c) - ord('A') + 1 + col_offset) for c in columnas_ignorar]

                                # Estilo de encabezados de columna
                                for col_idx in range(first_data_col, last_data_col + 1):
                                    cell = ws.cell(row=data_header_row, column=col_idx)
                                    cell.font = header_font_white
                                    cell.fill = header_fill
                                    cell.alignment = center_align
                                    cell.border = border

                                # Estilo de datos (sin bordes internos)
                                last_data_row = data_start_row + len(df_hoja) - 1
                                for row_idx in range(data_start_row, last_data_row + 1):
                                    for col_idx in range(first_data_col, last_data_col + 1):
                                        cell = ws.cell(row=row_idx, column=col_idx)
                                        cell.alignment = center_align
                                        if cell.column_letter not in columnas_ignorar_offset:
                                            if isinstance(cell.value, (int, float)):
                                                cell.number_format = money_fmt
                                    # Zebra verde
                                    if (row_idx - data_start_row) % 2 == 0:
                                        for col_idx in range(first_data_col, last_data_col + 1):
                                            ws.cell(row=row_idx, column=col_idx).fill = zebra_fill

                                # Fila TOTAL
                                total_row = last_data_row + 1
                                fc = first_data_col
                                fc_letter = get_column_letter(fc)
                                fc2_letter = get_column_letter(fc + 2)  # incluye Fecha Pago, Fecha Pres. y Nro Liquidacion
                                ws.merge_cells(f'{fc_letter}{total_row}:{fc2_letter}{total_row}')
                                ws[f'{fc_letter}{total_row}'] = "TOTAL"
                                ws[f'{fc_letter}{total_row}'].font = Font(bold=True, size=11, color='FFFFFF')
                                ws[f'{fc_letter}{total_row}'].fill = header_fill
                                ws[f'{fc_letter}{total_row}'].alignment = center_align
                                ws.cell(row=total_row, column=fc + 1).fill = header_fill
                                ws.cell(row=total_row, column=fc + 2).fill = header_fill

                                # SUM desde fc+3 en adelante (saltando Fecha Pago, Fecha Pres. y Nro Liquidacion)
                                for col_idx in range(fc + 3, last_data_col + 1):
                                    cell = ws.cell(row=total_row, column=col_idx)
                                    col_letter = get_column_letter(col_idx)
                                    cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
                                    cell.number_format = money_fmt
                                    cell.font = Font(bold=True, size=10, color='FFFFFF')
                                    cell.fill = header_fill
                                    cell.alignment = center_align

                                # ─── Borde negro intenso externo + verticales en header y total ───
                                thick_side = Side(border_style='thick', color='000000')
                                no_side = Side(border_style=None)
                                for row_i in range(data_header_row, total_row + 1):
                                    is_header = (row_i == data_header_row)
                                    is_total = (row_i == total_row)
                                    is_special_row = is_header or is_total
                                    for col_i in range(first_data_col, last_data_col + 1):
                                        cell = ws.cell(row=row_i, column=col_i)
                                        t = thick_side if is_header else (thick_side if is_total else no_side)
                                        b = thick_side if is_total else (thick_side if is_header else no_side)
                                        # Verticales thick en header y total, solo extremos en datos
                                        l = thick_side if (col_i == first_data_col or is_special_row) else no_side
                                        r = thick_side if (col_i == last_data_col or is_special_row) else no_side
                                        cell.border = Border(top=t, bottom=b, left=l, right=r)

                                # Auto-ajustar columnas (solo las de datos)
                                for col_idx in range(first_data_col, last_data_col + 1):
                                    col_letter = get_column_letter(col_idx)
                                    max_len = max(
                                        len(str(ws.cell(row=r, column=col_idx).value or ''))
                                        for r in range(data_header_row, min(total_row + 1, data_header_row + 50))
                                    )
                                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

                                # ─── CARGO TERMINAL y ACREDITACIONES highlight (siempre) ───────────────────
                                cols_hoja = list(df_hoja.columns)
                                cargo_terminal_cols = []
                                acreditaciones_cols = []
                                for i, col in enumerate(cols_hoja):
                                    real_col = i + 1 + col_offset  # +1 por offset de columna espaciadora
                                    if "CARGO TERMINAL" in col.upper():
                                        cargo_terminal_cols.append((real_col, get_column_letter(real_col)))
                                    if "ACREDITACIONES PAGO QRD" in col.upper():
                                        acreditaciones_cols.append((real_col, get_column_letter(real_col)))

                                yellow_fill = PatternFill('solid', fgColor='FFD600')
                                for col_idx_ct, col_letter_ct in cargo_terminal_cols:
                                    hdr_cell = ws.cell(row=data_header_row, column=col_idx_ct)
                                    hdr_cell.fill = PatternFill('solid', fgColor='FF6F00')
                                    hdr_cell.font = Font(bold=True, size=11, color='FFFFFF')
                                    from openpyxl.comments import Comment
                                    hdr_cell.comment = Comment("ATENCION: Cargo terminal detectado", "Sistema")
                                    for row_idx in range(data_start_row, total_row + 1):
                                        ws.cell(row=row_idx, column=col_idx_ct).fill = yellow_fill

                                for col_idx_ac, col_letter_ac in acreditaciones_cols:
                                    hdr_cell = ws.cell(row=data_header_row, column=col_idx_ac)
                                    hdr_cell.fill = yellow_fill
                                    hdr_cell.font = Font(bold=True, size=11, color='000000')
                                    for row_idx in range(data_start_row, total_row + 1):
                                        ws.cell(row=row_idx, column=col_idx_ac).fill = yellow_fill

                                # ─── Tabla resumen de impuestos (solo Liquidaciones) ──────
                                if mostrar_resumen:
                                    iva21_col_letters = []
                                    iva105_col_letters = []
                                    perc_iva_col_letters = []
                                    perc_iibb_col_letters = []
                                    sirtac_col_letters = []

                                    for i, col in enumerate(cols_hoja):
                                        col_upper = col.upper()
                                        col_letter = get_column_letter(i + 1 + col_offset)
                                        if "IVA" in col_upper and not col_upper.startswith("PER"):
                                            if "IVA RI" in col_upper or ("10,50" not in col and "10.50" not in col):
                                                iva21_col_letters.append(col_letter)
                                            else:
                                                iva105_col_letters.append(col_letter)
                                        if col_upper.startswith("PER") and "IVA" in col_upper:
                                            perc_iva_col_letters.append(col_letter)
                                        elif col_upper.startswith("PER") and "IVA" not in col_upper:
                                            perc_iibb_col_letters.append(col_letter)
                                        if "SIRTAC" in col_upper:
                                            sirtac_col_letters.append(col_letter)

                                    resumen_items = []
                                    tr = total_row

                                    if iva21_col_letters:
                                        if len(iva21_col_letters) == 1:
                                            iva_ref = f"{iva21_col_letters[0]}{tr}"
                                        else:
                                            iva_ref = "+".join(f"{cl}{tr}" for cl in iva21_col_letters)
                                        resumen_items.append(("NETO 21", f"=ABS({iva_ref})/0.21"))
                                        resumen_items.append(("IVA 21", f"=ABS({iva_ref})"))

                                    if iva105_col_letters:
                                        if len(iva105_col_letters) == 1:
                                            iva105_ref = f"{iva105_col_letters[0]}{tr}"
                                        else:
                                            iva105_ref = "+".join(f"{cl}{tr}" for cl in iva105_col_letters)
                                        resumen_items.append(("NETO 10.5", f"=ABS({iva105_ref})/0.105"))
                                        resumen_items.append(("IVA 10.5", f"=ABS({iva105_ref})"))

                                    if perc_iva_col_letters:
                                        if len(perc_iva_col_letters) == 1:
                                            p_ref = f"{perc_iva_col_letters[0]}{tr}"
                                        else:
                                            p_ref = "+".join(f"{cl}{tr}" for cl in perc_iva_col_letters)
                                        resumen_items.append(("PERC. IVA", f"=ABS({p_ref})"))

                                    if sirtac_col_letters:
                                        if len(sirtac_col_letters) == 1:
                                            s_ref = f"{sirtac_col_letters[0]}{tr}"
                                        else:
                                            s_ref = "+".join(f"{cl}{tr}" for cl in sirtac_col_letters)
                                        resumen_items.append(("SIRTAC", f"=ABS({s_ref})"))

                                    if perc_iibb_col_letters:
                                        if len(perc_iibb_col_letters) == 1:
                                            pi_ref = f"{perc_iibb_col_letters[0]}{tr}"
                                        else:
                                            pi_ref = "+".join(f"{cl}{tr}" for cl in perc_iibb_col_letters)
                                        resumen_items.append(("PERC. IIBB", f"=ABS({pi_ref})"))

                                    if resumen_items:
                                        resumen_start = total_row + 2
                                        # Resumen con 3 columnas (B:D)
                                        res_start_letter = get_column_letter(first_data_col)
                                        res_end_letter = get_column_letter(first_data_col + 2)
                                        ws.merge_cells(f'{res_start_letter}{resumen_start}:{res_end_letter}{resumen_start}')
                                        ws[f'{res_start_letter}{resumen_start}'] = "RESUMEN IMPOSITIVO"
                                        ws[f'{res_start_letter}{resumen_start}'].font = Font(bold=True, size=11, color='FFFFFF')
                                        ws[f'{res_start_letter}{resumen_start}'].fill = header_fill
                                        ws[f'{res_start_letter}{resumen_start}'].alignment = center_align

                                        for idx, (concepto, formula) in enumerate(resumen_items):
                                            r = resumen_start + 1 + idx
                                            merge_a = get_column_letter(first_data_col)
                                            merge_b = get_column_letter(first_data_col + 1)  # Concepto ocupa 2 cols (B:C)
                                            val_col = first_data_col + 2  # Valor en col D
                                            ws.merge_cells(f'{merge_a}{r}:{merge_b}{r}')
                                            ws[f'{merge_a}{r}'] = concepto
                                            ws[f'{merge_a}{r}'].font = Font(bold=True, size=10)
                                            ws[f'{merge_a}{r}'].alignment = center_align
                                            cell_val = ws.cell(row=r, column=val_col)
                                            cell_val.value = formula
                                            cell_val.number_format = money_fmt
                                            cell_val.alignment = center_align
                                            if idx % 2 == 0:
                                                ws[f'{merge_a}{r}'].fill = zebra_fill
                                                ws.cell(row=r, column=first_data_col + 1).fill = zebra_fill
                                                cell_val.fill = zebra_fill

                                        # Fila TOTAL del resumen
                                        r_total = resumen_start + 1 + len(resumen_items)
                                        first_val_row = resumen_start + 1
                                        last_val_row = r_total - 1
                                        merge_a = get_column_letter(first_data_col)
                                        merge_b = get_column_letter(first_data_col + 1)
                                        val_col = first_data_col + 2
                                        ws.merge_cells(f'{merge_a}{r_total}:{merge_b}{r_total}')
                                        ws[f'{merge_a}{r_total}'] = "TOTAL"
                                        ws[f'{merge_a}{r_total}'].font = Font(bold=True, size=11, color='FFFFFF')
                                        ws[f'{merge_a}{r_total}'].fill = header_fill
                                        ws[f'{merge_a}{r_total}'].alignment = center_align
                                        ws.cell(row=r_total, column=first_data_col + 1).fill = header_fill
                                        val_col_letter = get_column_letter(val_col)
                                        cell_total = ws.cell(row=r_total, column=val_col)
                                        cell_total.value = f"=SUM({val_col_letter}{first_val_row}:{val_col_letter}{last_val_row})"
                                        cell_total.number_format = money_fmt
                                        cell_total.font = Font(bold=True, size=10, color='FFFFFF')
                                        cell_total.fill = header_fill
                                        cell_total.alignment = center_align

                                        # ─── Borde negro intenso externo en resumen impositivo ───
                                        thick_side = Side(border_style='thick', color='000000')
                                        no_side = Side(border_style=None)
                                        res_first_col = first_data_col
                                        res_last_col = first_data_col + 2  # 3 columnas
                                        for row_i in range(resumen_start, r_total + 1):
                                            for col_i in range(res_first_col, res_last_col + 1):
                                                cell = ws.cell(row=row_i, column=col_i)
                                                t = thick_side if row_i == resumen_start else no_side
                                                b = thick_side if row_i == r_total else no_side
                                                l = thick_side if col_i == res_first_col else no_side
                                                r_s = thick_side if col_i == res_last_col else no_side
                                                cell.border = Border(top=t, bottom=b, left=l, right=r_s)

                            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                                df_movimientos.to_excel(writer, sheet_name="Liquidaciones", index=False)
                                if df_qr is not None:
                                    df_qr.to_excel(writer, sheet_name="QR", index=False)
                                if df_ajuste is not None:
                                    df_ajuste.to_excel(writer, sheet_name="AJUSTE", index=False)

                                # Formatear hojas
                                wb = writer.book
                                formatear_hoja_liq(wb["Liquidaciones"], df_movimientos, ["A", "B", "C"], encabezado_fc, banco, mostrar_resumen=True)
                                if df_qr is not None:
                                    formatear_hoja_liq(wb["QR"], df_qr, ["A", "B", "C"], encabezado_fc, "First Data", mostrar_resumen=False)
                                if df_ajuste is not None:
                                    formatear_hoja_liq(wb["AJUSTE"], df_ajuste, ["A", "B", "C"], encabezado_fc, banco, mostrar_resumen=False)

                            output.seek(0)

                        st.success("✓  Liquidaciones procesadas con éxito")

                        # Stats
                        st.markdown(f"""
                        <div class="stats-row">
                            <div class="stat-chip">
                                <span class="stat-val">{len(movimientos)}</span>
                                <span class="stat-lbl">Liquidaciones</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{len(columnas_qr)}</span>
                                <span class="stat-lbl">Cols. QR</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{len(columnas_ajuste)}</span>
                                <span class="stat-lbl">Cols. Ajuste</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                        st.info(
                            f"**{nombre_contribuyente}** · {tipo_tarjeta_final} · "
                            f"**{encabezado_fc}** · {banco} · "
                            f"{len(reader.pages)} páginas"
                        )

                        st.download_button(
                            label="↓  Descargar Excel de Liquidaciones",
                            data=output,
                            file_name=f"{liq_filename}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

                except Exception as e:
                    st.error(f"Error al procesar el archivo: {str(e)}")
                    st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            ESPERANDO ARCHIVO PDF · PASO 01
        </div>
        """, unsafe_allow_html=True)


elif herramienta == TOOL_DEDUCCIONES:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Limpieza Excel Deducciones IVA/Ganancias
    # ───────────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Archivo de Deducciones (.xls / .xlsx)</div>', unsafe_allow_html=True)
    uploaded_ded = st.file_uploader(
        "Subí el Excel descargado de Mis Retenciones/Percepciones de ARCA",
        type=["xls", "xlsx"],
        label_visibility="visible",
        key="deducciones_xls"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_ded:
        st.success(f"**{uploaded_ded.name}** listo para procesar")

        st.markdown('<div class="card"><div class="card-label">02 · Datos del contribuyente</div>', unsafe_allow_html=True)
        nombre_ded = st.text_input("Nombre / Razón Social del contribuyente", value="", key="nombre_deducciones")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><div class="card-label">03 · Procesar</div>', unsafe_allow_html=True)

        if st.button("⬡  Limpiar y Estilizar"):
            if not nombre_ded.strip():
                st.error("Ingresá el nombre del contribuyente para continuar.")
            else:
              try:
                with st.spinner("Procesando Excel de deducciones..."):
                    # Leer el archivo
                    df_ded = pd.read_excel(io.BytesIO(uploaded_ded.getvalue()))

                    if df_ded.empty:
                        st.error("El archivo está vacío.")
                        st.stop()

                    # ── Detectar tipo de impuesto ──
                    desc_imp_col = None
                    for c in df_ded.columns:
                        if 'descripci' in c.lower() and 'impuesto' in c.lower():
                            desc_imp_col = c
                            break
                    tipo_deduccion = 'DEDUCCIONES'
                    if desc_imp_col and not df_ded[desc_imp_col].dropna().empty:
                        primer_imp = str(df_ded[desc_imp_col].dropna().iloc[0]).upper()
                        if 'GANANCIA' in primer_imp:
                            tipo_deduccion = 'DEDUCCIONES GANANCIAS'
                        elif 'VALOR AGRE' in primer_imp or 'IVA' in primer_imp:
                            tipo_deduccion = 'DEDUCCIONES IVA'
                        elif 'SIRE' in primer_imp:
                            tipo_deduccion = 'SIRE IVA'

                    # ── Eliminar columnas vacías y redundantes ──
                    cols_drop = []
                    for c in df_ded.columns:
                        cl = c.lower()
                        if c.strip() == 'Impuesto' or c.strip() == 'Régimen':
                            cols_drop.append(c)
                    df_ded = df_ded.drop(columns=[c for c in cols_drop if c in df_ded.columns], errors='ignore')

                    # ── Renombrar columnas ──
                    # Detectar y renombrar la columna de Razón Social dinámicamente
                    for c in df_ded.columns:
                        cl = c.lower()
                        if 'denominaci' in cl and 'raz' in cl:
                            df_ded = df_ded.rename(columns={c: 'Razón Social'})
                            break

                    RENAME_DED = {
                        'CUIT Agente Ret./Perc.': 'CUIT',
                        'Descripción Impuesto': 'Impuesto',
                        'Descripción Régimen': 'Régimen',
                        'Fecha Ret./Perc.': 'Fecha',
                        'Número Certificado': 'Nro. Certificado',
                        'Descripción Operación': 'Operación',
                        'Importe Ret./Perc.': 'Importe',
                        'Número Comprobante': 'Nro. Comprobante',
                        'Fecha Comprobante': 'Fecha Comp.',
                        'Descripción Comprobante': 'Comprobante',
                        'Fecha Registración DJ Ag.Ret.': 'Fecha Reg. DJ',
                    }
                    df_ded = df_ded.rename(columns=RENAME_DED)

                    # ── Formatear CUIT como XX-XXXXXXXX-X ──
                    if 'CUIT' in df_ded.columns:
                        def format_cuit(val):
                            s = str(int(val)) if not pd.isna(val) else ''
                            if len(s) == 11:
                                return f"{s[:2]}-{s[2:10]}-{s[10]}"
                            return s
                        df_ded['CUIT'] = df_ded['CUIT'].apply(format_cuit)

                    # ── Ordenar por Fecha ascendente ──
                    if 'Fecha' in df_ded.columns:
                        try:
                            df_ded['_fecha_sort'] = pd.to_datetime(df_ded['Fecha'], format='%d/%m/%Y', errors='coerce')
                            df_ded = df_ded.sort_values('_fecha_sort', ascending=True).drop(columns=['_fecha_sort'])
                        except Exception:
                            pass

                    # ── Separar Retenciones y Percepciones ──
                    op_col = 'Operación'
                    df_ret = df_ded[df_ded[op_col].str.upper().str.contains('RETENCION', na=False)].copy() if op_col in df_ded.columns else pd.DataFrame()
                    df_per = df_ded[df_ded[op_col].str.upper().str.contains('PERCEPCION', na=False)].copy() if op_col in df_ded.columns else pd.DataFrame()
                    # Si no hay columna Operación, todo va a una hoja genérica
                    if op_col not in df_ded.columns:
                        df_ret = df_ded
                        df_per = pd.DataFrame()

                    # Eliminar columnas Impuesto y Operación (ya discriminadas por hoja)
                    for df_part in [df_ret, df_per]:
                        for drop_c in ['Impuesto', 'Operación']:
                            if drop_c in df_part.columns:
                                df_part.drop(columns=[drop_c], inplace=True)

                    # Mover Importe al final
                    for df_part in [df_ret, df_per]:
                        if 'Importe' in df_part.columns:
                            imp_data = df_part.pop('Importe')
                            df_part['Importe'] = imp_data

                    # ── Estilos dorados/ámbar ──
                    title_font = Font(bold=True, size=14, color='FFFFFF')
                    title_fill = PatternFill('solid', fgColor='BF8F00')
                    header_font = Font(bold=True, size=10, color='FFFFFF')
                    header_fill = PatternFill('solid', fgColor='D4A017')
                    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    center_align = Alignment(horizontal='center', vertical='center')
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    zebra_fill = PatternFill('solid', fgColor='FFF2CC')
                    accounting_fmt = '_-"$"* #,##0.00_-;-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

                    def _style_ded_sheet(ws, df_sheet, sheet_title, n_r, n_c, col_list):
                        """Aplica estilos dorados a una hoja de deducciones."""
                        # Fila 1: Nombre contribuyente
                        ws.merge_cells(f'A1:{get_column_letter(n_c)}1')
                        ws['A1'] = nombre_ded.strip().upper()
                        ws['A1'].font = title_font
                        ws['A1'].fill = title_fill
                        ws['A1'].alignment = center_align

                        # Fila 2: Título hoja + tipo y cantidad
                        ws.merge_cells(f'A2:{get_column_letter(n_c)}2')
                        ws['A2'] = f'{sheet_title} — {tipo_deduccion} — {n_r} registros'
                        ws['A2'].font = Font(italic=True, size=10, color='BF8F00')
                        ws['A2'].alignment = center_align

                        # Encabezados (fila 6)
                        for ci in range(1, n_c + 1):
                            cell = ws.cell(row=6, column=ci)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_align
                            cell.border = thin_border

                        # Columna Importe
                        imp_idx = col_list.index('Importe') + 1 if 'Importe' in col_list else None

                        # Datos (fila 7+)
                        for ri in range(7, n_r + 7):
                            for ci in range(1, n_c + 1):
                                cell = ws.cell(row=ri, column=ci)
                                cell.alignment = center_align
                                cell.border = thin_border
                                if ci == imp_idx:
                                    cell.number_format = accounting_fmt
                            if (ri - 7) % 2 == 0:
                                for ci in range(1, n_c + 1):
                                    ws.cell(row=ri, column=ci).fill = zebra_fill

                        # Fila TOTAL
                        if imp_idx:
                            tr = n_r + 7
                            ws.merge_cells(f'A{tr}:{get_column_letter(imp_idx - 1)}{tr}')
                            ws[f'A{tr}'] = 'TOTAL'
                            ws[f'A{tr}'].font = Font(bold=True)
                            ws[f'A{tr}'].alignment = Alignment(horizontal='right')
                            il = get_column_letter(imp_idx)
                            tc = ws.cell(row=tr, column=imp_idx)
                            tc.value = f'=SUM({il}7:{il}{tr - 1})'
                            tc.font = Font(bold=True)
                            tc.border = Border(top=Side(style='double'))
                            tc.number_format = accounting_fmt
                            tc.alignment = center_align

                        # Autofit
                        for ci in range(1, n_c + 1):
                            cl = get_column_letter(ci)
                            mx = len(str(ws.cell(row=6, column=ci).value or ''))
                            for ri in range(7, min(n_r + 7, 57)):
                                v = ws.cell(row=ri, column=ci).value
                                if v:
                                    mx = max(mx, len(str(v)))
                            ws.column_dimensions[cl].width = min(mx + 3, 45)

                    # ── Generar Excel ──
                    output = io.BytesIO()
                    sheets_written = []

                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        if not df_ret.empty:
                            df_ret.to_excel(writer, sheet_name='Retenciones', index=False, startrow=5)
                            ws_ret = writer.sheets['Retenciones']
                            ret_cols = list(df_ret.columns)
                            _style_ded_sheet(ws_ret, df_ret, 'RETENCIONES', len(df_ret), len(ret_cols), ret_cols)
                            sheets_written.append(('Retenciones', len(df_ret)))

                        if not df_per.empty:
                            df_per.to_excel(writer, sheet_name='Percepciones', index=False, startrow=5)
                            ws_per = writer.sheets['Percepciones']
                            per_cols = list(df_per.columns)
                            _style_ded_sheet(ws_per, df_per, 'PERCEPCIONES', len(df_per), len(per_cols), per_cols)
                            sheets_written.append(('Percepciones', len(df_per)))

                    output.seek(0)
                    n_rows = len(df_ded)

                st.success("✓  Proceso completado con éxito")

                # Stats
                stats_html = f'<div class="stats-row"><div class="stat-chip"><span class="stat-val">{n_rows}</span><span class="stat-lbl">Total</span></div>'
                for sname, scount in sheets_written:
                    stats_html += f'<div class="stat-chip"><span class="stat-val">{scount}</span><span class="stat-lbl">{sname}</span></div>'
                stats_html += '</div>'
                st.markdown(stats_html, unsafe_allow_html=True)

                ded_filename = f"{Path(uploaded_ded.name).stem}_limpio.xlsx"
                st.download_button(
                    label="↓  Descargar Excel Limpio",
                    data=output,
                    file_name=ded_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

              except Exception as e:
                st.error(f"Error al procesar el archivo: {str(e)}")
                st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            ESPERANDO ARCHIVO EXCEL · PASO 01
        </div>
        """, unsafe_allow_html=True)


elif herramienta == TOOL_ARBA:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Archivo Percepciones ARBA
    # ───────────────────────────────────────────────────────────────────────────────
    st.info("""
    **🔗 Diseños de Registro Oficiales ARBA (ARWeb)**
    * [Ver Formato para Percepciones (PDF)](https://www.arba.gov.ar/archivos/Publicaciones/Nuevo%20Disenio%20de%20Registro%20ARWeb%20Prod..pdf)
    * [Ver Formato para Retenciones (HTML)](https://www.arba.gov.ar/archivos/Tramites/Diseno-de-registros-de-lotes-de-importacion-A-122R.html)
    """)
    
    st.markdown('<div class="card"><div class="card-label">00 · Tipo de Presentación</div>', unsafe_allow_html=True)
    tipo_presentacion_arba = st.radio(
        "Seleccioná el régimen a declarar:",
        ["Percepciones", "Retenciones"],
        index=0,
        horizontal=True,
        key="arba_tipo"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="card"><div class="card-label">01 · Archivo fuente (Mendez o Excel)</div>', unsafe_allow_html=True)
    
    with st.expander("ℹ️ ¿Cómo debe estar estructurado el Excel?"):
        if "Retenciones" in tipo_presentacion_arba:
            st.markdown("""
            Para la importación de **Retenciones** (diseño A-122R), el archivo debe contener estas columnas (el orden no importa):
            - **Transacción Agente**: (o "Transaccion", "ID") Identificador interno (máx. 20 dígitos).
            - **CUIT**: CUIT del contribuyente retenido (con o sin guiones).
            - **Fecha**: Fecha de la retención (ej: `DD/MM/YYYY` o `YYYY-MM-DD`).
            - **Sucursal**: (o "PV", "Suc") Punto de venta.
            - **Base Imponible**: (o "Neto", "Base") Monto imponible sobre el que se retiene.
            - **Alicuota**: (o "Tasa") Porcentaje de retención (ej: `3.00`).
            - **Importe Retención**: Monto efectivamente retenido.
            """)
        else:
            st.markdown("""
            Si vas a cargar un Excel proporcionado por tu cliente para convertirlo al TXT de ARBA, el sistema buscará las siguientes columnas (el orden no importa):
            - **CUIT**: CUIT del contribuyente percibido/retenido (con o sin guiones).
            - **Fecha**: Fecha de la operación (ej: `DD/MM/YYYY` o `YYYY-MM-DD`).
            - **Tipo Comprobante**: (o "Tipo") Ej: `F`, `FC`, `NC`, `C`.
            - **Letra**: Letra del comprobante (ej: `A`, `B`).
            - **Sucursal**: (o "PV") Punto de venta.
            - **Nro Comprobante**: (o "Numero") Número del comprobante.
            - **Base Imponible**: (o "Neto") Monto base sobre el que se aplicó la alícuota.
            - **Alicuota**: (o "Tasa") Porcentaje aplicado (ej: `3.00`).
            - **Importe**: Monto percibido o retenido.
            """)

    texto_uploader = "Subí tu Excel de Retenciones (.xls, .xlsx)" if "Retenciones" in tipo_presentacion_arba else "Subí tu archivo de ventas (.txt) o tu Excel de deducciones (.xls, .xlsx)"
    tipos_uploader = ["xls", "xlsx"] if "Retenciones" in tipo_presentacion_arba else ["txt", "prn", "xls", "xlsx"]
    
    uploaded_arba = st.file_uploader(
        texto_uploader,
        type=tipos_uploader,
        label_visibility="visible",
        key="arba_file"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_arba:
        arba_filename = Path(uploaded_arba.name).stem
        ext = Path(uploaded_arba.name).suffix.lower()
        st.success(f"**{uploaded_arba.name}** listo para procesar")

        # ─── Card 02: Periodo y CUIT ────────────────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">02 · Datos de la Declaración</div>', unsafe_allow_html=True)
        col_p, col_q, col_c = st.columns(3)
        with col_p:
            periodo_arba = st.text_input(
                "Ingresá el periodo (MM/AAAA)",
                value="",
                placeholder="Ej: 03/2026",
                key="arba_periodo"
            )
        with col_q:
            quincena_arba = st.selectbox(
                "Quincena", 
                ["0 (Mensual)", "1 (Primera)", "2 (Segunda)"], 
                key="arba_q"
            )
        with col_c:
            cuit_agente_arba = st.text_input(
                "Ingresá el CUIT del Agente",
                value="",
                placeholder="Ej: 30-12345678-9",
                key="arba_cuit"
            )
            
        if "Retenciones" in tipo_presentacion_arba:
            col_a, col_l = st.columns(2)
            with col_a:
                actividad_arba = st.text_input("Cod. Actividad", placeholder="Ej: 68", key="arba_act")
            with col_l:
                lote_arba = st.text_input("Nro Lote", value="00001", key="arba_lote")

        st.markdown('</div>', unsafe_allow_html=True)

        # ─── Card 03: Procesar y Exportar ZIP ──────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">03 · Generar Archivo ARBA ZIP</div>', unsafe_allow_html=True)

        if st.button("⬡  Generar y Hashear ZIP"):
            periodo_limpio = periodo_arba.strip()
            cuit_limpio = cuit_agente_arba.strip().replace('-', '')

            if not re.match(r'^(\d{2})/(\d{4})$', periodo_limpio):
                st.error("El periodo es obligatorio y debe tener el formato **MM/AAAA** (ej: 03/2026)")
            elif not cuit_limpio or len(cuit_limpio) != 11:
                st.error("El CUIT del agente es obligatorio y debe ser válido (11 dígitos, ej: 30123456789)")
            else:
                mes_p, anio_p = periodo_limpio.split('/')
                
                try:
                    with st.spinner(f"Procesando {tipo_presentacion_arba}..."):
                        txt_arba = ""
                        n_lineas = 0
                        es_retencion = "Retenciones" in tipo_presentacion_arba

                        if es_retencion and (not actividad_arba.strip() or not lote_arba.strip()):
                            st.error("Para Retenciones, el Código de Actividad y el Número de Lote no pueden estar vacíos.")
                            st.stop()

                        if ext in ['.xls', '.xlsx']:
                            df_in = pd.read_excel(io.BytesIO(uploaded_arba.getvalue()))
                            if es_retencion:
                                txt_arba = generar_retenciones_arba_desde_excel(df_in)
                            else:
                                txt_arba = generar_arba_desde_excel(df_in)
                            n_lineas = len(df_in)
                        else:
                            if es_retencion:
                                st.error("Las Retenciones de ARBA solo pueden generarse a partir de un archivo Excel con las columnas requeridas.")
                                st.stop()
                            
                            raw_bytes = uploaded_arba.getvalue()
                            content_str = raw_bytes.decode('latin-1', errors='replace')
                            movimientos, metadata = parsear_archivo(content=content_str)
                            metadata['periodo'] = f"Desde el 01/{mes_p}/{anio_p} hasta el 28/{mes_p}/{anio_p}"
                            
                            txt_arba, _ = generar_percepciones_arba(movimientos, metadata)
                            n_lineas = len(txt_arba.splitlines())

                    if not txt_arba.strip():
                        st.warning(f"No se encontraron operaciones válidas de {tipo_presentacion_arba} en el archivo.")
                    else:
                        st.success("✓  Declaración procesada correctamente")

                        st.markdown(f"""
                        <div class="stats-row">
                            <div class="stat-chip">
                                <span class="stat-val">{n_lineas}</span>
                                <span class="stat-lbl">Registros Procesados</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                        # ─── Crear archivo ZIP en memoria ───
                        q_val = quincena_arba.split(" ")[0]
                        periodo_file = f"{anio_p}{mes_p}{q_val}"  # AAAAMMQ

                        if es_retencion:
                            lote_fmt = str(lote_arba).zfill(5)[:5]
                            nombre_txt = f"ER-{cuit_limpio}-{periodo_file}-{actividad_arba.strip()}-LOTE{lote_fmt}.txt"
                        else:
                            tipo_letra_perc = "D7" if q_val == "0" else "P7"
                            nombre_txt = f"AR-{cuit_limpio}-{periodo_file}-{tipo_letra_perc}-LOTE1.txt"

                        txt_bytes = txt_arba.encode("latin-1", errors="replace")
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                            zipf.writestr(nombre_txt, txt_bytes)
                        zip_bytes = zip_buffer.getvalue()

                        # ─── Nombres de salida (con/sin Hash) ───
                        if es_retencion:
                            # Según indicaciones, retenciones va sin hash MD5
                            nombre_zip = f"ER-{cuit_limpio}-{periodo_file}-{actividad_arba.strip()}-LOTE{lote_fmt}.ZIP"
                        else:
                            hash_md5 = hashlib.md5(zip_bytes).hexdigest().upper()
                            nombre_zip = f"AR-{cuit_limpio}-{periodo_file}-{tipo_letra_perc}-LOTE1_{hash_md5}.ZIP"

                        st.download_button(
                            label=f"↓  Descargar {nombre_zip}",
                            data=zip_bytes,
                            file_name=nombre_zip,
                            mime="application/zip",
                            use_container_width=True,
                        )

                except Exception as e:
                    st.error(f"Error al procesar el archivo: {str(e)}")
                    st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            ESPERANDO ARCHIVO · PASO 01
        </div>
        """, unsafe_allow_html=True)


elif herramienta == TOOL_CRUCE_CONCEPTO:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Cruce Concepto (TXT + Excel Sistema)
    # ───────────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Archivo TXT (Comprobantes)</div>', unsafe_allow_html=True)
    uploaded_txt_concepto = st.file_uploader(
        "Subí el .txt de Comprobantes de Compras (del sistema)",
        type=["txt", "prn"],
        label_visibility="visible",
        key="cruce_concepto_txt"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="card"><div class="card-label">02 · Archivo Excel Sistema (.xls)</div>', unsafe_allow_html=True)
    uploaded_xls_concepto = st.file_uploader(
        "Subí el Excel del sistema (.xls) con las compras",
        type=["xls", "xlsx"],
        label_visibility="visible",
        key="cruce_concepto_xls"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_txt_concepto and uploaded_xls_concepto:
        st.success(f"**{uploaded_txt_concepto.name}** + **{uploaded_xls_concepto.name}** listos para cruzar")

        st.markdown('<div class="card"><div class="card-label">03 · Formato de salida</div>', unsafe_allow_html=True)
        FMT_SISTEMA = "Formato Sistema (un impuesto por fila)"
        FMT_CONSOLIDADO = "Formato Consolidado (un comprobante por fila)"
        formato_cruce = st.radio(
            "Seleccioná el formato de salida:",
            options=[FMT_SISTEMA, FMT_CONSOLIDADO],
            index=0,
            help="**Sistema**: mantiene la estructura del Excel (varias filas por comprobante, una por tasa). "
                 "**Consolidado**: una fila por comprobante con columnas separadas para cada tasa de IVA.",
            key="cruce_formato"
        )
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card"><div class="card-label">04 · Procesar</div>', unsafe_allow_html=True)

        if st.button("⬡  Cruzar Concepto"):
            try:
                with st.spinner("Parseando TXT..."):
                    txt_content = uploaded_txt_concepto.getvalue().decode("latin-1")
                    transacciones, meta_txt = parsear_archivo(content=txt_content)

                if not transacciones:
                    st.error("No se encontraron transacciones en el TXT. Verificá el formato.")
                else:
                    with st.spinner("Leyendo Excel Sistema..."):
                        df_xls = pd.read_excel(io.BytesIO(uploaded_xls_concepto.getvalue()))

                        # Limpiar: quitar filas completamente vacías y fila de TOTALES
                        df_xls = df_xls.dropna(how='all')
                        # Buscar TOTALES en Fecha y Cond (donde suele estar)
                        for col_check in ['Fecha', 'Cond']:
                            if col_check in df_xls.columns:
                                df_xls = df_xls[~df_xls[col_check].astype(str).str.upper().str.contains('TOTAL', na=False)]
                        # En Nombre, solo eliminar la fila de resumen (valor exacto tipo "TOTALES" o "TOTAL GENERAL")
                        # NO eliminar empresas cuyo nombre contenga "TOTAL" (ej: "ELECTRICIDAD TOTAL S.A.")
                        if 'Nombre' in df_xls.columns:
                            df_xls = df_xls[~df_xls['Nombre'].astype(str).str.strip().str.upper().str.match(r'^TOTALE?S?\s*$|^TOTAL\s+GENERAL', na=False)]
                        df_xls = df_xls.reset_index(drop=True)

                    # ── Construir lookup de Concepto desde TXT ──────────────
                    # Clave: tipo + pv + nro (sin letra) + cuit (sin guiones)
                    concepto_lookup = {}
                    for t in transacciones:
                        numero_raw = t['Numero']
                        pv_txt = numero_raw.split('-')[0] if '-' in numero_raw else numero_raw[:5]
                        resto_num = numero_raw.split('-')[1] if '-' in numero_raw else numero_raw[5:]
                        # Quitar letra final del nro
                        nro_txt = resto_num[:-1] if resto_num and resto_num[-1].isalpha() else resto_num
                        cuit_txt = t['CUIT'].replace('-', '')

                        # Normalizar: quitar ceros a la izquierda del PV y Nro
                        try:
                            pv_norm = str(int(pv_txt))
                        except ValueError:
                            pv_norm = pv_txt
                        try:
                            nro_norm = str(int(nro_txt))
                        except ValueError:
                            nro_norm = nro_txt

                        key = f"{t['Tipo']}|{pv_norm}|{nro_norm}|{cuit_txt}"
                        concepto_lookup[key] = (t['Concepto'], t['Letra'])

                    # ── Parsear columnas del Excel ──────────────────────────
                    # El Excel tiene: Fecha, TC, Numero, Nombre, Cond, C.U.I.T., ...
                    # Numero tiene formato "PPPPP-NNNNNNNNNN/L" o similar
                    # Tipos del Excel del sistema (FCE, NCE, NDE) son equivalentes a FC/NC/ND en el TXT
                    _TC_NORMALIZE = {'FCE': 'FC', 'NCE': 'NC', 'NDE': 'ND'}

                    def extraer_key_xls(row):
                        tc = _TC_NORMALIZE.get(str(row.get('TC', '')).strip(), str(row.get('TC', '')).strip())
                        numero = str(row.get('Numero', '')).strip()
                        cuit = str(row.get('C.U.I.T.', '')).replace('-', '').replace('.', '').strip()

                        # Separar PV y Nro del campo Numero (ej: 00003-00021793/A)
                        if '-' in numero:
                            parts = numero.split('-', 1)
                            pv_raw = parts[0]
                            nro_raw = parts[1]
                        elif '/' in numero:
                            parts = numero.split('/', 1)
                            pv_raw = parts[0]
                            nro_raw = parts[1]
                        else:
                            pv_raw = numero[:5] if len(numero) >= 5 else numero
                            nro_raw = numero[5:] if len(numero) > 5 else ''

                        # Quitar letra y / del nro
                        nro_clean = re.sub(r'[/A-Za-z]+$', '', nro_raw).strip()

                        try:
                            pv_norm = str(int(pv_raw))
                        except ValueError:
                            pv_norm = pv_raw
                        try:
                            nro_norm = str(int(nro_clean))
                        except ValueError:
                            nro_norm = nro_clean

                        return f"{tc}|{pv_norm}|{nro_norm}|{cuit}"

                    with st.spinner("Cruzando datos..."):
                        # Filas "cabecera" son las que tienen Fecha; las demás son sub-filas del mismo movimiento
                        mask_header = df_xls['Fecha'].notna() & (df_xls['Fecha'] != '')

                        # Agregar columnas de Concepto y Jurisdicción
                        conceptos_cod = []
                        jurisdicciones = []
                        fecha_from_xls = {}  # key → fecha completa del Excel
                        matched = 0
                        last_concepto = ''
                        last_jur = ''
                        not_found_rows: list[dict] = []
                        for idx, row in df_xls.iterrows():
                            if mask_header.iloc[idx]:
                                # Es fila cabecera → buscar concepto y jurisdicción
                                key = extraer_key_xls(row)
                                result = concepto_lookup.get(key)
                                if result is not None:
                                    matched += 1
                                    last_concepto, last_jur = result
                                    # Guardar fecha completa del Excel para el formato consolidado
                                    fecha_from_xls[key] = str(row.get('Fecha', '')).strip()
                                else:
                                    last_concepto = ''
                                    last_jur = ''
                                    # Capturar los valores originales antes de las transformaciones posteriores
                                    not_found_rows.append({
                                        'Fecha': str(row.get('Fecha', '')).strip(),
                                        'Tipo': str(row.get('TC', '')).strip(),
                                        'Numero': str(row.get('Numero', '')).strip(),
                                        'Nombre': str(row.get('Nombre', '')).strip(),
                                        'CUIT': str(row.get('C.U.I.T.', '')).strip(),
                                    })
                                conceptos_cod.append(last_concepto)
                                jurisdicciones.append(last_jur)
                            else:
                                # Sub-fila → propagar de la cabecera
                                conceptos_cod.append(last_concepto)
                                jurisdicciones.append(last_jur)

                        # Insertar Concepto y Jurisdicción después de C.U.I.T.
                        cuit_pos = df_xls.columns.get_loc('C.U.I.T.') + 1 if 'C.U.I.T.' in df_xls.columns else len(df_xls.columns)
                        df_xls.insert(cuit_pos, 'Concepto', conceptos_cod)
                        df_xls.insert(cuit_pos + 1, 'Jur.', jurisdicciones)

                        # Forward-fill columnas identificatorias a sub-filas
                        for col_ff in ['Fecha', 'TC', 'Numero', 'Nombre', 'Cond', 'C.U.I.T.', 'Concepto', 'Jur.']:
                            if col_ff in df_xls.columns:
                                df_xls[col_ff] = df_xls[col_ff].ffill()

                        # Separar Fecha (dd/mm/yyyy) en Dia, Mes, Año
                        if 'Fecha' in df_xls.columns:
                            fecha_pos = df_xls.columns.get_loc('Fecha')
                            fecha_str = df_xls['Fecha'].astype(str)
                            # Intentar parsear como dd/mm/yyyy
                            partes_fecha = fecha_str.str.split('/', expand=True)
                            if partes_fecha.shape[1] >= 3:
                                df_xls.insert(fecha_pos, 'Dia', pd.to_numeric(partes_fecha[0], errors='coerce').fillna(0).astype(int))
                                df_xls.insert(fecha_pos + 1, 'Mes', pd.to_numeric(partes_fecha[1], errors='coerce').fillna(0).astype(int))
                                df_xls.insert(fecha_pos + 2, 'Año', pd.to_numeric(partes_fecha[2], errors='coerce').fillna(0).astype(int))
                                df_xls.drop(columns=['Fecha'], inplace=True)

                        if 'Numero' in df_xls.columns:
                            num_pos = df_xls.columns.get_loc('Numero')
                            def split_numero(val):
                                s = str(val).strip()
                                # Quitar /Letra o Letra final del numero
                                letra = ''
                                if '/' in s:
                                    parts = s.rsplit('/', 1)
                                    s = parts[0]
                                    letra = parts[1] if len(parts) > 1 else ''
                                elif s and s[-1].isalpha():
                                    letra = s[-1]
                                    s = s[:-1]
                                return s, letra

                            numero_list, letra_list = [], []
                            for val in df_xls['Numero']:
                                numero, letra = split_numero(val)
                                numero_list.append(numero)
                                letra_list.append(letra)

                            df_xls['Numero'] = numero_list
                            df_xls.insert(num_pos + 1, 'Letra', letra_list)

                        # Formatear CUIT con guiones (XX-XXXXXXXX-X)
                        if 'C.U.I.T.' in df_xls.columns:
                            def fmt_cuit(val):
                                s = str(val).replace('-', '').replace('.', '').replace(' ', '').strip()
                                # Quitar .0 si viene de float
                                if s.endswith('.0'):
                                    s = s[:-2]
                                if len(s) == 11 and s.isdigit():
                                    return f"{s[:2]}-{s[2:10]}-{s[10]}"
                                return s
                            df_xls['C.U.I.T.'] = df_xls['C.U.I.T.'].apply(fmt_cuit)

                        # Rellenar y convertir columnas monetarias a numérico
                        for col_fill in ['Neto', 'Iva', 'Sobretasa', 'Retenciones']:
                            if col_fill in df_xls.columns:
                                df_xls[col_fill] = pd.to_numeric(df_xls[col_fill], errors='coerce').fillna(0)

                    total_valid = mask_header.sum()
                    not_found = total_valid - matched

                    st.success("✓  Cruce completado")

                    # Stats
                    st.markdown(f"""
                    <div class="stats-row">
                        <div class="stat-chip">
                            <span class="stat-val">{total_valid}</span>
                            <span class="stat-lbl">Comprobantes</span>
                        </div>
                        <div class="stat-chip">
                            <span class="stat-val">{matched}</span>
                            <span class="stat-lbl">Matcheados</span>
                        </div>
                        <div class="stat-chip">
                            <span class="stat-val">{not_found}</span>
                            <span class="stat-lbl">No encontrados</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    if not_found > 0:
                        st.warning(f"**{not_found}** comprobantes del Excel no fueron encontrados en el TXT")
                        with st.expander(f"Ver detalle de los {not_found} comprobantes no encontrados"):
                            df_nf = pd.DataFrame(not_found_rows)
                            st.dataframe(df_nf, use_container_width=True, hide_index=True)

                    if formato_cruce == FMT_CONSOLIDADO:
                        # ── Formato Consolidado: usar crear_excel del TXT ──
                        # Una fila por comprobante, con columnas por cada tasa
                        # Enriquecer transacciones con la fecha completa del Excel
                        for t in transacciones:
                            numero_raw = t['Numero']
                            pv_txt = numero_raw.split('-')[0] if '-' in numero_raw else numero_raw[:5]
                            resto_num = numero_raw.split('-')[1] if '-' in numero_raw else numero_raw[5:]
                            nro_txt = resto_num[:-1] if resto_num and resto_num[-1].isalpha() else resto_num
                            cuit_txt = t['CUIT'].replace('-', '')
                            try:
                                pv_n = str(int(pv_txt))
                            except ValueError:
                                pv_n = pv_txt
                            try:
                                nro_n = str(int(nro_txt))
                            except ValueError:
                                nro_n = nro_txt
                            tkey = f"{t['Tipo']}|{pv_n}|{nro_n}|{cuit_txt}"
                            if tkey in fecha_from_xls:
                                t['Fecha'] = fecha_from_xls[tkey]

                        with st.spinner("Generando Excel consolidado..."):
                            output = io.BytesIO()
                            crear_excel_consolidado_simple(transacciones, meta_txt, output)
                            output.seek(0)

                        st.download_button(
                            label="↓  Descargar Consolidado",
                            data=output,
                            file_name="Movimientos_Consolidado.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

                    else:
                        # ── Formato Sistema: Excel enriquecido con Concepto/Jur ──
                        from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
                        from openpyxl.utils import get_column_letter

                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            # startrow=4 → encabezados columna en fila 5, datos desde fila 6
                            df_xls.to_excel(writer, sheet_name='Movimientos', index=False, startrow=4)
                            ws = writer.sheets['Movimientos']

                            total_cols = len(df_xls.columns)
                            last_col_letter = get_column_letter(total_cols)
                            center_align = Alignment(horizontal='center', vertical='center')

                            # ── Encabezado con datos del cliente ──────────────
                            ws.merge_cells(f'A1:{last_col_letter}1')
                            ws['A1'] = meta_txt.get('razon_social', 'CONTRIBUYENTE').upper()
                            ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
                            ws['A1'].fill = PatternFill('solid', fgColor='2F5496')
                            ws['A1'].alignment = center_align

                            ws.merge_cells(f'A2:{last_col_letter}2')
                            tipo_rep = meta_txt.get('tipo_reporte', 'COMPRAS')
                            ws['A2'] = tipo_rep.upper()
                            ws['A2'].font = Font(bold=True, size=12, color='C00000')
                            ws['A2'].alignment = center_align

                            ws.merge_cells(f'A3:{last_col_letter}3')
                            ws['A3'] = f"CUIT: {meta_txt.get('cuit_empresa', '')} | Periodo: {meta_txt.get('periodo', '')}"
                            ws['A3'].font = Font(bold=True, size=11, color='2F5496')
                            ws['A3'].alignment = center_align

                            # ── Estilo encabezados de columna (fila 5) ────────
                            header_font = Font(bold=True, size=10, color='FFFFFF')
                            header_fill = PatternFill('solid', fgColor='4472C4')
                            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            for col_idx in range(1, total_cols + 1):
                                cell = ws.cell(row=5, column=col_idx)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_align

                            # ── Formato numérico con 2 decimales, rojo si negativo ──
                            num_fmt_red = '$#,##0.00;[Red]-$#,##0.00'
                            col_list_xls = list(df_xls.columns)
                            money_cols_xls = ['Neto', 'Iva', 'Sobretasa', 'Retenciones', 'Total']
                            money_indices = [col_list_xls.index(c) + 1 for c in money_cols_xls if c in col_list_xls]

                            data_start_row = 6
                            for row in range(data_start_row, len(df_xls) + data_start_row):
                                for col_idx in money_indices:
                                    cell = ws.cell(row=row, column=col_idx)
                                    cell.number_format = num_fmt_red

                        output.seek(0)

                        st.download_button(
                            label="↓  Descargar Movimientos",
                            data=output,
                            file_name="Movimientos.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

            except Exception as e:
                st.error(f"Error al procesar: {str(e)}")
                st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            SUBÍ AMBOS ARCHIVOS · TXT + EXCEL SISTEMA
        </div>
        """, unsafe_allow_html=True)

elif herramienta == TOOL_CM05:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Papeles de Trabajo CM05
    # Cruce consolidado TXT + Excel Sistema + Resumen x Concepto + Resumen x Jur.
    # ───────────────────────────────────────────────────────────────────────────────

    # ─── Card 01: TXT ──────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Archivo TXT (Movimientos Mendez)</div>', unsafe_allow_html=True)
    uploaded_cm05_txt = st.file_uploader(
        "Arrastrá el TXT de movimientos o hacé click para seleccionarlo",
        type=["txt", "prn"],
        label_visibility="visible",
        key="cm05_txt"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    # ─── Card 02: Excel Sistema ─────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">02 · Archivo Excel Sistema (.xls / .xlsx)</div>', unsafe_allow_html=True)
    uploaded_cm05_xls = st.file_uploader(
        "Arrastrá el Excel del sistema con las compras o hacé click para seleccionarlo",
        type=["xls", "xlsx"],
        label_visibility="visible",
        key="cm05_xls"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_cm05_txt and uploaded_cm05_xls:
        st.success(
            f"**{uploaded_cm05_txt.name}** + **{uploaded_cm05_xls.name}** listos para procesar"
        )

        st.markdown('<div class="card"><div class="card-label">03 · Procesar</div>', unsafe_allow_html=True)

        if st.button("⬡  Generar Papeles de Trabajo CM05"):
            try:
                # ── 1. Parsear TXT ─────────────────────────────────────────────
                with st.spinner("Parseando TXT..."):
                    txt_content_cm05 = uploaded_cm05_txt.getvalue().decode("latin-1", errors="replace")
                    transacciones_cm05, meta_cm05 = parsear_archivo(content=txt_content_cm05)

                if not transacciones_cm05:
                    st.error("No se encontraron transacciones en el TXT. Verificá el formato del archivo.")
                else:
                    # ── 2. Leer Excel Sistema (primera hoja para el cruce) ──────
                    with st.spinner("Leyendo Excel Sistema..."):
                        df_xls_cm05 = pd.read_excel(
                            io.BytesIO(uploaded_cm05_xls.getvalue()),
                            sheet_name=0
                        )

                        # Limpiar: quitar filas vacías y totales
                        df_xls_cm05 = df_xls_cm05.dropna(how='all')
                        for col_check in ['Fecha', 'Cond']:
                            if col_check in df_xls_cm05.columns:
                                df_xls_cm05 = df_xls_cm05[
                                    ~df_xls_cm05[col_check].astype(str).str.upper().str.contains('TOTAL', na=False)
                                ]
                        if 'Nombre' in df_xls_cm05.columns:
                            df_xls_cm05 = df_xls_cm05[
                                ~df_xls_cm05['Nombre'].astype(str).str.strip().str.upper().str.match(
                                    r'^TOTALE?S?\s*$|^TOTAL\s+GENERAL', na=False
                                )
                            ]
                        df_xls_cm05 = df_xls_cm05.reset_index(drop=True)

                    # ── 3. Construir lookup de Concepto/Letra desde TXT ─────────
                    with st.spinner("Cruzando datos..."):
                        _TC_NORM_CM05 = {'FCE': 'FC', 'NCE': 'NC', 'NDE': 'ND'}

                        concepto_lookup_cm05 = {}
                        fecha_from_xls_cm05 = {}

                        for t in transacciones_cm05:
                            numero_raw = t['Numero']
                            pv_txt = numero_raw.split('-')[0] if '-' in numero_raw else numero_raw[:5]
                            resto_num = numero_raw.split('-')[1] if '-' in numero_raw else numero_raw[5:]
                            nro_txt = resto_num[:-1] if resto_num and resto_num[-1].isalpha() else resto_num
                            cuit_txt = t['CUIT'].replace('-', '')
                            try:
                                pv_norm = str(int(pv_txt))
                            except ValueError:
                                pv_norm = pv_txt
                            try:
                                nro_norm = str(int(nro_txt))
                            except ValueError:
                                nro_norm = nro_txt
                            key = f"{t['Tipo']}|{pv_norm}|{nro_norm}|{cuit_txt}"
                            concepto_lookup_cm05[key] = (t['Concepto'], t['Letra'])

                        def extraer_key_cm05(row):
                            tc = _TC_NORM_CM05.get(str(row.get('TC', '')).strip(), str(row.get('TC', '')).strip())
                            numero = str(row.get('Numero', '')).strip()
                            cuit = str(row.get('C.U.I.T.', '')).replace('-', '').replace('.', '').strip()
                            if '-' in numero:
                                parts = numero.split('-', 1)
                                pv_raw, nro_raw = parts[0], parts[1]
                            elif '/' in numero:
                                parts = numero.split('/', 1)
                                pv_raw, nro_raw = parts[0], parts[1]
                            else:
                                pv_raw = numero[:5] if len(numero) >= 5 else numero
                                nro_raw = numero[5:] if len(numero) > 5 else ''
                            nro_clean = re.sub(r'[/A-Za-z]+$', '', nro_raw).strip()
                            try:
                                pv_norm = str(int(pv_raw))
                            except ValueError:
                                pv_norm = pv_raw
                            try:
                                nro_norm = str(int(nro_clean))
                            except ValueError:
                                nro_norm = nro_clean
                            return f"{tc}|{pv_norm}|{nro_norm}|{cuit}"

                        # Filas cabecera (tienen Fecha)
                        mask_header_cm05 = df_xls_cm05['Fecha'].notna() & (df_xls_cm05['Fecha'] != '') \
                            if 'Fecha' in df_xls_cm05.columns else pd.Series([True] * len(df_xls_cm05))

                        matched_cm05 = 0
                        last_concepto_cm05 = ''
                        last_jur_cm05 = ''
                        conceptos_cod_cm05 = []
                        jurisdicciones_cm05 = []
                        not_found_rows_cm05: list[dict] = []

                        for idx, row in df_xls_cm05.iterrows():
                            if mask_header_cm05.iloc[idx]:
                                key = extraer_key_cm05(row)
                                result = concepto_lookup_cm05.get(key)
                                if result is not None:
                                    matched_cm05 += 1
                                    last_concepto_cm05, last_jur_cm05 = result
                                    fecha_from_xls_cm05[key] = str(row.get('Fecha', '')).strip()
                                else:
                                    last_concepto_cm05 = ''
                                    last_jur_cm05 = ''
                                    not_found_rows_cm05.append({
                                        'Fecha': str(row.get('Fecha', '')).strip(),
                                        'Tipo': str(row.get('TC', '')).strip(),
                                        'Numero': str(row.get('Numero', '')).strip(),
                                        'Nombre': str(row.get('Nombre', '')).strip(),
                                        'CUIT': str(row.get('C.U.I.T.', '')).strip(),
                                    })
                                conceptos_cod_cm05.append(last_concepto_cm05)
                                jurisdicciones_cm05.append(last_jur_cm05)
                            else:
                                conceptos_cod_cm05.append(last_concepto_cm05)
                                jurisdicciones_cm05.append(last_jur_cm05)

                        # Enriquecer transacciones con fecha del Excel
                        for t in transacciones_cm05:
                            numero_raw = t['Numero']
                            pv_txt = numero_raw.split('-')[0] if '-' in numero_raw else numero_raw[:5]
                            resto_num = numero_raw.split('-')[1] if '-' in numero_raw else numero_raw[5:]
                            nro_txt = resto_num[:-1] if resto_num and resto_num[-1].isalpha() else resto_num
                            cuit_txt = t['CUIT'].replace('-', '')
                            try:
                                pv_n = str(int(pv_txt))
                            except ValueError:
                                pv_n = pv_txt
                            try:
                                nro_n = str(int(nro_txt))
                            except ValueError:
                                nro_n = nro_txt
                            tkey = f"{t['Tipo']}|{pv_n}|{nro_n}|{cuit_txt}"
                            if tkey in fecha_from_xls_cm05:
                                t['Fecha'] = fecha_from_xls_cm05[tkey]

                    # ── 4. Generar Excel consolidado con resumenes via crear_excel ─
                    with st.spinner("Generando Excel de Papeles de Trabajo CM05..."):
                        output_cm05_raw = io.BytesIO()
                        crear_excel(
                            transacciones_cm05,
                            meta_cm05,
                            output_cm05_raw,
                            con_resumenes=True,
                            con_auxiliar=False,
                            cruce_arca=False,
                            df_arca=None,
                            con_asiento=False,
                        )
                        output_cm05_raw.seek(0)

                        # ── Filtrar: mantener solo las 3 hojas necesarias ────────
                        from openpyxl import load_workbook
                        HOJAS_CM05 = {"Movimientos", "Resumen x Concepto", "Resumen x Concepto y Jur."}
                        wb_cm05 = load_workbook(output_cm05_raw)
                        for titulo in [ws.title for ws in wb_cm05.worksheets]:
                            if titulo not in HOJAS_CM05:
                                del wb_cm05[titulo]

                        # ── Post-procesar "Resumen x Concepto" ────────────────────
                        # Encabezados en fila 6 (startrow=5 en pandas → fila 6 en Excel)
                        ws_rc_cm05 = wb_cm05["Resumen x Concepto"]
                        HDR_ROW_CM05 = 6
                        DATA_START_CM05 = HDR_ROW_CM05 + 1
                        last_row_cm05 = ws_rc_cm05.max_row
                        total_row_cm05 = last_row_cm05  # la última fila es TOTAL GENERAL

                        # Columnas fijas que NO se reclasifican
                        # (no incluimos 'cantidad' aquí para que pase al elif)
                        SKIP_NAMES_CM05 = {'concepto', 'descripcion', 'deducciones', 'total'}

                        # Mapear número de columna → nombre de header
                        col_hdr_cm05 = {}
                        for cell in ws_rc_cm05[HDR_ROW_CM05]:
                            if cell.value is not None:
                                col_hdr_cm05[cell.column] = str(cell.value)

                        # ── Clasificar columnas ───────────────────────────────────
                        iva_cols_cm05  = []   # IVA*
                        cant_cols_cm05 = []   # Cantidad → eliminar
                        neto_cols_cm05 = []   # Neto* / Exento / Monotributo → Neto
                        otro_cols_cm05 = []   # Otros impuestos → también van a Neto

                        for col_n, hdr in col_hdr_cm05.items():
                            hl = hdr.lower()
                            hu = hdr.upper()
                            if hl in SKIP_NAMES_CM05:
                                continue
                            elif hl == 'cantidad':
                                cant_cols_cm05.append(col_n)
                            elif hl.startswith('iva'):
                                iva_cols_cm05.append(col_n)
                            elif any(k in hu for k in ('PERC', 'RET.', 'SIRCREB')):
                                pass  # deducciones: no se tocan
                            elif any(k in hl for k in ('neto', 'exento', 'monotributo')):
                                neto_cols_cm05.append(col_n)
                            else:
                                # Otros impuestos (IMP.CIG, IMP.SELLO, etc.) → suman al Neto
                                otro_cols_cm05.append(col_n)

                        # Todos los que van al Neto
                        all_neto_src = neto_cols_cm05 + otro_cols_cm05

                        def _combinar_formula(ws, row_i, cols):
                            """
                            Construye '=part1+part2+...' combinando
                            fórmulas (string) y valores numéricos de varias celdas.
                            """
                            parts = []
                            for c in cols:
                                v = ws.cell(row=row_i, column=c).value
                                if v is None:
                                    continue
                                elif isinstance(v, str) and v.startswith('='):
                                    parts.append(v[1:])   # quitar el '='
                                elif isinstance(v, (int, float)):
                                    parts.append(str(v))
                            return ('=' + '+'.join(parts)) if parts else 0

                        from openpyxl.utils import get_column_letter as _gcl

                        # ── Neto: combinar fórmulas en primera col Neto ───────────
                        neto_dest_col = all_neto_src[0] if all_neto_src else None
                        if neto_dest_col:
                            ws_rc_cm05.cell(row=HDR_ROW_CM05, column=neto_dest_col).value = "Neto"
                            fmt_neto = ws_rc_cm05.cell(row=DATA_START_CM05, column=neto_dest_col).number_format
                            for row_i in range(DATA_START_CM05, total_row_cm05):  # excluye TOTAL
                                cell = ws_rc_cm05.cell(row=row_i, column=neto_dest_col)
                                cell.value = _combinar_formula(ws_rc_cm05, row_i, all_neto_src)
                                if fmt_neto:
                                    cell.number_format = fmt_neto

                        # ── IVA: combinar fórmulas en primera col IVA ─────────────
                        iva_dest_col = iva_cols_cm05[0] if iva_cols_cm05 else None
                        if iva_dest_col:
                            ws_rc_cm05.cell(row=HDR_ROW_CM05, column=iva_dest_col).value = "IVA"
                            fmt_iva = ws_rc_cm05.cell(row=DATA_START_CM05, column=iva_dest_col).number_format
                            for row_i in range(DATA_START_CM05, total_row_cm05):
                                cell = ws_rc_cm05.cell(row=row_i, column=iva_dest_col)
                                cell.value = _combinar_formula(ws_rc_cm05, row_i, iva_cols_cm05)
                                if fmt_iva:
                                    cell.number_format = fmt_iva

                        # ── Eliminar columnas sobrantes (de mayor a menor) ────────
                        cols_to_del_cm05 = sorted(
                            all_neto_src[1:]    # neto + otros sobrantes
                            + iva_cols_cm05[1:] # iva sobrantes
                            + cant_cols_cm05,   # Cantidad
                            reverse=True
                        )
                        for col_del in cols_to_del_cm05:
                            ws_rc_cm05.delete_cols(col_del)
                        # ── Corregir encabezados combinados tras el borrado ───────
                        # openpyxl no ajusta los merged ranges al borrar columnas;
                        # hay que deshacer todos y volver a combinar con el nuevo ancho.
                        new_max_col = ws_rc_cm05.max_column
                        new_last_col_l = _gcl(new_max_col)

                        # Disolver todos los merges actuales (iteramos sobre copia)
                        for mr in list(ws_rc_cm05.merged_cells.ranges):
                            try:
                                ws_rc_cm05.unmerge_cells(str(mr))
                            except KeyError:
                                # La celda ya fue borrada por delete_cols;
                                # el range ya fue removido del tracking → ignorar
                                pass

                        # Re-combinar filas de título (1–4) con el nuevo ancho
                        for title_row in range(1, HDR_ROW_CM05):  # filas 1,2,3,4,5
                            ws_rc_cm05.merge_cells(f"A{title_row}:{new_last_col_l}{title_row}")

                        # ── Reescribir fórmulas TOTAL GENERAL tras el borrado ─────
                        # (openpyxl NO actualiza referencias al borrar columnas)
                        new_col_hdr = {}
                        for cell in ws_rc_cm05[HDR_ROW_CM05]:
                            if cell.value is not None:
                                new_col_hdr[str(cell.value).lower()] = cell.column

                        money_fmt_rc = "#,##0.00"
                        n_new_cols = ws_rc_cm05.max_column

                        # Reescribir fila TOTAL (SUM de cada columna numérica)
                        for col_i in range(1, n_new_cols + 1):
                            hdr_v = ws_rc_cm05.cell(row=HDR_ROW_CM05, column=col_i).value
                            if hdr_v is None:
                                continue
                            hl2 = str(hdr_v).lower()
                            if hl2 in ('concepto', 'descripcion'):
                                continue  # texto, no se suma
                            col_l = _gcl(col_i)
                            tc = ws_rc_cm05.cell(row=total_row_cm05, column=col_i)
                            tc.value = f"=SUM({col_l}{DATA_START_CM05}:{col_l}{total_row_cm05 - 1})"
                            tc.number_format = money_fmt_rc

                        # Reescribir "Total" de cada fila de datos para que sume Neto+IVA+Deducciones
                        total_col_new = new_col_hdr.get('total')
                        neto_col_new  = new_col_hdr.get('neto')
                        iva_col_new   = new_col_hdr.get('iva')
                        ded_col_new   = new_col_hdr.get('deducciones')

                        if total_col_new and neto_col_new and ded_col_new:
                            total_col_l = _gcl(total_col_new)
                            neto_col_l  = _gcl(neto_col_new)
                            iva_col_l   = _gcl(iva_col_new) if iva_col_new else None
                            ded_col_l   = _gcl(ded_col_new)
                            for row_i in range(DATA_START_CM05, total_row_cm05):
                                if iva_col_l:
                                    formula = f"={neto_col_l}{row_i}+{iva_col_l}{row_i}+{ded_col_l}{row_i}"
                                else:
                                    formula = f"={neto_col_l}{row_i}+{ded_col_l}{row_i}"
                                cell = ws_rc_cm05.cell(row=row_i, column=total_col_new)
                                cell.value = formula
                                cell.number_format = money_fmt_rc

                        output_cm05 = io.BytesIO()
                        wb_cm05.save(output_cm05)
                        output_cm05.seek(0)




                    st.success("✓  Papeles de Trabajo CM05 generados con éxito")

                    from collections import Counter
                    tipos_cm05 = Counter(t.get("Tipo", "") for t in transacciones_cm05)
                    total_xls_rows = mask_header_cm05.sum()
                    not_found_cm05 = total_xls_rows - matched_cm05

                    st.markdown(f"""
                    <div class="stats-row">
                        <div class="stat-chip">
                            <span class="stat-val">{len(transacciones_cm05)}</span>
                            <span class="stat-lbl">Total TXT</span>
                        </div>
                        <div class="stat-chip">
                            <span class="stat-val">{matched_cm05}</span>
                            <span class="stat-lbl">Matcheados</span>
                        </div>
                        <div class="stat-chip">
                            <span class="stat-val">{not_found_cm05}</span>
                            <span class="stat-lbl">No encontrados</span>
                        </div>
                        <div class="stat-chip">
                            <span class="stat-val">{tipos_cm05.get('FC', 0)}</span>
                            <span class="stat-lbl">Facturas</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    if not_found_cm05 > 0:
                        st.warning(f"**{not_found_cm05}** comprobantes del Excel no fueron encontrados en el TXT")
                        with st.expander(f"Ver detalle de los {not_found_cm05} comprobantes no encontrados"):
                            df_nf_cm05 = pd.DataFrame(not_found_rows_cm05)
                            st.dataframe(df_nf_cm05, use_container_width=True, hide_index=True)

                    st.info(
                        f"**{meta_cm05.get('tipo_reporte', 'N/A')}** · "
                        f"{meta_cm05.get('razon_social', 'Contribuyente')} · "
                        f"{meta_cm05.get('periodo', '')}"
                    )

                    cm05_filename = f"{Path(uploaded_cm05_txt.name).stem}_CM05.xlsx"
                    st.download_button(
                        label="↓  Descargar Papeles de Trabajo CM05",
                        data=output_cm05,
                        file_name=cm05_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )


            except Exception as e:
                st.error(f"Error al procesar: {str(e)}")
                st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)

    elif uploaded_cm05_txt and not uploaded_cm05_xls:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            FALTA EL EXCEL SISTEMA · SUBILO EN EL PASO 02
        </div>
        """, unsafe_allow_html=True)

    elif not uploaded_cm05_txt and uploaded_cm05_xls:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            FALTA EL TXT · SUBILO EN EL PASO 01
        </div>
        """, unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            SUBÍ EL TXT Y EL EXCEL SISTEMA · PASOS 01 Y 02
        </div>
        """, unsafe_allow_html=True)

elif herramienta == TOOL_CRUCE_DEDUCCIONES:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Cruce de Deducciones
    # Cruza el TXT de Mendez con reportes/padrones de ARBA, AGIP e IVA
    # ───────────────────────────────────────────────────────────────────────────────
    
    st.markdown('<div class="card"><div class="card-label">00 · Configuración del Cruce</div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        organismo = st.selectbox("Organismo:", ["ARBA", "AGIP", "IVA", "Ganancias"], index=0)
    with col2:
        tipo_cruce_sel = st.selectbox("Tipo de Deducción:", ["Percepciones", "Retenciones"], index=0)
    with col3:
        periodo_estrategia = st.selectbox("Frecuencia:", ["Mensual", "Varios Periodos"], index=0)
    st.markdown('</div>', unsafe_allow_html=True)

    if organismo == "IVA":
        pass  # Funcionalidad IVA activa

    if periodo_estrategia == "Varios Periodos":
        st.info("🚧 El cruce para **Varios Períodos** simultáneos se encuentra en desarrollo. Temporalmente usá el modo Mensual.")
        st.stop()


    # ── Función de parsing de AGIP ──────────────────────────────────────────────────
    def parsear_agip_iibb(agip_content: bytes) -> dict:
        import pandas as pd
        import io
        
        text = agip_content.decode("latin-1", errors="replace")
        skip = 1 if text.startswith("sep=") else 0
        df = pd.read_csv(io.StringIO(text), sep=',', skiprows=skip, index_col=False)
        
        percepciones = []
        retenciones = []
        
        cols_map = {}
        for c in df.columns:
            c_low = str(c).lower().strip()
            if 'cuit' in c_low: cols_map['cuit'] = c
            elif 'fecha comprobante' in c_low or 'fecha retencion' in c_low or 'fecha percepcion' in c_low:
                if 'fecha' not in cols_map: cols_map['fecha'] = c
            elif 'n' in c_low and 'comprobante' in c_low and 'tipo' not in c_low: cols_map['nro'] = c
            elif 'tipo' in c_low and 'comprobante' in c_low: cols_map['tipo_comp'] = c
            elif 'monto percibido' in c_low: cols_map['monto_p'] = c
            elif 'monto retenido' in c_low: cols_map['monto_r'] = c
            elif 'razon social' in c_low: cols_map['razon_social'] = c
            
        if 'fecha' not in cols_map:
            for c in df.columns:
                if 'fecha' in str(c).lower().strip(): cols_map['fecha'] = c; break
                
        for idx, row in df.iterrows():
            cuit_raw = str(row.get(cols_map.get('cuit', 'CUIT'), ''))
            cuit_limpio = cuit_raw.replace('-', '').replace('.0', '').strip()
            if not cuit_limpio or cuit_limpio == 'nan':
                continue
                
            # Formatear el CUIT local para que tenga los guiones estándar: 30-70828643-1
            if len(cuit_limpio) == 11 and '-' not in cuit_raw:
                cuit_raw = f"{cuit_limpio[:2]}-{cuit_limpio[2:10]}-{cuit_limpio[10]}"
            
            monto_p_val = row.get(cols_map.get('monto_p', 'Monto Percibido'), 0)
            monto_r_val = row.get(cols_map.get('monto_r', 'Monto Retenido'), 0)
            
            # Limpiar por las dudas si entra como object string
            if isinstance(monto_p_val, str):
                monto_p_val = monto_p_val.replace('.', '').replace(',', '.') if ',' in monto_p_val else monto_p_val
            if isinstance(monto_r_val, str):
                monto_r_val = monto_r_val.replace('.', '').replace(',', '.') if ',' in monto_r_val else monto_r_val
                
            monto_p = float(monto_p_val or 0)
            monto_r = float(monto_r_val or 0)
            rz = str(row.get(cols_map.get('razon_social', 'Razon Social'), ''))
            
            if monto_p > 0:
                seccion = "P"
                monto_f = monto_p
            elif monto_r > 0:
                seccion = "R"
                monto_f = monto_r
            else:
                seccion = "R" if 'monto_r' in cols_map else "P"
                monto_f = 0.0
                
            fecha_s = str(row.get(cols_map.get('fecha', 'Fecha Comprobante'), ''))
            nro_s = str(row.get(cols_map.get('nro', 'N Comprobante'), ''))
            tipo_c = str(row.get(cols_map.get('tipo_comp', 'Tipo Comprobante'), ''))
            
            pv_norm = "0"
            nro_norm = nro_s.strip()
            try:
                nro_norm = str(int(float(nro_norm.replace(',', ''))))
            except ValueError:
                pass
                
            registro = {
                'tipo_reg':    seccion,
                'cod_jur':     '00000',
                'cuit':        cuit_raw,
                'cuit_limpio': cuit_limpio,
                'fecha':       fecha_s,
                'pv':          pv_norm,
                'pv_norm':     pv_norm,
                'nro':         nro_norm,
                'nro_norm':    nro_norm,
                'tipo_comp':   tipo_c.strip(),
                'monto':       monto_f,
                'razon_social': rz
            }
            
            if seccion == "P":
                percepciones.append(registro)
            else:
                retenciones.append(registro)
                
        return {'percepciones': percepciones, 'retenciones': retenciones}

    # ── Función de parsing de IVA ──────────────────────────────────────────────────
    def parsear_iva_xls(iva_content: bytes) -> dict:
        import pandas as pd
        import io
        
        # dtype=str: evita que pandas convierta números grandes a float (notación científica en Nro)
        df = pd.read_excel(io.BytesIO(iva_content), dtype=str)
        df = df.fillna('')
        
        percepciones = []
        retenciones = []
        
        cols_map = {}
        # Pasada 1: criterios estrictos (busca la palabra 'comprobante' en el nombre)
        for c in df.columns:
            c_low = str(c).lower().strip()
            if 'cuit' in c_low: cols_map['cuit'] = c
            elif 'fecha' in c_low and 'comprobante' in c_low: cols_map['fecha'] = c
            elif ('tipo' in c_low or 'descripci' in c_low) and 'comprobante' in c_low: cols_map['tipo_comp'] = c
            elif ('nro' in c_low or 'n°' in c_low or 'numero' in c_low or 'número' in c_low) and 'comprobante' in c_low: cols_map['nro'] = c
            elif 'importe' in c_low: cols_map['monto'] = c
            elif 'razon social' in c_low or 'denominaci' in c_low: cols_map['razon_social'] = c
            elif 'operaci' in c_low: cols_map['operacion'] = c
            # Fecha Ret./Perc.: columna con 'fecha' pero SIN 'comprobante' (ej: "Fecha Ret./Perc.")
            elif 'fecha' in c_low and 'comprobante' not in c_low and 'registra' not in c_low:
                cols_map['fecha_ret'] = c
        # Pasada 2: fallback para columnas sin 'comprobante' (ej: 'Tipo', 'Nro', 'Fecha')
        for c in df.columns:
            c_low = str(c).lower().strip()
            if 'fecha' not in cols_map and 'fecha' in c_low: cols_map['fecha'] = c
            elif 'tipo_comp' not in cols_map and c_low.startswith('tipo'): cols_map['tipo_comp'] = c
            elif 'nro' not in cols_map and ('nro' in c_low or 'n°' in c_low or 'numero' in c_low or 'número' in c_low): cols_map['nro'] = c

        for idx, row in df.iterrows():
            cuit_raw = str(row.get(cols_map.get('cuit', 'CUIT'), '')).strip()
            cuit_limpio = cuit_raw.replace('-', '').replace('.0', '').strip()
            if not cuit_limpio or cuit_limpio == 'nan':
                continue
                
            if len(cuit_limpio) == 11 and '-' not in cuit_raw:
                cuit_raw = f"{cuit_limpio[:2]}-{cuit_limpio[2:10]}-{cuit_limpio[10]}"
            
            # Monto: viene como str por dtype=str, convertir a float manualmente
            monto_str = str(row.get(cols_map.get('monto', 'Importe'), '0')).strip()
            monto_str = monto_str.replace('.', '').replace(',', '.') if ',' in monto_str else monto_str
            try:
                monto_f = float(monto_str or 0)
            except ValueError:
                monto_f = 0.0
            
            op_raw = str(row.get(cols_map.get('operacion', 'Operacion'), 'PERCEPCION'))
            seccion = "R" if 'reten' in op_raw.lower() else "P"

            rz = str(row.get(cols_map.get('razon_social', 'Razon Social'), '')).strip()
            fecha_s = str(row.get(cols_map.get('fecha', 'Fecha'), '')).strip()
            # Fecha de la retención/percepción (distinta a la fecha del comprobante)
            fecha_ret_s = str(row.get(cols_map.get('fecha_ret', ''), '')).strip()
            if fecha_ret_s in ('', 'nan', 'None'):
                fecha_ret_s = ''

            # Nro: viene como str limpio (sin notación científica gracias a dtype=str)
            nro_m = str(row.get(cols_map.get('nro', 'Nro'), '')).strip()
            tipo_c = str(row.get(cols_map.get('tipo_comp', 'Tipo'), '')).strip()

            pv_norm = "0"
            nro_norm = nro_m
            try:
                if '-' in nro_norm:
                    pv_norm = str(int(nro_norm.split('-')[0]))
                    nro_norm = str(int(nro_norm.split('-')[1]))
                elif nro_norm and nro_norm not in ('', 'nan'):
                    # Puede venir como "1234.0" si quedó algo de float
                    nro_norm = str(int(float(nro_norm)))
            except (ValueError, OverflowError):
                pass

            registro = {
                'tipo_reg':    seccion,
                'cod_jur':     '00000',
                'cuit':        cuit_raw,
                'cuit_limpio': cuit_limpio,
                'fecha':       fecha_s,
                'fecha_ret':   fecha_ret_s,
                'pv':          pv_norm,
                'pv_norm':     pv_norm,
                'nro':         nro_m,
                'nro_norm':    nro_norm,
                'tipo_comp':   tipo_c,
                'monto':       monto_f,
                'razon_social': rz
            }
            
            if seccion == "P":
                percepciones.append(registro)
            else:
                retenciones.append(registro)
                
        return {'percepciones': percepciones, 'retenciones': retenciones}

    # ── Función de parsing del TXT ARBA ─────────────────────────────────────────────
    def parsear_arba_iibb(content: str, tipo_default: str = None) -> dict:
        """
        Parsea el archivo TXT de ARBA de percepciones/retenciones de IIBB Buenos Aires.

        El archivo puede tener secciones marcadas por cabeceras de texto:
            PERCEPCIONES:  (o la línea contiene 'PERCEP')
            ... líneas de datos ...
            rETENCIONES:   (o la línea contiene 'RETEN')
            ... líneas de datos ...

        Si no se detectan cabeceras, usa tipo_default ('P' o 'R') para todos los registros.

        Formato de cada línea de datos:
            Posición  1- 3: Código jurisdicción (029=Bs.As. percep, 019=Bs.As. reten)
            Posición  4- 5: Código régimen ARBA (ej: 02)
            Posición  6-18: CUIT con guiones (XX-XXXXXXXX-X)
            Posición 19-28: Fecha DD/MM/YYYY
            Posición 29-33: PV sucursal (5 chars)
            Posición 34-52: Número comprobante (19 chars)
            Posición 53-54: Tipo comprobante (FC, FA, CA, R , O , etc.)
            Posición 55+  : Monto con ceros a la izquierda y coma decimal
        """
        percepciones = []
        retenciones  = []
        seccion_activa = tipo_default  # usar default si no hay headers

        for linea in content.splitlines():
            linea_strip = linea.strip()
            if not linea_strip:
                continue

            # Detectar cabecera de seción (case-insensitive, tolera espacios)
            linea_upper = linea_strip.upper()
            if 'PERCEP' in linea_upper and len(linea_strip) < 40:
                seccion_activa = "P"
                continue
            if 'RETEN' in linea_upper and len(linea_strip) < 40:
                seccion_activa = "R"
                continue

            # Ignorar si es muy corta o si empieza con letra (cabecera de texto)
            if seccion_activa is None or len(linea_strip) < 30:
                continue
            # Ignorar líneas que no empiezan con dígito (son cabeceras de texto)
            if not linea_strip[0].isdigit():
                continue

            try:
                # Campos comunes a ambas secciones
                # Formato (0-indexed):
                #   [0:5]   → código jurisdicción + régimen (ej: 02902)
                #   [5:18]  → CUIT con guiones: XX-XXXXXXXX-X (13 chars)
                #   [18:28] → Fecha DD/MM/YYYY (10 chars)
                #   [28:33] → PV sucursal (5 chars)
                #   [33:53] → Número comprobante (20 chars)
                #   [53:55] → Tipo comprobante (2 chars: FC, FA, CA, R , O …)
                #   [55:]   → Monto (con ceros a la izq. y coma decimal)
                cod_jur  = linea_strip[0:5]         # ej: "02902"
                cuit     = linea_strip[5:18]         # XX-XXXXXXXX-X
                fecha_s  = linea_strip[18:28]        # DD/MM/YYYY
                pv_s     = linea_strip[28:33]        # 5 chars
                nro_s    = linea_strip[33:53]        # 20 chars  ← era 33:52 (19), corregido
                tipo_c   = linea_strip[53:55].strip()  # FC, FA, CA, R, O…  ← era 52:54
                monto_s  = linea_strip[55:].strip()    # resto = monto  ← era 54:

                # Validar que la fecha tenga formato DD/MM/YYYY (sanity check)
                if len(fecha_s) < 10 or '/' not in fecha_s:
                    continue

                # Detectar formato CR (Constancias de Retención ARBA): tiene un campo
                # extra de 20 dígitos (nro_constancia) antes del importe real.
                # Ejemplo: "00000000000000020912000000628790,62"
                #           ←── 20 dígitos ──→←── monto ────→
                # Si los primeros 20 chars son todos dígitos sin coma → es formato CR.
                if len(monto_s) > 20 and monto_s[:20].isdigit() and ',' not in monto_s[:20]:
                    monto_s = monto_s[20:]

                # Convertir monto: quitar signo si es negativo, normalizar coma→punto
                signo = -1 if monto_s.startswith('-') else 1
                monto_limpio = monto_s.lstrip('-').replace('.', '').replace(',', '.')
                try:
                    monto_f = signo * float(monto_limpio)
                except ValueError:
                    monto_f = 0.0

                # Normalizar CUIT: quitar guiones para el cruce
                cuit_limpio = cuit.replace('-', '').strip()

                # Normalizar PV y nro: quitar ceros a la izquierda
                try:
                    pv_norm = str(int(pv_s))
                except ValueError:
                    pv_norm = pv_s.strip()
                try:
                    nro_norm = str(int(nro_s))
                except ValueError:
                    nro_norm = nro_s.strip()

                registro = {
                    'tipo_reg':    seccion_activa,   # 'P' o 'R'
                    'cod_jur':     cod_jur,          # 5 chars: ej "02902"
                    'cuit':        cuit.strip(),
                    'cuit_limpio': cuit_limpio,
                    'fecha':       fecha_s,
                    'pv':          pv_s.strip(),
                    'pv_norm':     pv_norm,
                    'nro':         nro_s.strip(),
                    'nro_norm':    nro_norm,
                    'tipo_comp':   tipo_c,
                    'monto':       monto_f,
                }

                if seccion_activa == "P":
                    percepciones.append(registro)
                else:
                    retenciones.append(registro)

            except Exception:
                continue   # Línea malformada → ignorar

        return {'percepciones': percepciones, 'retenciones': retenciones}


    # ── Card 01: TXT Mendez ─────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Archivo TXT Mendez (movimientos)</div>', unsafe_allow_html=True)
    uploaded_arba_txt_mendez = st.file_uploader(
        "Arrastrá el TXT de movimientos de Mendez o hacé click para seleccionarlo",
        type=["txt", "prn"],
        label_visibility="visible",
        key="cruce_arba_txt_mendez"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Card 02: TXT ARBA ───────────────────────────────────────────────────────────
    st.markdown(f'<div class="card"><div class="card-label">02 · Archivo TXT {organismo}</div>', unsafe_allow_html=True)

    uploaded_arba_txt_arba = st.file_uploader(
        f"Arrastrá el archivo de {organismo} o hacé click para seleccionarlo",
        type=["txt", "csv", "xlsx", "xls"],
        label_visibility="visible",
        key="cruce_arba_txt_arba"
    )

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Procesamiento con ambos archivos ────────────────────────────────────────────
    if uploaded_arba_txt_mendez and uploaded_arba_txt_arba:
        st.success(f"**{uploaded_arba_txt_mendez.name}** + **{uploaded_arba_txt_arba.name}** listos para analizar")

        # Configurar parámetro de cruce desde la UI
        tipo_default_arba = "P" if "Percepciones" in tipo_cruce_sel else "R"


        # ── Card 03: Acción ─────────────────────────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">03 · Procesar</div>', unsafe_allow_html=True)

        if st.button("⬡  Cruzar y exportar Excel"):
            try:
                # ── 1. Parsear TXT Mendez ────────────────────────────────────────────
                with st.spinner("Parseando TXT Mendez..."):
                    txt_mendez_content = uploaded_arba_txt_mendez.getvalue().decode("latin-1", errors="replace")
                    transacciones_arba, meta_arba = parsear_archivo(content=txt_mendez_content)

                    # Extraer mes y año de Mendez desde el encabezado (ej. "Desde 01/03/2026 hasta...")
                    import re
                    mes_anio_mendez = ""
                    if meta_arba and meta_arba.get('periodo'):
                        match_ma = re.search(r'/(\d{2}/\d{4}|\d{2}/\d{2})', meta_arba['periodo'])
                        if match_ma:
                            mes_anio_mendez = match_ma.group(1)


                if not transacciones_arba:
                    st.error("No se encontraron transacciones en el TXT Mendez. Verificá el formato.")
                else:
                    # ── 2. Parsear Archivo de Organismo ───────────────────────────────────
                    with st.spinner(f"Parseando archivo de {organismo}..."):
                        if organismo == "ARBA":
                            arba_content = uploaded_arba_txt_arba.getvalue().decode("latin-1", errors="replace")
                            arba_data    = parsear_arba_iibb(arba_content, tipo_default=tipo_default_arba)
                        elif organismo == "AGIP":
                            agip_content = uploaded_arba_txt_arba.getvalue()
                            arba_data    = parsear_agip_iibb(agip_content)
                        elif organismo in ("IVA", "Ganancias"):
                            iva_content = uploaded_arba_txt_arba.getvalue()
                            arba_data    = parsear_iva_xls(iva_content)

                    # Alias de visualización: IVA/Ganancias → ARCA en títulos y nombres de hojas
                    label_organismo = "ARCA" if organismo in ("IVA", "Ganancias") else organismo

                    percepciones = arba_data['percepciones']
                    retenciones  = arba_data['retenciones']
                    # Filtrar por tipo seleccionado: si el usuario elige "Retenciones"
                    # no incluir percepciones y viceversa (evita falsos cruces si el
                    # archivo del organismo contiene ambos tipos a la vez).
                    if tipo_default_arba == "P":
                        registros_activos = percepciones
                    elif tipo_default_arba == "R":
                        registros_activos = retenciones
                    else:
                        registros_activos = percepciones + retenciones

                    if tipo_default_arba == "P":
                        label_tipo = "Percepciones"
                        if organismo == "AGIP":
                            kw_mendez = ("CAP. FED.", "CAP", "FED", "CABA", "AGIP")
                            excl_mendez = ("ADUA", "I.V.A", "IVA", "GCIAS")
                        elif organismo == "IVA":
                            kw_mendez = ("IVA", "I.V.A", "AGREGADO")
                            excl_mendez = ("ADUA", "GCIAS")
                        elif organismo == "Ganancias":
                            kw_mendez = ("GCIAS", "GANANCIAS", "GCIA")
                            excl_mendez = ("ADUA", "I.V.A", "IVA", "SIRCREB", "SIRTAC")
                        else:
                            kw_mendez = ("BS.AS", "BSAS", "BS AS", "BUENOS AIRES")
                            excl_mendez = ("ADUA", "I.V.A", "IVA", "GCIAS")
                    elif tipo_default_arba == "R":
                        label_tipo = "Retenciones"
                        if organismo == "AGIP":
                            kw_mendez = ("CAP. FED.", "CAP", "FED", "CABA", "AGIP")
                            excl_mendez = ("SIRCREB", "SIRTAC", "BCO", "GCIAS", "IVA", "BANCO", "BANCAR")
                        elif organismo == "IVA":
                            kw_mendez = ("IVA", "I.V.A", "AGREGADO")
                            excl_mendez = ("SIRCREB", "SIRTAC", "BCO", "GCIAS", "BANCO", "BANCAR")
                        elif organismo == "Ganancias":
                            kw_mendez = ("GCIAS", "GANANCIAS", "GCIA")
                            excl_mendez = ("SIRCREB", "SIRTAC", "BCO", "IVA", "BANCO", "BANCAR")
                        else:
                            kw_mendez = ("BS.AS", "BSAS", "BS AS", "BUENOS AIRES")
                            excl_mendez = ("SIRCREB", "SIRTAC", "BCO", "GCIAS", "IVA", "BANCO", "BANCAR")
                    else:
                        label_tipo = "Percepciones + Retenciones"
                        if organismo == "AGIP":
                            kw_mendez = ("CAP. FED.", "CAP", "FED", "CABA", "AGIP")
                            excl_mendez = ("ADUA", "I.V.A", "IVA", "GCIAS", "SIRCREB", "SIRTAC", "BCO", "BANCO")
                        elif organismo == "IVA":
                            kw_mendez = ("IVA", "I.V.A", "AGREGADO")
                            excl_mendez = ("ADUA", "GCIAS", "SIRCREB", "SIRTAC", "BCO", "BANCO")
                        elif organismo == "Ganancias":
                            kw_mendez = ("GCIAS", "GANANCIAS", "GCIA")
                            excl_mendez = ("ADUA", "I.V.A", "IVA", "SIRCREB", "SIRTAC", "BCO", "BANCO")
                        else:
                            kw_mendez = ("BS.AS", "BSAS", "BS AS", "BUENOS AIRES")
                            excl_mendez = ("ADUA", "I.V.A", "IVA", "GCIAS", "SIRCREB", "SIRTAC", "BCO", "BANCO")

                    if not registros_activos:
                        st.warning(f"El TXT ARBA no contiene registros. Verificá el contenido del archivo.")
                    else:
                        # ── 3. Extraer percepciones/retenciones Bs.As. de Mendez ─────
                        with st.spinner("Cruzando por CUIT..."):
                            IVA_RATES_SET = {
                                'Tasa 21%', 'T.21%', 'C.F.21%', 'Tasa 27%', 'T.27%',
                                'Tasa 10.5%', 'Tasa 10,5%', 'T.10.5%', 'T.10,5%',
                                'C.F.10.5%', 'C.F.10,5%', 'Tasa 5%', 'T.5%',
                                'Tasa 2.5%', 'Tasa 2,5%', 'T.2.5%', 'T.2,5%',
                                'T.IMP 21%', 'T.IMP 10%', 'Exento', 'R.Monot21', 'R.Mont.10',
                            }
                            def _es_bsas_comp(nombre: str, tipo_reg: str) -> bool:
                                """Filtra sub-conceptos de Mendez que sean del Organismo actual."""
                                if not nombre or nombre in IVA_RATES_SET:
                                    return False
                                nu = nombre.upper()
                                # Debe contener la keyword de tipo (PERC o RET)
                                if tipo_reg == "P" and "PERC" not in nu:
                                    return False
                                if tipo_reg == "R" and "RET" not in nu:
                                    return False
                                if tipo_reg is None:
                                    if "PERC" not in nu and "RET" not in nu:
                                        return False
                                # Debe contener la keyword del organismo
                                if not any(kw in nu for kw in kw_mendez):
                                    return False
                                # Excluir no deseados
                                if any(x in nu for x in excl_mendez):
                                    return False
                                return True

                            # Acumular monto Bs.As. por CUIT desde Mendez
                            mendez_por_cuit = {}   # cuit_limpio → monto
                            mendez_detalle  = []   # registros para hoja detalle

                            for t in transacciones_arba:
                                cuit_raw = t.get('CUIT', '')
                                cuit_limpio = cuit_raw.replace('-', '').strip()
                                if not cuit_limpio:
                                    continue

                                monto_bsas = 0.0
                                concepto_encontrado = ''

                                # Desde la tasa principal
                                tasa = t.get('Tasa', '')
                                if tasa and _es_bsas_comp(tasa, tipo_default_arba):
                                    monto_bsas += t.get('Neto', 0.0)
                                    concepto_encontrado = tasa

                                # Desde sub-conceptos
                                for s in t.get('SubConceptos', []):
                                    nom = s.get('Concepto', '')
                                    if _es_bsas_comp(nom, tipo_default_arba):
                                        m = s['Neto'] if s.get('Neto', 0) != 0 else s.get('Percepcion', 0)
                                        monto_bsas += m
                                        if not concepto_encontrado:
                                            concepto_encontrado = nom

                                if monto_bsas != 0.0:
                                    # Invertir signo si es Nota de Crédito (NC)
                                    if t.get('Tipo') == 'NC':
                                        monto_bsas = -monto_bsas

                                    mendez_por_cuit[cuit_limpio] = mendez_por_cuit.get(cuit_limpio, 0.0) + monto_bsas

                                    # Separar PV y Nro del número (quitar letra final si existe)
                                    numero_raw = t.get('Numero', '')
                                    if '-' in numero_raw:
                                        pv_m = numero_raw.split('-')[0].lstrip('0') or '0'
                                        nro_m = numero_raw.split('-')[1]
                                    else:
                                        pv_m  = numero_raw[:5].lstrip('0') or '0'
                                        nro_m = numero_raw[5:]
                                    # Quitar letra final del número
                                    if nro_m and nro_m[-1].isalpha():
                                        nro_m = nro_m[:-1]
                                    try:
                                        nro_m = str(int(nro_m))
                                        pv_m  = str(int(pv_m))
                                    except ValueError:
                                        pass
                                    # CUIT con guiones para la hoja
                                    cuit_raw_fmt = cuit_raw if '-' in cuit_raw else (
                                        f"{cuit_raw[:2]}-{cuit_raw[2:10]}-{cuit_raw[10]}" if len(cuit_raw) == 11 else cuit_raw
                                    )
                                    
                                    # Formatear Fecha Mendez
                                    dia_raw = str(t.get('Fecha', '')).strip()
                                    fecha_fmt = dia_raw
                                    if dia_raw.isdigit() and mes_anio_mendez:
                                        fecha_fmt = f"{int(dia_raw):02d}/{mes_anio_mendez}"

                                    mendez_detalle.append({
                                        'CUIT':         cuit_raw_fmt,
                                        'Proveedor':    t.get('Proveedor', ''),
                                        'Fecha Emision': fecha_fmt,
                                        'Tipo Comp.': t.get('Tipo', ''),

                                        'PV':         int(pv_m) if str(pv_m).isdigit() else pv_m,
                                        'Nro':        int(nro_m) if str(nro_m).isdigit() else nro_m,
                                        'Monto':      round(monto_bsas, 2),
                                    })

                            # Acumular monto por CUIT desde ARBA
                            arba_por_cuit = {}   # cuit_limpio → monto
                            for r in registros_activos:
                                ck = r['cuit_limpio']
                                arba_por_cuit[ck] = arba_por_cuit.get(ck, 0.0) + r['monto']

                            # Construir tabla de cruce
                            all_cuits = set(mendez_por_cuit.keys()) | set(arba_por_cuit.keys())
                            # Construir lookup CUIT → proveedor desde Mendez
                            cuit_proveedor = {}
                            for t in transacciones_arba:
                                ck = t.get('CUIT', '').replace('-', '').strip()
                                if ck and ck not in cuit_proveedor:
                                    cuit_proveedor[ck] = t.get('Proveedor', '')

                            # Integrar razon social de AGIP o si está disponible en padrón
                            for r in registros_activos:
                                ck = r['cuit_limpio']
                                if ck not in cuit_proveedor and r.get('razon_social'):
                                    cuit_proveedor[ck] = r['razon_social']

                            # Buscar en CuitOnline los CUITs que no están en Mendez ni en el Padrón
                            cuits_faltantes = [ck for ck in arba_por_cuit.keys() if ck not in cuit_proveedor]
                            if cuits_faltantes and organismo == "ARBA":
                                prog_text = f"Buscando {len(cuits_faltantes)} proveedores en CuitOnline..."
                                progress_bar = st.progress(0, text=prog_text)
                                for idx_cf, ck in enumerate(cuits_faltantes):
                                    rz = obtener_razon_social_cuitonline(ck)
                                    if rz:
                                        cuit_proveedor[ck] = rz + " (CuitOnline)"
                                    else:
                                        cuit_proveedor[ck] = "⚠ SIN PROVEEDOR"
                                    progress_bar.progress((idx_cf + 1) / len(cuits_faltantes), text=prog_text)
                                progress_bar.empty()
                            elif cuits_faltantes:
                                for ck in cuits_faltantes:
                                    cuit_proveedor[ck] = "⚠ SIN PROVEEDOR"

                            filas_cruce = []
                            for ck in sorted(all_cuits):
                                m_mendez = mendez_por_cuit.get(ck, 0.0)
                                m_arba   = arba_por_cuit.get(ck, 0.0)
                                diff     = round(m_arba - m_mendez, 2)
                                # Formatear CUIT con guiones
                                cuit_fmt = f"{ck[:2]}-{ck[2:10]}-{ck[10]}" if len(ck) == 11 else ck
                                # Estado
                                if diff == 0.0:
                                    count_m = sum(1 for x in mendez_detalle if str(x.get('CUIT', '')).replace('-', '') == ck)
                                    count_a = sum(1 for r in registros_activos if r['cuit_limpio'] == ck)
                                    if count_m == 0 and count_a > 0:
                                        estado = "⚠ Falta en Mendez"
                                    elif count_a == 0 and count_m > 0:
                                        estado = f"⚠ Falta en {organismo}"
                                    elif count_m != count_a:
                                        estado = "⚠ Diferencia"
                                    else:
                                        estado = "✓ OK"
                                elif m_mendez == 0.0:
                                    estado = "⚠ Falta en Mendez"
                                elif m_arba == 0.0:
                                    estado = f"⚠ Falta en {organismo}"
                                else:
                                    estado = "⚠ Diferencia"
                                filas_cruce.append({
                                    'CUIT': cuit_fmt,
                                    'Proveedor': cuit_proveedor.get(ck, ''),
                                    f'Total {organismo} ({label_tipo})': round(m_arba, 2),
                                    'Total Mendez': round(m_mendez, 2),
                                    f'Diferencia ({organismo} - Mendez)': diff,
                                    'Estado': estado,
                                })

                            df_cruce = pd.DataFrame(filas_cruce)
                            # Helper subtotales
                            def agregar_subtotales_cuit(lista_dicts, col_monto='Monto'):
                                if not lista_dicts: return []
                                # Ordenar por CUIT y Fecha
                                lista_dicts.sort(key=lambda x: (x.get('CUIT', ''), x.get('Fecha Emision', '')))
                                filas_out = []
                                last_cuit = None
                                last_prov = ""
                                run_sum = 0.0
                                for r in lista_dicts:
                                    cuit = r.get('CUIT', '')
                                    prov = r.get('Proveedor', '')

                                    if last_cuit is not None and cuit != last_cuit:
                                        sub = {k: '' for k in lista_dicts[0].keys()}
                                        sub['CUIT'] = last_cuit
                                        sub['Proveedor'] = f"{last_prov} (SUBTOTAL)" if last_prov else "(SUBTOTAL)"
                                        sub[col_monto] = run_sum
                                        filas_out.append(sub)
                                        run_sum = 0.0
                                        
                                    if prov and "SIN PROVEEDOR" not in prov:
                                        last_prov = prov
                                    
                                    last_cuit = cuit
                                    run_sum += r.get(col_monto, 0.0)
                                    filas_out.append(r)
                                
                                if last_cuit is not None:
                                    sub = {k: '' for k in lista_dicts[0].keys()}
                                    sub['CUIT'] = last_cuit
                                    sub['Proveedor'] = f"{last_prov} (SUBTOTAL)" if last_prov else "(SUBTOTAL)"
                                    sub[col_monto] = run_sum

                                    filas_out.append(sub)
                                return filas_out

                            # Para IVA/Ganancias: la columna PV se reutiliza para mostrar la Fecha Ret./Perc.
                            # (el Nro ya incluye el número completo sin necesidad de PV separado)
                            _usa_fecha_ret = organismo in ('IVA', 'Ganancias')
                            _pv_col_lbl = 'Fecha Ret./Perc.' if _usa_fecha_ret else 'PV'
                            arba_detalle_list = [
                                {
                                    'CUIT':        r['cuit'],
                                    'Proveedor':   cuit_proveedor.get(r['cuit_limpio'], ''),
                                    'Fecha Emision': r['fecha'],
                                    'Tipo Comp.':    r['tipo_comp'],
                                    _pv_col_lbl:   r.get('fecha_ret', '') if _usa_fecha_ret
                                                   else (int(r['pv_norm']) if str(r['pv_norm']).isdigit() else r['pv_norm']),
                                    'Nro':         int(r['nro_norm']) if str(r['nro_norm']).isdigit() else r['nro_norm'],
                                    'Monto':       r['monto'],
                                }
                                for r in registros_activos
                            ]

                            # Consolidar registros del organismo que dividen una percepción
                            # en varios movimientos para el mismo (CUIT, Nro). Excepción: bancos,
                            # que informan percepciones de forma especial y deben mantenerse separados.
                            _grupos = {}
                            _orden = []
                            for _r in arba_detalle_list:
                                _prov = str(_r.get('Proveedor', '')).upper()
                                if 'BANCO' in _prov:
                                    _key = ('_uniq_', id(_r))
                                else:
                                    _key = (str(_r.get('CUIT', '')).replace('-', ''), str(_r.get('Nro', '')))
                                if _key not in _grupos:
                                    _grupos[_key] = []
                                    _orden.append(_key)
                                _grupos[_key].append(_r)
                            _consolidados = []
                            for _k in _orden:
                                _regs = _grupos[_k]
                                if len(_regs) == 1:
                                    _regs[0]['_original_rows'] = [_regs[0]]
                                    _consolidados.append(_regs[0])
                                else:
                                    _base = _regs[0].copy()
                                    _base['Monto'] = round(sum(_x.get('Monto', 0) for _x in _regs), 2)
                                    _base['_original_rows'] = _regs[:]
                                    _consolidados.append(_base)
                            arba_detalle_list = _consolidados

                            # Índice único por registro para tracking post-matching
                            for _i, _r in enumerate(mendez_detalle):    _r['_idx'] = _i
                            for _i, _r in enumerate(arba_detalle_list): _r['_idx'] = _i

                            # -------------------------------------------------------------
                            # Matriz de Diferencias (Separada por Mendez y Organismo)
                            # -------------------------------------------------------------
                            lista_m_dif = []
                            lista_a_dif = []
                            
                            for ck in sorted(all_cuits):
                                m_mendez = mendez_por_cuit.get(ck, 0.0)
                                m_arba   = arba_por_cuit.get(ck, 0.0)
                                diff     = round(m_arba - m_mendez, 2)
                                t_men = [x for x in mendez_detalle if str(x.get('CUIT','')).replace('-','') == ck]
                                t_arb = [x for x in arba_detalle_list if str(x.get('CUIT','')).replace('-','') == ck]
                                if diff != 0.0 or len(t_men) != len(t_arb):
                                    cuit_fmt = f"{ck[:2]}-{ck[2:10]}-{ck[10]}" if len(ck) == 11 else ck

                                    # Cancelar comprobantes en dos fases:
                                    # Fase 1: priorizar coincidencia por Nro (tolera que IVA XLS
                                    #         guarde PV+Nro juntos, ej: 11700025262 vs 25262).
                                    # Fase 2: fallback por monto para los sin pareja por Nro.
                                    def _nro(x):   return str(x.get('Nro', ''))
                                    def _monto(x): return round(x.get('Monto', 0), 2)

                                    def _nro_match(a, b):
                                        sa, sb = str(a).lstrip('0'), str(b).lstrip('0')
                                        if not sa or not sb: return sa == sb
                                        if sa == sb: return True
                                        # Suffix match con mínimo 5 chars para evitar falsos positivos
                                        if len(sa) >= 5 and sb.endswith(sa): return True
                                        if len(sb) >= 5 and sa.endswith(sb): return True
                                        return False

                                    # ── Dirección Mendez → Organismo ──────────────────────
                                    pool_arb = list(t_arb)
                                    sin_nro_men = []
                                    for t in t_men:
                                        idx = next((i for i, a in enumerate(pool_arb)
                                                    if _nro_match(_nro(a), _nro(t)) and _monto(a) == _monto(t)), None)
                                        if idx is not None:
                                            pool_arb.pop(idx)
                                        else:
                                            sin_nro_men.append(t)
                                    for t in sin_nro_men:
                                        idx = next((i for i, a in enumerate(pool_arb)
                                                    if _monto(a) == _monto(t)), None)
                                        if idx is not None:
                                            pool_arb.pop(idx)
                                        else:
                                            t_copy = t.copy()
                                            t_copy['CUIT'] = cuit_fmt
                                            lista_m_dif.append(t_copy)

                                    # ── Dirección Organismo → Mendez ──────────────────────
                                    pool_men = list(t_men)
                                    sin_nro_org = []
                                    for t in t_arb:
                                        idx = next((i for i, m in enumerate(pool_men)
                                                    if _nro_match(_nro(m), _nro(t)) and _monto(m) == _monto(t)), None)
                                        if idx is not None:
                                            pool_men.pop(idx)
                                        else:
                                            sin_nro_org.append(t)
                                    for t in sin_nro_org:
                                        idx = next((i for i, m in enumerate(pool_men)
                                                    if _monto(m) == _monto(t)), None)
                                        if idx is not None:
                                            pool_men.pop(idx)
                                        else:
                                            for orig in t.get('_original_rows', [t]):
                                                orig_copy = orig.copy()
                                                orig_copy['CUIT'] = cuit_fmt
                                                lista_a_dif.append(orig_copy)


                            # Estado por registro: ✓ Ok si fue emparejado, ⚠ Falta si quedó sin par
                            _unmatched_men_idx = {r['_idx'] for r in lista_m_dif if '_idx' in r}
                            _unmatched_arb_idx = {r['_idx'] for r in lista_a_dif if '_idx' in r}
                            for rec in mendez_detalle:
                                rec['_estado'] = f'⚠ Falta en {organismo}' if rec.get('_idx') in _unmatched_men_idx else '✓ Ok'
                            for rec in arba_detalle_list:
                                rec['_estado'] = '⚠ Falta en Mendez' if rec.get('_idx') in _unmatched_arb_idx else '✓ Ok'

                            cols_out_men = ['CUIT', 'Proveedor', 'Fecha Emision', 'Tipo Comp.', 'PV', 'Nro', 'Monto']
                            cols_out_org = ['CUIT', 'Proveedor', 'Fecha Emision', 'Tipo Comp.', _pv_col_lbl, 'Nro', 'Monto']
                            df_mas_mendez = pd.DataFrame(lista_m_dif)[cols_out_men] if lista_m_dif else pd.DataFrame(columns=cols_out_men)
                            df_mas_org    = pd.DataFrame(lista_a_dif)[cols_out_org] if lista_a_dif else pd.DataFrame(columns=cols_out_org)

                            df_cruce      = pd.DataFrame(filas_cruce)
                            df_mendez_det = pd.DataFrame(agregar_subtotales_cuit(mendez_detalle))
                            df_arba_det   = pd.DataFrame(agregar_subtotales_cuit(arba_detalle_list))

                            if df_mendez_det.empty:
                                df_mendez_det = pd.DataFrame(columns=['CUIT', 'Proveedor', 'Fecha Emision', 'Tipo Comp.', 'PV', 'Nro', 'Monto'])
                                df_mendez_det['Estado'] = ''
                            else:
                                df_mendez_det['Estado'] = df_mendez_det['_estado'].fillna('').astype(str).replace('nan', '')
                                df_mendez_det.drop(columns=['_idx', '_estado'], inplace=True, errors='ignore')
                            df_mendez_det['Cantidad'] = ''

                            if df_arba_det.empty:
                                df_arba_det = pd.DataFrame(columns=['CUIT', 'Proveedor', 'Fecha Emision', 'Tipo Comp.', 'PV', 'Nro', 'Monto'])
                                df_arba_det['Estado'] = ''
                            else:
                                df_arba_det['Estado'] = df_arba_det['_estado'].fillna('').astype(str).replace('nan', '')
                                df_arba_det.drop(columns=['_idx', '_estado', '_original_rows'], inplace=True, errors='ignore')
                            df_arba_det['Cantidad'] = ''



                        # ── 4. Stats ─────────────────────────────────────────────────
                        total_ok    = (df_cruce['Estado'] == '✓ OK').sum()
                        total_diff  = len(df_cruce) - total_ok
                        monto_total_arba   = sum(r['monto'] for r in registros_activos)
                        monto_total_mendez = sum(mendez_por_cuit.values())

                        st.success(f"✓  Cruce {organismo} · **{label_tipo}** completado")

                        st.markdown(f"""
                        <div class="stats-row">
                            <div class="stat-chip">
                                <span class="stat-val">{len(all_cuits):,}</span>
                                <span class="stat-lbl">CUITs totales</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val" style="color:#4ae8a0;">{total_ok:,}</span>
                                <span class="stat-lbl">Sin diferencia</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val" style="color:#f87171;">{total_diff:,}</span>
                                <span class="stat-lbl">Con diferencia</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{len(registros_activos):,}</span>
                                <span class="stat-lbl">Reg. {organismo}</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                        # ── 5. Generar Excel de 3 hojas ──────────────────────────────
                        with st.spinner("Generando Excel..."):
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            from openpyxl.formatting.rule import CellIsRule, FormulaRule
                            from openpyxl.utils import get_column_letter

                            output_cruce = io.BytesIO()
                            with pd.ExcelWriter(output_cruce, engine='openpyxl') as writer:

                                # — Helper estilos ——————————————————————————————————————
                                HDR_FONT   = Font(bold=True, size=10, color='FFFFFF')
                                CTR        = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                THIN       = Border(
                                    left=Side(style='thin', color='252935'),
                                    right=Side(style='thin', color='252935'),
                                    top=Side(style='thin', color='252935'),
                                    bottom=Side(style='thin', color='252935'),
                                )
                                FILL_TITLE = PatternFill('solid', fgColor='2F5496')
                                FILL_HDR_P = PatternFill('solid', fgColor='7B2D8B')   # violeta percepciones
                                FILL_HDR_R = PatternFill('solid', fgColor='C00000')   # rojo retenciones
                                FILL_HDR   = FILL_HDR_P if tipo_default_arba != "R" else FILL_HDR_R
                                FILL_OK    = PatternFill('solid', fgColor='E2EFDA')   # verde claro
                                FILL_DIFF  = PatternFill('solid', fgColor='FCE4D6')   # rojo claro
                                FILL_ZEBRA = PatternFill('solid', fgColor='F2F2F2')
                                FMT_MONEY  = '_-"$"* #,##0.00_-;[Red]-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
                                razon      = meta_arba.get('razon_social', 'CONTRIBUYENTE').upper()
                                periodo    = meta_arba.get('periodo', '')

                                def _encabezado(ws, n_cols, titulo2, info3):
                                    lc = get_column_letter(n_cols)
                                    ws.merge_cells(f'A1:{lc}1')
                                    ws['A1'] = razon
                                    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
                                    ws['A1'].fill = FILL_TITLE
                                    ws['A1'].alignment = CTR
                                    ws.row_dimensions[1].height = 22
                                    ws.merge_cells(f'A2:{lc}2')
                                    ws['A2'] = titulo2
                                    ws['A2'].font = Font(bold=True, size=12, color='FFFFFF')
                                    ws['A2'].fill = FILL_HDR
                                    ws['A2'].alignment = CTR
                                    ws.row_dimensions[2].height = 20
                                    ws.merge_cells(f'A3:{lc}3')
                                    ws['A3'] = info3
                                    ws['A3'].font = Font(bold=True, size=10, color='2F5496')
                                    ws['A3'].alignment = CTR
                                    ws.row_dimensions[3].height = 16

                                def _estilizar_hdr(ws, hdr_row, n_cols, fill=None):
                                    f = fill or FILL_HDR
                                    for c in range(1, n_cols + 1):
                                        cell = ws.cell(row=hdr_row, column=c)
                                        cell.font = HDR_FONT; cell.fill = f
                                        cell.alignment = CTR; cell.border = THIN

                                def _autofit_ws(ws, n_cols):
                                    for c in range(1, n_cols + 1):
                                        max_len = 0
                                        col_letter = get_column_letter(c)
                                        for row in ws.iter_rows(min_col=c, max_col=c):
                                            for cell in row:
                                                try:
                                                    if cell.value:
                                                        max_len = max(max_len, len(str(cell.value)))
                                                except:
                                                    pass
                                        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

                                # ══════ Hoja 1: CRUCE x CUIT (toda formulada) ══════════
                                # Las hojas de detalle deben crearse ANTES para que las
                                # referencias SUMIF/VLOOKUP funcionen
                                CRUCE_HEADERS = [
                                    'CUIT',
                                    'Proveedor',
                                    f'Total {organismo} ({label_tipo})',
                                    'Total Mendez',
                                    f'Diferencia ({organismo} - Mendez)',
                                    'Estado',
                                ]
                                n1 = len(CRUCE_HEADERS)
                                last_arba_row = max(len(df_arba_det) + 5, 6)
                                last_men_row  = max(len(df_mendez_det) + 5, 6)

                                ws1 = writer.book.create_sheet('Cruce x CUIT', 0)
                                _encabezado(ws1, n1,
                                    f'CRUCE {organismo} · {label_tipo.upper()}',
                                    f'CUIT: {meta_arba.get("cuit_empresa","")} | {periodo} | {len(all_cuits)} CUITs'
                                )
                                ws1.row_dimensions[4].height = 4
                                for c_i, h in enumerate(CRUCE_HEADERS, start=1):
                                    cell = ws1.cell(row=5, column=c_i, value=h)
                                    cell.font = HDR_FONT; cell.fill = FILL_HDR
                                    cell.alignment = CTR; cell.border = THIN

                                cuits_sorted = sorted(all_cuits)
                                for row_i, ck in enumerate(cuits_sorted, start=6):
                                    cuit_fmt = f"{ck[:2]}-{ck[2:10]}-{ck[10]}" if len(ck) == 11 else ck

                                    # A: CUIT (valor estático)
                                    ws1.cell(row_i, 1).value = cuit_fmt

                                    # B: Proveedor → VLOOKUP desde Detalle Mendez, fallback Detalle ARBA
                                    ws1.cell(row_i, 2).value = (
                                        f"=IFERROR(VLOOKUP(A{row_i},'Detalle Mendez'!$A$6:$B${last_men_row},2,0),"
                                        f"IFERROR(VLOOKUP(A{row_i},'Detalle {organismo}'!$A$6:$B${last_arba_row},2,0),\"\"))"
                                    )

                                    # C: Total ARBA → SUMIFS Detalle ARBA excluyendo subtotal
                                    ws1.cell(row_i, 3).value = (
                                        f"=SUMIFS('Detalle {organismo}'!$G$6:$G${last_arba_row},"
                                        f"'Detalle {organismo}'!$A$6:$A${last_arba_row},A{row_i},"
                                        f"'Detalle {organismo}'!$B$6:$B${last_arba_row},\"<>*(SUBTOTAL)*\")"
                                    )
                                    ws1.cell(row_i, 3).number_format = FMT_MONEY

                                    # D: Total Mendez → SUMIFS Detalle Mendez excluyendo subtotal
                                    ws1.cell(row_i, 4).value = (
                                        f"=SUMIFS('Detalle Mendez'!$G$6:$G${last_men_row},"
                                        f"'Detalle Mendez'!$A$6:$A${last_men_row},A{row_i},"
                                        f"'Detalle Mendez'!$B$6:$B${last_men_row},\"<>*(SUBTOTAL)*\")"
                                    )
                                    ws1.cell(row_i, 4).number_format = FMT_MONEY


                                    # E: Diferencia → =C-D
                                    ws1.cell(row_i, 5).value = f'=C{row_i}-D{row_i}'
                                    ws1.cell(row_i, 5).number_format = FMT_MONEY

                                    # F: Estado → IF anidado
                                    ws1.cell(row_i, 6).value = (
                                        f'=IF(E{row_i}=0,"✓ OK",'
                                        f'IF(D{row_i}=0,"⚠ Falta en Mendez",'
                                        f'IF(C{row_i}=0,"⚠ Falta en {organismo}","⚠ Diferencia")))'
                                    )

                                    # Estilar fila según diferencia calculada en Python
                                    m_a = arba_por_cuit.get(ck, 0.0)
                                    m_m = mendez_por_cuit.get(ck, 0.0)
                                    # Ya no usamos FILL_DIFF estatico, será condicional.
                                    for c in range(1, n1 + 1):
                                        cell = ws1.cell(row_i, c)
                                        if round(m_a - m_m, 2) == 0.0: cell.fill = FILL_OK
                                        cell.alignment = CTR
                                        cell.border = THIN

                                # Fila TOTAL GENERAL
                                tot_row = len(cuits_sorted) + 6
                                FILL_TOT = PatternFill('solid', fgColor='D9E1F2')
                                for c_i in range(1, n1 + 1):
                                    cell = ws1.cell(tot_row, c_i)
                                    cell.fill = FILL_TOT; cell.border = THIN; cell.alignment = CTR
                                ws1.cell(tot_row, 1).value = 'TOTAL GENERAL'
                                ws1.cell(tot_row, 1).font = Font(bold=True, size=10)
                                for c_i in (3, 4, 5):
                                    col_l = get_column_letter(c_i)
                                    cell = ws1.cell(tot_row, c_i)
                                    cell.value = f'=SUM({col_l}6:{col_l}{tot_row-1})'
                                    cell.number_format = FMT_MONEY
                                    cell.font = Font(bold=True)
                                ws1.row_dimensions[tot_row].height = 18
                                _autofit_ws(ws1, n1)

                                FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                                FILL_RED    = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                                FILL_ORANGE = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

                                f_dif = FormulaRule(formula=['$F6="⚠ Diferencia"'], stopIfTrue=False, fill=FILL_YELLOW)
                                f_fal_org = FormulaRule(formula=[f'$F6="⚠ Falta en {organismo}"'], stopIfTrue=False, fill=FILL_RED)
                                f_fal_men = FormulaRule(formula=['$F6="⚠ Falta en Mendez"'], stopIfTrue=False, fill=FILL_ORANGE)

                                _cf_last_1 = max(len(cuits_sorted) + 5, 6)
                                ws1.conditional_formatting.add(f'A6:F{_cf_last_1}', f_dif)
                                ws1.conditional_formatting.add(f'A6:F{_cf_last_1}', f_fal_org)
                                ws1.conditional_formatting.add(f'A6:F{_cf_last_1}', f_fal_men)
                                
                                # ══════ Hoja 2: DE MAS EN MENDEZ ═══════════════════════════
                                def _escribir_por_cuit(ws, lista_dif, titulo, subtitulo, col_labels=None):
                                    """Escribe mini-tablas por CUIT: encabezado rojo + datos + TOTAL + fila vacía.
                                    col_labels: dict opcional {key_col: label_display} para renombrar cabeceras."""
                                    # Las COLS_DIF usan las claves reales del dict de datos
                                    # (para IVA la clave 'PV' ya viene como 'Fecha Ret./Perc.' en los datos)
                                    primera = lista_dif[0] if lista_dif else {}
                                    COLS_DIF = [k for k in ['CUIT', 'Proveedor', 'Fecha Emision', 'Tipo Comp.', 'PV', 'Fecha Ret./Perc.', 'Nro', 'Monto'] if k in primera]
                                    if not COLS_DIF:
                                        COLS_DIF = ['CUIT', 'Proveedor', 'Fecha Emision', 'Tipo Comp.', 'PV', 'Nro', 'Monto']
                                    n_dif     = len(COLS_DIF)
                                    idx_monto = COLS_DIF.index('Monto') + 1
                                    idx_nro   = COLS_DIF.index('Nro') + 1
                                    col_m_l   = get_column_letter(idx_monto)
                                    # Labels de cabecera: usa col_labels si se provee, si no el nombre de clave
                                    _labels   = col_labels or {}

                                    _encabezado(ws, n_dif, titulo, subtitulo)
                                    ws.row_dimensions[4].height = 4

                                    # Agrupar por CUIT manteniendo orden de aparición
                                    grupos = {}
                                    for r in lista_dif:
                                        grupos.setdefault(r['CUIT'], []).append(r)

                                    cur = 5  # fila de inicio
                                    for cuit_v, registros in grupos.items():
                                        # — Mini encabezado rojo ——————————————————————
                                        for ci, h in enumerate(COLS_DIF, 1):
                                            cell = ws.cell(cur, ci, value=_labels.get(h, h))
                                            cell.font = HDR_FONT
                                            cell.fill = PatternFill('solid', fgColor='C00000')
                                            cell.alignment = CTR
                                            cell.border = THIN
                                        cur += 1

                                        data_start = cur
                                        # — Filas de detalle ——————————————————————————
                                        for ri, rec in enumerate(registros):
                                            fill = FILL_ZEBRA if ri % 2 == 1 else None
                                            for ci, col in enumerate(COLS_DIF, 1):
                                                val  = rec.get(col, '')
                                                cell = ws.cell(cur, ci, value=val)
                                                cell.alignment = CTR
                                                cell.border = THIN
                                                if fill: cell.fill = fill
                                                if ci == idx_monto:
                                                    cell.number_format = FMT_MONEY
                                                if ci == idx_nro:
                                                    # Forzar texto para evitar notación científica
                                                    cell.number_format = '@'
                                                    cell.value = str(val) if val != '' else ''
                                            cur += 1
                                        data_end = cur - 1

                                        # — Fila TOTAL ————————————————————————————————
                                        for ci in range(1, n_dif + 1):
                                            cell = ws.cell(cur, ci)
                                            cell.fill      = PatternFill('solid', fgColor='D9E1F2')
                                            cell.border    = THIN
                                            cell.alignment = CTR
                                            cell.font      = Font(bold=True)
                                        ws.cell(cur, 1).value = 'TOTAL'
                                        cel_tot = ws.cell(cur, idx_monto)
                                        cel_tot.value         = f'=SUM({col_m_l}{data_start}:{col_m_l}{data_end})'
                                        cel_tot.number_format = FMT_MONEY
                                        cel_tot.font          = Font(bold=True)
                                        cur += 2  # TOTAL + 1 fila vacía

                                    _autofit_ws(ws, n_dif)

                                if lista_m_dif:
                                    ws_mas_m = writer.book.create_sheet('DE MAS EN MENDEZ(CONTROLAR)')
                                    _escribir_por_cuit(
                                        ws_mas_m, lista_m_dif,
                                        'DE MAS EN MENDEZ(CONTROLAR)',
                                        f'Comprobantes enfrentados para CUITs con diferencias | {periodo}'
                                    )

                                # ══════ Hoja 2.5: DE MAS EN ORGANISMO ═══════════════════════════
                                if lista_a_dif:
                                    hoja_org_name = f'DE MAS EN {label_organismo[:10]}(PROCESAR)'
                                    ws_mas_o = writer.book.create_sheet(hoja_org_name)
                                    _escribir_por_cuit(
                                        ws_mas_o, lista_a_dif,
                                        f'DE MAS EN {label_organismo}(PROCESAR)',
                                        f'Comprobantes enfrentados para CUITs con diferencias | {periodo}'
                                    )


                                # Añadir formato condicional amarillo a diferencias





                                # ══════ Hoja 3: DETALLE ORGANISMO ═══════════════════════════
                                df_arba_det.to_excel(writer, sheet_name=f'Detalle {organismo}', index=False, startrow=4)
                                ws2 = writer.sheets[f'Detalle {organismo}']
                                n2 = len(df_arba_det.columns)
                                _encabezado(ws2, n2,
                                    f'DETALLE {organismo} · {label_tipo.upper()}',
                                    f'{len(registros_activos)} registros | {periodo}'
                                )
                                ws2.row_dimensions[4].height = 4
                                _estilizar_hdr(ws2, 5, n2)
                                # Para IVA: la columna PV del DataFrame ya se llama 'Fecha Ret./Perc.'
                                # (el nombre viene de _pv_col_lbl); no hace falta renombrar la celda.

                                idx_monto_arba = list(df_arba_det.columns).index('Monto') + 1
                                idx_prov_arba  = list(df_arba_det.columns).index('Proveedor') + 1
                                idx_match_arba = list(df_arba_det.columns).index('Estado') + 1
                                idx_cant_arba  = list(df_arba_det.columns).index('Cantidad') + 1
                                last_men_r     = max(len(mendez_detalle) + 6, 7) # aprox para countifs
                                _col_m_a = get_column_letter(idx_monto_arba)
                                _col_p_a = get_column_letter(idx_prov_arba)
                                _last_a  = len(df_arba_det) + 5  # última fila de datos (base 1)

                                for row_i in range(6, _last_a + 1):
                                    is_sub = str(ws2.cell(row=row_i, column=idx_prov_arba).value).endswith('(SUBTOTAL)')
                                    fill = FILL_ZEBRA if (row_i % 2 == 0) else None
                                    if is_sub:
                                        fill = PatternFill('solid', fgColor='D9E1F2')
                                        ws2.row_dimensions[row_i].outlineLevel = 0
                                        # Subtotal formulado: suma comprobantes del CUIT excluyendo la propia fila subtotal
                                        ws2.cell(row=row_i, column=idx_monto_arba).value = (
                                            f'=SUMIFS(${_col_m_a}$6:${_col_m_a}${_last_a},'
                                            f'$A$6:$A${_last_a},A{row_i},'
                                            f'${_col_p_a}$6:${_col_p_a}${_last_a},"<>*(SUBTOTAL)*")'
                                        )
                                    else:
                                        ws2.row_dimensions[row_i].outlineLevel = 1
                                        ws2.row_dimensions[row_i].hidden = True

                                    # Estado: subtotales → VLOOKUP al resumen por CUIT; filas individuales → pre-computado
                                    if is_sub:
                                        ws2.cell(row=row_i, column=idx_match_arba).value = (
                                            f'=IFERROR(VLOOKUP(A{row_i}, \'Cruce x CUIT\'!$A$6:$F${len(cuits_sorted)+5}, 6, 0), "")'
                                        )
                                    ws2.cell(row=row_i, column=idx_cant_arba).value = (
                                        f'=IF(COUNTIF(B{row_i}, "*(SUBTOTAL)*")>0, '
                                        f'COUNTIFS($A$6:$A${_last_a}, A{row_i}, $B$6:$B${_last_a}, "<>*(SUBTOTAL)*"), '
                                        f'"")'
                                    )

                                    for c in range(1, n2 + 1):
                                        cell = ws2.cell(row=row_i, column=c)
                                        cell.alignment = CTR; cell.border = THIN
                                        if fill: cell.fill = fill
                                        if is_sub: cell.font = Font(bold=True)
                                        if c == idx_monto_arba: cell.number_format = FMT_MONEY

                                # ── Fila TOTAL GENERAL organismo ──────────────────────
                                _tot_a = _last_a + 1
                                for ci in range(1, n2 + 1):
                                    cell = ws2.cell(_tot_a, ci)
                                    cell.fill = FILL_TITLE
                                    cell.font = Font(bold=True, size=10, color='FFFFFF')
                                    cell.alignment = CTR; cell.border = THIN
                                ws2.cell(_tot_a, 1).value = 'TOTAL GENERAL'
                                _cel_tot_a = ws2.cell(_tot_a, idx_monto_arba)
                                _cel_tot_a.value = (
                                    f'=SUMIFS({_col_m_a}6:{_col_m_a}{_last_a},'
                                    f'{_col_p_a}6:{_col_p_a}{_last_a},"<>*(SUBTOTAL)*")'
                                )
                                _cel_tot_a.number_format = FMT_MONEY
                                _cel_tot_a.font = Font(bold=True, color='FFFFFF')
                                ws2.row_dimensions[_tot_a].height = 18
                                _autofit_ws(ws2, n2)
                                
                                
                                FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                                FILL_ORANGE = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
                                FILL_OK     = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                                str_col = openpyxl.utils.get_column_letter(idx_match_arba)
                                r_dif = FormulaRule(formula=[f'${str_col}6="⚠ Diferencia"'], stopIfTrue=False, fill=FILL_YELLOW)
                                r_men = FormulaRule(formula=[f'${str_col}6="⚠ Falta en Mendez"'], stopIfTrue=False, fill=FILL_ORANGE)
                                r_ok  = FormulaRule(formula=[f'${str_col}6="✓ Ok"'], stopIfTrue=False, fill=FILL_OK)
                                _cf_last_a = max(len(df_arba_det) + 5, 6)
                                ws2.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n2)}{_cf_last_a}', r_dif)
                                ws2.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n2)}{_cf_last_a}', r_men)
                                ws2.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n2)}{_cf_last_a}', r_ok)




                                # ══════ Hoja 4: DETALLE MENDEZ ══════════════════════════
                                df_mendez_det.to_excel(writer, sheet_name='Detalle Mendez', index=False, startrow=4)
                                ws3 = writer.sheets['Detalle Mendez']
                                n3 = len(df_mendez_det.columns)
                                _encabezado(ws3, n3,
                                    'DETALLE MENDEZ · IIBB BS.AS.',
                                    f'{len(mendez_detalle)} registros con percepción/retención Bs.As. | {periodo}'
                                )
                                ws3.row_dimensions[4].height = 4
                                _estilizar_hdr(ws3, 5, n3, fill=PatternFill('solid', fgColor='2E75B6'))
                                
                                idx_monto_men = list(df_mendez_det.columns).index('Monto') + 1
                                idx_prov_men  = list(df_mendez_det.columns).index('Proveedor') + 1
                                idx_match_men = list(df_mendez_det.columns).index('Estado') + 1
                                idx_cant_men  = list(df_mendez_det.columns).index('Cantidad') + 1
                                _col_m_m = get_column_letter(idx_monto_men)
                                _col_p_m = get_column_letter(idx_prov_men)
                                _last_m  = len(df_mendez_det) + 5  # última fila de datos (base 1)

                                for row_i in range(6, _last_m + 1):
                                    is_sub = str(ws3.cell(row=row_i, column=idx_prov_men).value).endswith('(SUBTOTAL)')
                                    fill = FILL_ZEBRA if (row_i % 2 == 0) else None
                                    if is_sub:
                                        fill = PatternFill('solid', fgColor='D9E1F2')
                                        ws3.row_dimensions[row_i].outlineLevel = 0
                                        # Subtotal formulado: suma comprobantes del CUIT excluyendo la propia fila subtotal
                                        ws3.cell(row=row_i, column=idx_monto_men).value = (
                                            f'=SUMIFS(${_col_m_m}$6:${_col_m_m}${_last_m},'
                                            f'$A$6:$A${_last_m},A{row_i},'
                                            f'${_col_p_m}$6:${_col_p_m}${_last_m},"<>*(SUBTOTAL)*")'
                                        )
                                    else:
                                        ws3.row_dimensions[row_i].outlineLevel = 1
                                        ws3.row_dimensions[row_i].hidden = True

                                    # Estado: subtotales → VLOOKUP al resumen por CUIT; filas individuales → pre-computado
                                    if is_sub:
                                        ws3.cell(row=row_i, column=idx_match_men).value = (
                                            f'=IFERROR(VLOOKUP(A{row_i}, \'Cruce x CUIT\'!$A$6:$F${len(cuits_sorted)+5}, 6, 0), "")'
                                        )
                                    ws3.cell(row=row_i, column=idx_cant_men).value = (
                                        f'=IF(COUNTIF(B{row_i}, "*(SUBTOTAL)*")>0, '
                                        f'COUNTIFS($A$6:$A${_last_m}, A{row_i}, $B$6:$B${_last_m}, "<>*(SUBTOTAL)*"), '
                                        f'"")'
                                    )

                                    for c in range(1, n3 + 1):
                                        cell = ws3.cell(row=row_i, column=c)
                                        cell.alignment = CTR; cell.border = THIN
                                        if fill: cell.fill = fill
                                        if is_sub: cell.font = Font(bold=True)
                                        if c == idx_monto_men: cell.number_format = FMT_MONEY

                                # ── Fila TOTAL GENERAL Mendez ─────────────────────────
                                _tot_m = _last_m + 1
                                for ci in range(1, n3 + 1):
                                    cell = ws3.cell(_tot_m, ci)
                                    cell.fill = PatternFill('solid', fgColor='2E75B6')
                                    cell.font = Font(bold=True, size=10, color='FFFFFF')
                                    cell.alignment = CTR; cell.border = THIN
                                ws3.cell(_tot_m, 1).value = 'TOTAL GENERAL'
                                _cel_tot_m = ws3.cell(_tot_m, idx_monto_men)
                                _cel_tot_m.value = (
                                    f'=SUMIFS({_col_m_m}6:{_col_m_m}{_last_m},'
                                    f'{_col_p_m}6:{_col_p_m}{_last_m},"<>*(SUBTOTAL)*")'
                                )
                                _cel_tot_m.number_format = FMT_MONEY
                                _cel_tot_m.font = Font(bold=True, color='FFFFFF')
                                ws3.row_dimensions[_tot_m].height = 18
                                _autofit_ws(ws3, n3)
                                

                                FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                                FILL_RED    = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                                FILL_OK     = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                                str_col_m = openpyxl.utils.get_column_letter(idx_match_men)
                                r_dif_m = FormulaRule(formula=[f'${str_col_m}6="⚠ Diferencia"'], stopIfTrue=False, fill=FILL_YELLOW)
                                r_org_m = FormulaRule(formula=[f'${str_col_m}6="⚠ Falta en {organismo}"'], stopIfTrue=False, fill=FILL_RED)
                                r_ok_m  = FormulaRule(formula=[f'${str_col_m}6="✓ Ok"'], stopIfTrue=False, fill=FILL_OK)
                                _cf_last_m = max(len(df_mendez_det) + 5, 6)
                                ws3.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n3)}{_cf_last_m}', r_dif_m)
                                ws3.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n3)}{_cf_last_m}', r_org_m)
                                ws3.conditional_formatting.add(f'A6:{openpyxl.utils.get_column_letter(n3)}{_cf_last_m}', r_ok_m)





                            output_cruce.seek(0)

                        # ── Descarga ─────────────────────────────────────────────────
                        nombre_base = Path(uploaded_arba_txt_mendez.name).stem
                        st.download_button(
                            label=f"↓  Descargar Excel de Cruce {organismo}",
                            data=output_cruce,
                            file_name=f"{nombre_base}_Cruce{organismo}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

            except Exception as e:
                st.error(f"Error al procesar: {str(e)}")
                st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)


    elif uploaded_arba_txt_mendez and not uploaded_arba_txt_arba:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            FALTA EL TXT ARBA · SUBILO EN EL PASO 02
        </div>
        """, unsafe_allow_html=True)

    elif not uploaded_arba_txt_mendez and uploaded_arba_txt_arba:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            FALTA EL TXT MENDEZ · SUBILO EN EL PASO 01
        </div>
        """, unsafe_allow_html=True)

    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            SUBÍ AMBOS ARCHIVOS · TXT MENDEZ + TXT ARBA
        </div>
        """, unsafe_allow_html=True)

elif herramienta == TOOL_IMPORTACION:
    def _slug_concepto_imp(desc: str) -> str:
        s = re.sub(r'[\\/:*?"<>|]', '', desc).strip()
        s = re.sub(r'\s+', '_', s)
        return s or 'Concepto'

    def _prefijo_desde_zip(zip_filename: str) -> str:
        stem = Path(zip_filename).stem
        m = re.match(r'^(.*)_(\d{8})_(\d{3,4})$', stem)
        if m:
            return f"{m.group(1)}_{m.group(2)}"
        return stem

    def _localizar_cuit_col_imp(df):
        for c in df.columns:
            cl = c.strip().lower()
            if ('nro' in cl or 'mero' in cl) and 'doc' in cl and ('vendedor' in cl or 'comprador' in cl):
                return c
        for c in df.columns:
            if c.strip().lower() == 'cuit':
                return c
        return None

    def _find_col_imp(df, keywords):
        """Primera columna cuyo header (lower) contiene todas las keywords."""
        for c in df.columns:
            cl = c.strip().lower()
            if all(k in cl for k in keywords):
                return c
        return None

    def _parse_monto_imp(serie):
        """Serie de strings formato argentino ('1.234,56') → float (NaN si vacío)."""
        return pd.to_numeric(
            serie.astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False),
            errors='coerce',
        )

    st.markdown('<div class="card"><div class="card-label">01 · Archivo TXT Mendez</div>', unsafe_allow_html=True)
    uploaded_txt_imp = st.file_uploader(
        "Subí el TXT/PRN de movimientos del sistema Mendez",
        type=["txt", "prn"],
        key="imp_txt"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="card"><div class="card-label">02 · Archivo ARCA (.zip)</div>', unsafe_allow_html=True)
    uploaded_arca_imp = st.file_uploader(
        "Subí el .zip del Portal IVA de ARCA",
        type=["zip"],
        key="imp_arca"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_txt_imp is not None and uploaded_arca_imp is not None:
        st.markdown('<div class="card"><div class="card-label">03 · Procesar</div>', unsafe_allow_html=True)
        if st.button("⬡  Generar ZIPs por Concepto", use_container_width=True):
            try:
                from collections import Counter

                with st.spinner("Parseando TXT Mendez..."):
                    txt_content = uploaded_txt_imp.getvalue().decode("latin-1")
                    transacciones_imp, meta_imp = parsear_archivo(content=txt_content)

                if not transacciones_imp:
                    st.error("No se encontraron transacciones en el TXT Mendez.")
                    st.stop()

                cuit_counters = {}
                for t in transacciones_imp:
                    cuit_norm = re.sub(r'[^0-9]', '', str(t.get('CUIT') or ''))
                    if not cuit_norm:
                        continue
                    concepto = t.get('Concepto')
                    if concepto in (None, '', 0):
                        continue
                    numero = str(t.get('Numero') or '').strip()
                    if not numero or not numero[-1].isalpha():
                        continue
                    cuit_counters.setdefault(cuit_norm, Counter())[concepto] += 1

                concepto_por_cuit = {
                    cuit: counter.most_common(1)[0][0]
                    for cuit, counter in cuit_counters.items()
                }

                with st.spinner("Leyendo ZIP de ARCA..."):
                    with zipfile.ZipFile(io.BytesIO(uploaded_arca_imp.getvalue())) as zf_in:
                        archivos_arca = [f for f in zf_in.namelist() if not f.endswith('/')]
                        if not archivos_arca:
                            st.error("El .zip de ARCA está vacío.")
                            st.stop()
                        target_arca = archivos_arca[0]
                        raw_arca = zf_in.open(target_arca).read()

                csv_text_arca = raw_arca.decode('latin-1')
                sep_arca = ';' if csv_text_arca.count(';') > csv_text_arca.count(',') else ','
                df_arca_raw = pd.read_csv(
                    io.StringIO(csv_text_arca),
                    sep=sep_arca,
                    dtype=str,
                    keep_default_na=False,
                    na_values=[],
                    on_bad_lines='skip',
                )

                cuit_col_arca = _localizar_cuit_col_imp(df_arca_raw)
                if cuit_col_arca is None:
                    st.error("No se encontró la columna de CUIT del proveedor en el CSV de ARCA.")
                    st.stop()

                df_arca_raw['_cuit_norm'] = df_arca_raw[cuit_col_arca].astype(str).apply(
                    lambda v: re.sub(r'[^0-9]', '', v)
                )
                df_arca_raw['_concepto'] = df_arca_raw['_cuit_norm'].map(concepto_por_cuit)

                # ── Verificación: Importe Total NO debe coincidir con un Neto Gravado ──
                # (síntoma de IVA no sumado al cargar el comprobante). Solo avisa.
                df_flagged = pd.DataFrame()
                col_total = _find_col_imp(df_arca_raw, ['importe', 'total'])
                netos_chequear = [
                    ('10,5%', _find_col_imp(df_arca_raw, ['neto', 'gravado', '10,5'])),
                    ('21%', _find_col_imp(df_arca_raw, ['neto', 'gravado', '21'])),
                    ('27%', _find_col_imp(df_arca_raw, ['neto', 'gravado', '27'])),
                ]
                netos_chequear = [(al, c) for al, c in netos_chequear if c is not None]
                if col_total is not None and netos_chequear:
                    total_num = _parse_monto_imp(df_arca_raw[col_total])
                    alic_match = pd.Series([''] * len(df_arca_raw), index=df_arca_raw.index)
                    neto_match = pd.Series([float('nan')] * len(df_arca_raw), index=df_arca_raw.index)
                    flag_mask = pd.Series(False, index=df_arca_raw.index)
                    for alic, col in netos_chequear:
                        neto_num = _parse_monto_imp(df_arca_raw[col])
                        coincide = (total_num > 0) & (neto_num > 0) & ((total_num - neto_num).abs() < 0.01)
                        nuevos = coincide & (~flag_mask)
                        alic_match.loc[nuevos] = alic
                        neto_match.loc[nuevos] = neto_num.loc[nuevos]
                        flag_mask = flag_mask | coincide
                    if flag_mask.any():
                        cols_id = []
                        for kws in (['tipo', 'comprobante'], ['punto', 'venta'], ['mero', 'comprobante'], ['denominaci']):
                            c = _find_col_imp(df_arca_raw, kws)
                            if c is not None and c not in cols_id:
                                cols_id.append(c)
                        if cuit_col_arca not in cols_id:
                            cols_id.append(cuit_col_arca)
                        df_flagged = df_arca_raw.loc[flag_mask, cols_id + [col_total]].copy()
                        df_flagged['Alícuota Neto'] = alic_match.loc[flag_mask].values
                        df_flagged['Neto Gravado'] = neto_match.loc[flag_mask].values

                prefijo_zip = _prefijo_desde_zip(uploaded_arca_imp.name)

                container_buf = io.BytesIO()
                stats_rows = []
                with zipfile.ZipFile(container_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for concepto_val, grupo in df_arca_raw.groupby('_concepto', dropna=False):
                        grupo_export = grupo.drop(columns=['_concepto', '_cuit_norm'])
                        csv_io = io.StringIO()
                        grupo_export.to_csv(csv_io, sep=sep_arca, index=False, lineterminator='\n')
                        csv_bytes_out = csv_io.getvalue().encode('latin-1', errors='replace')

                        if pd.isna(concepto_val) or concepto_val is None:
                            nombre_base = f"{prefijo_zip}_0000"
                            label = "0000 · SIN CONCEPTO (revisar manualmente)"
                        else:
                            cod_int = int(concepto_val)
                            cod_padded = f"{cod_int:04d}"
                            descripcion = CONCEPTOS_MAP.get(str(cod_int), f'Concepto {cod_int}')
                            nombre_base = f"{prefijo_zip}_{cod_padded}"
                            label = f"{cod_padded} · {descripcion}"

                        zip_name = f"{nombre_base}.zip"
                        csv_inner_name = f"{nombre_base} (montos expresados en pesos).csv"

                        inner_buf = io.BytesIO()
                        with zipfile.ZipFile(inner_buf, 'w', zipfile.ZIP_DEFLATED) as zin:
                            zin.writestr(csv_inner_name, csv_bytes_out)
                        zout.writestr(zip_name, inner_buf.getvalue())

                        stats_rows.append({
                            'Concepto': label,
                            'Comprobantes': len(grupo_export),
                            'ZIP': zip_name,
                        })

                container_name = uploaded_arca_imp.name

                total_arca = len(df_arca_raw)
                total_cruzados = int(df_arca_raw['_concepto'].notna().sum())
                total_sin = total_arca - total_cruzados
                total_zips = len(stats_rows)

                st.success(
                    f"{total_zips} ZIPs generados · "
                    f"{total_cruzados}/{total_arca} comprobantes cruzados · "
                    f"{total_sin} sin concepto"
                )

                if not df_flagged.empty:
                    st.warning(
                        f"⚠ {len(df_flagged)} comprobante(s) con Importe Total = Neto Gravado "
                        f"(IVA no sumado · revisar carga del contribuyente)"
                    )
                    with st.expander("Comprobantes con posible error de carga", expanded=True):
                        st.dataframe(df_flagged, use_container_width=True, hide_index=True)

                stats_df = pd.DataFrame(stats_rows).sort_values(by='Comprobantes', ascending=False)
                with st.expander("Detalle de ZIPs generados", expanded=True):
                    st.dataframe(stats_df, use_container_width=True, hide_index=True)

                st.download_button(
                    label=f"↓  Descargar {container_name}",
                    data=container_buf.getvalue(),
                    file_name=container_name,
                    mime="application/zip",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Error al procesar: {e}")
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            SUBÍ AMBOS ARCHIVOS · TXT MENDEZ + ZIP ARCA
        </div>
        """, unsafe_allow_html=True)

elif herramienta == TOOL_VENTAS_CITI:
    def _periodo_desde_zip_citi(zip_filename: str) -> str:
        stem = Path(zip_filename).stem
        m = re.search(r'(\d{6})', stem)
        if m:
            p = m.group(1)
            return f"{p[:4]}-{p[4:6]}"
        return 'sin_periodo'

    st.markdown('<div class="card"><div class="card-label">01 · Archivo ARCA (.zip Portal IVA - Ventas)</div>', unsafe_allow_html=True)
    uploaded_arca_citi = st.file_uploader(
        "Subí el .zip del Portal IVA de ARCA con los comprobantes de ventas",
        type=["zip"],
        key="citi_arca",
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_arca_citi is not None:
        st.markdown('<div class="card"><div class="card-label">02 · Procesar</div>', unsafe_allow_html=True)
        if st.button("⬡  Generar Excel + ZIP CITI Ventas", use_container_width=True):
            try:
                with st.spinner("Leyendo ZIP del Portal IVA..."):
                    with zipfile.ZipFile(io.BytesIO(uploaded_arca_citi.getvalue())) as zf_in:
                        archivos = [f for f in zf_in.namelist() if not f.endswith('/') and f.lower().endswith('.csv')]
                        if not archivos:
                            st.error("El .zip no contiene ningún CSV.")
                            st.stop()
                        csv_name = archivos[0]
                        raw = zf_in.open(csv_name).read()

                csv_text = raw.decode('latin-1')
                sep = ';' if csv_text.count(';') > csv_text.count(',') else ','
                df_arca = pd.read_csv(
                    io.StringIO(csv_text),
                    sep=sep,
                    dtype=str,
                    keep_default_na=False,
                    na_values=[],
                    on_bad_lines='skip',
                )

                with st.spinner("Normalizando montos y fechas..."):
                    df_norm = normalizar_csv_ventas_arca(df_arca)

                with st.spinner("Consolidando tipo Ticket Z por (Fecha, PV, Tipo, Comprador)..."):
                    df_cons = consolidar_ventas_citi(df_norm)

                if df_cons.empty:
                    st.error("No se encontraron filas para consolidar.")
                    st.stop()

                ventas_lineas = generar_citi_ventas_lineas(df_cons)
                alic_lineas = generar_citi_alicuotas_lineas(df_cons)

                periodo = _periodo_desde_zip_citi(uploaded_arca_citi.name)

                # Excel
                excel_buf = io.BytesIO()
                crear_excel_ventas_citi(df_cons, periodo, excel_buf, df_original=df_norm)
                excel_bytes = excel_buf.getvalue()

                # ZIP con los dos TXT
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                    zout.writestr('VENTAS.txt', '\r\n'.join(ventas_lineas).encode('latin-1'))
                    zout.writestr('ALICUOTAS.txt', '\r\n'.join(alic_lineas).encode('latin-1'))
                zip_bytes = zip_buf.getvalue()

                st.session_state['citi_excel_bytes'] = excel_bytes
                st.session_state['citi_zip_bytes'] = zip_bytes
                st.session_state['citi_periodo'] = periodo

                # Métricas de verificación
                sum_total_csv = float(df_norm['importe_total'].sum())
                sum_total_cons = float(df_cons['importe_total'].sum())
                sum_iva_csv = float(df_norm['total_iva'].sum())
                sum_iva_cons = float(df_cons['total_iva'].sum())

                ok_total = abs(sum_total_csv - sum_total_cons) < 0.01

                st.success(
                    f"{len(df_cons)} consolidados · {int(df_cons['cant_cbtes'].sum())} comprobantes originales · "
                    f"{len(ventas_lineas)} líneas VENTAS.txt · {len(alic_lineas)} líneas ALICUOTAS.txt"
                )

                with st.expander("Verificación de totales", expanded=True):
                    col1, col2 = st.columns(2)
                    col1.metric("Σ Importe Total CSV", f"{sum_total_csv:,.2f}")
                    col2.metric(
                        "Σ Importe Total Consolidado",
                        f"{sum_total_cons:,.2f}",
                        delta=f"{sum_total_cons - sum_total_csv:+.2f}",
                    )
                    if ok_total:
                        st.success("✓ Totales coinciden al centavo")
                    else:
                        st.warning("⚠ Diferencia detectada — revisar datos del CSV")

                    st.caption(f"Σ Total IVA · CSV: {sum_iva_csv:,.2f}  |  Consolidado: {sum_iva_cons:,.2f}")
                    st.caption(f"Largo VENTAS.txt: {len(ventas_lineas[0])} chars (debe ser 266)")
                    st.caption(f"Largo ALICUOTAS.txt: {len(alic_lineas[0])} chars (debe ser 62)")

            except ValueError as ve:
                st.error(f"Error de validación: {ve}")
            except Exception as e:
                st.error(f"Error al procesar: {e}")
                import traceback
                st.code(traceback.format_exc())
        st.markdown('</div>', unsafe_allow_html=True)

        if 'citi_excel_bytes' in st.session_state and 'citi_zip_bytes' in st.session_state:
            st.markdown('<div class="card"><div class="card-label">03 · Descargar</div>', unsafe_allow_html=True)
            periodo = st.session_state.get('citi_periodo', 'sin_periodo')
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label=f"↓  Excel resumen ({periodo})",
                    data=st.session_state['citi_excel_bytes'],
                    file_name=f"citi_ventas_{periodo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col2:
                st.download_button(
                    label=f"↓  .zip CITI ({periodo})",
                    data=st.session_state['citi_zip_bytes'],
                    file_name=f"citi_ventas_{periodo}.zip",
                    mime="application/zip",
                    use_container_width=True,
                )
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            SUBÍ EL .ZIP DEL PORTAL IVA · VENTAS
        </div>
        """, unsafe_allow_html=True)

elif herramienta == TOOL_RETENCIONES:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Importacion Retenciones IVA / Ganancias (XLS ARCA -> .zip Portal IVA)
    # ───────────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Tipo de retención</div>', unsafe_allow_html=True)
    tipo_ret = st.radio(
        "Elegí el impuesto a procesar:",
        options=["IVA", "Ganancias"],
        horizontal=True,
        key="ret_tipo",
    )
    tipo_cbte_pos = st.radio(
        "Tipo de Comprobante (importes positivos):",
        options=["99", "80"],
        horizontal=True,
        key="ret_tipo_cbte",
        help="Los importes negativos se emiten siempre como Nota de Crédito (tipo 3).",
    )
    # Si cambia el tipo de retención o el tipo de comprobante, limpiar resultados
    # previos para no mostrar descargas viejas.
    if (st.session_state.get('ret_last_tipo') != tipo_ret
            or st.session_state.get('ret_last_tipo_cbte') != tipo_cbte_pos):
        for k in ('ret_zip_bytes', 'ret_zip_name', 'ret_count', 'ret_periodo', 'ret_tipo_generado'):
            st.session_state.pop(k, None)
        st.session_state['ret_last_tipo'] = tipo_ret
        st.session_state['ret_last_tipo_cbte'] = tipo_cbte_pos
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown(f'<div class="card"><div class="card-label">02 · Archivo {tipo_ret} (.xls / .xlsx)</div>', unsafe_allow_html=True)
    st.caption(f"Subí el XLS de Mis Retenciones de {tipo_ret} descargado de ARCA.")
    uploaded_ret = st.file_uploader(
        f"{tipo_ret}.xls",
        type=["xls", "xlsx"],
        label_visibility="collapsed",
        key=f"ret_xls_{tipo_ret.lower()}",
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_ret:
        st.markdown('<div class="card"><div class="card-label">03 · Generar .zip Portal IVA</div>', unsafe_allow_html=True)

        if st.button("⬡  Generar .zip"):
            try:
                with st.spinner(f"Procesando retenciones de {tipo_ret}..."):
                    df = parsear_arca_retenciones_xls(uploaded_ret.getvalue())
                    csv_text, periodo = transformar_retenciones_a_csv_arca(df, tipo_cbte_pos)
                    zip_bytes, zip_name = generar_zip_retenciones_arca(csv_text, periodo)
                    st.session_state['ret_zip_bytes'] = zip_bytes
                    st.session_state['ret_zip_name'] = zip_name
                    st.session_state['ret_count'] = len(df)
                    st.session_state['ret_periodo'] = periodo
                    st.session_state['ret_tipo_generado'] = tipo_ret

                st.success("✓ .zip generado correctamente")
                st.markdown(
                    f'<div class="stats-row">'
                    f'<div class="stat-chip"><div class="stat-val">{len(df)}</div><div class="stat-lbl">RET. {tipo_ret.upper()}</div></div>'
                    f'<div class="stat-chip"><div class="stat-val">{periodo}</div><div class="stat-lbl">PERIODO</div></div>'
                    f'<div class="stat-chip"><div class="stat-val">{tipo_cbte_pos}</div><div class="stat-lbl">TIPO CBTE</div></div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            except ValueError as ve:
                st.error(f"Error de validación: {ve}")
            except Exception as e:
                st.error(f"Error al procesar: {e}")
                import traceback
                st.code(traceback.format_exc())

        st.markdown('</div>', unsafe_allow_html=True)

        if 'ret_zip_bytes' in st.session_state:
            st.markdown('<div class="card"><div class="card-label">04 · Descargar</div>', unsafe_allow_html=True)
            tipo_dl = st.session_state.get('ret_tipo_generado', tipo_ret)
            st.download_button(
                label=f"↓  .zip Retenciones {tipo_dl} ({st.session_state['ret_periodo']})",
                data=st.session_state['ret_zip_bytes'],
                file_name=st.session_state['ret_zip_name'],
                mime="application/zip",
                use_container_width=True,
                key="dl_ret",
            )
            st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            SUBÍ EL XLS · {tipo_ret.upper()}
        </div>
        """, unsafe_allow_html=True)

